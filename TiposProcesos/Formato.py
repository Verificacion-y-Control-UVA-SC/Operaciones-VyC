# Formato.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def exportar_excel(df, ruta_salida):
    # Guardar DataFrame en Excel
    df.to_excel(ruta_salida, index=False, sheet_name="Tipo de Proceso")

    # Abrir con openpyxl
    wb = load_workbook(ruta_salida)
    ws = wb.active

    # Estilos para encabezados
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Aplicar formato a la fila de encabezados
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Ajustar ancho de columnas según contenido
    for col_num, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        col_letter = get_column_letter(col_num)
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Guardar archivo con formato aplicado
    wb.save(ruta_salida)
