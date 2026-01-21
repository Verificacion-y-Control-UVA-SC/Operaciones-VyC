import os, sys, json, re
import customtkinter as ctk
import pandas as pd
from tkinter import filedialog, messagebox
from datetime import datetime
from core.manejador_archivos import convertir_a_json, leer_json, guardar_config
from openpyxl.styles import PatternFill, numbers


def resource_path(relative_path):
    """Obtiene la ruta absoluta incluso si está empaquetado en .exe"""
    try:
        # PyInstaller crea una carpeta temporal y guarda la ruta en _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



# CONFIG
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("dark-blue")

COLORES = {
    "amarillo": "#ecd925",
    "negro": "#282828",
    "gris_oscuro": "#4d4d4d",
    "gris_claro": "#d8d8d8",
    "blanco": "#FFFFFF"
}

FUENTE_PRINCIPAL = "Inter"
FUENTE_SECUNDARIA = "Inter"

class VentanaBaseEtiquetado(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Base de Etiquetado ULTA - AXO")
        self.geometry("900x600")
        self.configure(fg_color=COLORES["blanco"])

        # --- Estado de archivos ---
        self.base_general = None
        self.layout = None

        # Persistencia visual de archivos cargados
        self.base_general_persistido = False
        self.layout_persistido = False

        # Cargar configuración de base_etiquetado.json
        theme_path = resource_path(os.path.join("theme", "base_etiquetado.json"))
        if os.path.exists(theme_path):
            with open(theme_path, "r", encoding="utf-8") as f:
                self.theme_config = json.load(f)

        # --- Persistencia: cargar BASE_GENERAL y LAYOUT si existen ---
        base_general_path = resource_path(os.path.join("data", "BASE_GENERAL_ULTA.json"))
        layout_path = resource_path(os.path.join("data", "LAYOUT.json"))
        if os.path.exists(base_general_path):
            with open(base_general_path, "r", encoding="utf-8") as f:
                self.base_general = json.load(f)
            self.base_general_persistido = True
        if os.path.exists(layout_path):
            with open(layout_path, "r", encoding="utf-8") as f:
                self.layout = json.load(f)
            self.layout_persistido = True


        # Crear UI (después de cargar persistencia)
        self.crear_interfaz()
        os.makedirs("data", exist_ok=True)
        self.protocol("WM_DELETE_WINDOW", self.on_cerrar)


# --- Diseño de la aplicacion Base de etiquetado ULTA --- #   
    def crear_interfaz(self):
        main = ctk.CTkFrame(self, fg_color=COLORES["blanco"])
        main.pack(fill="both", expand=True, padx=10, pady=10)

        # --- Encabezado ---
        header = ctk.CTkFrame(main, fg_color=COLORES["amarillo"], corner_radius=12)
        header.pack(fill="x", pady=(0, 10))
        header_content = ctk.CTkFrame(header, fg_color="transparent")
        header_content.pack(expand=True, fill="both", padx=10, pady=8)
        ctk.CTkLabel(
            header_content,
            text="🏷️ Base de Etiquetado ULTA - AXO",
            font=(FUENTE_PRINCIPAL, 22, "bold"),
            text_color=COLORES["negro"]
        ).pack(pady=(0, 2))
        ctk.CTkLabel(
            header_content,
            text="Sistema integrado para la gestión y generación de bases de etiquetado",
            font=(FUENTE_SECUNDARIA, 12),
            text_color=COLORES["gris_oscuro"]
        ).pack()

        # --- Panel de Progreso ---
        self.progress_frame = ctk.CTkFrame(main, fg_color=COLORES["gris_claro"], corner_radius=8)
        self.progress_frame.pack(fill="x", pady=(0, 8), padx=2)
        progress_content = ctk.CTkFrame(self.progress_frame, fg_color="transparent")
        progress_content.pack(fill="x", padx=8, pady=5)
        self.progress_bar = ctk.CTkProgressBar(progress_content, height=6, corner_radius=3,
                                       progress_color=COLORES["amarillo"],
                                       fg_color=COLORES["gris_oscuro"])
        self.progress_bar.pack(fill="x", pady=(0, 4))
        self.progress_bar.set(0)
        counters_frame = ctk.CTkFrame(progress_content, fg_color="transparent")
        counters_frame.pack(fill="x")
        self.lbl_loaded = ctk.CTkLabel(counters_frame, text="🟡 0/2 archivos cargados",
                        font=(FUENTE_SECUNDARIA, 11, "bold"),
                        text_color=COLORES["negro"])
        self.lbl_loaded.pack(side="left")
        self.lbl_status = ctk.CTkLabel(counters_frame, text="⏳ Listo para cargar archivos",
                        font=(FUENTE_SECUNDARIA, 11),
                        text_color=COLORES["gris_oscuro"])
        self.lbl_status.pack(side="right")

        # --- Sección Archivos ---
        archivos_frame = ctk.CTkFrame(main, fg_color=COLORES["blanco"], corner_radius=10, border_width=1, border_color=COLORES["gris_claro"])
        archivos_frame.pack(fill="x", pady=6, padx=2)
        seccion_header = ctk.CTkFrame(archivos_frame, fg_color="transparent")
        seccion_header.pack(fill="x", padx=10, pady=(8, 4))
        ctk.CTkLabel(
            seccion_header,
            text="📂 ARCHIVOS REQUERIDOS",
            font=(FUENTE_PRINCIPAL, 15, "bold"),
            text_color=COLORES["negro"]
        ).pack(side="left")
        ctk.CTkLabel(
            seccion_header,
            text="Carga los archivos en el orden sugerido",
            font=(FUENTE_SECUNDARIA, 10),
            text_color=COLORES["gris_oscuro"]
        ).pack(side="right")

        # --- Función helper para crear bloques de carga ---
        def crear_bloque_archivo(parent, texto, cargar_func, es_requerido=True, persistido=False, archivo_nombre=None):
            frame = ctk.CTkFrame(parent, fg_color=COLORES["gris_claro"], corner_radius=8)
            frame.pack(fill="x", padx=8, pady=4)
            content_frame = ctk.CTkFrame(frame, fg_color="transparent")
            content_frame.pack(fill="x", padx=6, pady=6)
            left_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
            left_frame.pack(side="left", fill="x", expand=True)
            text_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
            text_frame.pack(fill="x")

            ctk.CTkLabel(
                text_frame,
                text=texto,
                font=(FUENTE_PRINCIPAL, 12, "bold"),
                text_color=COLORES["negro"],
                anchor="w"
            ).pack(side="left", fill="x", expand=True)

            status_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
            status_frame.pack(fill="x", pady=(2, 0))

            status_label = ctk.CTkLabel(
                status_frame,
                text="⌛ Esperando carga...",
                text_color=COLORES["gris_oscuro"],
                font=(FUENTE_SECUNDARIA, 10),
                anchor="w"
            )
            status_label.pack(side="left")

            # 🔹 Etiqueta de “Requerido” (se eliminará dinámicamente)
            req_label = None
            if es_requerido and not persistido:
                req_label = ctk.CTkLabel(
                    status_frame,
                    text="• Requerido",
                    text_color="#e74c3c",
                    font=(FUENTE_SECUNDARIA, 8, "bold")
                )
                req_label.pack(side="right", padx=(8, 0))

            # --- Botones ---
            btn_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
            btn_frame.pack(side="right", padx=(6, 0))

            btn_reemplazar = ctk.CTkButton(
                btn_frame,
                text="🔄 Cambiar",
                fg_color=COLORES["negro"],
                text_color=COLORES["blanco"],
                width=80,
                height=28,
                corner_radius=8,
                font=(FUENTE_PRINCIPAL, 10),
                hover_color=COLORES["gris_oscuro"]
            )
            btn_reemplazar.pack(side="right", padx=(6, 0))
            btn_reemplazar.configure(state="disabled")

            btn_seleccionar = ctk.CTkButton(
                btn_frame,
                text="📂 Seleccionar Archivo",
                fg_color=COLORES["negro"],
                text_color=COLORES["blanco"],
                width=110,
                height=30,
                corner_radius=8,
                font=(FUENTE_PRINCIPAL, 11, "bold"),
                hover_color=COLORES["gris_oscuro"]
            )
            btn_seleccionar.pack(side="right")

            # --- Estado si ya hay archivo persistido ---
            if persistido and archivo_nombre:
                status_label.configure(text=f"✅ {archivo_nombre}", text_color=COLORES["negro"])
                btn_reemplazar.configure(state="normal")
                if req_label:
                    req_label.destroy()  # 🔹 Quita el texto de “Requerido”

            # --- Manejo de carga ---
            def manejar_carga():
                status_label.configure(text="✅ Archivo cargado", text_color="#27ae60")
                btn_reemplazar.configure(state="normal")
                if req_label and req_label.winfo_exists():
                    req_label.destroy()  # 🔹 Quita el texto rojo cuando se carga correctamente
                self.update()
                cargar_func(status_label, btn_seleccionar, btn_reemplazar, frame)

            btn_seleccionar.configure(command=manejar_carga)
            btn_reemplazar.configure(command=manejar_carga)

            return status_label, btn_seleccionar, btn_reemplazar


        # --- Bloques de carga --- #
        base_persistido = self.base_general_persistido
        layout_persistido = self.layout_persistido
        base_nombre = "BASE_GENERAL_ULTA.json" if base_persistido else None
        layout_nombre = "LAYOUT.json" if layout_persistido else None
        self.lbl_base_status, self.btn_base_sel, self.btn_base_rep = crear_bloque_archivo(
            archivos_frame, "📚 Base General ULTA", self.cargar_base_general, True, base_persistido, base_nombre)
        self.lbl_layout_status, self.btn_layout_sel, self.btn_layout_rep = crear_bloque_archivo(
             archivos_frame, "📐 Layout", self.cargar_layout, True, layout_persistido, layout_nombre)


        # --- Panel de Acciones ---
        acciones = ctk.CTkFrame(main, fg_color=COLORES["blanco"])
        acciones.pack(fill="x", pady=(10, 6))
        info_frame = ctk.CTkFrame(acciones, fg_color=COLORES["gris_claro"], corner_radius=8)
        info_frame.pack(fill="x", pady=(0, 8))
        self.info_label = ctk.CTkLabel(
            info_frame,
            text="💡 Cargue los tres archivos requeridos para habilitar la generación de la base",
            font=(FUENTE_SECUNDARIA, 11),
            text_color=COLORES["gris_oscuro"]
        )
        self.info_label.pack(padx=10, pady=8)
        self.btn_generar = ctk.CTkButton(
            acciones,
            text="🚀 GENERAR BASE DE ETIQUETADO ULTA",
            command=self.generar_base,
            fg_color=COLORES["negro"],
            width=220,
            height=36,
            corner_radius=10,
            text_color=COLORES["blanco"],
            font=(FUENTE_PRINCIPAL, 13, "bold"),
            hover_color=COLORES["gris_oscuro"],
            state="disabled"
        )
        self.btn_generar.pack(pady=6)

        # --- Footer ---
        footer = ctk.CTkFrame(main, fg_color=COLORES["blanco"])
        footer.pack(fill="x", pady=(10, 0))
        footer_content = ctk.CTkFrame(footer, fg_color="transparent")
        footer_content.pack(fill="x", pady=4)
        ctk.CTkLabel(
            footer_content,
            text="Base de Etiquetado ULTA - AXO © 2025 • v1.0.0",
            font=(FUENTE_SECUNDARIA, 9),
            text_color=COLORES["gris_oscuro"]
        ).pack(side="left")

        # Inicializar estado
        self.actualizar_contadores()

# --- Persistencia de datos para los archivos BASE GENERAL ULTA y REL ULTA EMBARQUES --- #
    def actualizar_contadores(self):
        """Actualiza contadores, barra de progreso y panel de acciones"""
        archivos_cargados = 0
        archivos_faltantes = []

        # Base General
        if self.base_general is not None:
            archivos_cargados += 1
            self.lbl_base_status.configure(text=f"✅ BASE_GENERAL_ULTA.json", text_color=COLORES["negro"])
            self.btn_base_sel.configure(state="disabled")
            self.btn_base_rep.configure(state="normal")
        else:
            archivos_faltantes.append("Base General")
            self.lbl_base_status.configure(text="⌛ Esperando carga...", text_color=COLORES["gris_oscuro"])
            self.btn_base_sel.configure(state="normal")
            self.btn_base_rep.configure(state="disabled")

        # Layout
        if self.layout is not None:
            archivos_cargados += 1
            self.lbl_layout_status.configure(text=f"✅ LAYOUT.json", text_color=COLORES["negro"])
            self.btn_layout_sel.configure(state="disabled")
            self.btn_layout_rep.configure(state="normal")
        else:
            archivos_faltantes.append("Layout")
            self.lbl_layout_status.configure(text="⌛ Esperando carga...", text_color=COLORES["gris_oscuro"])
            self.btn_layout_sel.configure(state="normal")
            self.btn_layout_rep.configure(state="disabled")

        # Actualizar barra de progreso (ahora sobre 2 archivos)
        self.progress_bar.set(archivos_cargados / 2)
        self.lbl_loaded.configure(text=f"🟢 {archivos_cargados}/2 archivos cargados")

        # --- Actualizar panel de acciones ---
        if archivos_faltantes:
            self.info_label.configure(
                text=f"⚠️ Faltan archivos: {', '.join(archivos_faltantes)}",
                text_color="#e74c3c"
            )
            self.btn_generar.configure(state="disabled")
        else:
            self.info_label.configure(
                text="✅ Todos los archivos cargados. Puede generar la base",
                text_color=COLORES["negro"]
            )
            self.btn_generar.configure(state="normal")

    def verificar_estado(self):
        # Habilitar generar solo si layout está cargado (requisito mínimo)
        if self.layout is not None:
            self.btn_generar.configure(state="normal")
            self.info_label.configure(text="✅ Layout cargado. Listo para generar base de etiquetado")
        else:
            self.btn_generar.configure(state="disabled")
            self.info_label.configure(text="⏳ Esperando Layout:")



# --- Se borra el Layout al cerrar el programa --- #
    def on_cerrar(self):
        """Se ejecuta al cerrar la aplicación"""
        layout_path = resource_path(os.path.join("data", "LAYOUT.json"))
        if os.path.exists(layout_path):
            try:
                os.remove(layout_path)
                print("Archivo LAYOUT.json eliminado al cerrar la app.")
            except Exception as e:
                print(f"No se pudo eliminar Layout.json: {e}")
        self.destroy()


# --- Archivos para generar el armado de las bases de etiquetado ULTA - AXO --- #
    def cargar_base_general(self, status_label=None, btn_sel=None, btn_rep=None, *_):
        archivo = filedialog.askopenfilename(title="Seleccionar BASE_GENERAL_ULTA.xlsx", filetypes=[("Excel files","*.xls *.xlsx"),("All","*.*")])
        if archivo:
            data = convertir_a_json(archivo, sheet_name=0, nombre_json="BASE_GENERAL_ULTA.json", persist=True)
            if data is not None:
                self.base_general = data
                if status_label:
                    status_label.configure(text=f"✅ {os.path.basename(archivo)}", text_color=COLORES["negro"])
                self.info_label.configure(text="✅ Base General cargada y guardada correctamente")
                guardar_config({"base_general_excel": archivo})
                if btn_sel:
                    btn_sel.configure(state="normal", fg_color=COLORES["negro"], text_color=COLORES["blanco"])
                if btn_rep:
                    btn_rep.configure(state="normal", fg_color=COLORES["negro"], text_color=COLORES["blanco"])
            else:
                messagebox.showerror("Error", "No se pudo convertir Base General a JSON")
        self.verificar_estado()
        self.actualizar_contadores()
    
    def cargar_layout(self, status_label=None, btn_sel=None, btn_rep=None, *_):
        archivo = filedialog.askopenfilename(title="Seleccionar LAYOUT.xlsx", filetypes=[("Excel files","*.xls *.xlsx"),("All","*.*")])
        if archivo:
            data = convertir_a_json(archivo, sheet_name="Layout 1", nombre_json="LAYOUT.json")
            if data is not None:
                self.layout = data
                if status_label:
                    status_label.configure(text=f"✅ {os.path.basename(archivo)}", text_color=COLORES["negro"])
                self.info_label.configure(text="✅ Layout cargado correctamente")
                guardar_config({"layout_excel": archivo})
                if btn_sel:
                    btn_sel.configure(state="normal", fg_color=COLORES["negro"], text_color=COLORES["blanco"])
                if btn_rep:
                    btn_rep.configure(state="normal", fg_color=COLORES["negro"], text_color=COLORES["blanco"])
            else:
                messagebox.showerror("Error", "No se pudo convertir Layout a JSON")
        self.verificar_estado()
        self.actualizar_contadores()



# --- Armado para las BASES DE ETIQUETADO ULTA - AXO  --- # 
    def generar_base(self):
        try:
            messagebox.showinfo("Armado de Base", "Iniciando armado de la Base de Etiquetado ULTA...")

            # --- Validar archivos cargados ---
            base_general = pd.DataFrame(self.base_general)
            layout = pd.DataFrame(self.layout)

            # Normalizar encabezados
            base_general.columns = base_general.columns.str.strip().str.upper()
            layout.columns = layout.columns.str.strip().str.upper()


            # --- Columnas finales ---
            columnas = [
                "CATEGORIA", "UPC", "DENOMINACION", "DENOMINACION AXO", "MARCA",
                "LEYENDAS PRECAUTORIAS", "INSTRUCCIONES DE USO", "OBSERVACIONES",
                "TAMAÑO DE LA DECLARACION DE CONTENIDO", "CONTENIDO", "PAIS ORIGEN",
                "IMPORTADOR", "NORMA", "INGREDIENTES", "MEDIDAS", "TIPO DE ETIQUETA"
            ]
            base_etiquetado = pd.DataFrame(columns=columnas)

            # --- Validar columnas requeridas ---
            columnas_requeridas_layout = ["PARTE", "DENOMINACION SOCIAL O NOMBRE"]
            for col in columnas_requeridas_layout:
                if col not in layout.columns:
                    raise ValueError(f"No se encontró la columna '{col}' en el LAYOUT")

            # --- Llenado de información ---
            for _, fila in layout.iterrows():
                codigo = str(fila["PARTE"]).strip()
                fila_final = {col: "" for col in columnas}

                bg = base_general[base_general["UPC"].astype(str) == codigo]

                if not bg.empty:
                    fila_final["CATEGORIA"] = bg.iloc[0].get("CATEGORIA", "")

                    upc_val = bg.iloc[0].get("UPC", "")
                    if pd.notna(upc_val) and str(upc_val).strip() != "":
                        try:
                            fila_final["UPC"] = int(float(upc_val))  # 🔹 entero real
                        except:
                            fila_final["UPC"] = str(upc_val).strip()  # si no se puede convertir
                    else:
                        fila_final["UPC"] = "N/A"

                    fila_final["DENOMINACION AXO"] = bg.iloc[0].get("DENOMINACION AXO", "")
                    fila_final["LEYENDAS PRECAUTORIAS"] = bg.iloc[0].get("LEYENDAS PRECAUTORIAS", "")
                    fila_final["INSTRUCCIONES DE USO"] = bg.iloc[0].get("INSTRUCCIONES DE USO", "")
                    fila_final["OBSERVACIONES"] = bg.iloc[0].get("OBSERVACIONES", "")
                    fila_final["TAMAÑO DE LA DECLARACION DE CONTENIDO"] = bg.iloc[0].get("TAMAÑO DE LA DECLARACION DE CONTENIDO", "")
                    fila_final["IMPORTADOR"] = bg.iloc[0].get("IMPORTADOR", "")
                    fila_final["NORMA"] = bg.iloc[0].get("NORMA", "")
                    fila_final["INGREDIENTES"] = bg.iloc[0].get("INGREDIENTES Y LOTE", "")
                    fila_final["MEDIDAS"] = bg.iloc[0].get("MEDIDAS", "")
                    fila_final["TIPO DE ETIQUETA"] = bg.iloc[0].get("TIPO DE ETIQUETA", "")
                    fila_final["DENOMINACION"] = bg.iloc[0].get("DESCRIPCION", "")
                    fila_final["CONTENIDO"] = bg.iloc[0].get("CONTENIDO", "")
                    fila_final["PAIS ORIGEN"] = bg.iloc[0].get("PAIS DE ORIGEN", "")

                fila_final["MARCA"] = fila.get("DENOMINACION SOCIAL O NOMBRE", "")

                for k, v in fila_final.items():
                    if v == "" or pd.isna(v):
                        fila_final[k] = "N/A"

                if fila_final["CATEGORIA"] != "N/A" or fila_final["UPC"] != "N/A":
                    base_etiquetado = pd.concat([base_etiquetado, pd.DataFrame([fila_final])], ignore_index=True)

            # --- Seleccionar ruta de guardado ---
            salida = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Guardar Base de Etiquetado ULTA.xlsx"
            )
            if not salida:
                return

            # --- Crear el archivo con ambas hojas ---
            with pd.ExcelWriter(salida, engine="openpyxl") as writer:
                base_etiquetado.to_excel(writer, index=False, sheet_name="Base Etiquetado Completa")
                pd.DataFrame(columns=base_etiquetado.columns).to_excel(writer, index=False, sheet_name="Muestras")

            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
            from openpyxl.styles import numbers

            wb = load_workbook(salida)
            ws = wb["Base Etiquetado Completa"]
            ws_muestras = wb["Muestras"]

            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin'))
            amarillo_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            verde_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")

            # --- Ajuste de columnas y alto ---
            row_height = 80
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = row_height

            column_widths = self.theme_config.get("column_widths", {}) if hasattr(self, "theme_config") else {}
            for idx, col_name in enumerate(base_etiquetado.columns, start=1):
                width = column_widths.get(col_name, 15)
                ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
                ws_muestras.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

            filas_amarillas = []

            for row_idx in range(2, ws.max_row + 1):
                valor = ws.cell(row=row_idx, column=15).value  # MEDIDAS
                valor_str = str(valor).strip().upper() if valor is not None else ""

                # Determinar color
                if "REQUIERE ETIQUETADO ESPECIAL" in valor_str or "NO IMPRIMIR HASTA TENER VISTO BUENO DE V&C" in valor_str:
                    fill = amarillo_fill  # 🟡 Especial
                    fila_valores = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
                    # 🔹 Convertir UPC a entero antes de guardar en filas_amarillas
                    if fila_valores[1] != "N/A":
                        try:
                            fila_valores[1] = str(int(float(fila_valores[1])))
                        except:
                            pass
                    filas_amarillas.append(fila_valores)
                else:
                    fill = verde_fill  # 🟢 Sin especial

                # Aplicar formato a hoja principal
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(name='Calibri', size=11)
                    cell.border = thin_border

                    # 🔹 Si es UPC (columna 2), forzar formato de número entero
                    if col == 2 and cell.value != "N/A":
                        try:
                            cell.value = int(cell.value)
                            cell.number_format = numbers.FORMAT_NUMBER
                        except:
                            pass

            # --- Copiar encabezado y filas AMARILLAS a hoja 2 ---
            headers = [cell.value for cell in ws[1]]
            for c, header in enumerate(headers, start=1):
                ws_muestras.cell(row=1, column=c, value=header)

            for r, fila in enumerate(filas_amarillas, start=2):
                for c, valor in enumerate(fila, start=1):
                    cell = ws_muestras.cell(row=r, column=c, value=valor)
                    cell.fill = amarillo_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(name='Calibri', size=9)
                    cell.border = thin_border

                    if c == 2 and valor != "N/A":
                        try:
                            cell.value = int(valor)
                            cell.number_format = numbers.FORMAT_NUMBER
                        except:
                            pass

                ws_muestras.row_dimensions[r].height = row_height

            wb.save(salida)
            wb.close()


            messagebox.showinfo("Éxito", f"✅ Base generada correctamente.\n\n- Hoja 1: Base completa\n- Hoja 2: Muestras")

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error al generar la base:\n{e}")


if __name__ == "__main__":
    try:
        app = VentanaBaseEtiquetado()
        app.mainloop()
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo iniciar la aplicación:\n{e}")
