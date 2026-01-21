import customtkinter as ctk
import tkinter.filedialog as fd
import tkinter.messagebox as messagebox
import json
import pandas as pd
from pathlib import Path
import os

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

class ComparadorJSONExcel(ctk.CTk):
    def __init__(self):
        super().__init__()

        self.title("🧩 Comparador de JSONs ULTA → Excel")
        self.geometry("700x500")
        self.resizable(False, False)

        self.json_paths = []
        
        # Columnas específicas para ULTA
        self.columnas_ulta = [
            "CATEGORIA", "UPC", "DENOMINACION", "DENOMINACION AXO", "MARCA",
            "LEYENDAS PRECAUTORIAS", "INSTRUCCIONES DE USO", "OBSERVACIONES",
            "TAMAÑO DE LA DECLARACION DE CONTENIDO", "CONTENIDO", "PAIS ORIGEN",
            "IMPORTADOR", "NORMA", "INGREDIENTES", "MEDIDAS", "TIPO DE ETIQUETA"
        ]

        # --- Título ---
        ctk.CTkLabel(
            self, text="🔍 COMPARADOR ULTA - JSONs → EXCEL",
            font=("Arial", 20, "bold"), text_color="#282828"
        ).pack(pady=(25, 10))

        ctk.CTkLabel(
            self, text="Compara archivos JSON y exporta SOLO los registros con diferencias\nBasado en UPC y CATEGORIA iguales pero otros campos diferentes",
            font=("Arial", 12), text_color="#4d4d4d"
        ).pack(pady=(0, 20))

        # --- Frame principal ---
        main_frame = ctk.CTkFrame(self, fg_color="transparent")
        main_frame.pack(pady=10, padx=20, fill="both")

        # --- Botón de selección ---
        ctk.CTkButton(
            main_frame, text="📂 SELECCIONAR ARCHIVOS JSON",
            command=self.seleccionar_archivos,
            fg_color="#ecd925", hover_color="#d4bf1f",
            text_color="#282828", corner_radius=10,
            height=40, font=("Arial", 14, "bold")
        ).pack(pady=10, fill="x")

        # --- Label archivos seleccionados ---
        self.label_archivos = ctk.CTkLabel(
            main_frame, text="Ningún archivo seleccionado",
            font=("Arial", 11), text_color="#4d4d4d", 
            wraplength=600, justify="left"
        )
        self.label_archivos.pack(pady=10)

        # --- Frame para botones de acción ---
        action_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        action_frame.pack(pady=20, fill="x")

        # Botón comparar
        ctk.CTkButton(
            action_frame, text="🧾 COMPARAR Y EXPORTAR EXCEL",
            command=self.comparar_y_exportar,
            fg_color="#282828", hover_color="#4d4d4d",
            text_color="white", corner_radius=10,
            height=40, font=("Arial", 14, "bold")
        ).pack(pady=5, fill="x")

        # Botón limpiar
        ctk.CTkButton(
            action_frame, text="🔄 LIMPIAR SELECCIÓN",
            command=self.limpiar_seleccion,
            fg_color="#6c757d", hover_color="#5a6268",
            text_color="white", corner_radius=10,
            height=35, font=("Arial", 12)
        ).pack(pady=5, fill="x")

        # --- Resultado ---
        self.resultado_label = ctk.CTkLabel(
            main_frame, text="", 
            font=("Arial", 12, "bold"), text_color="#4d4d4d"
        )
        self.resultado_label.pack(pady=15)

        # --- Información ---
        info_text = """
ℹ️  INSTRUCCIONES:
1. Selecciona múltiples archivos JSON
2. El PRIMER archivo se usará como BASE
3. Los demás se compararán contra el base
4. Solo se mostrarán registros con:
   - Mismo UPC y CATEGORIA
   - Diferente información en otros campos
5. El Excel incluirá ambos registros para comparación
        """
        ctk.CTkLabel(
            main_frame, text=info_text,
            font=("Arial", 10), text_color="#6c757d",
            justify="left"
        ).pack(pady=10)

    def normalizar_clave(self, clave):
        """Normaliza claves para comparación"""
        return str(clave).strip().upper().replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")

    def normalizar_valor(self, valor):
        """Normaliza valores para comparación"""
        if valor is None:
            return ""
        return str(valor).strip()

    def cargar_y_normalizar_json(self, path):
        """Carga un JSON y normaliza sus datos"""
        try:
            with open(path, "r", encoding="utf-8") as f:
                datos = json.load(f)
            
            datos_normalizados = []
            for item in datos:
                if isinstance(item, dict):
                    item_normalizado = {}
                    # Normalizar todas las columnas esperadas
                    for columna in self.columnas_ulta:
                        # Buscar la columna en diferentes formatos
                        valor = item.get(columna) or item.get(columna.upper()) or item.get(columna.lower()) or ""
                        item_normalizado[columna] = self.normalizar_valor(valor)
                    datos_normalizados.append(item_normalizado)
            return datos_normalizados
        except Exception as e:
            raise Exception(f"Error cargando {Path(path).name}: {str(e)}")

    def seleccionar_archivos(self):
        """Selecciona archivos JSON"""
        archivos = fd.askopenfilenames(
            title="Seleccionar archivos JSON",
            filetypes=[("Archivos JSON", "*.json")]
        )
        if archivos:
            self.json_paths = list(archivos)
            nombres = [Path(p).name for p in self.json_paths]
            
            if len(nombres) > 0:
                texto = f"📁 ARCHIVO BASE: {nombres[0]}\n"
                if len(nombres) > 1:
                    texto += f"📊 ARCHIVOS A COMPARAR ({len(nombres)-1}):\n" + "\n".join([f"   • {name}" for name in nombres[1:]])
                self.label_archivos.configure(text=texto)
        else:
            self.label_archivos.configure(text="Ningún archivo seleccionado")

    def limpiar_seleccion(self):
        """Limpia la selección de archivos"""
        self.json_paths = []
        self.label_archivos.configure(text="Ningún archivo seleccionado")
        self.resultado_label.configure(text="")

    def comparar_y_exportar(self):
        """Compara los JSON y exporta SOLO los registros con diferencias"""
        if len(self.json_paths) < 2:
            messagebox.showwarning("Advertencia", "Selecciona al menos 2 archivos JSON para comparar.")
            return

        try:
            # Mostrar progreso
            self.resultado_label.configure(text="🔄 Cargando y comparando archivos...", text_color="#282828")
            self.update()

            # Cargar archivo base
            base_path = self.json_paths[0]
            base_data = self.cargar_y_normalizar_json(base_path)
            base_name = Path(base_path).name

            # Crear diccionario de base para búsqueda rápida
            base_dict = {}
            for item in base_data:
                upc = item.get('UPC', '')
                categoria = item.get('CATEGORIA', '')
                if upc and categoria:
                    clave = f"{upc}_{categoria}"
                    base_dict[clave] = item

            # Lista para almacenar SOLO los registros con diferencias
            registros_con_diferencias = []

            # Comparar cada archivo con el base
            for i in range(1, len(self.json_paths)):
                archivo_actual = self.json_paths[i]
                archivo_nombre = Path(archivo_actual).name
                
                # Cargar datos del archivo actual
                datos_actual = self.cargar_y_normalizar_json(archivo_actual)
                
                # Buscar registros con mismo UPC y CATEGORIA pero diferentes en otros campos
                for item_actual in datos_actual:
                    upc = item_actual.get('UPC', '')
                    categoria = item_actual.get('CATEGORIA', '')
                    
                    if not upc or not categoria:
                        continue
                    
                    clave = f"{upc}_{categoria}"
                    item_base = base_dict.get(clave)
                    
                    # Solo procesar si existe en el base (mismo UPC y CATEGORIA)
                    if item_base:
                        # Verificar si hay diferencias en ALGÚN campo (excluyendo UPC y CATEGORIA)
                        tiene_diferencias = False
                        campos_con_diferencias = []
                        
                        for columna in self.columnas_ulta:
                            # No comparar UPC y CATEGORIA ya que son iguales por definición
                            if columna in ['UPC', 'CATEGORIA']:
                                continue
                                
                            valor_base = item_base.get(columna, '')
                            valor_actual = item_actual.get(columna, '')
                            
                            if valor_base != valor_actual:
                                tiene_diferencias = True
                                campos_con_diferencias.append(columna)
                        
                        # Solo agregar si tiene diferencias en otros campos
                        if tiene_diferencias:
                            # Agregar registro BASE
                            registro_base = item_base.copy()
                            registro_base['ORIGEN'] = 'BASE'
                            registro_base['ARCHIVO'] = base_name
                            registro_base['CAMPOS_DIFERENTES'] = ', '.join(campos_con_diferencias)
                            registros_con_diferencias.append(registro_base)
                            
                            # Agregar registro CONFLICTO
                            registro_conflicto = item_actual.copy()
                            registro_conflicto['ORIGEN'] = 'CONFLICTO'
                            registro_conflicto['ARCHIVO'] = archivo_nombre
                            registro_conflicto['CAMPOS_DIFERENTES'] = ', '.join(campos_con_diferencias)
                            registros_con_diferencias.append(registro_conflicto)

            # Exportar a Excel
            if registros_con_diferencias:
                # Crear DataFrame con los datos
                df = pd.DataFrame(registros_con_diferencias)
                
                # Reordenar columnas: ORIGEN, ARCHIVO y CAMPOS_DIFERENTES primero, luego las columnas ULTA
                columnas_orden = ['ORIGEN', 'ARCHIVO', 'CAMPOS_DIFERENTES'] + self.columnas_ulta
                # Asegurarse de que todas las columnas existan en el DataFrame
                for col in columnas_orden:
                    if col not in df.columns:
                        df[col] = ''
                
                df = df[columnas_orden]
                
                # Guardar archivo
                output_path = fd.asksaveasfilename(
                    title="Guardar reporte de diferencias",
                    defaultextension=".xlsx",
                    filetypes=[("Archivos Excel", "*.xlsx")],
                    initialfile="diferencias_ulta.xlsx"
                )
                
                if output_path:
                    # Crear Excel con formato mejorado
                    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='REGISTROS_CON_DIFERENCIAS')
                        
                        # Obtener el libro y hoja para formatear
                        workbook = writer.book
                        worksheet = writer.sheets['REGISTROS_CON_DIFERENCIAS']
                        
                        # Formatear encabezados
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                        
                        # Estilo para encabezados
                        header_font = Font(bold=True, color="FFFFFF", size=12)
                        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                        
                        for col_num, value in enumerate(df.columns.values):
                            cell = worksheet.cell(1, col_num + 1)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center", vertical="center")
                        
                        # Ajustar ancho de columnas
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                        
                        # Resaltar filas de CONFLICTO y campos diferentes
                        conflict_fill = PatternFill(start_color="FFEAA7", end_color="FFEAA7", fill_type="solid")
                        base_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                        
                        for row_num, row in enumerate(worksheet.iter_rows(min_row=2), 2):
                            origen_cell = worksheet.cell(row=row_num, column=1)  # Columna ORIGEN
                            fill_color = conflict_fill if origen_cell.value == 'CONFLICTO' else base_fill
                            
                            for cell in row:
                                cell.fill = fill_color
                                # Resaltar bordes para mejor visualización
                                cell.border = Border(
                                    left=Side(style='thin'), 
                                    right=Side(style='thin'), 
                                    top=Side(style='thin'), 
                                    bottom=Side(style='thin')
                                )
                    
                    self.resultado_label.configure(
                        text=f"✅ EXCEL GENERADO EXITOSAMENTE\n"
                             f"📊 {len(registros_con_diferencias)//2} registros con diferencias encontrados\n"
                             f"📁 {Path(output_path).name}",
                        text_color="#28a745"
                    )
                    
                    # Preguntar si abrir el archivo
                    if messagebox.askyesno("Éxito", 
                                         f"Se encontraron {len(registros_con_diferencias)//2} registros con diferencias.\n"
                                         f"¿Deseas abrir el archivo Excel?"):
                        try:
                            os.startfile(output_path)  # Windows
                        except:
                            try:
                                os.system(f'open "{output_path}"')  # macOS
                            except:
                                try:
                                    os.system(f'xdg-open "{output_path}"')  # Linux
                                except:
                                    pass
                else:
                    self.resultado_label.configure(text="❌ Exportación cancelada", text_color="#dc3545")
            else:
                self.resultado_label.configure(
                    text="✅ NO SE ENCONTRARON DIFERENCIAS\nNo hay registros con mismo UPC y CATEGORIA pero diferentes en otros campos",
                    text_color="#28a745"
                )
                messagebox.showinfo("Sin diferencias", 
                                  "No se encontraron registros con diferencias.\n"
                                  "Todos los registros con mismo UPC y CATEGORIA son idénticos.")

        except Exception as e:
            self.resultado_label.configure(text=f"❌ ERROR: {str(e)}", text_color="#dc3545")
            messagebox.showerror("Error", f"Ocurrió un error durante la comparación:\n{str(e)}")

if __name__ == "__main__":
    app = ComparadorJSONExcel()
    app.mainloop()