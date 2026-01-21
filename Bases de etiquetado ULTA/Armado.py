import os
import re
import unicodedata
import pandas as pd
import json
import datetime
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, numbers
from core.manejador_archivos import convertir_a_json

# ─── Funciones auxiliares ───────────────────────────
def cargar_json_como_df(ruta_json):
    if not os.path.exists(ruta_json):
        raise FileNotFoundError(f"No se encontró el archivo: {ruta_json}")
    with open(ruta_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    if isinstance(data, list):
        df = pd.DataFrame(data)
    elif isinstance(data, dict):
        primera_hoja = list(data.keys())[0]
        df = pd.DataFrame(data[primera_hoja])
    else:
        raise ValueError("Formato de JSON no soportado")
    
    return df

def normalize_str(s):
    if isinstance(s, str):
        s = s.strip().upper()
        s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    return s

def sanitize_header(s):
    """Normaliza cabeceras eliminando BOMs, espacios no imprimibles, acentos y caracteres extra.
    Devuelve una versión en MAYÚSCULAS y sin acentos apta para comparaciones."""
    if s is None:
        return ""
    s = str(s)
    # eliminar BOM y caracteres invisibles comunes
    for ch in ['\ufeff', '\u200b', '\u200c', '\u200d', '\xa0']:
        s = s.replace(ch, ' ')
    # normalizar unicode y quitar acentos
    s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
    # mantener letras y números y espacios; convertir a mayúsculas
    s = re.sub(r"[^0-9A-Za-zÑñ ]+", ' ', s)
    s = re.sub(r"\s+", ' ', s).strip().upper()
    return s

# ─── Función principal ───────────────────────────
def generar_base():
    try:
        # --- Cargar archivos JSON ---
        base_general = cargar_json_como_df(os.path.join("data", "BASE_GENERAL_ULTA_ETIQUETADO.json"))
        base_ulta = cargar_json_como_df(os.path.join("data", "BASE_ULTA.json"))

        # --- Seleccionar Layout ---
        layout_file = filedialog.askopenfilename(title="Selecciona el Layout ULTA",
                                                 filetypes=[("Excel files", "*.xlsx *.xls")])
        if not layout_file:
            messagebox.showwarning("Layout", "No se seleccionó ningún layout.")
            return
        # Leer layout: la estructura viene con encabezado en la fila 3 (header=2).
        # Intentamos abrir la hoja 'Layout 1' primero y si no existe, usamos la primera hoja.
        try:
            layout = pd.read_excel(layout_file, sheet_name="Layout 1", header=2, dtype=str)
        except Exception:
            # Fallback: leer primer sheet pero manteniendo header en la fila 3
            layout = pd.read_excel(layout_file, header=2, dtype=str)

        # Limpiar columnas vacías que pandas marque como Unnamed
        if hasattr(layout, 'columns'):
            layout = layout.loc[:, ~layout.columns.str.contains('^Unnamed')]
        layout.dropna(how='all', inplace=True)
        layout.fillna('', inplace=True)

        # --- Normalizar encabezados (sanitizar para comparaciones robustas) ---
        for df in [base_general, base_ulta, layout]:
            df.columns = [sanitize_header(c) for c in df.columns]

        # Buscar columna PARTE de manera flexible (usa sanitized headers y fallback por substring)
        def find_column(df, target):
            target_norm = sanitize_header(target)
            # 1) búsqueda exacta
            for col in df.columns:
                try:
                    if sanitize_header(col) == target_norm:
                        return col
                except Exception:
                    continue
            # 2) búsqueda por substring (p. ej. 'NUMERO PARTE', 'NRO PARTE')
            for col in df.columns:
                try:
                    if target_norm in sanitize_header(col):
                        return col
                except Exception:
                    continue
            return None

        col_parte = find_column(layout, "PARTE")
        if col_parte is None:
            # Mostrar columnas detectadas para ayudar al diagnóstico
            columnas_disp = ', '.join(list(layout.columns[:20]))
            messagebox.showerror("Error - Columna no encontrada",
                                 f"No se encontró la columna 'PARTE' en el layout.\nColumnas detectadas (ejemplo):\n{columnas_disp}")
            raise ValueError("No se encontró la columna 'PARTE' en el layout")

        columnas = [
            "CATEGORIA", "UPC", "DENOMINACION", "DENOMINACION AXO", "MARCA",
            "LEYENDAS PRECAUTORIAS", "INSTRUCCIONES DE USO", "OBSERVACIONES",
            "TAMAÑO DE LA DECLARACION DE CONTENIDO", "CONTENIDO", "PAIS ORIGEN",
            "IMPORTADOR", "NORMA", "INGREDIENTES", "MEDIDAS", "TIPO DE ETIQUETA"
        ]
        base_etiquetado = pd.DataFrame(columns=columnas)

        # --- Preparar sets de UPC ---
        upcs_base_general_set = set(base_general["UPC"].astype(str).str.strip().str.upper())
        upcs_base_ulta_set = set(base_ulta["UPC"].astype(str).str.strip().str.upper())

        filas_base_general_idx = []
        filas_ulta_idx = []
        filas_amarillas = []

        # --- Armado de filas ---
        for idx, fila in layout.iterrows():
            codigo = str(fila[col_parte]).strip().upper()
            if not codigo or codigo == "NAN":
                continue

            fila_final = {col: "" for col in columnas}

            # Prioridad Base General
            if codigo in upcs_base_general_set:
                bg = base_general[base_general["UPC"].astype(str).str.strip().str.upper() == codigo]
                filas_base_general_idx.append(len(base_etiquetado))
            elif codigo in upcs_base_ulta_set:
                bg = base_ulta[base_ulta["UPC"].astype(str).str.strip().str.upper() == codigo]
                filas_ulta_idx.append(len(base_etiquetado))
            else:
                continue

            fila_bg = bg.iloc[0]

            # --- Asignar valores por columna ---
            fila_final["CATEGORIA"] = fila_bg.get("CATEGORIA", "")
            upc_val = fila_bg.get("UPC", "")
            fila_final["UPC"] = str(upc_val).strip() if pd.notna(upc_val) and str(upc_val).strip() != "" else "N/A"

            col_denominacion = [c for c in fila_bg.index if "DENOMINACION" in normalize_str(c) and "GENERICA" in normalize_str(c)]
            fila_final["DENOMINACION"] = fila_bg[col_denominacion[0]].strip() if col_denominacion and pd.notna(fila_bg[col_denominacion[0]]) else fila_bg.get("DENOMINACION", "N/A")
            fila_final["DENOMINACION AXO"] = fila_bg.get("DENOMINACION AXO", "N/A")
            fila_final["LEYENDAS PRECAUTORIAS"] = fila_bg.get("LEYENDAS PRECAUTORIAS", "N/A")
            fila_final["INSTRUCCIONES DE USO"] = fila_bg.get("INSTRUCCIONES DE USO", "N/A")
            fila_final["OBSERVACIONES"] = fila_bg.get("OBSERVACIONES", "N/A")

            
            target_tamano = "TAMAÑO DE LA DECLARACION DE CONTENIDO"
            tam_col = None
            for col in fila_bg.index:
                try:
                    if sanitize_header(col) == sanitize_header(target_tamano):
                        tam_col = col
                        break
                except Exception:
                    continue

            # Si no encontramos por igualdad, intentar búsqueda por inclusión
            if tam_col is None:
                for col in fila_bg.index:
                    try:
                        if sanitize_header(target_tamano) in sanitize_header(col):
                            tam_col = col
                            break
                    except Exception:
                        continue

            fila_final["TAMAÑO DE LA DECLARACION DE CONTENIDO"] = fila_bg.get(tam_col, "") if tam_col else fila_bg.get("TAMAÑO DE LA DECLARACION DE CONTENIDO", "")

            fila_final["CONTENIDO"] = fila_bg.get("CONTENIDO", "N/A")
            fila_final["PAIS ORIGEN"] = fila_bg.get("PAIS ORIGEN", fila_bg.get("PAIS DE ORIGEN", "N/A"))
            fila_final["IMPORTADOR"] = fila_bg.get("IMPORTADOR", "N/A")
            fila_final["NORMA"] = fila_bg.get("NORMA", "N/A")
            fila_final["INGREDIENTES"] = fila_bg.get("INGREDIENTES Y LOTE", "N/A")
            fila_final["MEDIDAS"] = fila_bg.get("MEDIDAS", "N/A")
            fila_final["TIPO DE ETIQUETA"] = fila_bg.get("TIPO DE ETIQUETA", "N/A")

            col_marca = [c for c in fila_bg.index if "MARCA" in normalize_str(c)]
            fila_final["MARCA"] = fila_bg[col_marca[0]].strip() if col_marca and pd.notna(fila_bg[col_marca[0]]) else "N/A"

            # Rellenar N/A donde falten valores
            for k, v in fila_final.items():
                if v == "" or pd.isna(v):
                    fila_final[k] = "N/A"

            base_etiquetado = pd.concat([base_etiquetado, pd.DataFrame([fila_final])], ignore_index=True)

        # --- Guardar archivo y aplicar colores ---
        salida = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                              filetypes=[("Excel files", "*.xlsx")],
                                              title="Guardar Base de Etiquetado ULTA.xlsx")
        if not salida:
            return
        
        # --- Al final, después de guardar el Excel ---
        historial_file = os.path.join("data", "historial_bases.json")

        # Cargar historial existente
        if os.path.exists(historial_file):
            with open(historial_file, "r", encoding="utf-8") as f:
                historial = json.load(f)
        else:
            historial = []

        # Agregar nuevo registro con nombre, ruta y fecha
        registro = {
            "nombre_archivo": os.path.basename(salida),
            "ruta_archivo": os.path.abspath(salida),
            "fecha_generacion": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
        }

        historial.append(registro)

        # Guardar historial actualizado
        with open(historial_file, "w", encoding="utf-8") as f:
            json.dump(historial, f, indent=4, ensure_ascii=False)

        with pd.ExcelWriter(salida, engine="openpyxl") as writer:
            base_etiquetado.to_excel(writer, index=False, sheet_name="Base Etiquetado Completa")
            pd.DataFrame(columns=base_etiquetado.columns).to_excel(writer, index=False, sheet_name="Muestras")

        # Colores y formato MEJORADO
        wb = load_workbook(salida)
        ws = wb["Base Etiquetado Completa"]
        ws_muestras = wb["Muestras"]

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        amarillo_fill = PatternFill(start_color="FBF871", end_color="FBF871", fill_type="solid")
        verde_fill = PatternFill(start_color="9AE494", end_color="9AE494", fill_type="solid")
        azul_fill = PatternFill(start_color="69CFF3", end_color="69CFF3", fill_type="solid")
        rosa_fill = PatternFill(start_color="F890DD", end_color="F890DD", fill_type="solid")  # 🔸 NUEVO COLOR NARANJA

        # --- FUNCIÓN PARA AUTO-AJUSTE ---
        def auto_ajustar_columnas_y_filas(worksheet):
            """Ajusta automáticamente el ancho de columnas y el alto de filas para mejor visualización."""

            # --- Ajustar ancho de columnas ---
            for col in worksheet.columns:
                max_length = 0
                column_letter = col[0].column_letter

                for cell in col:
                    if cell.value:
                        # Calcular longitud considerando saltos de línea
                        cell_lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in cell_lines)
                        max_length = max(max_length, max_line_length)

                # Ajustar ancho con márgenes
                adjusted_width = min(max(max_length + 2, 10), 50)  # Mínimo 10, máximo 50
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # --- Ajustar altura de filas ---
            BASE_HEIGHT = 18       # Altura mínima base
            LINE_HEIGHT = 16       # Altura adicional por cada línea
            MAX_HEIGHT = 70        # Límite máximo de altura

            for row in worksheet.iter_rows():
                max_lines = 1
                for cell in row:
                    if cell.value:
                        # Contar líneas en cada celda
                        line_count = str(cell.value).count('\n') + 1
                        max_lines = max(max_lines, line_count)

                # Aplicar altura proporcional con base y límite
                row_height = max(BASE_HEIGHT, min(max_lines * LINE_HEIGHT, MAX_HEIGHT))
                worksheet.row_dimensions[row[0].row].height = row_height

        # --- APLICAR COLORES Y CAPTURAR FILAS AMARILLAS ---
        for row_idx in range(2, ws.max_row + 1):
            df_idx = row_idx - 2

            tipo_etiqueta_val = ws.cell(row=row_idx, column=16).value  # Columna TIPO DE ETIQUETA (16)
            tipo_etiqueta_str = str(tipo_etiqueta_val).strip().upper() if tipo_etiqueta_val else ""

            if ("ETIQUETA TRANSPARENTE" in tipo_etiqueta_str or
                "ETIQUETA NEGRA CON LETRAS BLANCAS" in tipo_etiqueta_str):
                fill = rosa_fill
                print(f"🔸 Fila {row_idx} marcada como NARANJA - TIPO DE ETIQUETA: {tipo_etiqueta_str}")

            else:
                valor = ws.cell(row=row_idx, column=15).value
                valor_str = str(valor).strip().upper() if valor else ""

                if "REQUIERE ETIQUETADO ESPECIAL" in valor_str or "NO IMPRIMIR HASTA TENER VISTO BUENO DE V&C" in valor_str:
                    fill = amarillo_fill
                    # 📋 Copiar esta fila a la hoja "Muestras"
                    fila_valores = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
                    ws_muestras.append(fila_valores)

                    # 🔶 Aplicar formato amarillo y estilos en la hoja "Muestras"
                    new_row_idx = ws_muestras.max_row
                    for col in range(1, ws_muestras.max_column + 1):
                        cell_m = ws_muestras.cell(row=new_row_idx, column=col)
                        cell_m.fill = amarillo_fill
                        cell_m.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell_m.font = Font(name='Calibri', size=10)
                        cell_m.border = thin_border

                elif df_idx in filas_base_general_idx:
                    fill = verde_fill
                elif df_idx in filas_ulta_idx:
                    fill = azul_fill
                else:
                    fill = verde_fill

            # --- Aplicar formato visual a la hoja principal ---
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border

                # Convertir a número si aplica
                if col == 2 and cell.value != "N/A":
                    try:
                        cell.value = int(cell.value)
                        cell.number_format = numbers.FORMAT_NUMBER
                    except:
                        pass

        # 🔄 Autoajuste para ambas hojas
        print("🔄 Aplicando auto-ajuste a las columnas y filas...")
        auto_ajustar_columnas_y_filas(ws)
        auto_ajustar_columnas_y_filas(ws_muestras)




        # Formato adicional para la hoja de muestras
        for row in ws_muestras.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(name='Calibri', size=10)
                cell.border = thin_border

        # Encabezados en negrita para ambas hojas
        for worksheet in [ws, ws_muestras]:
            for cell in worksheet[1]:  # Fila 1 son los encabezados
                cell.font = Font(name='Calibri', size=10, bold=True)
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

        # Guardar cambios
        wb.save(salida)
        wb.close()
        
        messagebox.showinfo("Éxito", f"✅ Base generada correctamente.\n\n- Hoja 1: Base completa\n- Hoja 2: Muestras")

        # --- Intentar actualizar el Dashboard si está abierto ---
        try:
            from Dashboard import VentanaDashboard
            if hasattr(VentanaDashboard, "instancia_activa") and VentanaDashboard.instancia_activa:
                VentanaDashboard.instancia_activa.cargar_historial_bases()
                VentanaDashboard.instancia_activa.actualizar_dashboard()
                print("🔄 Dashboard actualizado automáticamente.")
        except Exception as e:
            print(f"⚠️ No se pudo actualizar el Dashboard automáticamente: {e}")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al generar la base:\n{e}")

