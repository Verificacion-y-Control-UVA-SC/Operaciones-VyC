import os, re
import json
import tkinter as tk
from tkinter import ttk
import customtkinter as ctk
import pandas as pd
from openpyxl import load_workbook
from tkinter import messagebox
from datetime import datetime
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, numbers
from openpyxl.utils.dataframe import dataframe_to_rows

COLORES = {
    "primario": "#ECD925",
    "secundario": "#282828",
    "exito": "#008d53",
    "peligro": "#d74a3d",
    "fondo": "#F8F9FA",
    "surface": "#FFFFFF",
    "texto_oscuro": "#282828",
    "texto_claro": "#4b4b4b",
    "borde": "#BDC3C7",
}

FUENTE = "Inter"

def load_config():
    cfg_path = os.path.join("data", "config.json")
    if os.path.exists(cfg_path):
        try:
            with open(cfg_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}

def save_config(cfg):
    os.makedirs("data", exist_ok=True)
    cfg_path = os.path.join("data", "config.json")
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=4, ensure_ascii=False)

class EditorWindow(ctk.CTkToplevel):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.title("Editor de Bases - ULTA/AXO")
        self.geometry("1200x600")  # 🔸 Aumenté el tamaño para mejor visualización
        self.configure(fg_color="#F8F9FA")

        # load config and file paths
        self.cfg = load_config()
        self.path_base_general = self.cfg.get("base_general_path") or os.path.join("data", "BASE_GENERAL_ULTA_ETIQUETADO.json")
        self.path_base_norma_024 = self.cfg.get("base_norma_024_path") or os.path.join("data", "BASE_NORMA_024.json")

        # dataframes
        self.df_general = pd.DataFrame()
        self.df_norma_024 = pd.DataFrame()

        # recent actions
        self.recent_general = []
        self.recent_norma_024 = []

        # --- Control de paginación para mostrar las bases --- #
        self.page_size = 1000
        self.current_page_general = 0
        self.current_page_norma_024 = 0

        # Inicializar conjuntos de revisados
        self.reviewed_general = set()
        self.reviewed_norma_024 = set()

        self.update_stats()
        self.create_ui()
        self.load_files()

    def create_ui(self):
        header = ctk.CTkFrame(self, fg_color=COLORES["fondo"], height=80)
        header.pack(fill="x")

        # --- Título principal ---
        title = ctk.CTkLabel(header, text="📊 Editor de Bases", font=(FUENTE, 22, "bold"), text_color="#282828")
        title.pack(anchor="w", padx=16, pady=18)

        # ---  SIMBOLOGÍA DE COLORES (parte superior derecha) ---
        simbologia_frame = ctk.CTkFrame(header, fg_color="transparent")
        simbologia_frame.pack(anchor="e", padx=16, pady=10, side="right")

        def crear_item_simbologia(color_hex, texto):
            """Crea un recuadro de color con su etiqueta descriptiva."""
            item_frame = ctk.CTkFrame(simbologia_frame, fg_color="transparent")
            item_frame.pack(side="right", padx=(8, 0))

            color_box = ctk.CTkLabel(item_frame, text="", width=22, height=22, fg_color=color_hex, corner_radius=4)
            color_box.pack(side="left", padx=(0, 6))

            label = ctk.CTkLabel(item_frame, text=texto, text_color="#282828", font=(FUENTE, 13))
            label.pack(side="left")

        # Crear los tres ítems de simbología
        crear_item_simbologia("#f6e96b", "AMARILLO: Se requiere muestra (Requiere visto bueno / Etiquetas Especiales / Etiquetas Metalicas)")
        crear_item_simbologia("#b6e3a8", "VERDE: Tamaño definido")
        # crear_item_simbologia("#f7b0b7", "ROSA: Etiqueta negra o transparente")

        # --- CUERPO PRINCIPAL ---
        body = ctk.CTkFrame(self, fg_color="transparent")
        body.pack(fill="both", expand=True, padx=12, pady=12)

        # --- Estadísticas ---
        stats_row = ctk.CTkFrame(body, fg_color='transparent')
        stats_row.pack(fill='x', pady=(0,6))
        
        # Contador Base General
        self.lbl_stats_general = ctk.CTkLabel(stats_row, text='Base General: categorías: 0 | UPC únicos: 0', text_color='#282828')
        self.lbl_stats_general.pack(side='left', padx=(6,12))

        # Contador Base Norma 024
        self.lbl_stats_norma_024 = ctk.CTkLabel(stats_row, text='Base Norma 024: asignaciones: 0 | UPC únicos: 0 | Modelo: 0', text_color='#282828')
        self.lbl_stats_norma_024.pack(side='left', padx=(6,12))

        # --- BARRAS DE BÚSQUEDA EN UNA SOLA FILA ---
        search_main_frame = ctk.CTkFrame(body, fg_color='transparent')
        search_main_frame.pack(fill='x', pady=(0, 10))

        # Configuración de tamaños uniformes
        ENTRY_WIDTH = 100
        LABEL_WIDTH = 90

        # Contenedor principal (una sola fila)
        search_row = ctk.CTkFrame(search_main_frame, fg_color='transparent')
        search_row.pack(fill='x', pady=4)

        # Buscar por CATEGORÍA
        ctk.CTkLabel(search_row, text="Categoría:", width=LABEL_WIDTH, anchor='w').pack(side='left', padx=(0, 4))
        self.search_categoria = ctk.CTkEntry(
            search_row,
            placeholder_text='Categoría',
            width=ENTRY_WIDTH,
            fg_color='#FFFFFF',
            text_color='#000000'
        )
        self.search_categoria.pack(side='left', padx=(0, 10))
        self.search_categoria.bind('<Return>', lambda e: self.perform_multi_search())

        # Buscar por UPC
        ctk.CTkLabel(search_row, text="UPC:", width=LABEL_WIDTH, anchor='w').pack(side='left', padx=(0, 4))
        self.search_upc = ctk.CTkEntry(
            search_row,
            placeholder_text='UPC',
            width=ENTRY_WIDTH,
            fg_color='#FFFFFF',
            text_color='#000000'
        )
        self.search_upc.pack(side='left', padx=(0, 10))
        self.search_upc.bind('<Return>', lambda e: self.perform_multi_search())

        # Buscar por ASIGNACIÓN
        ctk.CTkLabel(search_row, text="Asignación:", width=LABEL_WIDTH, anchor='w').pack(side='left', padx=(0, 4))
        self.search_asignacion = ctk.CTkEntry(
            search_row,
            placeholder_text='Asignación',
            width=ENTRY_WIDTH,
            fg_color='#FFFFFF',
            text_color='#000000'
        )
        self.search_asignacion.pack(side='left', padx=(0, 10))
        self.search_asignacion.bind('<Return>', lambda e: self.perform_multi_search())

        # Buscar por MEDIDAS
        ctk.CTkLabel(search_row, text="Medidas:", width=LABEL_WIDTH, anchor='w').pack(side='left', padx=(0, 4))
        self.search_medidas = ctk.CTkEntry(
            search_row,
            placeholder_text='Medidas',
            width=ENTRY_WIDTH,
            fg_color='#FFFFFF',
            text_color='#000000'
        )
        self.search_medidas.pack(side='left', padx=(0, 10))
        self.search_medidas.bind('<Return>', lambda e: self.perform_multi_search())



        # --- BOTONES Y CONTADOR A LA DERECHA ---
        ctk.CTkButton(
            search_row, text='🔍 Buscar', command=self.perform_multi_search,
            fg_color='#282828', hover_color='#4b4b4b', text_color='white',
            width=90
        ).pack(side='left', padx=(6, 6))

        ctk.CTkButton(
            search_row, text='🧹 Limpiar', command=self.clear_search,
            fg_color='#6c6c6c', hover_color='#8a8a8a', text_color='white',
            width=90
        ).pack(side='left', padx=(0, 10))

        # Contador de resultados
        self.lbl_resultados = ctk.CTkLabel(search_row, text='📊 Resultados: 0', 
                                         text_color='#E67E22', font=('Inter', 13, 'bold'))
        self.lbl_resultados.pack(side='left', padx=(12,0))

        # --- Estilos del Treeview ---
        try:
            style = ttk.Style()
            try:
                style.theme_use('clam')
            except Exception:
                pass

            style.configure(
                'Treeview',
                background='#FFFFFF',
                foreground='#000000',
                fieldbackground='#FFFFFF',
                rowheight=35,
                font=('Inter', 12)
            )

            style.configure(
                'Treeview.Heading',
                background='#ECD925',
                foreground='#282828',
                font=('Inter', 13, 'bold'),
                padding=(10, 8)
            )

            style.map('Treeview',
                    background=[('selected', '#BEE3F8')],
                    foreground=[('selected', '#000000')])
        except Exception:
            pass

        # --- Pestañas ---
        self.tabs = ctk.CTkTabview(body)
        self.tabs.pack(fill="both", expand=True, padx=6, pady=6)
        self.tabs.add("Base General")
        self.tabs.add("Base Norma 024")

        # Frame para Base General
        frame_general = ctk.CTkFrame(self.tabs.tab("Base General"), fg_color="transparent")
        frame_general.pack(fill="both", expand=True)

        # Frame para Base Norma 024
        frame_norma_024 = ctk.CTkFrame(self.tabs.tab("Base Norma 024"), fg_color="transparent")
        frame_norma_024.pack(fill="both", expand=True)

        # Treeviews con scrollbars
        self.tree_general = self._make_tree(frame_general)
        self.tree_norma_024 = self._make_tree(frame_norma_024)

        # --- Botones inferiores ---
        btns = ctk.CTkFrame(self, fg_color='#F8F9FA')
        btns.pack(fill="x", padx=12, pady=(0,12))

        ctk.CTkButton(btns, text="➕ Añadir fila", command=self.on_add,
                    fg_color="#282828", hover_color="#4b4b4b", text_color="white").pack(side="left", padx=6)
        ctk.CTkButton(btns, text="✏️ Editar seleccionado", command=self.on_edit,
                    fg_color="#282828", hover_color="#4b4b4b", text_color="white").pack(side="left", padx=6)
        ctk.CTkButton(btns, text="🗑️ Borrar seleccionado", command=self.on_delete,
                    fg_color="#d74a3d", hover_color="#d57067", text_color="white").pack(side="left", padx=6)
        
        ctk.CTkButton(btns, text="❌ Cerrar", command=self.destroy,
                    fg_color="#d74a3d", hover_color="#d57067", text_color="white").pack(side="right", padx=6)
        
        # --- Navegación de páginas ---
        nav_frame = ctk.CTkFrame(body, fg_color="transparent")
        nav_frame.pack(fill='x', pady=(4, 0))

        ctk.CTkButton(
            nav_frame,
            text="⏮ Anterior",
            width=120,
            fg_color="#ecd925",
            text_color="#282828",
            hover_color="#f5e85a",
            command=self.prev_page
        ).pack(side="left", padx=4)

        ctk.CTkButton(
            nav_frame,
            text="Siguiente ⏭",
            width=120,
            fg_color="#ecd925",
            text_color="#282828",
            hover_color="#f5e85a",
            command=self.next_page
        ).pack(side="right", padx=4)

    def perform_quick_search(self):
        """Búsqueda rápida en todos los campos"""
        query = self.quick_search.get().strip()
        if not query:
            self._refresh_trees()
            self.lbl_resultados.configure(text='📊 Resultados: Mostrando todo')
            return

        scope = "Todas"
        query_norm = query.upper()
        
        targets = []
        if scope in ('Todas', 'Base General'):
            targets.append(('Base General', self.tree_general, self.df_general))
        if scope in ('Todas', 'Base Norma 024'):
            targets.append(('Base Norma 024', self.tree_norma_024, self.df_norma_024))

        total_results = 0

        for name, tree, df in targets:
            if df is None or df.empty:
                continue

            for iid in tree.get_children():
                tree.delete(iid)

            cols = list(df.columns)
            tree.config(columns=cols)
            for col in cols:
                tree.heading(col, text=col)
                tree.column(col, width=self._calculate_column_width(df, col), anchor='center')

            # Búsqueda en todas las columnas
            mask = df.apply(lambda row: row.astype(str).str.upper().str.contains(query_norm, na=False).any(), axis=1)
            df_filtered = df[mask]
            total_results += len(df_filtered)

            # Mostrar resultados
            for idx, row in df_filtered.iterrows():
                values = [self._format_cell(row.get(c, '')) for c in cols]
                tag = self._row_tag(row)
                
                upc_col = self._find_column(df, ['UPC', 'Upc', 'Upc/Ean'])
                if upc_col:
                    upc_val = str(row.get(upc_col, '')).strip().upper()
                    if name == "Base General" and upc_val in self.reviewed_general:
                        values[0] = "✅ " + values[0]
                    elif name == "Base Norma 024" and upc_val in self.reviewed_norma_024:
                        values[0] = "✅ " + values[0]
                    else:
                        self.marcar_como_revisado(upc_val, name)

                tree.insert('', 'end', iid=str(idx), values=values, tags=(tag,))

        self.lbl_resultados.configure(text=f'📊 Resultados: {total_results}')

    def clear_search(self):
        """Limpia todas las búsquedas"""
        self.search_categoria.delete(0, 'end')
        self.search_upc.delete(0, 'end')
        self.search_asignacion.delete(0, 'end')
        self.search_medidas.delete(0, 'end')
        self.quick_search.delete(0, 'end')
        
        self.lbl_resultados.configure(text='📊 Resultados: 0')
        self._refresh_trees()

    def _make_tree(self, parent):
        # use a frame that expands and contains both scrollbars
        frame = ctk.CTkFrame(parent, fg_color="transparent")
        frame.pack(fill="both", expand=True)

        # horizontal and vertical scrollbars
        hsb = ttk.Scrollbar(frame, orient="horizontal")
        vsb = ttk.Scrollbar(frame, orient="vertical")

        tv = ttk.Treeview(frame, show="headings", xscrollcommand=hsb.set, yscrollcommand=vsb.set)

        # layout: tree in center, vscroll on right, hscroll at bottom
        tv.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')

        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        vsb.configure(command=tv.yview)
        hsb.configure(command=tv.xview)

        return tv

    def load_files(self):
        # load JSON if exists
        default_general = os.path.join('data', 'BASE_GENERAL_ULTA_ETIQUETADO.json')
        default_norma_024 = os.path.join('data', 'BASE_NORMA_024.json')
        
        try:
            if not os.path.exists(self.path_base_general) and os.path.exists(default_general):
                old = self.path_base_general
                self.path_base_general = os.path.abspath(default_general)
                try:
                    self.cfg['base_general_path'] = self.path_base_general
                    save_config(self.cfg)
                except Exception:
                    pass
                try:
                    messagebox.showinfo('Ruta corregida', f"Se usará la ruta local por defecto para Base General:\n{self.path_base_general}\n(la ruta anterior no existe:\n{old})")
                except Exception:
                    print(f"[Editor] Ruta corregida: using {self.path_base_general} (old: {old})")
        except Exception:
            pass
        
        try:
            if os.path.exists(self.path_base_general):
                with open(self.path_base_general, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    found = None
                    for v in data.values():
                        if isinstance(v, list):
                            found = v
                            break
                    if found is not None:
                        data = found
                    else:
                        data = [data]
                self.df_general = pd.DataFrame(data)
                try:
                    count = len(data) if isinstance(data, (list,tuple)) else (0 if data is None else 1)
                    print(f"[Editor] Base General cargada desde {self.path_base_general} - registros detectados: {count}")
                except Exception:
                    pass
            else:
                self.df_general = pd.DataFrame()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer Base General:\n{e}")
            self.df_general = pd.DataFrame()

        # Cargar Base Norma 024
        try:
            if not os.path.exists(self.path_base_norma_024) and os.path.exists(default_norma_024):
                old = self.path_base_norma_024
                self.path_base_norma_024 = os.path.abspath(default_norma_024)
                try:
                    self.cfg['base_norma_024_path'] = self.path_base_norma_024
                    save_config(self.cfg)
                except Exception:
                    pass
        except Exception:
            pass

        try:
            if os.path.exists(self.path_base_norma_024):
                with open(self.path_base_norma_024, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, dict):
                    found = None
                    for v in data.values():
                        if isinstance(v, list):
                            found = v
                            break
                    if found is not None:
                        data = found
                    else:
                        data = [data]
                self.df_norma_024 = pd.DataFrame(data)
                try:
                    count = len(data) if isinstance(data, (list,tuple)) else (0 if data is None else 1)
                    print(f"[Editor] Base Norma 024 cargada desde {self.path_base_norma_024} - registros detectados: {count}")
                except Exception:
                    pass
            else:
                self.df_norma_024 = pd.DataFrame()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer Base Norma 024:\n{e}")
            self.df_norma_024 = pd.DataFrame()

        # normalize columns: ensure strings
        for df in [self.df_general, self.df_norma_024]:
            if df is not None and not df.empty:
                cols = list(df.columns)
                df.columns = [str(c) for c in cols]
                for col in df.columns:
                    df[col] = df[col].astype(object)

        self._refresh_trees()
        try:
            self.update_stats()
        except Exception:
            pass

    def _refresh_trees(self):
        # --- Columnas de ambas bases ---
        cols_g = list(self.df_general.columns) if not self.df_general.empty else []
        cols_n24 = list(self.df_norma_024.columns) if not self.df_norma_024.empty else []

        # --- Mostrar página actual ---
        self._populate_tree_paginated(self.tree_general, self.df_general, cols_g, self.current_page_general)
        self._populate_tree_paginated(self.tree_norma_024, self.df_norma_024, cols_n24, self.current_page_norma_024)

        self.update_stats()

    def _populate_tree_paginated(self, tree, df, columns, page):
        for c in tree.get_children():
            tree.delete(c)

        tree.config(columns=columns)
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=self._calculate_column_width(df, col), anchor='center', stretch=False)

        if df is None or df.empty:
            return

        start = page * self.page_size
        end = start + self.page_size
        subset = df.iloc[start:end]

        # 🔸 Detectar si estamos mostrando Base Norma 024
        is_norma_024 = (tree == self.tree_norma_024)

        for idx, row in subset.iterrows():
            values = [self._format_cell(row.get(c, "")) for c in columns]
            tag = self._row_tag(row)

            # 🔹 Fuerza color amarillo para Base Norma 024
            if is_norma_024:
                tag = 'n024'

            tree.insert("", "end", iid=str(idx), values=values, tags=(tag,))

        # 🔸 Configurar los colores
        tree.tag_configure('special', background='#FFFF80', foreground='#000000')
        tree.tag_configure('normal', background='#DFF0D8', foreground='#000000')
        tree.tag_configure('n024', background='#FBF871', foreground='#000000')

    def _calculate_column_width(self, df, column_name):
        """Calcula el ancho óptimo para una columna basado en encabezado y contenido"""
        if df is None or df.empty or column_name not in df.columns:
            return 150  # Ancho por defecto
        
        # Longitud del encabezado
        header_width = len(str(column_name)) * 9
        
        # Longitud máxima del contenido en la columna (muestra solo las primeras filas para eficiencia)
        sample_size = min(50, len(df))
        content_width = 0
        for i in range(sample_size):
            if i < len(df):
                content = str(df.iloc[i][column_name]) if not pd.isna(df.iloc[i][column_name]) else ""
                content_width = max(content_width, len(content) * 8)
        
        # Tomar el máximo entre encabezado y contenido, con límites
        width = min(max(header_width, content_width, 120) + 20, 400)  # 🔸 Mínimo 120, máximo 400
        return width

    def _format_cell(self, v):
        if pd.isna(v):
            return ""
        # 🔸 MEJORA: Limitar longitud de texto muy largo para mejor visualización
        text = str(v)
        if len(text) > 100:
            return text[:100] + "..."
        return text

    def _row_tag(self, row):
        medidas = ''
        for key in row.index:
            try:
                if str(key).strip().upper() == 'MEDIDAS':
                    medidas = str(row.get(key, '')).upper()
                    break
            except Exception:
                continue
        # Normalizar medidas y leyendas para detección
        medidas_norm = (medidas or '').replace('×', 'x').replace(',', '.').strip()
        if any(phrase in medidas_norm for phrase in ["REQUIERE ETIQUETADO ESPECIAL", "NO IMPRIMIR HASTA TENER VISTO BUENO DE V&C"]):
            return 'special'
        # detectar medidas tipo '17 x 25' o '17mm x 25mm'
        try:
            if re.search(r'\d+\s*(mm)?\s*[x×]\s*\d+', medidas_norm, flags=re.IGNORECASE):
                return 'normal'
        except Exception:
            pass
        return 'normal'

# --- Search and stats helpers --- #
    def _find_column(self, df, candidates):
        # Try to find a column name in df matching any of the candidate names (case-insensitive)
        if df is None or df.empty:
            return None
        cols = {c.strip().upper(): c for c in df.columns}
        for cand in candidates:
            key = cand.strip().upper()
            if key in cols:
                return cols[key]
        return None

    def update_stats(self):
        try:
            # --- BASE GENERAL ---
            if not self.df_general.empty:
                cat_col_g = self._find_column(self.df_general, ['CATEGORIA', 'CATEGORÍA', 'CATEGORY', 'Categoria'])
                upc_col_g = self._find_column(self.df_general, ['UPC', 'Upc', 'Upc/Ean'])
                cat_count_g = int(self.df_general[cat_col_g].dropna().astype(str).nunique()) if cat_col_g else 0
                upc_count_g = int(self.df_general[upc_col_g].dropna().astype(str).nunique()) if upc_col_g else 0

            else:
                cat_count_g = upc_count_g  = 0

            self.lbl_stats_general.configure(
                text=f'Base General: categorías: {cat_count_g} | UPC únicos: {upc_count_g}'
            )

            # --- BASE NORMA 024 ---
            if not self.df_norma_024.empty:
                cat_col_n24 = self._find_column(self.df_norma_024, ['ASIGNACION', 'ASIGNACIÓN', 'asignacion', 'asignación'])
                upc_col_n24 = self._find_column(self.df_norma_024, ['UPC', 'Upc', 'Upc/Ean'])
                mod_col_n24 = self._find_column(self.df_norma_024, ['MODELO','modelo'])
                cat_count_n24 = int(self.df_norma_024[cat_col_n24].dropna().astype(str).nunique()) if cat_col_n24 else 0
                upc_count_n24 = int(self.df_norma_024[upc_col_n24].dropna().astype(str).nunique()) if upc_col_n24 else 0
                mod_count_n24 = int(self.df_norma_024[mod_col_n24].dropna().astype(str).nunique()) if mod_col_n24 else 0  # 🔸 CORRECCIÓN: usar mod_col_n24

            else:
                cat_count_n24 = upc_count_n24 = mod_count_n24 = 0

            self.lbl_stats_norma_024.configure(
                text=f'Base Norma 024: Asignacion: {cat_count_n24} | UPC únicos: {upc_count_n24} | Modelo: {mod_count_n24}'
            )

        except Exception as e:
            print(f"[update_stats] Error: {e}")

    def marcar_como_revisado(self, upc, base):
        upc = str(upc).strip().upper()
        if not upc:
            return

        if base == "Base Norma 024":
            self.reviewed_norma_024.add(upc)
        elif base == "Base General":
            self.reviewed_general.add(upc)

        self.update_stats()

    def _record_recent(self, tab_name, action, rowdata):
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        key = ''
        if isinstance(rowdata, dict):
            possible = ['UPC', 'Upc', 'Upc/Ean', 'ID']
            for p in possible:
                if p in rowdata and rowdata.get(p):
                    key = str(rowdata.get(p))
                    break
            if not key:
                for v in rowdata.values():
                    if v is not None and str(v).strip():
                        key = str(v)
                        break
        preview = key if key else 'sin referencia'
        entry = {'time': ts, 'action': action, 'key': preview}
        if tab_name == 'Base General':
            self.recent_general.insert(0, entry)
            self.recent_general = self.recent_general[:10]
        else:
            self.recent_norma_024.insert(0, entry)
            self.recent_norma_024 = self.recent_norma_024[:10]

# --- Funcion que permite buscar por medio del Scanner --- #
    def perform_search(self):
        q = self.search_entry.get().strip()
        scope = self.scope_menu.get()
        if not q:
            return

        q_norm = q.strip().upper()
        is_upc_search = bool(re.fullmatch(r'\d+', q_norm))

        targets = []
        if scope in ('Todas', 'Base General'):
            targets.append(('Base General', self.tree_general, self.df_general))
        if scope in ('Todas', 'Base Norma 024'):
            targets.append(('Base Norma 024', self.tree_norma_024, self.df_norma_024))

        for name, tree, df in targets:
            if df is None or df.empty:
                continue

            for iid in tree.get_children():
                tree.delete(iid)

            cols = list(df.columns)
            tree.config(columns=cols)
            for col in cols:
                tree.heading(col, text=col)
                tree.column(col, width=120, anchor='w')

            # --- Búsqueda ---
            if is_upc_search:
                upc_col = self._find_column(df, ['UPC', 'Upc', 'Upc/Ean'])
                cat_col = self._find_column(df, ['CATEGORIA', 'Categoria', 'category'])
                asi_col = self._find_column(df, ['ASIGNACIÓN', 'ASINGACION', 'Asignación', 'Asignacion'])

                mask_upc = pd.Series([False] * len(df))
                mask_cat = pd.Series([False] * len(df))
                mask_asi = pd.Series([False] * len(df))

                if upc_col:
                    series_upc = df[upc_col].astype(str).fillna('').str.strip().str.replace(r'\.0$', '', regex=True).str.upper()
                    mask_upc = series_upc.str.contains(q_norm, na=False)

                if cat_col:
                    series_cat = df[cat_col].astype(str).fillna('').str.strip().str.upper()
                    mask_cat = series_cat.str.contains(q_norm, na=False)
                
                if asi_col:
                    series_asi = df[asi_col].astype(str).fillna('').str.strip().str.replace(r'\.0$', '', regex=True).str.upper()
                    mask_asi = series_asi.str.contains(q_norm, na=False)

                mask = mask_upc | mask_cat | mask_asi
                df_filtered = df[mask]
            else:
                mask = df.apply(lambda row: row.astype(str).str.upper().str.contains(q_norm, na=False).any(), axis=1)
                df_filtered = df[mask]

            # --- Mostrar resultados con check de revisados ---
            for idx, row in df_filtered.iterrows():
                values = [self._format_cell(row.get(c, '')) for c in cols]
                tag = self._row_tag(row)

                upc_val = ''
                if is_upc_search and upc_col:
                    upc_val = str(row.get(upc_col, '')).strip().upper()

                    # Verificar si ya está revisado
                    if name == "Base General" and upc_val in self.reviewed_general:
                        values[0] = "✅ " + values[0]
                    elif name == "Base Norma 024" and upc_val in self.reviewed_norma_024:
                        values[0] = "✅ " + values[0]
                    else:
                        # Nuevo revisado → agregar y registrar
                        self.marcar_como_revisado(upc_val, name)

                tree.insert('', 'end', iid=str(idx), values=values, tags=(tag,))

    def clear_search(self):
        for entry in [self.search_categoria, self.search_upc, self.search_asignacion, self.search_medidas]:
            try:
                entry.delete(0, 'end')
            except Exception:
                pass
        self.lbl_resultados.configure(text='Resultados: 0')
        self._refresh_trees()

    def next_page(self):
        tab = self.tabs.get()
        if tab == 'Base General':
            if (self.current_page_general + 1) * self.page_size < len(self.df_general):
                self.current_page_general += 1
                self._refresh_trees()
        else:
            if (self.current_page_norma_024 + 1) * self.page_size < len(self.df_norma_024):
                self.current_page_norma_024 += 1
                self._refresh_trees()

    def prev_page(self):
        tab = self.tabs.get()
        if tab == 'Base General' and self.current_page_general > 0:
            self.current_page_general -= 1
            self._refresh_trees()
        elif tab == 'Base Norma 024' and self.current_page_norma_024 > 0:
            self.current_page_norma_024 -= 1
            self._refresh_trees()

    def on_add(self):
        tab = self.tabs.get()
        if tab == 'Base General':
            df = self.df_general
            cols = list(df.columns) if not df.empty else []
            self._open_row_editor(df, None, cols, tab, mode='add')
        else:
            df = self.df_norma_024
            cols = list(df.columns) if not df.empty else []
            self._open_row_editor(df, None, cols, tab, mode='add')

    def on_edit(self):
        tab = self.tabs.get()
        if tab == 'Base General':
            tree = self.tree_general
            df = self.df_general
        else:
            tree = self.tree_norma_024
            df = self.df_norma_024

        sel = tree.selection()
        if not sel:
            messagebox.showinfo('Editar', 'Selecciona una fila para editar')
            return

        iid = int(sel[0])
        self._open_row_editor(df, iid, list(df.columns), tab, mode='edit')

    def _open_row_editor(self, df, row_idx, columns, tab_name, mode='edit'):
        editor = ctk.CTkToplevel(self)
        editor.geometry('700x600')
        editor.transient(self)
        editor.grab_set()
        editor.title('Añadir Fila' if mode == 'add' else 'Editar Fila')

        # Cabecera
        header = ctk.CTkFrame(editor, fg_color='#ECD925', height=48)
        header.pack(fill='x', pady=(0,8))
        ctk.CTkLabel(header, text=('Añadir Fila' if mode == 'add' else 'Editar Fila'),
                     font=('Inter', 14, 'bold'), text_color='#282828').pack(side='left', padx=8)

        body = ctk.CTkScrollableFrame(editor, fg_color='#FFFFFF')
        body.pack(fill='both', expand=True, padx=10, pady=5)

        entries = {}
        for i, col in enumerate(columns):
            ctk.CTkLabel(body, text=col, text_color='#282828').grid(row=i, column=0, sticky='w', pady=4)
            val = str(df.at[row_idx, col]) if (row_idx is not None and col in df.columns) else ''
            e = ctk.CTkEntry(body, fg_color='#F7F7F7', text_color='#000000')
            e.insert(0, val)
            e.grid(row=i, column=1, sticky='we', padx=6, pady=4)
            entries[col] = e
        body.grid_columnconfigure(1, weight=1)

        # Botones
        btn_frame = ctk.CTkFrame(editor, fg_color='transparent')
        btn_frame.pack(fill='x', pady=(10, 6))
        ctk.CTkButton(btn_frame, text='Cancelar', command=editor.destroy,
                      fg_color='#282828', text_color='white').pack(side='right', padx=6)
        ctk.CTkButton(btn_frame, text='Guardar',
                      command=lambda: self._save_row(editor, df, row_idx, entries, columns, tab_name, mode),
                      fg_color='#27AE60', text_color='white').pack(side='right', padx=6)

    def _save_to_file(self, tab_name):
        if tab_name == 'Base General':
            path = self.path_base_general
            df = self.df_general
            excel_path = self.cfg.get("base_general_excel")  # 🔄 Ruta del Excel original
        else:
            path = self.path_base_norma_024
            df = self.df_norma_024
            excel_path = self.cfg.get("base_norma_024_excel")  # 🔄 Ruta del Excel original
            
        try:
            os.makedirs('data', exist_ok=True)
            
            # 🔄 GUARDAR EN JSON (manteniendo estructura original)
            if os.path.exists(path):
                with open(path, 'r', encoding='utf-8') as f:
                    original_data = json.load(f)
                
                if isinstance(original_data, dict) and len(original_data) == 1:
                    main_key = list(original_data.keys())[0]
                    data_to_save = {main_key: df.to_dict(orient='records')}
                else:
                    data_to_save = df.to_dict(orient='records')
            else:
                data_to_save = df.to_dict(orient='records')
            
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(data_to_save, f, indent=4, ensure_ascii=False, sort_keys=False)
                
            print(f"✅ Archivo JSON guardado: {path}")
            
            # 🔄 GUARDAR EN EXCEL ORIGINAL
            if excel_path and os.path.exists(excel_path):
                self._guardar_en_excel(df, excel_path, tab_name)
            else:
                print(f"⚠️ No se encontró la ruta del Excel original para {tab_name}")
                
        except Exception as e:
            messagebox.showerror("Error al guardar", f"No se pudo guardar el archivo:\n{e}")

    def _guardar_en_excel(self, df, excel_path, tab_name):
        """Guarda el DataFrame en el archivo Excel original manteniendo formato y estructura"""
        try:
            # Cargar el workbook existente
            book = load_workbook(excel_path)
            
            # Determinar el nombre de la hoja basado en el archivo original
            sheet_name = self._obtener_nombre_hoja(excel_path, tab_name)
            
            # Si la hoja existe, limpiarla y escribir nuevos datos
            if sheet_name in book.sheetnames:
                ws = book[sheet_name]
                # Limpiar la hoja excepto los encabezados
                ws.delete_rows(2, ws.max_row)  # Conserva la fila 1 (encabezados)
                
                # Escribir los datos
                for idx, row in df.iterrows():
                    row_data = [row.get(col, '') for col in df.columns]
                    ws.append(row_data)
            else:
                # Si no existe la hoja, crear una nueva
                ws = book.create_sheet(sheet_name)
                # Escribir encabezados
                ws.append(list(df.columns))
                # Escribir datos
                for idx, row in df.iterrows():
                    row_data = [row.get(col, '') for col in df.columns]
                    ws.append(row_data)
            
            # Guardar el workbook
            book.save(excel_path)
            print(f"✅ Archivo Excel actualizado: {excel_path} (Hoja: {sheet_name})")
            
        except Exception as e:
            print(f"❌ Error al guardar Excel: {e}")

    def _obtener_nombre_hoja(self, excel_path, tab_name):
        """Obtiene el nombre de la hoja del archivo Excel original"""
        try:
            # Intentar leer el archivo Excel para obtener el nombre de la primera hoja
            with pd.ExcelFile(excel_path) as xls:
                sheet_names = xls.sheet_names
                if sheet_names:
                    return sheet_names[0]  # Devolver la primera hoja
        except:
            pass
        
        # Nombres por defecto basados en el tipo de base
        if tab_name == 'Base General':
            return 'BASE_GENERAL'
        else:
            return 'NORMA_024'

    def _save_row(self, editor, df, row_idx, entries, columns, tab_name, mode):
        rowdata = {col: entries[col].get() for col in columns}

        if tab_name == 'Base General':
            tree = self.tree_general
        else:
            tree = self.tree_norma_024

        if mode == 'add':
            new_idx = len(df)
            df.loc[new_idx] = rowdata
            values = [self._format_cell(rowdata.get(c, "")) for c in columns]
            tag = self._row_tag(pd.Series(rowdata))
            tree.insert("", "end", iid=str(new_idx), values=values, tags=(tag,))
        else:
            for col, val in rowdata.items():
                df.at[row_idx, col] = val
            values = [self._format_cell(rowdata.get(c, "")) for c in columns]
            tag = self._row_tag(pd.Series(rowdata))
            tree.item(str(row_idx), values=values, tags=(tag,))

        self._record_recent(tab_name, "Añadido" if mode == 'add' else "Editado", rowdata)
        
        self.dirty = True

        # 🔄 GUARDAR EN JSON Y EXCEL ORIGINAL
        self._save_to_file(tab_name)
        self._notify_dashboards()
        
        # 🔄 ACTUALIZAR INTERFAZ LOCAL
        self._refresh_trees()
        self.update_stats()
        
        editor.destroy()

    def on_delete(self):
        tab = self.tabs.get()
        if tab == 'Base General':
            tree = self.tree_general
            df = self.df_general
        else:
            tree = self.tree_norma_024
            df = self.df_norma_024

        sel = tree.selection()
        if not sel:
            messagebox.showinfo('Borrar', 'Selecciona una fila para borrar')
            return
        if not messagebox.askyesno('Confirmar', '¿Borrar la(s) fila(s) seleccionada(s)?'):
            return

        for iid in sel:
            idx = int(iid)
            if idx in df.index:
                df.drop(index=idx, inplace=True)
            tree.delete(iid)

        df.reset_index(drop=True, inplace=True)
        

        self.dirty = True
        self.on_delete()

        # 🔄 GUARDAR EN JSON Y EXCEL ORIGINAL
        self._save_to_file(tab)
        self._notify_dashboards()
        
        # 🔄 ACTUALIZAR INTERFAZ LOCAL
        self._refresh_trees()
        self.update_stats()

    def destroy(self):
        # Guardar solo si hay cambios
        if getattr(self, 'dirty', False):
            messagebox.showinfo("Guardando cambios", "Se detectaron cambios. Guardando archivos...")
            self._save_to_file("Base General")
            self._save_to_file("Base Norma 024")
            self._notify_dashboards()
        
        super().destroy()

    def perform_multi_search(self):
        """Permite buscar por múltiples campos específicos: Categoría, UPC, Asignación o Medidas."""
        query_categoria = self.search_categoria.get().strip().upper()
        query_upc = self.search_upc.get().strip().upper()
        query_asignacion = self.search_asignacion.get().strip().upper()
        query_medidas = self.search_medidas.get().strip().upper()

        total_resultados = 0
        targets = [
            ('Base General', self.tree_general, self.df_general),
            ('Base Norma 024', self.tree_norma_024, self.df_norma_024)
        ]

        for name, tree, df in targets:
            if df.empty:
                continue

            for iid in tree.get_children():
                tree.delete(iid)

            cols = list(df.columns)
            tree.config(columns=cols)
            for col in cols:
                tree.heading(col, text=col)
                tree.column(col, width=120, anchor='center')

            mask = pd.Series([True] * len(df))

            def col(name_variants):
                return self._find_column(df, name_variants)

            # Filtrar según las entradas
            if query_categoria:
                col_cat = col(['CATEGORIA', 'CATEGORÍA', 'Categoria'])
                if col_cat:
                    mask &= df[col_cat].astype(str).str.upper().str.contains(query_categoria, na=False)

            if query_upc:
                col_upc = col(['UPC', 'Upc', 'Upc/Ean'])
                if col_upc:
                    mask &= df[col_upc].astype(str).str.upper().str.contains(query_upc, na=False)

            if query_asignacion:
                col_asi = col(['ASIGNACION', 'ASIGNACIÓN', 'Asignación'])
                if col_asi:
                    mask &= df[col_asi].astype(str).str.upper().str.contains(query_asignacion, na=False)

            if query_medidas:
                col_med = col(['MEDIDAS', 'Medidas'])
                if col_med:
                    mask &= df[col_med].astype(str).str.upper().str.contains(query_medidas, na=False)

            df_filtered = df[mask]
            total_resultados += len(df_filtered)

            for idx, row in df_filtered.iterrows():
                values = [self._format_cell(row.get(c, "")) for c in cols]
                tag = self._row_tag(row)
                tree.insert('', 'end', iid=str(idx), values=values, tags=(tag,))

        self.lbl_resultados.configure(text=f'Resultados: {total_resultados}')

    def on_save(self, show_message=True):
        """Guarda respaldo JSON y configuración."""
        try:
            os.makedirs('data', exist_ok=True)
            with open(self.path_base_general, 'w', encoding='utf-8') as f:
                json.dump(self.df_general.to_dict(orient='records'), f, indent=4, ensure_ascii=False)
            with open(self.path_base_norma_024, 'w', encoding='utf-8') as f:
                json.dump(self.df_norma_024.to_dict(orient='records'), f, indent=4, ensure_ascii=False)

            save_config(self.cfg)
            if show_message:
                messagebox.showinfo('Guardado', 'Datos guardados correctamente.')

        except Exception as e:
            messagebox.showerror('Error', f'No se pudo guardar:\n{e}')

    def _notify_dashboards(self):
        # intentar notificar ventanas Dashboard abiertas para que refresquen
        try:
            import Dashboard
            root = tk._default_root
            if not root:
                return
            for w in root.winfo_children():
                try:
                    if isinstance(w, Dashboard.VentanaDashboard):
                        try:
                            w.actualizar_dashboard()
                        except Exception:
                            pass
                except Exception:
                    continue
        except Exception:
            pass
        # intentar notificar ventanas Dashboard abiertas para que refresquen
        try:
            import Dashboard
            root = tk._default_root
            if not root:
                return
            for w in root.winfo_children():
                try:
                    if isinstance(w, Dashboard.VentanaDashboard):
                        try:
                            w.actualizar_dashboard()
                        except Exception:
                            pass
                except Exception:
                    continue
        except Exception:
            pass

