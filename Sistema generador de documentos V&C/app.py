# -- SISTEMA V&C - GENERADOR DE DICTÁMENES -- #
import os, re
import sys
import uuid
import shutil
import json
import pandas as pd
import customtkinter as ctk
from tkinter import filedialog, messagebox, simpledialog
from tkinter import ttk
import tkinter as tk
import tkinter.font as tkfont
import threading
import subprocess
import importlib
import importlib.util
from datetime import datetime
try:
    from tkcalendar import Calendar
except Exception:
    Calendar = None
import folio_manager
from plantillaPDF import cargar_tabla_relacion
import time
import platform
import unicodedata
from PyPDF2 import PdfReader
from PIL import Image

# ---------- ESTILO VISUAL V&C ---------- #
STYLE = {
    "primario": "#ECD925",
    "secundario": "#282828",
    "exito": "#008D53",
    "advertencia": "#ff1500",
    "peligro": "#d74a3d",
    "fondo": "#F8F9FA",
    # Make surface match the background to avoid visible white cards/borders
    "surface": "#F8F9FA",
    "texto_oscuro": "#282828",
    "texto_claro": "#ffffff",
    # Border color should match surface so thin borders are not visible
    "borde": "#F8F9FA"
}

FONT_TITLE = ("Inter", 22, "bold")
FONT_SUBTITLE = ("Inter", 17, "bold")
FONT_LABEL = ("Inter", 13)
FONT_SMALL = ("Inter", 12)

# Ruta base compatible con PyInstaller (.exe)
if getattr(sys, 'frozen', False):
    BASE_DIR = getattr(sys, '_MEIPASS', os.path.abspath("."))
else:
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))

if getattr(sys, 'frozen', False):
    APP_DIR = os.path.dirname(sys.executable)
else:
    APP_DIR = os.path.abspath(os.path.dirname(__file__))

DATA_DIR = os.getenv('IMAGENESVC_DATA_DIR')
if DATA_DIR:
    DATA_DIR = os.path.abspath(DATA_DIR)
else:
    if getattr(sys, 'frozen', False):
        # Prefer a `data` folder next to the executable when running the
        # bundled .exe. Fall back to AppData if the local folder cannot be
        # created for permission reasons.
        preferred_data = os.path.join(APP_DIR, 'data')
        try:
            os.makedirs(preferred_data, exist_ok=True)
        except Exception:
            preferred_data = os.path.join(os.path.expanduser("~"), 'AppData', 'Local', 'Sistema_Generador_VC', 'data')
            try:
                os.makedirs(preferred_data, exist_ok=True)
            except Exception:
                # Last-resort fallback
                preferred_data = os.path.join(APP_DIR, 'data')

        def _bundle_candidates(subpath):
            candidates = [
                os.path.join(APP_DIR, subpath),
                os.path.join(APP_DIR, '_internal', subpath),
                os.path.join(BASE_DIR, subpath),
                os.path.join(BASE_DIR, '_internal', subpath),
            ]
            return [p for p in candidates if os.path.exists(p)]

        bundled_data = None
        cands = _bundle_candidates('data')
        if cands:
            bundled_data = cands[0]

        # Debug info into the chosen preferred_data
        try:
            bdbg = os.path.join(preferred_data, 'startup_exe_debug.log')
            with open(bdbg, 'a', encoding='utf-8') as _bd:
                _bd.write(f"bundle_candidates={cands}\nchosen={bundled_data}\npreferred={preferred_data}\n")
        except Exception:
            pass

        try:
            FORCE_REFRESH = os.environ.get('IMAGENESVC_FORCE_REFRESH') == '1'
            if bundled_data and os.path.exists(bundled_data):
                for root, dirs, files in os.walk(bundled_data):
                    rel = os.path.relpath(root, bundled_data)
                    target_root = os.path.join(preferred_data, rel) if rel != '.' else preferred_data
                    os.makedirs(target_root, exist_ok=True)
                    for f in files:
                        src = os.path.join(root, f)
                        dst = os.path.join(target_root, f)
                        try:
                            copy_it = False
                            if FORCE_REFRESH:
                                copy_it = True
                            elif not os.path.exists(dst):
                                copy_it = True
                            else:
                                try:
                                    dst_size = os.path.getsize(dst)
                                except Exception:
                                    dst_size = None
                                if dst_size == 0:
                                    copy_it = True
                                else:
                                    try:
                                        src_mtime = os.path.getmtime(src)
                                        dst_mtime = os.path.getmtime(dst)
                                        if src_mtime > dst_mtime + 1:
                                            copy_it = True
                                    except Exception:
                                        pass

                            if copy_it:
                                shutil.copy2(src, dst)
                        except Exception:
                            pass
        except Exception:
            pass

        DATA_DIR = os.path.abspath(preferred_data)
    else:
        DATA_DIR = os.path.join(APP_DIR, 'data')
try:
    os.environ['FOLIO_DATA_DIR'] = DATA_DIR
except Exception:
    pass

# DEBUG: volcar información de rutas y existencia de archivos clave al iniciar
try:
    dbg_files = [
        'Clientes.json', 'tabla_de_relacion.json', 'Firmas.json',
        'folio_counter.json', 'historial_visitas.json', 'pending_folios.json'
    ]
    dbg_path = None
    try:
        dbg_path = os.path.join(DATA_DIR, 'startup_exe_debug.log')
    except Exception:
        dbg_path = os.path.join(os.path.expanduser('~'), 'startup_exe_debug.log')

    with open(dbg_path, 'a', encoding='utf-8') as dbg:
        import time
        dbg.write('\n===== STARTUP DEBUG: ' + time.strftime('%Y-%m-%d %H:%M:%S') + ' =====\n')
        try:
            dbg.write(f"frozen={getattr(sys,'frozen',False)}\n")
            dbg.write(f"BASE_DIR={BASE_DIR}\n")
            dbg.write(f"APP_DIR={APP_DIR}\n")
            dbg.write(f"DATA_DIR={DATA_DIR}\n")
            dbg.write(f"CWD={os.path.abspath(os.getcwd())}\n")
            dbg.write(f"ENV.IMAGENESVC_DATA_DIR={os.environ.get('IMAGENESVC_DATA_DIR')}\n")
            dbg.write(f"ENV.FOLIO_DATA_DIR={os.environ.get('FOLIO_DATA_DIR')}\n")
        except Exception as e:
            dbg.write(f"ERROR writing basic info: {e}\n")

        for fn in dbg_files:
            try:
                p = os.path.join(DATA_DIR, fn)
                exists = os.path.exists(p)
                size = os.path.getsize(p) if exists and os.path.isfile(p) else None
                dbg.write(f"{fn}: exists={exists} size={size} path={p}\n")
            except Exception as e:
                dbg.write(f"{fn}: error checking: {e}\n")

        # list top-level data files
        try:
            for root, dirs, files in os.walk(DATA_DIR):
                rel = os.path.relpath(root, DATA_DIR)
                dbg.write(f"DIR: {rel}\n")
                for f in files:
                    try:
                        fp = os.path.join(root, f)
                        dbg.write(f"  - {f} ({os.path.getsize(fp)} bytes)\n")
                    except Exception:
                        dbg.write(f"  - {f} (size error)\n")
        except Exception as e:
            dbg.write(f"ERROR walking DATA_DIR: {e}\n")

        dbg.write('===== END STARTUP DEBUG =====\n')
except Exception:
    pass


class SistemaDictamenesVC(ctk.CTk):
    # --- PAGINACIÓN HISTORIAL ---
    HISTORIAL_PAGINA_ACTUAL = 1
    HISTORIAL_REGS_POR_PAGINA = 100

    def __init__(self):
        super().__init__()

        # Configuración general
        self.title("Generador de Dictámenes")
        self.geometry("1275x600")
        self.minsize(1275, 600)
        # Establecer icono de la ventana si existe
        try:
            icon_path = os.path.join(APP_DIR, 'img', 'icono.ico')
            if os.path.exists(icon_path):
                try:
                    self.iconbitmap(icon_path)
                except Exception:
                    try:
                        img_icon = tk.PhotoImage(file=icon_path)
                        self.iconphoto(False, img_icon)
                    except Exception:
                        pass
        except Exception:
            pass
        ctk.set_appearance_mode("light")
        self.configure(fg_color=STYLE["fondo"])

        # Variables de estado
        self.archivo_excel_cargado = None
        self.archivo_json_generado = None
        self.json_filename = None
        self.generando_dictamenes = False
        # Progreso: watcher para animar barra cuando el generador no emite actualizaciones frecuentes
        self._progress_watcher_thread = None
        self._progress_watcher_stop_event = None
        self._last_progress_value = 0.0
        self._last_progress_ts = 0.0
        self.clientes_data = []
        self.cliente_seleccionado = None
        self.domicilio_seleccionado = None
        self.archivo_etiquetado_json = None
        # Flag para edición de clientes desde el formulario de Reportes
        self.editing_cliente_rfc = None

        # Variables para nueva visita
        self.current_folio = "000001"

        # ===== NUEVAS VARIABLES PARA HISTORIAL =====
        self.historial_data = []
        self.historial_data_original = []
        self.historial_path = os.path.join(DATA_DIR, "historial_visitas.json")
        
        # INICIALIZAR self.historial COMO DICCIONARIO
        self.historial = {"visitas": []}

        # ===== NUEVA VARIABLE PARA FOLIOS POR VISITA =====
        data_dir = DATA_DIR
        if getattr(sys, 'frozen', False):
            try:
                # asegúrate de que exista data (sincronizar desde bundle cuando procede)
                if not os.path.exists(data_dir):
                    os.makedirs(data_dir, exist_ok=True)

                FORCE_REFRESH = os.environ.get('IMAGENESVC_FORCE_REFRESH') == '1'
                embedded_data = os.path.join(BASE_DIR, 'data')
                if os.path.exists(embedded_data):
                    for root, dirs, files in os.walk(embedded_data):
                        rel = os.path.relpath(root, embedded_data)
                        target_root = os.path.join(data_dir, rel) if rel != '.' else data_dir
                        os.makedirs(target_root, exist_ok=True)
                        for f in files:
                            sfile = os.path.join(root, f)
                            dfile = os.path.join(target_root, f)
                            try:
                                copy_it = False
                                if FORCE_REFRESH:
                                    copy_it = True
                                elif not os.path.exists(dfile):
                                    copy_it = True
                                else:
                                    try:
                                        dsize = os.path.getsize(dfile)
                                    except Exception:
                                        dsize = None
                                    if dsize == 0:
                                        copy_it = True
                                    else:
                                        try:
                                            sm = os.path.getmtime(sfile)
                                            dm = os.path.getmtime(dfile)
                                            if sm > dm + 1:
                                                copy_it = True
                                        except Exception:
                                            pass

                                if copy_it:
                                    shutil.copy2(sfile, dfile)
                            except Exception:
                                pass

                # Asegurar que otros recursos estén disponibles junto al exe.
                resource_folders = [
                    'Plantillas PDF',
                    'Pegado de Evidenvia Fotografica',
                    'Documentos Inspeccion',
                    'Firmas',
                    'img'
                ]
                for rf in resource_folders:
                    # buscar candidatos en BASE_DIR y en BASE_DIR/_internal
                    candidates = [os.path.join(BASE_DIR, rf), os.path.join(BASE_DIR, '_internal', rf)]
                    src = None
                    for c in candidates:
                        if os.path.exists(c):
                            src = c
                            break
                    if not src:
                        continue
                    dst = os.path.join(APP_DIR, rf)
                    os.makedirs(dst, exist_ok=True)
                    # copiar solo archivos/carpetas faltantes
                    for root, dirs, files in os.walk(src):
                        rel = os.path.relpath(root, src)
                        target_root = os.path.join(dst, rel) if rel != '.' else dst
                        os.makedirs(target_root, exist_ok=True)
                        for f in files:
                            sfile = os.path.join(root, f)
                            dfile = os.path.join(target_root, f)
                            try:
                                if (not os.path.exists(dfile)) or (os.path.exists(dfile) and os.path.getsize(dfile) == 0):
                                    shutil.copy2(sfile, dfile)
                            except Exception:
                                pass
            except Exception:
                try:
                    os.makedirs(data_dir, exist_ok=True)
                except Exception:
                    pass
        else:
            os.makedirs(data_dir, exist_ok=True)
        
        self.folios_visita_path = os.path.join(data_dir, "folios_visitas")
        os.makedirs(self.folios_visita_path, exist_ok=True)
        # Cargar reservas persistentes
        self.pending_folios = []
        try:
            self._load_pending_folios()
            # Iniciar watcher para detectar cambios externos en pending_folios.json
            self._start_pending_folios_watcher()
        except Exception:
            self.pending_folios = []
        # Directorio donde están los generadores/documentos (ReportLab, tablas, etc.)
        self.documentos_dir = os.path.join(BASE_DIR, "Documentos Inspeccion")

        # ===== NUEVA ESTRUCTURA DE NAVEGACIÓN =====
        self.crear_navegacion()
        self.crear_area_contenido()

        # ===== FOOTER =====
        self.crear_footer()

        # Cargar configuración de exportación Excel (persistente)
        self._cargar_config_exportacion()

        # Cargar clientes al iniciar
        self.cargar_clientes_desde_json()
        self.cargar_ultimo_folio()
        try:
            # Mostrar ventana de login / aplicar permisos según usuarios
            self._ensure_user_authenticated()
        except Exception:
            pass
        try:
            self._generar_datos_exportable()
        except Exception:
            pass
        
    # ----------------- Overlay de Acciones (botones interactivos) -----------------
    def _create_actions_overlay(self, parent, actions_col=None):
        """Crea un frame flotante con botones que se posiciona sobre la columna 'Acciones'."""
        try:
            overlay = tk.Frame(parent, bg=STYLE.get('fondo', '#fff'))
            overlay.place_forget()
            self._actions_overlay = overlay

            # Botones: Folios, Archivos, Editar, Borrar
            try:
                self._btn_folios = ctk.CTkButton(overlay, text="Folios", width=80, height=26, corner_radius=6, command=lambda: self._overlay_action('folios'))
                self._btn_archivos = ctk.CTkButton(overlay, text="Archivos", width=80, height=26, corner_radius=6, command=lambda: self._overlay_action('archivos'))
                self._btn_editar = ctk.CTkButton(overlay, text="Editar", width=60, height=26, corner_radius=6, command=lambda: self._overlay_action('editar'))
                self._btn_borrar = ctk.CTkButton(overlay, text="Borrar", width=60, height=26, corner_radius=6, fg_color=STYLE['peligro'], command=lambda: self._overlay_action('borrar'))
            except Exception:
                self._btn_folios = tk.Button(overlay, text="Folios", width=8, command=lambda: self._overlay_action('folios'))
                self._btn_archivos = tk.Button(overlay, text="Archivos", width=8, command=lambda: self._overlay_action('archivos'))
                self._btn_editar = tk.Button(overlay, text="Editar", width=6, command=lambda: self._overlay_action('editar'))
                self._btn_borrar = tk.Button(overlay, text="Borrar", width=6, command=lambda: self._overlay_action('borrar'))

            # Empacar botones para ocupar el ancho del overlay equitativamente
            for w in (self._btn_folios, self._btn_archivos, self._btn_editar, self._btn_borrar):
                try:
                    w.pack(side='left', fill='both', expand=True, padx=(2, 2), pady=2)
                except Exception:
                    w.pack(side='left', padx=(2, 2), pady=2)

            # Overlay bindings removed to disable hover animation showing action buttons
            # (kept functions for compatibility but not bound)
        except Exception:
            pass

    def _on_tree_motion(self, event):
        try:
            iid = self.hist_tree.identify_row(event.y)
            if not iid:
                self._hide_actions_overlay()
                return
            bbox = self.hist_tree.bbox(iid, column=self.hist_tree['columns'][-1])
            if not bbox:
                self._hide_actions_overlay()
                return
            x, y, w, h = bbox
            tree_x = self.hist_tree.winfo_x()
            tree_y = self.hist_tree.winfo_y()
            abs_x = tree_x + x
            abs_y = tree_y + y
            try:
                self._actions_overlay.place(x=abs_x, y=abs_y, width=w, height=h)
                self._actions_overlay.lift()
                self._overlay_iid = iid
            except Exception:
                pass
        except Exception:
            pass

    def _hide_actions_overlay(self):
        try:
            if hasattr(self, '_actions_overlay'):
                self._actions_overlay.place_forget()
                self._overlay_iid = None
        except Exception:
            pass

    def _on_tree_click(self, event):
        try:
            col = self.hist_tree.identify_column(event.x)
            iid = self.hist_tree.identify_row(event.y)
            if not iid:
                self._hide_actions_overlay()
                return
            last_col = f"#{len(self.hist_tree['columns'])}"
            if col == last_col:
                bbox = self.hist_tree.bbox(iid, column=self.hist_tree['columns'][-1])
                if bbox:
                    x, y, w, h = bbox
                    tree_x = self.hist_tree.winfo_x()
                    tree_y = self.hist_tree.winfo_y()
                    abs_x = tree_x + x
                    abs_y = tree_y + y
                    self._actions_overlay.place(x=abs_x, y=abs_y, width=w, height=h)
                    self._actions_overlay.lift()
                    self._overlay_iid = iid
            else:
                self._hide_actions_overlay()
        except Exception:
            pass

    def _overlay_action(self, action):
        iid = getattr(self, '_overlay_iid', None)
        if not iid:
            return
        reg = self._hist_map.get(iid)
        if not reg:
            return
        try:
            if action == 'folios':
                self.descargar_folios_visita(reg)
            elif action == 'archivos':
                self.mostrar_opciones_documentos(reg)
            elif action == 'editar':
                self.hist_editar_registro(reg)
            elif action == 'borrar':
                self.hist_eliminar_registro(reg)
        except Exception:
            pass

        # --------------------------- ICONO ---------------------------- #
        def resource_path(relative_path):
            try:
                base_path = sys._MEIPASS
            except Exception:
                base_path = os.path.abspath(".")
            # comprobar además la posible carpeta _internal donde PyInstaller coloca datas
            candidates = [os.path.join(base_path, relative_path), os.path.join(base_path, '_internal', relative_path)]
            for p in candidates:
                if os.path.exists(p):
                    return p
            # por defecto, devolver primer candidato
            return candidates[0]

        try:
            icon_path = resource_path("img/icono.ico")
            if os.path.exists(icon_path):
                self.iconbitmap(icon_path)
                print(f"🟡 Icono cargado: {icon_path}")
            else:
                print("⚠ No se encontró icono.ico")
        except Exception as e:
            print(f"⚠ Error cargando icono.ico: {e}")

    def centerwindow(self):
        self.update_idletasks()
        ancho_ventana = self.winfo_width()
        alto_ventana = self.winfo_height()
        ancho_pantalla = self.winfo_screenwidth()
        alto_pantalla = self.winfo_screenheight()
        x = (ancho_pantalla // 2) - (ancho_ventana // 2)
        y = (alto_pantalla // 2) - (alto_ventana // 2)
        self.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    def crear_navegacion(self):
        """Crea la barra de navegación con botones mejorados"""
        nav_frame = ctk.CTkFrame(self, fg_color=STYLE["surface"], height=48)
        nav_frame.pack(fill="x", padx=20, pady=(0, 0))
        nav_frame.pack_propagate(False)
        
        # Contenedores para botones: izquierda para navegación, derecha para info/acciones de usuario
        self.botones_left_frame = ctk.CTkFrame(nav_frame, fg_color="transparent")
        self.botones_left_frame.pack(side="left", fill="x", expand=True, padx=0, pady=2)
        self.botones_right_frame = ctk.CTkFrame(nav_frame, fg_color="transparent")
        self.botones_right_frame.pack(side="right", padx=0, pady=2)
        
        # Botón Principal con estilo mejorado
        self.btn_principal = ctk.CTkButton(
            self.botones_left_frame,
            text="🏠 Principal",
            command=self.mostrar_principal,
            font=("Inter", 14, "bold"),
            fg_color=STYLE["primario"],
            hover_color="#D4BF22",
            text_color=STYLE["secundario"],
            height=38,
            width=130,
            corner_radius=10,
            border_width=0,
            border_color=STYLE["secundario"]
        )
        self.btn_principal.pack(side="left", padx=(0, 10))
        
        # Botón Historial con estilo mejorado
        self.btn_historial = ctk.CTkButton(
            self.botones_left_frame,
            text="📊 Historial",
            command=self.mostrar_historial,
            font=("Inter", 14, "bold"),
            fg_color=STYLE["surface"],
            hover_color=STYLE["primario"],
            text_color=STYLE["secundario"],
            height=38,
            width=130,
            corner_radius=10,
            border_width=0,
            border_color=STYLE["secundario"]
        )
        self.btn_historial.pack(side="left", padx=(0, 10))

        # Botón Reportes
        self.btn_reportes = ctk.CTkButton(
            self.botones_left_frame,
            text="📑Clientes",
            command=self._cmd_mostrar_clientes,
            font=("Inter", 14, "bold"),
            fg_color=STYLE["surface"],
            hover_color=STYLE["primario"],
            text_color=STYLE["secundario"],
            height=38,
            width=130,
            corner_radius=10,
            border_width=0,
            border_color=STYLE["secundario"]
        )
        # Empaque controlado desde _apply_permissions
        
        # Botón Inspectores
        self.btn_inspectores = ctk.CTkButton(
            self.botones_left_frame,
            text="👥 Inspectores",
            command=self._cmd_mostrar_inspectores,
            font=("Inter", 14, "bold"),
            fg_color=STYLE["surface"],
            hover_color=STYLE["primario"],
            text_color=STYLE["secundario"],
            height=38,
            width=140,
            corner_radius=10,
            border_width=0,
            border_color=STYLE["secundario"]
        )
        # Empaque controlado desde _apply_permissions
        
        # Botón Reportes (ejecutivos)
        self.btn_reportes_ej = ctk.CTkButton(
            self.botones_left_frame,
            text="📈 Reportes",
            command=self.mostrar_reportes_ejecutivo,
            font=("Inter", 14, "bold"),
            fg_color=STYLE["surface"],
            hover_color=STYLE["primario"],
            text_color=STYLE["secundario"],
            height=38,
            width=130,
            corner_radius=10,
            border_width=0,
            border_color=STYLE["secundario"]
        )
        # no empaquetar por defecto: _apply_permissions decidirá si mostrar

        # Botón Reporte EMA (se controla su visibilidad por permisos)
        self.btn_reporte_ema = ctk.CTkButton(
            self.botones_left_frame,
            text="📊 Reporte EMA",
            command=self.descargar_excel_anual,
            font=("Inter", 14, "bold"),
            fg_color=STYLE["surface"],
            hover_color=STYLE["primario"],
            text_color=STYLE["secundario"],
            height=38,
            width=140,
            corner_radius=10,
            border_width=0,
            border_color=STYLE["secundario"]
        )
        
        # Información del sistema (derecha)
        self.lbl_info_sistema = ctk.CTkLabel(
            self.botones_right_frame,
            text="Sistema de Dictámenes - V&C",
            font=("Inter", 12),
            text_color=STYLE["texto_claro"]
        )
        # Botón Cerrar sesión (inicialmente no empaquetado)
        self.btn_logout = ctk.CTkButton(
            self.botones_right_frame,
            text="Cerrar sesión",
            command=self._logout if hasattr(self, '_logout') else (lambda: None),
            font=("Inter", 12),
            fg_color=STYLE["secundario"],
            hover_color="#444444",
            text_color=STYLE["texto_claro"],
            height=34,
            width=140,
            corner_radius=8
        )
        # Nota: no empaquetar aquí; se empaquetará en _apply_permissions cuando haya sesión
        
        # Mostrar etiqueta de info y logout en la derecha cuando haya sesión
        try:
            # ocultar por defecto hasta autenticación; se empaquetará desde _apply_permissions
            self.lbl_info_sistema.pack_forget()
        except Exception:
            pass
        
        # Botón Backup en la barra de navegación (no mostrar por defecto)
        # try:
        #     self.btn_backup = ctk.CTkButton(
        #         botones_frame,
        #         text="💾 Backup",
        #         command=self.hist_hacer_backup,
        #         height=34, width=110, corner_radius=8,
        #   except Exception:
        #      return {}
        # except Exception:
        #     self.btn_backup = None

        # self.lbl_info_sistema.pack(side="right")

    def crear_area_contenido(self):
        """Crea el área de contenido donde se muestran las secciones"""
        # Frame contenedor del contenido
        self.contenido_frame = ctk.CTkFrame(self, fg_color="transparent")
        self.contenido_frame.pack(fill="both", expand=True, padx=20, pady=(0, 0))
        
        # Frame para el contenido principal
        self.frame_principal = ctk.CTkFrame(self.contenido_frame, fg_color="transparent")
        
        # Frame para el historial
        self.frame_historial = ctk.CTkFrame(self.contenido_frame, fg_color="transparent")

        # Frame para reportes
        self.frame_reportes = ctk.CTkFrame(self.contenido_frame, fg_color="transparent")
        # Frame exclusivo para ejecutivos: reportes
        self.frame_reportes_ejecutivo = ctk.CTkFrame(self.contenido_frame, fg_color="transparent")
        # Frame para inspectores
        self.frame_inspectores = ctk.CTkFrame(self.contenido_frame, fg_color="transparent")
        
        # Construir el contenido de cada sección
        self._construir_tab_principal(self.frame_principal)
        self._construir_tab_historial(self.frame_historial)
        self._construir_tab_clientes(self.frame_reportes)
        # construir pestaña de reportes ejecutivos
        try:
            self._construir_tab_reportes_ejecutivo(self.frame_reportes_ejecutivo)
        except Exception:
            pass
        self._construir_tab_inspectores(self.frame_inspectores)
        
        # Mostrar la sección principal por defecto
        self.mostrar_principal()

    # ----------------- Autenticación y permisos -----------------
    def _load_users(self):
        """Carga `data/usuarios.json`. Si no existe crea un archivo por defecto."""
        fn = os.path.join(DATA_DIR, "usuarios.json")
        try:
            with open(fn, 'r', encoding='utf-8') as f:
                return json.load(f).get('users', [])
        except Exception:
            # crear archivo por defecto a partir del bundle 'data/usuarios.json' si existe
            try:
                default_path = os.path.join(BASE_DIR, 'data', 'usuarios.json')
                if os.path.exists(default_path):
                    with open(default_path, 'r', encoding='utf-8') as f:
                        return json.load(f).get('users', [])
            except Exception:
                pass
        return []

    def _ensure_user_authenticated(self):
        """Muestra el diálogo de login y aplica permisos en UI."""
        self.users = self._load_users()
        self.current_user = None
        self.current_role = None
        # Mostrar un diálogo modal de login
        try:
            self._show_login_dialog()
        except Exception:
            pass
        # Aplicar permisos según rol (si no autenticado, ocultar tabs)
        try:
            self._apply_permissions()
        except Exception:
            pass

    def _find_user(self, username):
        if not username:
            return None
        uname = username.strip().lower()
        for u in getattr(self, 'users', []):
            # aceptar tanto el username como el nombre completo (case-insensitive)
            if str(u.get('username', '')).lower() == uname:
                return u
            if str(u.get('name', '')).strip().lower() == uname:
                return u
        return None

    def _show_login_dialog(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Acceso al sistema")
        # modal + transient para que se vea encima de la app
        try:
            dlg.transient(self)
        except Exception:
            pass
        dlg.grab_set()
        # Tamaño ajustado para una mejor proporción (ahora el logo va arriba)
        dlg.geometry("600x400")
        dlg.resizable(False, False)

        # --- Estilo mejorado (puedes modificar STYLE si lo prefieres) ---
        # Se mantiene el uso de STYLE pero con algunos valores por defecto mejorados
        primary_color = STYLE.get("primario", "#2A7A62")       # Verde azulado
        secondary_color = STYLE.get("secundario", "#F5F5F5")   # Gris muy claro
        danger_color = STYLE.get("peligro", "#D32F2F")         # Rojo
        text_dark = STYLE.get("texto_oscuro", "#333333")
        text_light = STYLE.get("texto_claro", "#FFFFFF")
        surface_color = STYLE.get("surface", "#FFFFFF")
        
        # --- Tarjeta principal con sombra simulada (borde + padding) ---
        card = ctk.CTkFrame(
            dlg,
            fg_color=surface_color,
            corner_radius=0,
            border_width=0,
            border_color="#DDDDDD"  # borde suave que da sensación de sombra
        )
        card.pack(padx=0, pady=0, fill='both', expand=True)

        # --- Cabecera con título (más prominente y con nombre del sistema) ---
        header = ctk.CTkFrame(
            card,
            fg_color=STYLE.get("surface", surface_color),
            height=70,
            corner_radius=12
        )
        header.pack(fill='x', pady=(16, 20), padx=16)
        header.pack_propagate(False)
        
        # Título del sistema
        ctk.CTkLabel(
            header,
            text="Inicio de sesión",
            font=("Inter", 22, "bold"),
            text_color=STYLE.get("texto_oscuro", text_dark)
        ).pack(expand=True)

        # --- Cuerpo: logo arriba, formulario abajo (mejor centrado) ---
        body = ctk.CTkFrame(card, fg_color="transparent")
        body.pack(fill='both', expand=True, padx=20, pady=(0, 20))

        # --- Layout en dos columnas: logo a la izquierda, formulario centrado a la derecha ---
        left_col = ctk.CTkFrame(body, width=320, fg_color="transparent")
        left_col.pack(side='left', fill='y', padx=(12, 8), pady=6)
        left_col.pack_propagate(False)

        right_col = ctk.CTkFrame(body, fg_color="transparent")
        right_col.pack(side='left', fill='both', expand=True, padx=(8, 20), pady=6)

        # Contenedor del logo (izquierda)
        logo_frame = ctk.CTkFrame(left_col, fg_color="transparent")
        logo_frame.pack(expand=True)
        logo_frame.pack_propagate(False)
        try:
            logo_path = os.path.join(BASE_DIR, 'img', 'logo.png')
            if not os.path.exists(logo_path):
                logo_path = os.path.join(APP_DIR, 'img', 'logo.png')
            pil_img = Image.open(logo_path)
            # Ajustar a un tamaño visible pero no excesivo
            max_dim = 200
            w, h = pil_img.size
            scale = min(max_dim / w, max_dim / h, 1.0)
            new_size = (max(1, int(w * scale)), max(1, int(h * scale)))
            pil_img = pil_img.convert('RGBA').resize(new_size, Image.LANCZOS)
            logo_img = ctk.CTkImage(pil_img, size=new_size)
            lbl_logo = ctk.CTkLabel(logo_frame, image=logo_img, text="")
            lbl_logo.pack(expand=True)
        except Exception:
            ctk.CTkLabel(logo_frame, text=" ", fg_color="transparent").pack(expand=True)

        # Contenedor del formulario (derecha) y centrar verticalmente
        form_outer = ctk.CTkFrame(right_col, fg_color="transparent")
        form_outer.pack(fill='both', expand=True)
        form_inner = ctk.CTkFrame(form_outer, fg_color="transparent")
        form_inner.place(relx=0.5, rely=0.5, anchor='center')

        # ---- Formulario con campos estilizados dentro de form_inner ----
        form = form_inner

        # Texto del sistema arriba de los campos (solicitud del usuario)
        ctk.CTkLabel(
            form,
            text="Sistema Generador de Documentos V&C",
            font=("Inter", 12, "bold"),
            text_color=text_dark,
            anchor='w'
        ).pack(anchor='w', pady=(0, 8))

        # Etiqueta Usuario con icono (opcional, usando unicode)
        ctk.CTkLabel(
            form,
            text="👤  Usuario",
            font=("Inter", 12),
            text_color=text_dark,
            anchor='w'
        ).pack(anchor='w', pady=(6, 4))

        username_entry = ctk.CTkEntry(
            form,
            width=300,
            placeholder_text="usuario",
            corner_radius=8,
            border_width=0,
            border_color="#CCCCCC",
            fg_color="#FAFAFA"
        )
        username_entry.pack(fill='x', pady=(0, 12))

        # Etiqueta Contraseña con icono
        ctk.CTkLabel(
            form,
            text="🔒  Contraseña",
            font=("Inter", 12),
            text_color=text_dark,
            anchor='w'
        ).pack(anchor='w', pady=(2, 4))

        password_entry = ctk.CTkEntry(
            form,
            width=300,
            show='*',
            placeholder_text="contraseña",
            corner_radius=8,
            border_width=0,
            border_color="#CCCCCC",
            fg_color="#FAFAFA"
        )
        password_entry.pack(fill='x')

        # Mensaje de error con estilo mejorado (fondo suave y borde)
        # Mensaje de error con estilo mejorado (fondo suave y borde)
        # No mostrar el recuadro hasta que haya texto de error
        msg_frame = ctk.CTkFrame(form, fg_color="#FFEBEE", corner_radius=6, height=30)
        msg_frame.pack_propagate(False)
        
        msg = ctk.CTkLabel(
            msg_frame,
            text="",
            font=("Inter", 11),
            text_color=danger_color,
            anchor='w'
        )
        msg.pack(padx=8, pady=4, fill='both', expand=True)

        def _show_message(text):
            try:
                if text:
                    msg.configure(text=text)
                    # sólo empacar cuando haya texto
                    try:
                        msg_frame.pack(fill='x', pady=(12, 0))
                    except Exception:
                        pass
                else:
                    try:
                        msg_frame.pack_forget()
                    except Exception:
                        pass
                    msg.configure(text="")
            except Exception:
                pass

        # --- Botón Ingresar (destacado, con sombra simulada) ---
        btns = ctk.CTkFrame(card, fg_color='transparent')
        btns.pack(fill='x', pady=(10, 20), padx=20)

        # Función de login (sin cambios en la lógica)
        def _attempt_login(event=None):
            u = username_entry.get().strip()
            p = password_entry.get()
            user = self._find_user(u)
            if user and user.get('password') == p:
                self.current_user = user
                self.current_role = user.get('role')
                try:
                    display_name = user.get('name') or user.get('username') or u
                    self.title(f"Generador de Dictámenes - {display_name}")
                except Exception:
                    pass
                try:
                    dlg.grab_release()
                except Exception:
                    pass
                dlg.destroy()
                return
            else:
                _show_message("✗ Usuario o contraseña incorrectos")

        btn_ingresar = ctk.CTkButton(
            btns,
            text="Ingresar",
            command=_attempt_login,
            width=180,
            height=42,
            fg_color=STYLE["secundario"],
            hover_color="#4b4b4b",          # tono más oscuro del primario
            text_color=text_light,
            corner_radius=10,
            font=("Inter", 14, "bold")
        )
        # Centrar el botón de ingresar para una mejor experiencia
        btn_ingresar.pack(anchor='center', pady=(6,0))

        # --- Mejoras en la interacción (sin cambios lógicos) ---
        try:
            username_entry.bind('<Return>', lambda e: password_entry.focus_set())
            password_entry.bind('<Return>', _attempt_login)
            username_entry.focus_set()
        except Exception:
            pass

        # --- Centrar ventana y traer al frente (igual que antes) ---
        try:
            dlg.update_idletasks()
            w = dlg.winfo_width()
            h = dlg.winfo_height()
            sw = dlg.winfo_screenwidth()
            sh = dlg.winfo_screenheight()
            x = (sw // 2) - (w // 2)
            y = (sh // 2) - (h // 2)
            dlg.geometry(f"{w}x{h}+{x}+{y}")
            dlg.focus_force()
            dlg.attributes('-topmost', True)
            dlg.after(200, lambda: dlg.attributes('-topmost', False))
        except Exception:
            pass

        # --- Cierre de ventana (exactamente igual) ---
        def _on_close():
            try:
                self.current_user = None
            except Exception:
                pass
            try:
                dlg.grab_release()
            except Exception:
                pass
            # Destruir el diálogo y programar un quit() del mainloop
            try:
                dlg.destroy()
            except Exception:
                pass
            try:
                # programar quit() para salir del mainloop de forma ordenada
                self.after(50, lambda: self.quit())
            except Exception:
                try:
                    self.quit()
                except Exception:
                    pass

        try:
            dlg.protocol('WM_DELETE_WINDOW', _on_close)
        except Exception:
            pass

        self.wait_window(dlg)







    def _apply_permissions(self):
        """Oculta/muneda botones según rol del usuario autenticado."""
        role = getattr(self, 'current_role', None)
        # Mostrar/ocultar info de usuario y botón de cerrar sesión
        try:
            if getattr(self, 'current_user', None):
                # mostrar etiqueta con nombre
                display_name = (self.current_user.get('name') or self.current_user.get('username')) if self.current_user else ''
                try:
                    self.lbl_info_sistema.configure(text=f"{display_name}")
                except Exception:
                    pass
                try:
                    self.lbl_info_sistema.pack(side="right", padx=(0, 10))
                except Exception:
                    pass
                try:
                    self.btn_logout.pack(side="right", padx=(0, 10))
                except Exception:
                    pass
            else:
                try:
                    self.lbl_info_sistema.pack_forget()
                except Exception:
                    pass
                try:
                    self.btn_logout.pack_forget()
                except Exception:
                    pass
        except Exception:
            pass

        # Mostrar/ocultar botón 'Reportes' (ejecutivos)
        try:
            if role == 'ejecutivo':
                try:
                    self.btn_reportes_ej.pack(side="left", padx=(0, 10))
                except Exception:
                    pass
            else:
                try:
                    self.btn_reportes_ej.pack_forget()
                except Exception:
                    pass
        except Exception:
            pass
        # Controlar visibilidad de pestañas según rol
        try:
            # Siempre asegurarse de limpiar estados anteriores
            try:
                self.btn_reportes.pack_forget()
            except Exception:
                pass
            try:
                self.btn_inspectores.pack_forget()
            except Exception:
                pass
            try:
                self.btn_reporte_ema.pack_forget()
            except Exception:
                pass
            try:
                self.btn_reportes_ej.pack_forget()
            except Exception:
                pass

            # Mostrar para 'ejecutivo': Principal, Historial, Reportes (ejecutivo)
            if role == 'ejecutivo':
                try:
                    self.btn_reportes_ej.pack(side="left", padx=(0, 10))
                except Exception:
                    pass
                # Ejecutivos no mostrarán el Reporte EMA (solo administradores)

            # Mostrar para 'admin': Principal, Historial, Reportes (clientes), Inspectores, Reportes ejecutivos
            if role == 'admin':
                try:
                    self.btn_reportes.pack(side="left", padx=(0, 10))
                except Exception:
                    pass
                try:
                    self.btn_inspectores.pack(side="left", padx=(0, 10))
                except Exception:
                    pass
                try:
                    # también permitir acceso a reportes ejecutivos desde admin
                    self.btn_reportes_ej.pack(side="left", padx=(0, 10))
                except Exception:
                    pass
                try:
                    # Admin también puede ver/generar Reporte EMA
                    self.btn_reporte_ema.pack(side="left", padx=(0, 10))
                except Exception:
                    pass
        except Exception:
            pass

    def _ensure_admin_or_block(self):
        """Utilitario: llamar antes de mostrar pantallas privadas."""
        if getattr(self, 'current_role', None) != 'admin':
            messagebox.showwarning("Acceso denegado", "Necesita permisos de administrador para esta acción.")
            return False
        return True

    def _logout(self):
        """Cerrar sesión: limpiar estado y volver a mostrar el login."""
        try:
            # limpiar estado
            self.current_user = None
            self.current_role = None
        except Exception:
            pass
        try:
            # Restaurar título por defecto
            self.title("Generador de Dictámenes")
        except Exception:
            pass
        try:
            # actualizar permisos/UI antes de ocultar
            self._apply_permissions()
        except Exception:
            pass

        # Ocultar la ventana principal y pedir nuevas credenciales.
        # Si el usuario cierra el diálogo, el manejador del diálogo terminará la app.
        try:
            self.withdraw()
        except Exception:
            pass

        try:
            self._show_login_dialog()
        except Exception:
            pass

        # Si después del diálogo hay un usuario válido, restaurar la ventana
        try:
            if getattr(self, 'current_user', None):
                try:
                    display_name = self.current_user.get('name') or self.current_user.get('username')
                    self.title(f"Generador de Dictámenes - {display_name}")
                except Exception:
                    pass
                try:
                    self.deiconify()
                except Exception:
                    pass
                try:
                    self._apply_permissions()
                except Exception:
                    pass
            else:
                # si no hay usuario (posible si _show_login_dialog cerró la app), intentar salir limpio
                try:
                    self.destroy()
                except Exception:
                    pass
        except Exception:
            pass

    def _descargar_reporte_ejecutivo(self):
        """Genera y descarga un CSV con las visitas del día filtradas por el usuario actual."""
        try:
            if not getattr(self, 'current_user', None):
                messagebox.showwarning('Acceso', 'Debe iniciar sesión para descargar reportes.')
                return
            # Reporte global: incluir todas las visitas (de todos los usuarios),
            # opcionalmente filtradas por fecha si se proporciona.
            # Preferir la fecha seleccionada en el calendario si existe
            fecha = None
            try:
                if getattr(self, 'reporte_calendar', None) and Calendar is not None:
                    fecha = self.reporte_calendar.get_date()
                else:
                    fecha = (getattr(self, 'entry_reporte_fecha', None) and self.entry_reporte_fecha.get()) or ''
            except Exception:
                fecha = ''
            # normalizar fecha seleccionada
            fecha = self._normalize_fecha_str(fecha) or datetime.now().strftime('%d/%m/%Y')

            role = getattr(self, 'current_role', None)
            user = getattr(self, 'current_user', None)

            encontrados = []
            # Si el usuario es admin o supervisor, leer todos los folders en data/produccion
            scan_all = role in ('admin', 'supervisor')

            if not scan_all:
                # comportamiento por usuario (ejecutivo u otros)
                try:
                    uname = ''
                    if isinstance(user, dict):
                        uname = user.get('username') or ''
                    else:
                        uname = str(user or '')
                    owner_safe = re.sub(r"[^A-Za-z0-9_.-]", '_', str(uname))
                    owner_dir = os.path.join(DATA_DIR, 'produccion', owner_safe)
                    if not os.path.exists(owner_dir):
                        try:
                            self.reporte_log.configure(text='No existen registros guardados para su usuario.')
                        except Exception:
                            pass
                        return

                    targets = [owner_dir]
                except Exception:
                    targets = []
            else:
                # admin/supervisor: scan all owner dirs
                prod_root = os.path.join(DATA_DIR, 'produccion')
                targets = []
                try:
                    for name in os.listdir(prod_root):
                        p = os.path.join(prod_root, name)
                        if os.path.isdir(p):
                            targets.append(p)
                except Exception:
                    targets = []

            for owner_dir in targets:
                try:
                    if not os.path.exists(owner_dir):
                        continue
                    for fn in os.listdir(owner_dir):
                        if not fn.lower().endswith('.json'):
                            continue
                        fp = os.path.join(owner_dir, fn)
                        try:
                            with open(fp, 'r', encoding='utf-8') as jf:
                                obj = json.load(jf)

                            recs = []
                            if isinstance(obj, dict) and 'records' in obj:
                                recs = obj.get('records') or []
                            elif isinstance(obj, list):
                                recs = obj
                            else:
                                recs = [obj]

                            for rec in recs:
                                try:
                                    fval_raw = (rec.get('fecha_inicio') or '') if isinstance(rec, dict) else ''
                                    fval = self._normalize_fecha_str(fval_raw)
                                    if fecha and fval != fecha:
                                        continue
                                    encontrados.append(rec)
                                except Exception:
                                    continue
                        except Exception:
                            continue
                except Exception:
                    continue

            if not encontrados:
                try:
                    self.reporte_log.configure(text='No se encontraron visitas para la fecha/usuario seleccionados.')
                except Exception:
                    pass
                return

            # Pedir ruta para guardar
            suggested = f"reporte_produccion_{fecha.replace('/','-')}.csv"
            save_path = filedialog.asksaveasfilename(defaultextension='.csv', initialfile=suggested, filetypes=[('CSV','*.csv')])
            if not save_path:
                return

            import csv
            # Añadir columnas extra solicitadas: `folios_utilizados` y `total_folios_numericos`
            keys = ['folio_visita','folio_acta','fecha_inicio','hora_inicio','cliente','tipo_documento','estatus','folios_utilizados','total_folios_numericos','creado_por','creado_nombre']
            try:
                # Use UTF-8 with BOM so Excel on Windows detects accents correctly
                with open(save_path, 'w', newline='', encoding='utf-8-sig') as cf:
                    writer = csv.writer(cf)
                    writer.writerow(keys)
                    for rec in encontrados:
                        # Asegurar que los campos adicionales existan en el registro
                        row = [rec.get(k,'') for k in keys]
                        writer.writerow(row)
            except Exception as e:
                messagebox.showerror('Error', f'No se pudo guardar el reporte: {e}')
                return

                try:
                    self.reporte_log.configure(text=f'Reporte guardado: {save_path} — registros: {len(encontrados)}')
                except Exception:
                    pass
            try:
                # Guardar copia de producción en data/produccion
                try:
                    self._save_produccion_report(encontrados, save_path, fecha, getattr(self, 'current_user', None))
                except Exception:
                    pass
                messagebox.showinfo('OK', f'Reporte guardado: {save_path}')
            except Exception:
                pass
        except Exception as e:
            print(f'Error generando reporte ejecutivo: {e}')
            try:
                messagebox.showerror('Error', str(e))
            except Exception:
                pass

    def _exportar_reporte_ejecutivo_excel(self):
        """Exporta a Excel (.xlsx) las visitas del ejecutivo y una hoja con inspectores acreditados.

        Columnas visitas: CP, Fecha Creacion, Rango de Folios, Tipo de Documento, Estatus, Registrado Por
        Hoja Inspectores: Nombre, Correo, Puesto, Firma (ruta). Intento insertar imágenes si es posible.
        """
        try:
            if not getattr(self, 'current_user', None):
                messagebox.showwarning('Acceso', 'Debe iniciar sesión para exportar reportes.')
                return
            # Reporte global a Excel: incluir todas las visitas (de todos los usuarios),
            # opcionalmente filtradas por fecha.
            fecha = ''
            try:
                fecha = (self.entry_reporte_fecha.get() or '').strip()
            except Exception:
                fecha = datetime.now().strftime('%d/%m/%Y')

            fuente = getattr(self, 'historial_data', []) or []
            encontrados = []
            role = getattr(self, 'current_role', None)
            user = getattr(self, 'current_user', None)
            for r in fuente:
                try:
                    f = (r.get('fecha_inicio') or '').strip()
                    f = self._normalize_fecha_str(f) or f
                    fecha_norm = self._normalize_fecha_str(fecha) or fecha
                    if fecha and f != fecha_norm:
                        continue
                    # permitir que supervisores vean todo también
                    if role not in ('admin', 'supervisor'):
                        creador = (r.get('creado_por') or '').strip()
                        if creador != (user or ''):
                            continue
                    encontrados.append(r)
                except Exception:
                    continue

            if not encontrados:
                try:
                    self.reporte_log.configure(text='No se encontraron visitas para la fecha seleccionada.')
                except Exception:
                    pass
                return

            suggested = f"reporte_produccion_{fecha.replace('/','-')}.xlsx"
            save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', initialfile=suggested, filetypes=[('Excel','*.xlsx')])
            if not save_path:
                return

            # Construir DataFrame de visitas
            rows = []
            for rec in encontrados:
                cp = rec.get('cp') or rec.get('CP') or ''
                fecha_creacion = rec.get('fecha_inicio') or ''
                rango = rec.get('folios_utilizados') or f"{rec.get('folio_visita','')}"
                tipo = rec.get('tipo_documento') or ''
                estatus = rec.get('estatus') or ''
                registrado = rec.get('creado_nombre') or rec.get('creado_por') or ''
                rows.append({'CP': cp, 'Fecha Creacion': fecha_creacion, 'Rango Folios': rango, 'Tipo Documento': tipo, 'Estatus': estatus, 'Registrado Por': registrado})

            try:
                import pandas as _pd
                df = _pd.DataFrame(rows)
            except Exception:
                # fallback: write CSV instead
                import csv
                csv_path = save_path.replace('.xlsx', '.csv')
                with open(csv_path, 'w', newline='', encoding='utf-8') as cf:
                    w = csv.DictWriter(cf, fieldnames=['CP','Fecha Creacion','Rango Folios','Tipo Documento','Estatus','Registrado Por'])
                    w.writeheader()
                    for r in rows:
                        w.writerow(r)
                messagebox.showinfo('OK', f'Reporte guardado (CSV): {csv_path}')
                return

            # Inspectores
            try:
                self.cargar_inspectores_desde_json()
            except Exception:
                pass
            inspect_rows = []
            for ip in getattr(self, 'inspectores_data', []) or []:
                try:
                    nombre = ip.get('NOMBRE DE INSPECTOR') or ip.get('NOMBRE') or ip.get('nombre') or ''
                    correo = ip.get('CORREO') or ip.get('correo') or ''
                    puesto = ip.get('Puesto') or ip.get('puesto') or ''
                    firma = ip.get('FIRMA') or ip.get('firma') or ip.get('firma_path') or ''
                    inspect_rows.append({'Nombre': nombre, 'Correo': correo, 'Puesto': puesto, 'Firma': firma})
                except Exception:
                    continue

            try:
                with _pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Visitas', index=False)
                    _pd.DataFrame(inspect_rows).to_excel(writer, sheet_name='Inspectores', index=False)

                # Intentar insertar imágenes en la hoja Inspectores
                try:
                    from openpyxl import load_workbook
                    from openpyxl.drawing.image import Image as XLImage
                    wb = load_workbook(save_path)
                    ws = wb.get_sheet_by_name('Inspectores') if 'Inspectores' in wb.sheetnames else wb[wb.sheetnames[-1]]
                    # insertar imágenes empezando en columna D (4), fila 2
                    r = 2
                    for ir in inspect_rows:
                        firma_path = ir.get('Firma') or ''
                        if firma_path and os.path.exists(firma_path):
                            try:
                                img = XLImage(firma_path)
                                img.width = 120
                                img.height = 60
                                cell = f'D{r}'
                                ws.add_image(img, cell)
                            except Exception:
                                pass
                        r += 1
                    wb.save(save_path)
                except Exception:
                    # no crítico, continuar sin imágenes
                    pass

                try:
                    # Guardar copia de producción en data/produccion
                    try:
                        self._save_produccion_report(encontrados, save_path, fecha, getattr(self, 'current_user', None))
                    except Exception:
                        pass
                except Exception:
                    pass
                messagebox.showinfo('OK', f'Reporte Excel guardado: {save_path}')
                try:
                    self.reporte_log.configure(text=f'Reporte guardado: {save_path} — registros: {len(rows)}')
                except Exception:
                    pass
            except Exception as e:
                messagebox.showerror('Error', f'No se pudo generar el Excel: {e}')
                return
        except Exception as e:
            print(f'Error exportando reporte ejecutivo a Excel: {e}')
            try:
                messagebox.showerror('Error', str(e))
            except Exception:
                pass

    def _cmd_mostrar_clientes(self):
        if not self._ensure_admin_or_block():
            return
        try:
            self.mostrar_clientes()
        except Exception:
            pass
    
    def _save_produccion_report(self, records, export_path, fecha_filter, owner_username=None):
        """Guarda en JSON los datos de producción en DATA_DIR/produccion/<owner>/produccion_<ts>.json

        También mantiene un índice global en DATA_DIR/produccion/produccion.json con metadatos.
        """
        try:
            prod_root = os.path.join(DATA_DIR, 'produccion')
            os.makedirs(prod_root, exist_ok=True)

            owner = owner_username or getattr(self, 'current_user', 'unknown') or 'unknown'
            # normalizar nombre de carpeta
            owner_safe = re.sub(r"[^A-Za-z0-9_.-]", '_', str(owner))
            owner_dir = os.path.join(prod_root, owner_safe)
            os.makedirs(owner_dir, exist_ok=True)

            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            fname = f"produccion_{ts}.json"
            fpath = os.path.join(owner_dir, fname)

            payload = {
                'fecha_generacion': datetime.now().isoformat(),
                'fecha_filtro': fecha_filter,
                'creado_por': getattr(self, 'current_user', ''),
                'owner': owner_safe,
                'export_path': os.path.relpath(export_path, DATA_DIR) if export_path and isinstance(export_path, str) and export_path.startswith(DATA_DIR) else export_path,
                'count': len(records) if isinstance(records, (list, tuple)) else 0,
                'records': records,
            }

            with open(fpath, 'w', encoding='utf-8') as jf:
                json.dump(payload, jf, ensure_ascii=False, indent=2)

            # actualizar índice global
            idx_path = os.path.join(prod_root, 'produccion.json')
            index = []
            try:
                if os.path.exists(idx_path):
                    with open(idx_path, 'r', encoding='utf-8') as ix:
                        index = json.load(ix) or []
            except Exception:
                index = []

            entry = {
                'timestamp': payload['fecha_generacion'],
                'file': os.path.relpath(fpath, DATA_DIR),
                'owner': owner_safe,
                'count': payload['count'],
                'fecha_filtro': fecha_filter,
            }
            index.append(entry)
            try:
                with open(idx_path, 'w', encoding='utf-8') as ix:
                    json.dump(index, ix, ensure_ascii=False, indent=2)
            except Exception:
                pass

            return fpath
        except Exception as e:
            print('Error guardando produccion:', e)
            return None
        try:
            self.mostrar_clientes()
        except Exception:
            pass

    def mostrar_reportes_ejecutivo(self):
        """Muestra la pestaña de reportes para ejecutivos y oculta las demás."""
        try:
            # ocultar otras secciones
            try:
                self.frame_principal.pack_forget()
            except Exception:
                pass
            try:
                self.frame_historial.pack_forget()
            except Exception:
                pass
            try:
                self.frame_reportes.pack_forget()
            except Exception:
                pass
            try:
                self.frame_inspectores.pack_forget()
            except Exception:
                pass

            # mostrar reportes ejecutivos
            try:
                self.frame_reportes_ejecutivo.pack(fill='both', expand=True)
            except Exception:
                pass

            # actualizar estilo de botones
            try:
                # reset others
                for b in (getattr(self, 'btn_principal', None), getattr(self, 'btn_historial', None), getattr(self, 'btn_reportes', None), getattr(self, 'btn_inspectores', None), getattr(self, 'btn_reporte_ema', None)):
                    try:
                        if b:
                            b.configure(fg_color=STYLE['surface'], text_color=STYLE['secundario'], border_color=STYLE['secundario'])
                    except Exception:
                        pass
            except Exception:
                pass
            try:
                if getattr(self, 'btn_reportes_ej', None):
                    self.btn_reportes_ej.configure(fg_color=STYLE['primario'], text_color=STYLE['secundario'], border_color=STYLE['primario'])
            except Exception:
                pass
        except Exception:
            pass

    def _cmd_mostrar_inspectores(self):
        if not self._ensure_admin_or_block():
            return
        try:
            self.mostrar_inspectores()
        except Exception:
            pass


    def mostrar_principal(self):
        """Muestra la sección principal y oculta las demás"""
        # Ocultar todos los frames primero
        self.frame_principal.pack_forget()
        self.frame_historial.pack_forget()
        # Asegurarse de ocultar la pestaña de reportes ejecutivos también
        try:
            self.frame_reportes_ejecutivo.pack_forget()
        except Exception:
            pass
        # Asegurarse de ocultar la pestaña Reportes también
        try:
            self.frame_reportes.pack_forget()
        except Exception:
            pass
        
        # Mostrar el frame principal
        self.frame_principal.pack(fill="both", expand=True)
        
        # Actualizar estado de los botones con mejor contraste
        self.btn_principal.configure(
            fg_color=STYLE["primario"],
            text_color=STYLE["secundario"],
            border_color=STYLE["primario"]
        )
        self.btn_historial.configure(
            fg_color=STYLE["surface"],
            text_color=STYLE["secundario"],
            border_color=STYLE["secundario"]
        )
        try:
            self.btn_reportes.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
        except Exception:
            pass
        try:
            self.btn_reporte_ema.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
        except Exception:
            pass
        # Asegurar que la pestaña Inspectores quede oculta al mostrar Principal
        try:
            self.frame_inspectores.pack_forget()
        except Exception:
            pass
        # Ocultar backup nav cuando no estemos en Historial
        try:
            if getattr(self, 'btn_backup', None):
                self.btn_backup.pack_forget()
        except Exception:
            pass

    def mostrar_historial(self):
            """Muestra la sección de historial y oculta las demás"""
            # Ocultar todos los frames primero
            self.frame_principal.pack_forget()
            self.frame_historial.pack_forget()
            # Asegurarse de ocultar las pestañas de Reportes también
            try:
                self.frame_reportes.pack_forget()
            except Exception:
                pass
            try:
                self.frame_reportes_ejecutivo.pack_forget()
            except Exception:
                pass
            
            # Mostrar el frame de historial
            self.frame_historial.pack(fill="both", expand=True)
            
            # Actualizar estado de los botones con mejor contraste
            self.btn_principal.configure(
                fg_color=STYLE["surface"],
                text_color=STYLE["secundario"],
                border_color=STYLE["secundario"]
            )
            self.btn_historial.configure(
                fg_color=STYLE["primario"],
                text_color=STYLE["secundario"],
                border_color=STYLE["primario"]
            )
            try:
                self.btn_reportes.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            except Exception:
                pass
            try:
                self.btn_reporte_ema.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            except Exception:
                pass
            # Asegurar que la pestaña Inspectores quede oculta al mostrar Historial
            try:
                self.frame_inspectores.pack_forget()
            except Exception:
                pass
            
            # Verificar y reparar datos existentes al mostrar historial
            self.verificar_datos_folios_existentes()
            
            # Refrescar el historial si es necesario
            self._cargar_historial()
            self._poblar_historial_ui()

            # Mostrar backup nav cuando estemos en Historial
            try:
                if getattr(self, 'btn_backup', None):
                    self.btn_backup.pack(side="right", padx=(0, 10))
            except Exception:
                pass

    def mostrar_clientes(self):
        """Muestra la sección de reportes y oculta las demás"""
        # Ocultar todos los frames primero
        self.frame_principal.pack_forget()
        self.frame_historial.pack_forget()
        # Ocultar reportes ejecutivos si estaban visibles
        try:
            self.frame_reportes_ejecutivo.pack_forget()
        except Exception:
            pass
        self.frame_reportes.pack(fill="both", expand=True)
        # Asegurarse de ocultar backup nav en la pestaña Reportes
        try:
            if getattr(self, 'btn_backup', None):
                self.btn_backup.pack_forget()
        except Exception:
            pass

        # Actualizar estado de los botones
        try:
            self.btn_principal.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            self.btn_historial.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            self.btn_reportes.configure(fg_color=STYLE["primario"], text_color=STYLE["secundario"], border_color=STYLE["primario"])
            try:
                self.btn_inspectores.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            except Exception:
                pass
            try:
                # Al mostrar la pestaña Clientes, el botón EMA no queda activo
                self.btn_reporte_ema.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            except Exception:
                pass
        except Exception:
            pass
        # Asegurar que la pestaña Inspectores quede oculta al mostrar Clientes
        try:
            self.frame_inspectores.pack_forget()
        except Exception:
            pass

    def _construir_tab_principal(self, parent):
        """Construye la interfaz principal con dos tarjetas en proporción 30%/70%"""
        # ===== CONTENEDOR PRINCIPAL CON 2 COLUMNAS =====
        main_frame = ctk.CTkFrame(parent, fg_color="transparent")
        main_frame.pack(fill="both", expand=True)

        # Configurar grid para 2 columnas con proporción ~20%/80% (ligeramente más pequeña la izquierda)
        main_frame.grid_columnconfigure(0, weight=2)  # ~20%
        main_frame.grid_columnconfigure(1, weight=8)  # ~80%
        # Mantener ambas tarjetas (izquierda/derecha) con una altura mínima reducida
        # cuando se muestran/ocultan widgets según el tipo de documento.
        main_frame.grid_rowconfigure(0, weight=1, minsize=420)

        # ===== TARJETA INFORMACIÓN DE VISITA (IZQUIERDA) - 30% =====
        card_visita = ctk.CTkFrame(main_frame, fg_color=STYLE["surface"], corner_radius=12)
        card_visita.grid(row=0, column=0, padx=(0, 10), pady=0, sticky="nsew")
        try:
            card_visita.grid_propagate(False)
        except Exception:
            pass

        ctk.CTkLabel(
            card_visita,
            text="📋 Información de Visita",
            font=FONT_SUBTITLE,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", padx=20, pady=(20, 15))

        visita_frame = ctk.CTkFrame(card_visita, fg_color="transparent")
        visita_frame.pack(fill="both", expand=True, padx=0, pady=(0, 0))

        # Contenedor para el formulario con scrollbar
        scroll_form = ctk.CTkScrollableFrame(
            visita_frame,
            fg_color="transparent",
            scrollbar_button_color="#ecd925",
            scrollbar_button_hover_color="#ecd925"
        )
        scroll_form.pack(fill="both", expand=True, padx=0, pady=(0, 0))

        # === Tipo de Documento (Dictamen, Negación de dictamen, Constancia, Negación de Constancia) ===
        tipo_doc_frame = ctk.CTkFrame(scroll_form, fg_color="transparent")
        tipo_doc_frame.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            tipo_doc_frame,
            text="Tipo de documento:",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 5))

        self.combo_tipo_documento = ctk.CTkComboBox(
            tipo_doc_frame,
            values=["Dictamen", "Negación de Dictamen", "Constancia", "Negación de Constancia"],
            font=FONT_SMALL,
            dropdown_font=FONT_SMALL,
            state="readonly",
            command=self.actualizar_tipo_documento,
            height=35,
            corner_radius=8
        )
        self.combo_tipo_documento.pack(fill="x", pady=(0, 5))
        self.combo_tipo_documento.set("Dictamen")

        # Folio de visita (automático)
        folio_frame = ctk.CTkFrame(scroll_form, fg_color="transparent")
        folio_frame.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            folio_frame,
            text="Folio Visita:",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 5))

        self.entry_folio_visita = ctk.CTkEntry(
            folio_frame,
            placeholder_text="CP0001",
            font=FONT_SMALL,
            height=35,
            corner_radius=8
        )
        self.entry_folio_visita.pack(fill="x", pady=(0, 5))
        folio_con_prefijo = f"CP{self.current_folio}"
        self.entry_folio_visita.insert(0, folio_con_prefijo)
        self.entry_folio_visita.configure(state="normal")

        # Folio de acta (automático - AC + folio visita)
        acta_frame = ctk.CTkFrame(scroll_form, fg_color="transparent")
        acta_frame.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            acta_frame,
            text="Folio Acta:",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 5))

        self.entry_folio_acta = ctk.CTkEntry(
            acta_frame,
            placeholder_text="AC0001",
            font=FONT_SMALL,
            height=35,
            corner_radius=8
        )
        self.entry_folio_acta.pack(fill="x", pady=(0, 5))
        self.entry_folio_acta.insert(0, f"AC{self.current_folio}")
        self.entry_folio_acta.configure(state="normal")

        # Fecha Inicio
        fecha_inicio_frame = ctk.CTkFrame(scroll_form, fg_color="transparent")
        fecha_inicio_frame.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            fecha_inicio_frame,
            text="Fecha Inicio (dd/mm/yyyy):",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 5))

        self.entry_fecha_inicio = ctk.CTkEntry(
            fecha_inicio_frame,
            placeholder_text="dd/mm/yyyy",
            font=FONT_SMALL,
            height=35,
            corner_radius=8
        )
        self.entry_fecha_inicio.pack(fill="x", pady=(0, 5))
        self.entry_fecha_inicio.insert(0, datetime.now().strftime("%d/%m/%Y"))

        # Hora Inicio
        hora_inicio_frame = ctk.CTkFrame(scroll_form, fg_color="transparent")
        hora_inicio_frame.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            hora_inicio_frame,
            text="Hora Inicio (HH:MM):",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 5))

        self.entry_hora_inicio = ctk.CTkEntry(
            hora_inicio_frame,
            placeholder_text="HH:MM",
            font=FONT_SMALL,
            height=35,
            corner_radius=8
        )
        self.entry_hora_inicio.pack(fill="x", pady=(0, 5))
        self.entry_hora_inicio.insert(0, datetime.now().strftime("%H:%M"))
        self.entry_hora_inicio.configure(state="readonly")

        # Fecha Termino
        fecha_termino_frame = ctk.CTkFrame(scroll_form, fg_color="transparent")
        fecha_termino_frame.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            fecha_termino_frame,
            text="Fecha Termino (dd/mm/yyyy):",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 5))

        self.entry_fecha_termino = ctk.CTkEntry(
            fecha_termino_frame,
            placeholder_text="dd/mm/yyyy",
            font=FONT_SMALL,
            height=35,
            corner_radius=8
        )
        self.entry_fecha_termino.pack(fill="x", pady=(0, 5))

        # Hora Termino
        hora_termino_frame = ctk.CTkFrame(scroll_form, fg_color="transparent")
        hora_termino_frame.pack(fill="x", pady=(0, 10))

        ctk.CTkLabel(
            hora_termino_frame,
            text="Hora Termino (HH:MM):",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 5))

        self.entry_hora_termino = ctk.CTkEntry(
            hora_termino_frame,
            placeholder_text="18:00",
            font=FONT_SMALL,
            height=35,
            corner_radius=8
        )
        self.entry_hora_termino.pack(fill="x", pady=(0, 5))
        self.entry_hora_termino.insert(0, "18:00")
        self.entry_hora_termino.configure(state="readonly")

        
        # Supervisor field removed from UI: supervisor is derived from the loaded tabla de relación

        # ===== TARJETA GENERADOR (DERECHA) - 70% =====
        card_generacion = ctk.CTkFrame(main_frame, fg_color=STYLE["surface"], corner_radius=12)
        card_generacion.grid(row=0, column=1, padx=(10, 0), pady=0, sticky="nsew")
        try:
            card_generacion.grid_propagate(False)
        except Exception:
            pass

        self.generacion_title = ctk.CTkLabel(
            card_generacion,
            text="🚀 Generador de Dictámenes",
            font=FONT_SUBTITLE,
            text_color=STYLE["texto_oscuro"]
        )
        self.generacion_title.pack(anchor="w", padx=20, pady=(10, 8))

        generacion_frame = ctk.CTkFrame(card_generacion, fg_color="transparent")
        generacion_frame.pack(fill="both", expand=True, padx=20, pady=(0, 8))

        # Contenedor principal de generador con scrollbar
        scroll_generacion = ctk.CTkScrollableFrame(
            generacion_frame,
            fg_color="transparent",
            scrollbar_button_color="#ecd925",
            scrollbar_button_hover_color="#ecd925"
        )
        scroll_generacion.pack(fill="both", expand=True)

        # --- SELECCIONAR CLIENTE ---
        cliente_section = ctk.CTkFrame(scroll_generacion, fg_color="transparent")
        cliente_section.pack(fill="x", pady=(0, 4))

        ctk.CTkLabel(
            cliente_section,
            text="👤 Seleccionar Cliente",
            font=FONT_LABEL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 10))

        cliente_controls_frame = ctk.CTkFrame(cliente_section, fg_color="transparent")
        cliente_controls_frame.pack(fill="x", pady=(0, 6))

        # Usamos grid dentro de cliente_controls_frame para que el combo de
        # cliente y el combo de domicilios compartan el mismo ancho.
        self.combo_cliente = ctk.CTkComboBox(
            cliente_controls_frame,
            values=["Seleccione un cliente..."],
            font=FONT_SMALL,
            dropdown_font=FONT_SMALL,
            state="readonly",
            height=35,
            corner_radius=8,
            command=self.actualizar_cliente_seleccionado
        )
        self.combo_cliente.grid(row=0, column=0, sticky="nsew", padx=(0, 10))

        # Encabezado para selector de domicilio (columna central)
        self.lbl_domicilio_header = ctk.CTkLabel(
            cliente_controls_frame,
            text="Domicilio:",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"]
        )
        self.lbl_domicilio_header.grid(row=0, column=1, padx=(8, 4), sticky="w")

        # --- DOMICILIOS DEL CLIENTE (se rellena al seleccionar cliente) ---
        self.combo_domicilios = ctk.CTkComboBox(
            cliente_controls_frame,
            values=["Seleccione un domicilio..."],
            font=FONT_SMALL,
            dropdown_font=FONT_SMALL,
            state="disabled",
            height=35,
            corner_radius=8,
            command=self._seleccionar_domicilio
        )
        self.combo_domicilios.grid(row=0, column=2, sticky="nsew", padx=(8, 0))

        self.boton_limpiar_cliente = ctk.CTkButton(
            cliente_controls_frame,
            text="✕",
            command=self.limpiar_cliente,
            font=("Inter", 14, "bold"),
            fg_color=STYLE["primario"],
            hover_color="#D4BF22",
            text_color=STYLE["secundario"],
            height=35,
            width=35,
            corner_radius=8,
            state="disabled"
        )
        self.boton_limpiar_cliente.grid(row=0, column=3, padx=(8, 0))

        # Columnas 0 y 2 (combos) se expanden por igual; la columna 1 (label)
        # y 3 (botón limpiar) mantienen su tamaño.
        cliente_controls_frame.grid_columnconfigure(0, weight=1)
        cliente_controls_frame.grid_columnconfigure(1, weight=0)
        cliente_controls_frame.grid_columnconfigure(2, weight=1)
        cliente_controls_frame.grid_columnconfigure(3, weight=0)

        self.info_cliente = ctk.CTkLabel(
            cliente_section,
            text="No se ha seleccionado ningún cliente",
            font=FONT_SMALL,
            text_color=STYLE["texto_claro"],
            wraplength=350
        )
        self.info_cliente.pack(anchor="w", fill="x", pady=(0,4))

        # --- PEGADO DE EVIDENCIA (siempre visible, arriba de Cargar Tabla de Relación) ---
        pegado_section = ctk.CTkFrame(scroll_generacion, fg_color="transparent")
        pegado_section.pack(fill="x", pady=(0, 4))

        ctk.CTkLabel(
            pegado_section,
            text="🧩 Pegado de evidencia",
            font=FONT_LABEL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 8))

        pegado_botones_frame = ctk.CTkFrame(pegado_section, fg_color="transparent")
        pegado_botones_frame.pack(fill="x", pady=(0, 4))

        # --- FOLIOS RESERVADOS (Debajo del selector de cliente) ---
        self.cliente_folios_frame = ctk.CTkFrame(cliente_section, fg_color="transparent")
        self.cliente_folios_frame.pack(fill="x", pady=(4, 4))

        self.lbl_folios_pendientes = ctk.CTkLabel(
            self.cliente_folios_frame,
            text="Visitas Reservadas:",
            font=("Inter", 10),
            text_color=STYLE["texto_oscuro"]
        )
        self.lbl_folios_pendientes.pack(side="left", padx=(0,8))

        self.combo_folios_pendientes = ctk.CTkComboBox(
            self.cliente_folios_frame,
            values=[],
            font=FONT_SMALL,
            dropdown_font=FONT_SMALL,
            state="normal",
            height=30,
            corner_radius=8,
            command=self._seleccionar_folio_pendiente
        )
        self.combo_folios_pendientes.pack(side="left", fill="x", expand=True, padx=(0, 8))

        # Botones compactos: desmarcar y eliminar
        self.btn_desmarcar_folio = ctk.CTkButton(
            self.cliente_folios_frame,
            text="Desmarcar",
            width=32,
            height=30,
            corner_radius=6,
            fg_color=STYLE["secundario"],
            text_color=STYLE["surface"],
            command=self._desmarcar_folio_seleccionado
        )
        self.btn_desmarcar_folio.pack(side="left", padx=(0, 6))

        self.btn_eliminar_folio_pendiente = ctk.CTkButton(
            self.cliente_folios_frame,
            text="Eliminar Folio",
            width=32,
            height=30,
            corner_radius=6,
            fg_color=STYLE["peligro"],
            text_color=STYLE["surface"],
            command=self._eliminar_folio_pendiente
        )
        self.btn_eliminar_folio_pendiente.pack(side="left")

        # Por defecto ocultar la sección de folios reservados hasta que haya al menos uno
        try:
            if not self._get_folios_pendientes():
                try:
                    self.cliente_folios_frame.pack_forget()
                except Exception:
                    pass
        except Exception:
            pass

        # Botón para configurar carpetas de evidencias (abre modal para elegir grupo y carpeta)
        # Tres botones para modos de pegado (siempre están visibles en la card de pegado)
        self.boton_pegado_simple = ctk.CTkButton(
            pegado_botones_frame,
            text="🖼️ Una sola Carpeta",
            command=self.handle_pegado_simple,
            font=("Inter", 11, "bold"),
            fg_color=STYLE["primario"],
            hover_color="#D4BF22",
            text_color=STYLE["secundario"],
            height=32,
            width=150,
            corner_radius=8
        )

        self.boton_pegado_carpetas = ctk.CTkButton(
            pegado_botones_frame,
            text="📁 Carpetas con más carpetas",
            command=self.handle_pegado_carpetas,
            font=("Inter", 11, "bold"),
            fg_color=STYLE["primario"],
            hover_color="#D4BF22",
            text_color=STYLE["secundario"],
            height=32,
            width=150,
            corner_radius=8
        )

        # El botón "Índice en excel y cargar carpeta" fue eliminado.
        # Para compatibilidad, el handler `handle_pegado_indice` se mantiene
        # en el código pero el botón ya no se crea ni se muestra.
        # Botón para limpiar rutas de evidencias guardadas
        self.boton_limpiar_rutas_evidencias = ctk.CTkButton(
            pegado_botones_frame,
            text="🧹 Limpiar",
            command=self.handle_clear_evidence_paths,
            font=("Inter", 11, "bold"),
            fg_color=STYLE["peligro"],
            hover_color="#c84a3d",
            text_color=STYLE["surface"],
            height=32,
            width=180,
            corner_radius=8
        )
        # Empacar los botones en la card de pegado para que siempre estén visibles.
        try:
            # pack horizontalmente con separación uniforme
            for btn in (self.boton_pegado_simple, self.boton_pegado_carpetas, self.boton_limpiar_rutas_evidencias):
                if btn:
                    try:
                        btn.pack(side="left", padx=(0, 12))
                    except Exception:
                        pass
        except Exception:
            pass

        # Indicador + Label de estado del pegado: muestra si hay ruta cargada y el modo seleccionado
        self.pegado_path_loaded_var = ctk.BooleanVar(value=False)
        self.pegado_checkbox = ctk.CTkCheckBox(
            pegado_section,
            text="Ruta cargada",
            variable=self.pegado_path_loaded_var,
            state="disabled",
            text_color=STYLE["texto_oscuro"]
        )
        self.pegado_checkbox.pack(side="left", padx=(0, 8))

        self.pegado_status_label = ctk.CTkLabel(
            pegado_section,
            text="Ruta: Ninguna   ·   Modo: Ninguno",
            font=FONT_SMALL,
            text_color=STYLE["texto_oscuro"],
            wraplength=720
        )
        self.pegado_status_label.pack(anchor="w", pady=(6, 0))
        try:
            saved = self._load_evidence_paths()
            first_path = None
            if saved and isinstance(saved, dict):
                for k, v in saved.items():
                    if v:
                        first_path = v[0]
                        break
            if first_path:
                self._update_pegado_status(mode="Sin seleccionar", path=first_path)
        except Exception:
            pass


        # --- CARGAR TABLA DE RELACIÓN ---
        carga_section = ctk.CTkFrame(scroll_generacion, fg_color="transparent")
        carga_section.pack(fill="x", pady=(0, 4))

        ctk.CTkLabel(
            carga_section,
            text="📊 Cargar Tabla de Relación",
            font=FONT_LABEL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 10))

        self.info_archivo = ctk.CTkLabel(
            carga_section,
            text="No se ha cargado ningún archivo",
            font=FONT_SMALL,
            text_color=STYLE["texto_claro"],
            wraplength=350
        )
        self.info_archivo.pack(anchor="w", pady=(0, 10))

        botones_carga_frame = ctk.CTkFrame(carga_section, fg_color="transparent")
        botones_carga_frame.pack(fill="x", pady=(0, 10))

        botones_fila1 = ctk.CTkFrame(botones_carga_frame, fg_color="transparent")
        botones_fila1.pack(fill="x", pady=(0, 4))

        self.boton_cargar_excel = ctk.CTkButton(
            botones_fila1,
            text="Subir archivo",
            command=self.cargar_excel,
            font=("Inter", 13, "bold"),
            fg_color=STYLE["primario"],
            hover_color="#D4BF22",
            text_color=STYLE["secundario"],
            height=35,
            width=110,
            corner_radius=8
        )
        self.boton_cargar_excel.pack(side="left", padx=(0, 12))

        # Botón Verificar Datos movido aquí
        ctk.CTkButton(
            botones_fila1,
            text="🔍 Verificar Datos",
            command=self.verificar_integridad_datos,
            font=("Inter", 11, "bold"),
            fg_color=STYLE["advertencia"],
            hover_color="#b85a52",
            text_color=STYLE["surface"],
            height=35,
            width=100,
            corner_radius=8
        ).pack(side="left", padx=(0, 12))

        self.boton_limpiar = ctk.CTkButton(
            botones_fila1,
            text="Limpiar",
            command=self.limpiar_archivo,
            font=("Inter", 13, "bold"),
            fg_color=STYLE["secundario"],
            hover_color="#1a1a1a",
            text_color=STYLE["surface"],
            height=35,
            width=70,
            corner_radius=8,
            state="disabled"
        )
        self.boton_limpiar.pack(side="left", padx=(0, 12))

        self.boton_subir_etiquetado = ctk.CTkButton(
            botones_fila1,
            text="📦 Subir Base de Etiquetado",
            command=self.cargar_base_etiquetado,
            font=("Inter", 12, "bold"),
            fg_color=STYLE["primario"],
            hover_color="#D4BF22",
            text_color=STYLE["secundario"],
            height=35,
            width=160,
            corner_radius=8
        )
        # Inicialmente no se muestra
        self.boton_subir_etiquetado.pack(side="left", padx=(12, 12))
        self.boton_subir_etiquetado.pack_forget()  # Ocultar inicialmente

        self.info_etiquetado = ctk.CTkLabel(
            botones_carga_frame,
            text="",
            font=FONT_SMALL,
            text_color=STYLE["texto_claro"],
            wraplength=350
        )
        self.info_etiquetado.pack(anchor="w", pady=(5, 0))

        estado_carga_frame = ctk.CTkFrame(carga_section, fg_color="transparent")
        estado_carga_frame.pack(fill="x", pady=(0, 15))

        self.etiqueta_estado = ctk.CTkLabel(
            estado_carga_frame,
            text="",
            font=FONT_SMALL,
            text_color=STYLE["texto_claro"]
        )
        self.etiqueta_estado.pack(side="left")

        self.check_label = ctk.CTkLabel(
            estado_carga_frame,
            text="",
            font=("Inter", 16, "bold"),
            text_color=STYLE["exito"]
        )
        self.check_label.pack(side="right")

        # --- GENERAR DICTÁMENES ---
        generar_section = ctk.CTkFrame(scroll_generacion, fg_color="transparent")
        generar_section.pack(fill="x", pady=(0, 0))

        ctk.CTkLabel(
            generar_section,
            text="🧾 Generar Documentos PDF",
            font=FONT_LABEL,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w", pady=(0, 10))

        self.info_generacion = ctk.CTkLabel(
            generar_section,
            text="Seleccione un cliente y cargue la tabla para habilitar",
            font=FONT_SMALL,
            text_color=STYLE["texto_claro"]
        )
        self.info_generacion.pack(anchor="w", pady=(0, 4))

        # Barra de progreso
        self.barra_progreso = ctk.CTkProgressBar(
            generar_section,
            progress_color=STYLE["primario"],
            height=10,
            corner_radius=5
        )
        self.barra_progreso.pack(fill="x", pady=(5, 8))
        self.barra_progreso.set(0)

        self.etiqueta_progreso = ctk.CTkLabel(
            generar_section,
            text="",
            font=("Inter", 11),
            text_color=STYLE["texto_claro"]
        )
        self.etiqueta_progreso.pack(pady=(0, 8))

        # Label para avisar si hay folio pendiente para el tipo seleccionado
        self.info_folio_pendiente = ctk.CTkLabel(
            generar_section,
            text="",
            font=("Inter", 11),
            text_color=STYLE["advertencia"]
        )
        self.info_folio_pendiente.pack(pady=(0, 6))

        # Botón para guardar un folio incompleto / reservado en el historial
        # Lo creamos en el frame de folios del cliente para mostrarlo junto
        # al selector/acciones de folios y al botón Limpiar.
        # Crear el botón de reservar folio dentro de la fila de botones de carga
        # para que siempre esté visible junto a 'Subir archivo', 'Verificar Datos' y 'Limpiar'.
        self.boton_guardar_folio = ctk.CTkButton(
            botones_fila1,
            text="Reservar Visita",
            command=self.reservar_folios_tabla,
            font=("Inter", 12, "bold"),
            fg_color=STYLE["primario"],
            hover_color="#D4BF22",
            text_color=STYLE["secundario"],
            height=34,
            corner_radius=8,
            state="disabled"
        )
        try:
            self.boton_guardar_folio.pack(side="left", padx=(0, 12))
        except Exception:
            try:
                self.boton_guardar_folio.pack(pady=(0, 6))
            except Exception:
                pass
        # NOTE: no empacamos el botón aquí. Se mostrará junto al botón
        # de "Limpiar rutas evidencias" desde `actualizar_cliente_seleccionado`
        # para garantizar que ambos estén en el mismo contenedor y orden.

        # Botón de generación
        self.boton_generar_dictamen = ctk.CTkButton(
            generar_section,
            text="🧾 Generar Dictámenes",
            command=self.generar_dictamenes,
            font=("Inter", 13, "bold"),
            fg_color=STYLE["exito"],
            hover_color="#1f8c4d",
            text_color=STYLE["surface"],
            height=38,
            corner_radius=8,
            state="disabled"
        )
        self.boton_generar_dictamen.pack(pady=(0, 5))
        # Aplicar estado inicial de UI según tipo de documento seleccionado
        try:
            self.actualizar_tipo_documento()
        except Exception:
            pass

    def _construir_tab_historial(self, parent):
        cont = ctk.CTkFrame(parent, fg_color=STYLE["surface"], corner_radius=8)
        cont.pack(fill="both", expand=True, padx=10, pady=(0,5))

        # ===========================================================
        # BARRA SUPERIOR EN UNA SOLA LÍNEA (COMO EN LA IMAGEN)
        # ===========================================================
        barra_superior = ctk.CTkFrame(cont, fg_color="transparent", height=50)
        barra_superior.pack(fill="x", pady=(0, 10))
        barra_superior.pack_propagate(False)


        # --- FOLIO Y BÚSQUEDA EN MISMA LÍNEA ---
        linea_busqueda = ctk.CTkFrame(barra_superior, fg_color="transparent")
        linea_busqueda.pack(fill="x", pady=5)

        # Folio (izquierda)
        ctk.CTkLabel(
            linea_busqueda, text="Folio visita:", 
            font=("Inter", 11), text_color=STYLE["texto_oscuro"]
        ).pack(side="left", padx=(0, 8))

        self.entry_buscar_folio = ctk.CTkEntry(
            linea_busqueda, width=100, height=25,
            corner_radius=6, placeholder_text="CP0001"
        )
        self.entry_buscar_folio.pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            linea_busqueda, text="Buscar",
            command=self.hist_buscar_por_folio,
            width=40, height=25, corner_radius=6,
            fg_color=STYLE["secundario"], text_color=STYLE["surface"]
        ).pack(side="left", padx=(0, 8))

        # Botón Limpiar búsqueda
        ctk.CTkButton(
            linea_busqueda, text="Limpiar",
            command=self.hist_limpiar_busqueda,
            width=60, height=25, corner_radius=6,
            fg_color=STYLE["secundario"], text_color=STYLE["surface"]
        ).pack(side="left", padx=(0, 8))

        # (Se eliminó el botón global 'Borrar' aquí; el borrado ahora está disponible por fila)

       

        # Búsqueda general (derecha)
        ctk.CTkLabel(
            linea_busqueda, text="Búsqueda general:",
            font=("Inter", 11), text_color=STYLE["texto_oscuro"]
        ).pack(side="left", padx=(30, 8))
        # --- Controles de filtro adicionales ---
        try:
            # Filtro por supervisor
            self.combo_filtrar_supervisor = ctk.CTkComboBox(linea_busqueda, values=[""], font=("Inter", 10), state="readonly", height=25, width=180, command=lambda v: self.hist_buscar_general())
            self.combo_filtrar_supervisor.set("")
            self.combo_filtrar_supervisor.pack(side="left", padx=(8,6))
        except Exception:
            self.combo_filtrar_supervisor = None

        try:
            # Filtro por tipo de documento
            tipos_vals = ["", "Dictamen", "Constancia", "Negación de Dictamen", "Negación de Constancia"]
            self.combo_filtrar_tipo = ctk.CTkComboBox(linea_busqueda, values=tipos_vals, font=("Inter", 10), state="readonly", height=25, width=180, command=lambda v: self.hist_buscar_general())
            self.combo_filtrar_tipo.set("")
            self.combo_filtrar_tipo.pack(side="left", padx=(8,6))
        except Exception:
            self.combo_filtrar_tipo = None

        try:
            # Filtro por estatus
            estados = ["", "Pendiente", "Completada", "Cancelada"]
            self.combo_filtrar_estatus = ctk.CTkComboBox(linea_busqueda, values=estados, font=("Inter", 10), state="readonly", height=25, width=140, command=lambda v: self.hist_buscar_general())
            self.combo_filtrar_estatus.set("")
            self.combo_filtrar_estatus.pack(side="left", padx=(8,6))
        except Exception:
            self.combo_filtrar_estatus = None

        try:
            # Rango de fechas: desde / hasta (ahora con selector de calendario si tkcalendar está instalado)
            self.entry_hist_fecha_desde = ctk.CTkEntry(linea_busqueda, width=110, height=25, placeholder_text="Desde dd/mm/yyyy")
            self.entry_hist_fecha_desde.pack(side="left", padx=(8,4))
            self.entry_hist_fecha_desde.bind("<KeyRelease>", lambda e: self.hist_buscar_general())
            try:
                self.btn_hist_fecha_desde = ctk.CTkButton(linea_busqueda, text="📅", width=28, height=25, corner_radius=6, command=lambda e=self.entry_hist_fecha_desde: self._open_calendar_for_entry(e))
                self.btn_hist_fecha_desde.pack(side="left", padx=(0,4))
            except Exception:
                # fallback a tkinter Button si CTkButton no funciona
                self.btn_hist_fecha_desde = tk.Button(linea_busqueda, text="📅", width=3, command=lambda e=self.entry_hist_fecha_desde: self._open_calendar_for_entry(e))
                self.btn_hist_fecha_desde.pack(side="left", padx=(0,4))

            self.entry_hist_fecha_hasta = ctk.CTkEntry(linea_busqueda, width=110, height=25, placeholder_text="Hasta dd/mm/yyyy")
            self.entry_hist_fecha_hasta.pack(side="left", padx=(4,8))
            self.entry_hist_fecha_hasta.bind("<KeyRelease>", lambda e: self.hist_buscar_general())
            try:
                self.btn_hist_fecha_hasta = ctk.CTkButton(linea_busqueda, text="📅", width=28, height=25, corner_radius=6, command=lambda e=self.entry_hist_fecha_hasta: self._open_calendar_for_entry(e))
                self.btn_hist_fecha_hasta.pack(side="left", padx=(0,8))
            except Exception:
                self.btn_hist_fecha_hasta = tk.Button(linea_busqueda, text="📅", width=3, command=lambda e=self.entry_hist_fecha_hasta: self._open_calendar_for_entry(e))
                self.btn_hist_fecha_hasta.pack(side="left", padx=(0,8))
        except Exception:
            self.entry_hist_fecha_desde = None
            self.entry_hist_fecha_hasta = None
            self.btn_hist_fecha_desde = None
            self.btn_hist_fecha_hasta = None

        # Filtro por cliente (reemplaza la barra de búsqueda general)
        try:
            self.combo_filtrar_cliente = ctk.CTkComboBox(linea_busqueda, values=[""], font=("Inter", 10), state="readonly", height=25, width=250, command=lambda v: self.hist_buscar_general())
            self.combo_filtrar_cliente.set("")
            self.combo_filtrar_cliente.pack(side="left", padx=(0, 8))
            try:
                ctk.CTkButton(linea_busqueda, text="Limpiar filtros", command=self.hist_limpiar_filtros, width=120, height=25, corner_radius=6, fg_color=STYLE["secundario"], text_color=STYLE["surface"]).pack(side="left", padx=(0,8))
            except Exception:
                tk.Button(linea_busqueda, text="Limpiar filtros", command=self.hist_limpiar_filtros, width=15).pack(side="left", padx=(0,8))
        except Exception:
            # fallback a entry si CTkComboBox no está disponible
            self.combo_filtrar_cliente = None
            self.entry_buscar_general = ctk.CTkEntry(
                linea_busqueda, width=250, height=25,
                corner_radius=6, placeholder_text="Cliente, folio, fecha, supervisor..."
            )
            self.entry_buscar_general.pack(side="left", padx=(0, 8))
            self.entry_buscar_general.bind("<KeyRelease>", self.hist_buscar_general)

        ctk.CTkButton(
            linea_busqueda, text="X",
            command=self.hist_limpiar_busqueda,
            width=40, height=25, corner_radius=6,
            fg_color=STYLE["advertencia"], text_color=STYLE["surface"]
        ).pack(side="left")
        # Botones de generación rápidos en la parte superior (derecha)
        try:
            # Contenedor derecho que agrupa la leyenda (arriba) y botones rápidos (abajo)
            right_side_frame = ctk.CTkFrame(barra_superior, fg_color='transparent')
            right_side_frame.pack(side='right')

            # --- Leyenda de colores (arriba, encima del botón Reporte EMA) ---
            legend_frame = ctk.CTkFrame(right_side_frame, fg_color='transparent')
            legend_frame.pack(side='top', pady=(4, 0), padx=(8, 12))

            def _add_legend_item(parent, color, text):
                f = ctk.CTkFrame(parent, fg_color='transparent')
                f.pack(side='left', padx=(6, 4))
                box = ctk.CTkFrame(f, width=14, height=14, corner_radius=4, fg_color=color)
                box.pack(side='left', padx=(0, 6))
                ctk.CTkLabel(f, text=text, font=("Inter", 10), text_color=STYLE["texto_oscuro"]).pack(side='left')

            # Orden: Amarillo = Pendiente, Rojo = Canceladas, Verde = Completada
            _add_legend_item(legend_frame, "#F0AD4E", "Amarillo = Pendiente")
            _add_legend_item(legend_frame, "#D9534F", "Rojo = Canceladas")
            _add_legend_item(legend_frame, "#28a745", "Verde = Completada")

            # --- Botones rápidos (debajo de la leyenda) ---
            gen_top = ctk.CTkFrame(right_side_frame, fg_color='transparent')
            gen_top.pack(side='top')

            ctk.CTkButton(
                gen_top, text="📊 Reporte EMA",
                command=self.descargar_excel_anual,
                height=28, width=120, corner_radius=8,
                fg_color=STYLE["primario"], hover_color="#D4BF22", text_color=STYLE["secundario"], font=("Inter", 11, "bold")
            ).pack(side='right')
        except Exception:
            pass

        # Espaciador para empujar todo a la izquierda (opcional)
        # ctk.CTkFrame(linea_busqueda, fg_color="transparent").pack(side="left", expand=True)

        # ===========================================================
        # TABLA CON ENCABEZADOS CORREGIDOS (como en la imagen)
        # ===========================================================
        tabla_container = ctk.CTkFrame(cont, fg_color="transparent", corner_radius=8)
        tabla_container.pack(fill="both", expand=True)

        # Encabezados: usamos los headings del Treeview directamente

        # ANCHOS MEJORADOS Y ENCABEZADOS COMO EN LA IMAGEN
        column_widths = [
                40,    # Folio (más pequeño)
                40,    # Acta (más pequeño)
                40,    # Inicio (más pequeño)
                40,    # Término (más pequeño)
                40,    # Hora Ini
                40,    # Hora Fin
                150,   # Cliente (más ancho)
                80,    # Supervisor (ligeramente más pequeño)
                90,    # Tipo de documento
                50,    # Estatus
                60,    # Folios (más compacto)
                400    # Acciones (ajustado: más ancho para mostrar todas las acciones)
            ]

        # Encabezados exactamente como en la imagen
        headers = [
            "Folio", "Acta", "Inicio", "Término", 
            "Hora Inicio", "Hora Fin", "Cliente", 
            "Supervisor", "Tipo de documento", "Estatus", "Folios", "Acciones"
        ]

        # Reemplazar cabecera y scroll por un Treeview virtualizado (más eficiente)
        # Configurar estilo del Treeview para que combine con tema
        style = ttk.Style()
        try:
            # Usar tema 'clam' para permitir colorear encabezados en la mayoría de plataformas
            try:
                style.theme_use('clam')
            except Exception:
                pass
            style.configure('clientes.Treeview', font=("Inter", 12), rowheight=30,
                            background=STYLE["surface"], fieldbackground=STYLE["surface"], foreground=STYLE["texto_oscuro"]) 
            style.configure('clientes.Treeview.Heading', font=("Inter", 10, "bold"), background=STYLE["secundario"], foreground=STYLE["surface"], relief='flat')
            style.map('clientes.Treeview.Heading', background=[('active', STYLE['secundario'])], foreground=[('active', STYLE['surface'])])
            # Selección menos intrusiva para mantener legibilidad
            style.map('clientes.Treeview', background=[('selected', '#d9f0ff')], foreground=[('selected', STYLE['texto_oscuro'])])
            # Scrollbar neutral para no abusar del color primario
            style.configure('Vertical.TScrollbar', troughcolor=STYLE['surface'], background=STYLE['borde'], arrowcolor=STYLE['texto_oscuro'])
        except Exception:
            pass

        # Contenedor para el Treeview
        tree_container = ctk.CTkFrame(tabla_container, fg_color=STYLE["fondo"])
        tree_container.pack(fill="both", expand=True)

        cols = [f"c{i}" for i in range(len(column_widths))]
        self.hist_tree = ttk.Treeview(tree_container, columns=cols, show='headings', style='clientes.Treeview')
        # Configurar encabezados y anchos (permitir estirar columnas excepto la de Acciones)
        last_idx = len(headers) - 1
        for i, h in enumerate(headers):
            self.hist_tree.heading(cols[i], text=h)
            try:
                stretch = False if i == last_idx else True
                anchor = 'w' if i == last_idx else 'center'
                self.hist_tree.column(cols[i], width=column_widths[i], anchor=anchor, stretch=stretch)
            except Exception:
                self.hist_tree.column(cols[i], width=100, anchor='center')

        # Asegurar que la columna 'Acciones' sea siempre visible y no se reduzca
        try:
            # Aumentar el ancho por defecto y el ancho mínimo de la columna 'Acciones'
            # para que siempre muestre las cuatro opciones sin recortarse.
            self.hist_tree.column(cols[-1], width=360, minwidth=300, stretch=False, anchor='w')
        except Exception:
            try:
                # Fallback seguro
                self.hist_tree.column(cols[-1], width=300, anchor='w')
            except Exception:
                pass

        # Scrollbars: vertical y horizontal — usar grid para posicionar correctamente
        try:
            vsb = ctk.CTkScrollbar(tree_container, orientation="vertical", command=self.hist_tree.yview)
            # CTkScrollbar uses theme colors; set appearance if supported
            try:
                vsb.configure(button_color=STYLE['borde'], fg_color=STYLE['borde'])
            except Exception:
                pass
            self.hist_tree.configure(yscrollcommand=vsb.set)
        except Exception:
            # Fallback to ttk if CTkScrollbar not available
            vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.hist_tree.yview, style='Vertical.TScrollbar')
            try:
                style.configure('Vertical.TScrollbar', troughcolor=STYLE['surface'], background=STYLE['borde'], arrowcolor=STYLE['texto_oscuro'])
            except Exception:
                pass
            try:
                vsb.configure(background=STYLE['borde'], troughcolor=STYLE['surface'], activebackground=STYLE['borde'])
            except Exception:
                pass
            self.hist_tree.configure(yscrollcommand=vsb.set)

        # Layout con grid: tree en (0,0), vsb en (0,1), hsb en (1,0) colspan 2
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)
        self.hist_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')

        # Crear overlay de acciones (botones) y enlazarlo al Treeview
        try:
            self._create_actions_overlay(tree_container, cols[-1])
        except Exception:
            pass

        # Map para acceder al registro original por iid
        self._hist_map = {}

        # Menú contextual para acciones por fila
        self.hist_context_menu = tk.Menu(self, tearoff=0)
        self.hist_context_menu.add_command(label="Folios", command=lambda: self._hist_menu_action('folios'))
        self.hist_context_menu.add_command(label="Archivos", command=lambda: self._hist_menu_action('archivos'))
        self.hist_context_menu.add_command(label="Editar", command=lambda: self._hist_menu_action('editar'))
        self.hist_context_menu.add_command(label="Borrar", command=lambda: self._hist_menu_action('borrar'))

        # Bind derecho y doble-click
        self.hist_tree.bind("<Button-3>", self._hist_show_context_menu)
        self.hist_tree.bind("<Double-1>", self._hist_on_double_click)
        # Click izquierdo en columna de acciones abrirá el menú contextual
        self.hist_tree.bind("<Button-1>", self._hist_on_left_click)

        # ===========================================================
        # PIE DE PÁGINA (COMO EN LA IMAGEN) - Layout mejorado
        # ===========================================================
        footer = ctk.CTkFrame(cont, fg_color="transparent", height=60)
        footer.pack(fill="x", pady=(10, 0))
        footer.pack_propagate(False)

        footer_content = ctk.CTkFrame(footer, fg_color="transparent")
        footer_content.pack(expand=True, fill="both", padx=12, pady=10)


        pag_left = ctk.CTkFrame(footer_content, fg_color="transparent")
        pag_center = ctk.CTkFrame(footer_content, fg_color="transparent")
        pag_right = ctk.CTkFrame(footer_content, fg_color="transparent")

        # Fijar anchos laterales para que actúen como 'zonas' pegadas a bordes
        pag_left.configure(width=130)
        pag_right.configure(width=130)
        pag_left.pack_propagate(False)
        pag_right.pack_propagate(False)

        pag_left.pack(side='left', fill='y')
        pag_center.pack(side='left', expand=True, fill='both')
        pag_right.pack(side='right', fill='y')

        # Botón Anterior pegado al borde izquierdo dentro del subframe izquierdo
        self.btn_hist_prev = ctk.CTkButton(
            pag_left, text="⏪ Anterior",
            command=self.hist_pagina_anterior,
            height=28, width=100, corner_radius=6,
            fg_color=STYLE["secundario"], text_color=STYLE["surface"],
            hover_color="#1a1a1a"
        )
        self.btn_hist_prev.pack(side='left', anchor='w', padx=(6,0))

        # Contador centrado en el área central: asegurar centrado absoluto
        # Quitamos widgets laterales que desalineen la etiqueta y la centramos
        self.hist_pagina_label = ctk.CTkLabel(
            pag_center, text="Página 1",
            font=("Inter", 10), text_color=STYLE["texto_oscuro"]
        )
        # pack con expand=True y sin side para centrar horizontalmente
        self.hist_pagina_label.pack(expand=True)

        # Botón Siguiente pegado al borde derecho dentro del subframe derecho
        self.btn_hist_next = ctk.CTkButton(
            pag_right, text="Siguiente ⏩",
            command=self.hist_pagina_siguiente,
            height=28, width=100, corner_radius=6,
            fg_color=STYLE["secundario"], text_color=STYLE["surface"],
            hover_color="#1a1a1a"
        )
        # Generación de reportes: botones movidos al área superior de búsqueda

        self.btn_hist_next.pack(side='right', anchor='e', padx=(0,6))

        # Note: los botones de EMA/Anual y Backup se muestran en la pestaña "Reportes".

        # Cargar data
        self._cargar_historial()
        self._poblar_historial_ui()
        # Asegurar que el dropdown de folios pendientes se rellene al iniciar
        try:
            if hasattr(self, '_refresh_pending_folios_dropdown'):
                self._refresh_pending_folios_dropdown()
        except Exception:
            pass

    def _construir_tab_clientes(self, parent):
        """Consulta de clientes y agregar nuevos clientes"""
        cont = ctk.CTkFrame(parent, fg_color=STYLE["surface"], corner_radius=8)
        cont.pack(fill="both", expand=True, padx=10, pady=(0,0), side="top", anchor="n")

        contenido = ctk.CTkFrame(cont, fg_color="transparent")
        # Permitir que el contenido central expanda para aprovechar todo el alto
        contenido.pack(fill="both", expand=True, pady=(0,0), side="top", anchor="n")

        # ===== Sección: Clientes - formulario + tabla =====
        clientes_frame = ctk.CTkFrame(contenido, fg_color="transparent")
        clientes_frame.pack(fill="both", expand=True, padx=10, pady=(0,0))
     
        clientes_frame.grid_columnconfigure(0, weight=6, minsize=600)
        clientes_frame.grid_columnconfigure(1, weight=6)
        clientes_frame.grid_rowconfigure(0, weight=1)
        
        form_frame = ctk.CTkScrollableFrame(
            clientes_frame,
            fg_color=STYLE["surface"],
            corner_radius=8,
            scrollbar_button_color=STYLE["borde"],
            scrollbar_button_hover_color=STYLE["borde"]
        )
        form_frame.grid(row=0, column=0, sticky="nsew", padx=(0,8), pady=(1,1))

        ctk.CTkLabel(form_frame, text="Agregar nuevo cliente", font=FONT_SUBTITLE, text_color=STYLE["texto_oscuro"]).pack(anchor="w", padx=10, pady=(4,4))

        # Campos mínimos sugeridos
        self.cliente_campos = {}
        campos = [
            ("RFC", 25), ("CLIENTE", 35), ("No. CONTRATO", 30), ("FECHA DE CONTRATO", 18), ("ACTIVIDAD", 20),
            ("CURP", 18)
        ]

        for k, w in campos:
            frame_k = ctk.CTkFrame(form_frame, fg_color="transparent")
            frame_k.pack(fill="x", padx=12, pady=(4,4))
            ctk.CTkLabel(frame_k, text=f"{k}:", font=FONT_SMALL, width=140, anchor="w", text_color=STYLE["texto_oscuro"]).pack(side="left")
            ent = ctk.CTkEntry(frame_k, placeholder_text=k, font=FONT_SMALL, height=30)
            ent.pack(side="left", fill="x", expand=True)
            self.cliente_campos[k] = ent

        # ===== Domicilios de almacén (seleccione 0,1 o 2 y rellene formularios) =====
        dom_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        dom_frame.pack(fill="x", padx=12, pady=(6,6))
        ctk.CTkLabel(dom_frame, text="Número de domicilios de almacén:", font=FONT_SMALL, text_color=STYLE["texto_oscuro"]).pack(anchor="w")

        dom_select_frame = ctk.CTkFrame(dom_frame, fg_color="transparent")
        dom_select_frame.pack(fill="x", pady=(6,4))
        ctk.CTkLabel(dom_select_frame, text="Domicilios de almacén (puede agregar varios):", font=FONT_SMALL, text_color=STYLE["texto_oscuro"]).pack(anchor="w")

        controls = ctk.CTkFrame(dom_frame, fg_color="transparent")
        controls.pack(fill="x", pady=(6,4))
        # Contenedor donde se añadirán subformularios dinámicos
        self.dom_container = ctk.CTkFrame(dom_frame, fg_color="transparent")
        self.dom_container.pack(fill="both", pady=(4,4))

        # Lista de campos por domicilio
        self.dom_fields = []
        self.max_domicilios = 20

        def _crear_subform(parent, idx):
            frame = ctk.CTkFrame(parent, fg_color=STYLE["surface"], corner_radius=6)
            frame.pack(fill="x", pady=(6,6), padx=6)
            campos_dom = {}
            campos = [
                ("CALLE Y No", 'CALLE_Y_NO'),
                ("COLONIA O POBLACION", 'COLONIA_O_POBLACION'),
                ("MUNICIPIO O ALCADIA", 'MUNICIPIO_O_ALCADIA'),
                ("CIUDAD O ESTADO", 'CIUDAD_O_ESTADO'),
                ("CP", 'CP')
            ]
            for label_text, key in campos:
                f = ctk.CTkFrame(frame, fg_color="transparent")
                f.pack(fill="x", padx=6, pady=(4,2))
                ctk.CTkLabel(f, text=f"{label_text}:", font=FONT_SMALL, width=140, anchor="w", text_color=STYLE["texto_oscuro"]).pack(side="left")
                e = ctk.CTkEntry(f, placeholder_text=label_text, font=FONT_SMALL, height=30)
                e.pack(side="left", fill="x", expand=True)
                campos_dom[key] = e
            svc_f = ctk.CTkFrame(frame, fg_color="transparent")
            svc_f.pack(fill="x", padx=6, pady=(4,6))
            ctk.CTkLabel(svc_f, text="SERVICIO:", font=FONT_SMALL, width=140, anchor="w", text_color=STYLE["texto_oscuro"]).pack(side="left")
            svc = ctk.CTkComboBox(svc_f, values=["DICTAMEN","CONSTANCIA"], font=FONT_SMALL, dropdown_font=FONT_SMALL, state="readonly", height=30)
            svc.set("DICTAMEN")
            svc.pack(side="left", fill="x", expand=True)
            campos_dom['SERVICIO'] = svc
            # Botón para quitar este subform
            btn_rm = ctk.CTkButton(frame, text="Quitar domicilio", fg_color=STYLE["peligro"], hover_color=STYLE["peligro"], text_color=STYLE["surface"], font=("Inter", 11, "bold"), command=lambda f=frame: self._remove_domicilio_by_frame(f))
            btn_rm.pack(anchor="e", padx=6, pady=(0,6))
            return frame, campos_dom

        # Botones agregar/remover globales
        self.btn_add_domicilio = ctk.CTkButton(controls, text="Añadir domicilio", fg_color=STYLE["primario"], hover_color=STYLE["primario"], text_color=STYLE["secundario"], font=("Inter", 11, "bold"), command=self._add_domicilio)
        self.btn_add_domicilio.pack(side="left")
        self.lbl_dom_count = ctk.CTkLabel(controls, text="0 domicilios", font=FONT_SMALL, text_color=STYLE["texto_oscuro"])
        self.lbl_dom_count.pack(side="left", padx=(8,0))

        # store creator for use
        self._crear_domicilio_subform = _crear_subform

        btns_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        btns_frame.pack(fill="x", padx=12, pady=(8,12))
        # Guardar/Actualizar (se actualiza dinámicamente cuando se edita)
        self.btn_guardar_cliente = ctk.CTkButton(btns_frame, text="Guardar cliente", command=self._guardar_cliente_desde_form, fg_color=STYLE["primario"], hover_color=STYLE["primario"], text_color=STYLE["secundario"], font=("Inter", 11, "bold"), height=34, corner_radius=8)
        self.btn_guardar_cliente.pack(side="left")
        self.btn_limpiar_cliente = ctk.CTkButton(btns_frame, text="Limpiar", command=self._limpiar_formulario_cliente, fg_color=STYLE["advertencia"], hover_color=STYLE["advertencia"], text_color=STYLE["surface"], font=("Inter", 11, "bold"), height=34, corner_radius=8)
        self.btn_limpiar_cliente.pack(side="left", padx=(8,0))

        # Tabla de clientes
        tabla_frame = ctk.CTkFrame(clientes_frame, fg_color=STYLE["surface"], corner_radius=8)
        tabla_frame.grid(row=0, column=1, sticky="nsew", padx=(10,0), pady=(0,5))
        # Aumentar ligeramente la altura mínima de la tabla para mejor visibilidad
        tabla_frame.grid_rowconfigure(0, weight=1, minsize=420)
        header_frame = ctk.CTkFrame(tabla_frame, fg_color="transparent")
        header_frame.pack(fill='x', padx=12, pady=(8,6))
        ctk.CTkLabel(header_frame, text="Clientes registrados", font=FONT_SUBTITLE, text_color=STYLE["texto_oscuro"]).pack(side='left')
        self.lbl_total_clientes = ctk.CTkLabel(header_frame, text="Total: 0", font=FONT_SMALL, text_color=STYLE["texto_oscuro"]) 
        self.lbl_total_clientes.pack(side='right')

        cols = ("RFC","CLIENTE","NÚMERO DE CONTRATO","ACTIVIDAD","SERVICIO","ACCIONES")

        # Barra de búsqueda y filtros para la tabla de clientes
        search_frame = ctk.CTkFrame(tabla_frame, fg_color="transparent")
        search_frame.pack(fill="x", padx=12, pady=(4,6))
        # Filtro por servicio (p. ej. DICTAMEN / CONSTANCIA)
        try:
            self.combo_filtrar_servicio_clientes = ctk.CTkComboBox(search_frame, values=["", "DICTAMEN", "CONSTANCIA"], font=FONT_SMALL, state="readonly", height=30, command=lambda v: self._refrescar_tabla_clientes())
            self.combo_filtrar_servicio_clientes.set("")
            self.combo_filtrar_servicio_clientes.pack(side="left", padx=(0,8))
        except Exception:
            # fallback: mantener compatibilidad si CTkComboBox falla
            self.combo_filtrar_servicio_clientes = None

        self.entry_buscar_cliente = ctk.CTkEntry(search_frame, placeholder_text="Buscar por cliente, RFC, servicio o contrato", font=FONT_SMALL, height=30)
        self.entry_buscar_cliente.pack(side="left", fill="x", expand=True)
        ctk.CTkButton(search_frame, text="Buscar", command=self._buscar_clientes, fg_color=STYLE["primario"], hover_color=STYLE["primario"], text_color=STYLE["secundario"], font=("Inter", 11, "bold"), height=32, corner_radius=8).pack(side="left", padx=(8,0))
        ctk.CTkButton(search_frame, text="Limpiar", command=self._limpiar_busqueda_clientes, fg_color=STYLE["advertencia"], hover_color=STYLE["advertencia"], text_color=STYLE["surface"], font=("Inter", 11, "bold"), height=32, corner_radius=8).pack(side="left", padx=(8,0))

        tree_container = tk.Frame(tabla_frame)
        # Exponer contenedor para poder ajustar columnas al cambiar tamaño
        self.tree_clientes_container = tree_container
        # Ajustar espacio para que la tabla ocupe la mayor parte del panel
        tree_container.pack(fill="both", expand=True, padx=12, pady=(0,0))

        # Estilo similar al historial
        style = ttk.Style()
        try:
            style.theme_use('clam')
        except Exception:
            pass
        try:
            style.configure('clientes.Treeview', font=("Inter", 12), background=STYLE["surface"], foreground=STYLE["texto_oscuro"], rowheight=30, fieldbackground=STYLE["surface"])
            style.configure('clientes.Treeview.Heading', background=STYLE["secundario"], foreground=STYLE["surface"], font=("Inter", 10, "bold"))
            # Selección menos intrusiva: color claro para destacar sin ocultar texto
            style.map('clientes.Treeview', background=[('selected', '#d9f0ff')], foreground=[('selected', STYLE['texto_oscuro'])])
            # Evitar usar el color primario en el scrollbar para no abusar del color
            style.configure('Vertical.TScrollbar', troughcolor=STYLE['surface'], background=STYLE['borde'], arrowcolor=STYLE['texto_oscuro'])
        except Exception:
            pass

        self.tree_clientes = ttk.Treeview(tree_container, columns=cols, show='headings', selectmode='browse', style='clientes.Treeview')
        # Establecer anchos más estrechos por columna para reducir el ancho total
        col_widths = {
            'RFC': 110,
            'CLIENTE': 220,
            'NÚMERO DE CONTRATO': 180,
            'ACTIVIDAD': 120,
            'SERVICIO': 120,
            'ACCIONES': 120
        }
        for c in cols:
            self.tree_clientes.heading(c, text=c)
            try:
                w = col_widths.get(c, 120)
                self.tree_clientes.column(c, width=w, anchor='w')
            except Exception:
                self.tree_clientes.column(c, width=120, anchor='w')

        # Bind para detectar clicks en la columna de acciones (última columna)
        try:
            self.tree_clientes.bind('<Button-1>', self._on_client_tree_click)
        except Exception:
            pass
        # Ajustar columnas al redimensionar el contenedor para mantener ACCIONES visible
        try:
            self.tree_clientes_container.bind('<Configure>', lambda e: self._adjust_clientes_columns())
        except Exception:
            pass

        # Vertical scrollbar con color primario; quitar scrollbar horizontal externa
        try:
            vsb = ctk.CTkScrollbar(tree_container, orientation="vertical", command=self.tree_clientes.yview)
            try:
                vsb.configure(button_color=STYLE['borde'], fg_color=STYLE['borde'])
            except Exception:
                pass
            self.tree_clientes.configure(yscrollcommand=vsb.set)
            self.tree_clientes.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
        except Exception:
            vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree_clientes.yview, style='Vertical.TScrollbar')
            try:
                style.configure('Vertical.TScrollbar', troughcolor=STYLE['surface'], background=STYLE['borde'], arrowcolor=STYLE['texto_oscuro'])
            except Exception:
                pass
            self.tree_clientes.configure(yscrollcommand=vsb.set)
            self.tree_clientes.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")

        # Botones de gestión de tabla
        tbl_btns = ctk.CTkFrame(tabla_frame, fg_color="transparent")
        tbl_btns.pack(fill="x", padx=12, pady=(0,10))
        # Subframes para separar botones: left, center (expand), right
        left_actions = ctk.CTkFrame(tbl_btns, fg_color="transparent")
        center_actions = ctk.CTkFrame(tbl_btns, fg_color="transparent")
        right_actions = ctk.CTkFrame(tbl_btns, fg_color="transparent")
        left_actions.pack(side="left")
        center_actions.pack(side="left", expand=True, fill="both")
        right_actions.pack(side="right")

        # Botón Refrescar a la izquierda
        ctk.CTkButton(left_actions, text="Refrescar", command=self._refrescar_tabla_clientes, fg_color=STYLE["exito"], hover_color=STYLE["exito"], text_color=STYLE["surface"], font=("Inter", 11, "bold"), height=32, corner_radius=8).pack(side="left")

        # Botones centrales (centrados) con espacio entre ellos
        ctk.CTkButton(center_actions, text="Editar", fg_color=STYLE["primario"], hover_color="#D4BF22", text_color=STYLE["secundario"], height=32, corner_radius=8, command=self._editar_cliente_seleccionado).pack(side="left", padx=12)
        # Botón para desmarcar / salir de modo edición
        ctk.CTkButton(center_actions, text="Desmarcar", fg_color=STYLE["advertencia"], hover_color="#d2693e", text_color=STYLE["surface"], height=32, corner_radius=8, command=self.desmarcar_cliente).pack(side="left", padx=12)
        ctk.CTkButton(center_actions, text="Eliminar", fg_color=STYLE["peligro"], hover_color="#c84a3d", text_color=STYLE["surface"], height=32, corner_radius=8, command=self._eliminar_cliente_seleccionado).pack(side="left", padx=12)

        # Botón para exportar todo el catálogo de clientes a Excel (derecha)
        try:
            ctk.CTkButton(right_actions, text="Catálogo de clientes", fg_color=STYLE['exito'], hover_color=STYLE['exito'], text_color=STYLE['surface'], height=36, corner_radius=8, font=("Inter", 11, "bold"), command=self._export_catalogo_clientes).pack(side="right")
        except Exception:
            pass

        # Poblar la tabla inicialmente
        try:
            self._refrescar_tabla_clientes()
        except Exception:
            pass

    def mostrar_inspectores(self):
        """Muestra la sección de Inspectores y oculta las demás"""
        # Ocultar todos los frames primero
        try:
            self.frame_principal.pack_forget()
        except Exception:
            pass
        try:
            self.frame_historial.pack_forget()
        except Exception:
            pass
        try:
            self.frame_reportes.pack_forget()
        except Exception:
            pass
        # Asegurarse de ocultar reportes ejecutivos si estaban visibles
        try:
            self.frame_reportes_ejecutivo.pack_forget()
        except Exception:
            pass
        # Mostrar inspectores
        try:
            self.frame_inspectores.pack(fill="both", expand=True)
        except Exception:
            pass

        # Actualizar estado de los botones
        try:
            self.btn_principal.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            self.btn_historial.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            self.btn_reportes.configure(fg_color=STYLE["surface"], text_color=STYLE["secundario"], border_color=STYLE["secundario"])
            self.btn_inspectores.configure(fg_color=STYLE["primario"], text_color=STYLE["secundario"], border_color=STYLE["primario"])
        except Exception:
            pass

    def _construir_tab_inspectores(self, parent):
        """Formulario simple para registrar inspectores y tabla de listado (almacena en data/Firmas.json)"""
        cont = ctk.CTkFrame(parent, fg_color=STYLE["surface"], corner_radius=8)
        cont.pack(fill="both", expand=True, padx=10, pady=(0,0), side="top", anchor="n")

        contenido = ctk.CTkFrame(cont, fg_color="transparent")
        contenido.pack(fill="both", expand=True, pady=(0,0), side="top", anchor="n")

        inspect_frame = ctk.CTkFrame(contenido, fg_color="transparent")
        inspect_frame.pack(fill="both", expand=True, padx=10, pady=(0,0))
        # Usar la misma proporción y tamaño mínimo que la pestaña Clientes
        inspect_frame.grid_columnconfigure(0, weight=6, minsize=600)
        inspect_frame.grid_columnconfigure(1, weight=6)
        inspect_frame.grid_rowconfigure(0, weight=1)

        form_frame = ctk.CTkScrollableFrame(inspect_frame, fg_color=STYLE["surface"], corner_radius=8)
        form_frame.grid(row=0, column=0, sticky="nsew", padx=(0,8), pady=(1,1))

        ctk.CTkLabel(form_frame, text="Agregar nuevo inspector", font=FONT_SUBTITLE, text_color=STYLE["texto_oscuro"]).pack(anchor="w", padx=10, pady=(4,4))


        self.inspector_campos = {}
        # Campos básicos (sin IMAGEN). Las normas se muestran como checkboxes más abajo.
        campos = [
            ("NOMBRE DE INSPECTOR", 40), ("CORREO", 30), ("FIRMA", 30), ("Puesto", 25), ("VIGENCIA", 20), ("Fecha de acreditación", 24), ("Referencia", 30)
        ]

        for k, w in campos:
            frame_k = ctk.CTkFrame(form_frame, fg_color="transparent")
            frame_k.pack(fill="x", padx=12, pady=(4,4))
            # usar el mismo ancho de etiqueta que Clientes
            ctk.CTkLabel(frame_k, text=f"{k}:", font=FONT_SMALL, width=140, anchor="w", text_color=STYLE["texto_oscuro"]).pack(side="left")
            ent = ctk.CTkEntry(frame_k, placeholder_text=k, font=FONT_SMALL, height=30)
            ent.pack(side="left", fill="x", expand=True)
            self.inspector_campos[k] = ent

        # --- Normas acreditadas (checkboxes) ---
        normas_label = ctk.CTkLabel(form_frame, text="Normas acreditadas:", font=FONT_SMALL, text_color=STYLE["texto_oscuro"]) 
        normas_label.pack(anchor='w', padx=12, pady=(8,2))
        normas_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        normas_frame.pack(fill="x", padx=12, pady=(2,8))
        # Cargar catálogo de normas desde data/Normas.json
        self.inspector_normas_vars = {}
        normas_path = os.path.join(DATA_DIR, 'Normas.json')
        normas_list = []
        try:
            if os.path.exists(normas_path):
                with open(normas_path, 'r', encoding='utf-8') as nf:
                    normas_json = json.load(nf)
                    for rec in normas_json:
                        nom = rec.get('NOM') or rec.get('NOMBRE')
                        if nom:
                            normas_list.append(nom)
        except Exception:
            normas_list = []
        # Crear checkboxes en dos columnas
        col = 0
        row = 0
        for i, nom in enumerate(normas_list):
            var = tk.BooleanVar(value=False)
            cb = ctk.CTkCheckBox(normas_frame, text=nom, variable=var)
            cb.grid(row=row, column=col, sticky='w', padx=6, pady=2)
            self.inspector_normas_vars[nom] = var
            col += 1
            if col >= 2:
                col = 0
                row += 1

        btns_frame = ctk.CTkFrame(form_frame, fg_color="transparent")
        btns_frame.pack(fill="x", padx=12, pady=(8,12))
        self.btn_guardar_inspector = ctk.CTkButton(btns_frame, text="Guardar inspector", command=self._guardar_inspector_desde_form, fg_color=STYLE["primario"], hover_color=STYLE["primario"], text_color=STYLE["secundario"], font=("Inter", 11, "bold"), height=34, corner_radius=8)
        self.btn_guardar_inspector.pack(side="left")
        self.btn_limpiar_inspector = ctk.CTkButton(btns_frame, text="Limpiar", command=self._limpiar_formulario_inspector, fg_color=STYLE["advertencia"], hover_color=STYLE["advertencia"], text_color=STYLE["surface"], font=("Inter", 11, "bold"), height=34, corner_radius=8)
        self.btn_limpiar_inspector.pack(side="left", padx=(8,0))

        # Tabla de inspectores
        tabla_frame = ctk.CTkFrame(inspect_frame, fg_color=STYLE["surface"], corner_radius=8)
        tabla_frame.grid(row=0, column=1, sticky="nsew", padx=(10,0), pady=(0,5))
        tabla_frame.grid_rowconfigure(0, weight=1, minsize=420)
        header_frame = ctk.CTkFrame(tabla_frame, fg_color="transparent")
        header_frame.pack(fill='x', padx=12, pady=(8,6))
        ctk.CTkLabel(header_frame, text="Inspectores registrados", font=FONT_SUBTITLE, text_color=STYLE["texto_oscuro"]).pack(side='left')
        self.lbl_total_inspectores = ctk.CTkLabel(header_frame, text="Total: 0", font=FONT_SMALL, text_color=STYLE["texto_oscuro"]) 
        self.lbl_total_inspectores.pack(side='right')

        # Barra de búsqueda y filtro para Inspectores
        search_frame = ctk.CTkFrame(tabla_frame, fg_color="transparent")
        search_frame.pack(fill="x", padx=12, pady=(4,6))
        self.entry_buscar_inspector = ctk.CTkEntry(search_frame, placeholder_text="Buscar por nombre, correo o firma", font=FONT_SMALL, height=30)
        self.entry_buscar_inspector.pack(side="left", fill="x", expand=True)
        # Filtro por puesto (texto libre)
        try:
            self.entry_filtro_puesto_inspector = ctk.CTkEntry(search_frame, placeholder_text="Filtrar por puesto", font=FONT_SMALL, height=30)
            self.entry_filtro_puesto_inspector.pack(side="left", padx=(8,0))
        except Exception:
            self.entry_filtro_puesto_inspector = None
        ctk.CTkButton(search_frame, text="Buscar", command=self._buscar_inspectores, fg_color=STYLE["primario"], hover_color=STYLE["primario"], text_color=STYLE["secundario"], font=("Inter", 11, "bold"), height=32, corner_radius=8).pack(side="left", padx=(8,0))
        ctk.CTkButton(search_frame, text="Limpiar", command=self._limpiar_busqueda_inspectores, fg_color=STYLE["advertencia"], hover_color=STYLE["advertencia"], text_color=STYLE["surface"], font=("Inter", 11, "bold"), height=32, corner_radius=8).pack(side="left", padx=(8,0))

        cols = ("NOMBRE","CORREO","FIRMA","PUESTO","VIGENCIA")
        tree_container = tk.Frame(tabla_frame)
        tree_container.pack(fill="both", expand=True, padx=12, pady=(0,0))

        style = ttk.Style()
        try:
            style.theme_use('clam')
        except Exception:
            pass
        try:
            # Usar el mismo estilo que la pestaña Clientes para apariencia idéntica
            style.configure('clientes.Treeview', font=("Inter", 12), background=STYLE["surface"], foreground=STYLE["texto_oscuro"], rowheight=30, fieldbackground=STYLE["surface"])
            style.configure('clientes.Treeview.Heading', background=STYLE["secundario"], foreground=STYLE["surface"], font=("Inter", 10, "bold"))
        except Exception:
            pass

        self.tree_inspectores = ttk.Treeview(tree_container, columns=cols, show='headings', selectmode='browse', style='clientes.Treeview')
        # Anchuras iniciales sugeridas (similar proporción a Clientes)
        col_widths = {
            'NOMBRE': 150,
            'CORREO': 180,
            'FIRMA': 120,
            'PUESTO': 200,
            'VIGENCIA': 120
        }
        for c in cols:
            self.tree_inspectores.heading(c, text=c)
            try:
                w = col_widths.get(c, 120)
                self.tree_inspectores.column(c, width=w, anchor='w')
            except Exception:
                self.tree_inspectores.column(c, width=120, anchor='w')

        try:
            vsb = ctk.CTkScrollbar(tree_container, orientation="vertical", command=self.tree_inspectores.yview)
            self.tree_inspectores.configure(yscrollcommand=vsb.set)
            self.tree_inspectores.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")
        except Exception:
            vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree_inspectores.yview)
            self.tree_inspectores.configure(yscrollcommand=vsb.set)
            self.tree_inspectores.pack(side="left", fill="both", expand=True)
            vsb.pack(side="right", fill="y")

        tbl_btns = ctk.CTkFrame(tabla_frame, fg_color="transparent")
        tbl_btns.pack(fill="x", padx=12, pady=(0,10))
        left_actions = ctk.CTkFrame(tbl_btns, fg_color="transparent")
        center_actions = ctk.CTkFrame(tbl_btns, fg_color="transparent")
        right_actions = ctk.CTkFrame(tbl_btns, fg_color="transparent")
        left_actions.pack(side="left")
        center_actions.pack(side="left", expand=True)
        right_actions.pack(side="right")

        ctk.CTkButton(left_actions, text="Refrescar", command=self._refrescar_tabla_inspectores, fg_color=STYLE["exito"], hover_color=STYLE["exito"], text_color=STYLE["surface"], font=("Inter", 11, "bold"), height=32, corner_radius=8).pack(side="left")
        # Centro: editar / desmarcar / eliminar
        ctk.CTkButton(center_actions, text="Editar", fg_color=STYLE["primario"], hover_color="#D4BF22", text_color=STYLE["secundario"], height=32, corner_radius=8, command=self._editar_inspector_seleccionado).pack(side="left", padx=12)
        ctk.CTkButton(center_actions, text="Desmarcar", fg_color=STYLE["advertencia"], hover_color="#d2693e", text_color=STYLE["surface"], height=32, corner_radius=8, command=self.desmarcar_inspector).pack(side="left", padx=12)
        ctk.CTkButton(center_actions, text="Eliminar", fg_color=STYLE["peligro"], hover_color="#c84a3d", text_color=STYLE["surface"], height=32, corner_radius=8, command=self._eliminar_inspector_seleccionado).pack(side="left", padx=12)
        try:
            ctk.CTkButton(right_actions, text="Exportar catálogo", fg_color=STYLE['exito'], hover_color=STYLE['exito'], text_color=STYLE['surface'], height=36, corner_radius=8, font=("Inter", 11, "bold"), command=self._export_catalogo_inspectores).pack(side="right")
        except Exception:
            pass

        # Poblar inicialmente
        try:
            self._refrescar_tabla_inspectores()
        except Exception:
            pass

        # Exponer contenedor para poder ajustar columnas al cambiar tamaño (igual que Clientes)
        try:
            self.tree_inspectores_container = tree_container
            self.tree_inspectores_container.bind('<Configure>', lambda e: self._adjust_inspectores_columns())
        except Exception:
            pass


    # -------------------- Inspectores helpers --------------------
    def cargar_inspectores_desde_json(self):
        posibles = [os.path.join(DATA_DIR, 'Firmas.json'), 'data/Firmas.json', 'Firmas.json']
        ruta = None
        for p in posibles:
            try:
                if os.path.exists(p):
                    ruta = p
                    break
            except Exception:
                continue
        if not ruta:
            self.inspectores_data = []
            return
        try:
            with open(ruta, 'r', encoding='utf-8') as f:
                datos = json.load(f)
        except Exception:
            datos = []
        self.inspectores_data = datos if isinstance(datos, list) else []

    def _guardar_inspector_desde_form(self):
        nuevo = {}
        for k, ent in (self.inspector_campos or {}).items():
            try:
                v = ent.get().strip()
            except Exception:
                try:
                    v = ent.get() or ""
                except Exception:
                    v = ""
            nuevo[k] = v

        # Normas seleccionadas desde checkboxes
        try:
            normas_sel = [n for n, var in (self.inspector_normas_vars or {}).items() if getattr(var, 'get', lambda: False)()]
            nuevo['Normas acreditadas'] = normas_sel
        except Exception:
            pass

        ruta = os.path.join(DATA_DIR, 'Firmas.json')
        datos = []
        try:
            if os.path.exists(ruta):
                with open(ruta, 'r', encoding='utf-8') as f:
                    datos = json.load(f) or []
        except Exception:
            datos = []

        actualizado = False
        # Si estamos editando, reemplazar el registro coincidente por FIRMA
        try:
            edit_key = getattr(self, 'editing_inspector_firma', None)
            if edit_key:
                for i, rec in enumerate(datos):
                    try:
                        if str(rec.get('FIRMA','')) == str(edit_key):
                            # Merge existing record with updated fields from form
                            # to preserve keys like 'IMAGEN' when not provided in form.
                            merged = rec.copy() if isinstance(rec, dict) else {}
                            merged.update(nuevo)
                            datos[i] = merged
                            actualizado = True
                            break
                    except Exception:
                        continue
        except Exception:
            pass

        if not actualizado:
            datos.append(nuevo)

        try:
            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(datos, f, ensure_ascii=False, indent=2)
            if actualizado:
                messagebox.showinfo('Inspector actualizado', 'El inspector se ha actualizado en Firmas.json')
            else:
                messagebox.showinfo('Inspector guardado', 'El inspector se ha guardado en Firmas.json')
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo guardar el inspector: {e}')
            return

        try:
            # reset editing flag
            if getattr(self, 'editing_inspector_firma', None):
                self.editing_inspector_firma = None
                try:
                    self.btn_guardar_inspector.configure(text="Guardar inspector")
                except Exception:
                    pass
        except Exception:
            pass

        try:
            self._refrescar_tabla_inspectores()
        except Exception:
            pass
        self._limpiar_formulario_inspector()

    def _limpiar_formulario_inspector(self):
        try:
            for ent in (self.inspector_campos or {}).values():
                try:
                    ent.delete(0, 'end')
                except Exception:
                    pass
        except Exception:
            pass
        # Resetear checkboxes de normas
        try:
            for var in (self.inspector_normas_vars or {}).values():
                try:
                    var.set(False)
                except Exception:
                    pass
        except Exception:
            pass

    def _refrescar_tabla_inspectores(self):
        self.cargar_inspectores_desde_json()
        try:
            for i in self.tree_inspectores.get_children():
                self.tree_inspectores.delete(i)
        except Exception:
            pass
        # Obtener filtros desde UI (si existen)
        q = ''
        try:
            q = (self.entry_buscar_inspector.get() or '').strip().lower()
        except Exception:
            q = ''
        puesto_filtro = ''
        try:
            if getattr(self, 'entry_filtro_puesto_inspector', None):
                puesto_filtro = (self.entry_filtro_puesto_inspector.get() or '').strip().lower()
        except Exception:
            puesto_filtro = ''

        total = 0
        for rec in (self.inspectores_data or []):
            try:
                nombre = (rec.get('NOMBRE DE INSPECTOR') or rec.get('NOMBRE') or '')
                correo = (rec.get('CORREO','') or '')
                firma = (rec.get('FIRMA','') or '')
                puesto = (rec.get('Puesto','') or '')
                vig = (rec.get('VIGENCIA','') or '')

                # Aplicar filtros: búsqueda por texto y filtro por puesto
                aceptar = True
                if q:
                    hay = False
                    for field in (nombre, correo, firma, puesto):
                        try:
                            if q in str(field).lower():
                                hay = True
                                break
                        except Exception:
                            continue
                    if not hay:
                        aceptar = False
                if aceptar and puesto_filtro:
                    try:
                        if puesto_filtro not in str(puesto).lower():
                            aceptar = False
                    except Exception:
                        aceptar = False

                if not aceptar:
                    continue

                self.tree_inspectores.insert('', 'end', values=(nombre, correo, firma, puesto, vig))
                total += 1
            except Exception:
                continue
        try:
            self.lbl_total_inspectores.configure(text=f"Total: {total}")
        except Exception:
            pass
        # Ajustar anchos de columnas según el contenido para que los textos sean visibles
        try:
            # Primero intentar auto-ajustar según contenido usando helper común
            self._auto_resize_tree_columns(self.tree_inspectores)
        except Exception:
            try:
                # como fallback, usar el ajuste para el contenedor
                self._adjust_inspectores_columns()
            except Exception:
                pass

    def _buscar_inspectores(self):
        """Busca inspectores por texto en nombre, correo, firma o puesto."""
        try:
            # delegar a _refrescar_tabla_inspectores que ya usa el texto del entry
            self._refrescar_tabla_inspectores()
        except Exception:
            pass

    def _limpiar_busqueda_inspectores(self):
        """Limpia la caja de búsqueda y el filtro de puesto para inspectores."""
        try:
            if getattr(self, 'entry_buscar_inspector', None):
                try:
                    self.entry_buscar_inspector.delete(0, tk.END)
                except Exception:
                    try:
                        self.entry_buscar_inspector.delete(0, 'end')
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            if getattr(self, 'entry_filtro_puesto_inspector', None):
                try:
                    self.entry_filtro_puesto_inspector.delete(0, tk.END)
                except Exception:
                    try:
                        self.entry_filtro_puesto_inspector.delete(0, 'end')
                    except Exception:
                        pass
        except Exception:
            pass
        try:
            self._refrescar_tabla_inspectores()
        except Exception:
            pass

    def _export_catalogo_inspectores(self):
        """Exporta `data/Firmas.json` (inspectores) a un archivo Excel.
        Incluye las normas acreditadas concatenadas en la columna 'NORMAS'.
        """
        ruta = os.path.join(DATA_DIR, 'Firmas.json')
        datos = []
        try:
            if os.path.exists(ruta):
                with open(ruta, 'r', encoding='utf-8') as f:
                    datos = json.load(f) or []
        except Exception as e:
            messagebox.showerror('Exportar', f'No se pudo leer Firmas.json: {e}')
            return

        if not datos:
            messagebox.showinfo('Exportar', 'No hay datos de inspectores para exportar.')
            return

        # Procesar normas para que cada norma ocupe su propia columna (NORMA 1, NORMA 2, ...)
        processed = []
        # usar regex para separar cadenas de normas cuando sea necesario
        import re
        max_normas = 0
        for rec in datos:
            try:
                normas = rec.get('Normas acreditadas') or rec.get('Normas') or rec.get('NORMAS') or []
                if not isinstance(normas, (list, tuple)):
                    # intentar deserializar separadores comunes
                    if isinstance(normas, str):
                        normas_list = [s.strip() for s in re.split(r';|,|\n', normas) if s.strip()]
                    else:
                        normas_list = [str(normas)]
                else:
                    normas_list = [str(n).strip() for n in normas]
                    # ordenar natural (por números dentro de la cadena) de menor a mayor
                    def _natural_key(x):
                        parts = re.split(r'(\d+)', x)
                        return [int(p) if p.isdigit() else p.lower() for p in parts]
                    try:
                        normas_list = sorted(normas_list, key=_natural_key)
                    except Exception:
                        normas_list = sorted(normas_list)
                max_normas = max(max_normas, len(normas_list))
                processed.append((rec, normas_list))
            except Exception:
                processed.append((rec, []))

        try:
            save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel','*.xlsx')], title='Guardar catálogo de inspectores')
            if not save_path:
                return

            # Construir filas con columnas de normas separadas
            rows = []
            for rec, normas_list in processed:
                nombre = rec.get('NOMBRE DE INSPECTOR') or rec.get('NOMBRE') or ''
                correo = rec.get('CORREO') or rec.get('EMAIL') or ''
                firma = rec.get('FIRMA','')
                puesto = rec.get('Puesto') or rec.get('PUESTO') or ''
                vig = rec.get('VIGENCIA','')
                row = {
                    'NOMBRE': nombre,
                    'CORREO': correo,
                    'FIRMA': firma,
                    'PUESTO': puesto,
                    'VIGENCIA': vig
                }
                # Añadir columnas NORMA 1..N
                for i in range(max_normas):
                    key = f'NORMA {i+1}'
                    row[key] = normas_list[i] if i < len(normas_list) else ''
                rows.append(row)

            # Columnas base + dinamicas de normas
            base_cols = ['NOMBRE','CORREO','FIRMA','PUESTO','VIGENCIA']
            norma_cols = [f'NORMA {i+1}' for i in range(max_normas)]
            df = pd.DataFrame(rows, columns=base_cols + norma_cols)

            try:
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Inspectores')
                    try:
                        ws = writer.sheets['Inspectores']
                        from openpyxl.utils import get_column_letter
                        for i, col in enumerate(df.columns, 1):
                            col_letter = get_column_letter(i)
                            try:
                                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                                ws.column_dimensions[col_letter].width = max_len
                            except Exception:
                                pass
                    except Exception:
                        pass
            except Exception:
                pass
            except Exception:
                # fallback sin ajustar anchos
                df.to_excel(save_path, index=False)

            messagebox.showinfo('Exportar', f'Catálogo de inspectores exportado correctamente a:\n{save_path}')
        except Exception as e:
            messagebox.showerror('Exportar', f'Error al exportar: {e}')

    def _adjust_inspectores_columns(self):
        """Ajusta anchos de columnas en `self.tree_inspectores` para mantener apariencia consistente con Clientes."""
        try:
            tree = getattr(self, 'tree_inspectores', None)
            cont = getattr(self, 'tree_inspectores_container', None)
            if not tree or not cont:
                return
            total_w = cont.winfo_width() or tree.winfo_width()
            if not total_w or total_w < 100:
                return

            padding = 8
            min_col = 80
            cols = list(tree['columns'])

            # mínimos por columna para inspectores
            desired_mins = {
                'NOMBRE': 150,
                'CORREO': 180,
                'FIRMA': 120,
                'PUESTO': 200,
                'VIGENCIA': 100
            }

            available = max(50, total_w - padding)
            sum_mins = sum(desired_mins.get(c, 80) for c in cols)

            new_widths = {}
            if sum_mins <= available:
                extra = available - sum_mins
                for c in cols:
                    base = desired_mins.get(c, 80)
                    add = 0
                    if c == 'NOMBRE' and extra > 0:
                        add = extra
                    new_widths[c] = max(min_col, int(base + add))
            else:
                min_hard = 60
                total_weight = sum(desired_mins.get(c, 80) for c in cols)
                for c in cols:
                    weight = desired_mins.get(c, 80)
                    w = int(max(min_hard, available * (weight / total_weight)))
                    new_widths[c] = w

            for c, w in new_widths.items():
                try:
                    tree.column(c, width=w)
                except Exception:
                    pass
        except Exception:
            pass

    def _editar_inspector_seleccionado(self):
        sel = None
        try:
            sel = self.tree_inspectores.selection()
        except Exception:
            sel = None
        if not sel:
            messagebox.showinfo('Seleccionar', 'Seleccione un inspector en la tabla primero.')
            return
        iid = sel[0]
        vals = self.tree_inspectores.item(iid).get('values') or []
        if not vals:
            messagebox.showinfo('Seleccionar', 'No se pudo obtener el registro seleccionado.')
            return
        # Buscar registro en datos por FIRMA o NOMBRE
        key_firma = vals[2] if len(vals) > 2 else None
        target = None
        for rec in (self.inspectores_data or []):
            try:
                if key_firma and str(rec.get('FIRMA','')) == str(key_firma):
                    target = rec
                    break
                if str(rec.get('NOMBRE DE INSPECTOR','')) == str(vals[0]):
                    target = rec
                    break
            except Exception:
                continue
        if not target:
            messagebox.showinfo('Error', 'No se encontró el registro en los datos.')
            return
        # Poblar formulario
        try:
            for k, ent in (self.inspector_campos or {}).items():
                try:
                    val = target.get(k) or target.get(k.upper()) or ''
                    ent.delete(0, 'end')
                    ent.insert(0, str(val))
                except Exception:
                    continue
        except Exception:
            pass
        # Poblar checkboxes de normas
        try:
            normas = target.get('Normas acreditadas') or target.get('Normas') or []
            for n, var in (self.inspector_normas_vars or {}).items():
                try:
                    var.set(n in normas)
                except Exception:
                    continue
        except Exception:
            pass
        # Marcar modo edición
        try:
            self.editing_inspector_firma = target.get('FIRMA') or target.get('FIRMA','')
            try:
                self.btn_guardar_inspector.configure(text="Actualizar inspector")
            except Exception:
                pass
        except Exception:
            pass

    def desmarcar_inspector(self):
        try:
            # quitar selección en la tabla
            sels = self.tree_inspectores.selection()
            for s in sels:
                try:
                    self.tree_inspectores.selection_remove(s)
                except Exception:
                    pass
        except Exception:
            pass
        # limpiar formulario y reset edición
        try:
            self._limpiar_formulario_inspector()
        except Exception:
            pass
        try:
            if getattr(self, 'editing_inspector_firma', None):
                self.editing_inspector_firma = None
                try:
                    self.btn_guardar_inspector.configure(text="Guardar inspector")
                except Exception:
                    pass
        except Exception:
            pass

    def _eliminar_inspector_seleccionado(self):
        sel = None
        try:
            sel = self.tree_inspectores.selection()
        except Exception:
            sel = None
        if not sel:
            messagebox.showinfo('Seleccionar', 'Seleccione un inspector en la tabla primero.')
            return
        iid = sel[0]
        vals = self.tree_inspectores.item(iid).get('values') or []
        key_firma = vals[2] if len(vals) > 2 else None
        if not key_firma:
            messagebox.showinfo('Error', 'No se pudo identificar el inspector seleccionado.')
            return
        if not messagebox.askyesno('Confirmar', f'¿Eliminar el inspector {vals[0]}?'):
            return

        ruta = os.path.join(DATA_DIR, 'Firmas.json')
        try:
            with open(ruta, 'r', encoding='utf-8') as f:
                datos = json.load(f) or []
        except Exception:
            datos = []

        nuevo_arr = []
        eliminado = False
        for rec in (datos or []):
            try:
                if str(rec.get('FIRMA', '')).strip().upper() != str(key_firma).strip().upper():
                    nuevo_arr.append(rec)
                else:
                    eliminado = True
            except Exception:
                # si falla la comparación, mantener el registro para no perder datos
                nuevo_arr.append(rec)

        try:
            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(nuevo_arr, f, indent=2, ensure_ascii=False)
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo actualizar {ruta}: {e}')
            return

        if eliminado:
            messagebox.showinfo('Eliminado', 'Inspector eliminado correctamente')
        else:
            messagebox.showwarning('No encontrado', 'No se encontró el inspector en el archivo')

        try:
            self._refrescar_tabla_inspectores()
        except Exception:
            pass
            # Limpiar y estandarizar la cadena
            hora_str = str(hora_str).strip()
            
            # Si ya contiene AM/PM, devolver tal cual (pero limpiando espacios)
            hora_str_upper = hora_str.upper()
            if "AM" in hora_str_upper or "PM" in hora_str_upper:
                # Ya está en formato 12h, solo limpiar
                # Asegurar que AM/PM estén separados correctamente
                if "AM" in hora_str_upper:
                    hora_str = hora_str_upper.replace("AM", " AM")
                elif "PM" in hora_str_upper:
                    hora_str = hora_str_upper.replace("PM", " PM")
                return hora_str.strip()
            
            # Reemplazar punto por dos puntos (por si viene como "17.25")
            hora_str = hora_str.replace(".", ":")
            
            # Parsear la hora
            if ":" in hora_str:
                partes = hora_str.split(":")
                hora = int(partes[0].strip())
                minutos = partes[1].strip()[:2]  # Tomar solo los primeros 2 dígitos
                
                # Formatear minutos a 2 dígitos
                if len(minutos) == 1:
                    minutos = f"0{minutos}"
                
                # Determinar AM/PM
                if hora == 0:
                    return f"12:{minutos} AM"
                elif hora < 12:
                    return f"{hora}:{minutos} AM"
                elif hora == 12:
                    return f"12:{minutos} PM"
                else:
                    return f"{hora-12}:{minutos} PM"
            else:
                # Si no tiene formato de hora, devolver tal cual
                return hora_str
        except Exception as e:
            print(f"⚠️ Error formateando hora {hora_str}: {e}")
            return hora_str

    def crear_footer(self):
        footer = ctk.CTkFrame(self, fg_color=STYLE["fondo"], corner_radius=0, height=40)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)

        footer_content = ctk.CTkFrame(footer, fg_color="transparent")
        footer_content.pack(expand=True, fill="both", padx=25, pady=10)

        ctk.CTkLabel(
            footer_content,
            text="Sistema V&C - Generador de Dictámenes de Cumplimiento",
            font=("Inter", 10),
            text_color=STYLE["secundario"]
        ).pack(side="left")
        # Label para mostrar el siguiente folio de documento (visual, no persiste salvo confirmación)
        try:
            self.lbl_siguiente_folio_doc = ctk.CTkLabel(
                footer_content,
                text="Siguiente folio documento: ---",
                font=("Inter", 10),
                text_color=STYLE["secundario"]
            )
            self.lbl_siguiente_folio_doc.pack(side="right")
            # Actualizar valor inicial
            try:
                self._update_siguiente_folio_label()
            except Exception:
                pass
        except Exception:
            pass

    def _open_calendar_for_entry(self, entry_widget):
        """Abre un calendario emergente (tkcalendar.Calendar) y coloca la fecha seleccionada en `entry_widget`.

        Si `tkcalendar` no está disponible muestra un aviso informando cómo instalarlo.
        """
        try:
            if Calendar is None:
                messagebox.showinfo("Selector de fecha no disponible", "Para usar el selector de calendario instale la librería 'tkcalendar' (pip install tkcalendar).\nMientras tanto puede escribir la fecha en formato dd/mm/yyyy.")
                return

            top = tk.Toplevel(self)
            top.transient(self)
            top.grab_set()
            # Centrar la ventana pequeña sobre la app principal
            try:
                x = self.winfo_rootx() + 50
                y = self.winfo_rooty() + 80
                top.geometry(f"+{x}+{y}")
            except Exception:
                pass

            cal = Calendar(top, selectmode='day', date_pattern='dd/mm/yyyy')
            cal.pack(padx=8, pady=8)

            btn_frame = tk.Frame(top)
            btn_frame.pack(pady=(0,8))

            def on_ok():
                try:
                    sel = cal.get_date()
                    # Insertar en el entry (compatible con CTkEntry)
                    try:
                        entry_widget.delete(0, tk.END)
                        entry_widget.insert(0, sel)
                    except Exception:
                        # Si CTkEntry no soporta delete/insert directamente
                        entry_widget.set(sel)
                    top.destroy()
                    # disparar búsqueda para actualizar la lista
                    try:
                        self.hist_buscar_general()
                    except Exception:
                        pass
                except Exception:
                    top.destroy()

            def on_cancel():
                top.destroy()

            ok = tk.Button(btn_frame, text="OK", width=8, command=on_ok)
            ok.pack(side="left", padx=6)
            cancel = tk.Button(btn_frame, text="Cancelar", width=8, command=on_cancel)
            cancel.pack(side="left", padx=6)

        except Exception as e:
            try:
                messagebox.showerror("Error", f"No se pudo abrir el selector de fecha: {e}")
            except Exception:
                pass

    # -----------------------------------------------------------
    # MÉTODOS PARA GESTIÓN DE CLIENTES
    # -----------------------------------------------------------
    def cargar_clientes_desde_json(self):
        """Carga `data/Clientes.json` y rellena el combo de clientes.

        Esta implementación es tolerante a variaciones en la clave del nombre
        (`CLIENTE` o `RAZÓN SOCIAL `) y no modifica el archivo en disco. Los
        datos leídos se guardan en `self.clientes_data` (lista original de dicts)
        y el combobox `self.combo_cliente` se rellena con los nombres detectados.
        """
        posibles_rutas = [
            os.path.join(DATA_DIR, 'Clientes.json'),
            os.path.join(BASE_DIR, 'Clientes.json'),
            'data/Clientes.json',
            'Clientes.json',
            '../data/Clientes.json'
        ]

        archivo_encontrado = None
        for ruta in posibles_rutas:
            try:
                if os.path.exists(ruta):
                    archivo_encontrado = ruta
                    break
            except Exception:
                continue

        if not archivo_encontrado:
            # No hay archivo; dejar combo con valor por defecto
            try:
                self.combo_cliente.configure(values=['Seleccione un cliente...'])
                self.combo_cliente.set('Seleccione un cliente...')
            except Exception:
                pass
            self.clientes_data = []
            return

        try:
            with open(archivo_encontrado, 'r', encoding='utf-8') as f:
                datos = json.load(f)
            try:
                dbg = os.path.join(DATA_DIR, 'startup_exe_debug.log')
                with open(dbg, 'a', encoding='utf-8') as _d:
                    _d.write(f"cargar_clientes: opened {archivo_encontrado} size={os.path.getsize(archivo_encontrado)}\n")
            except Exception:
                pass
        except Exception:
            datos = []
            try:
                dbg = os.path.join(DATA_DIR, 'startup_exe_debug.log')
                with open(dbg, 'a', encoding='utf-8') as _d:
                    _d.write(f"cargar_clientes: FAILED to open {archivo_encontrado}\n")
            except Exception:
                pass

        # Guardar lista original en memoria
        self.clientes_data = datos if isinstance(datos, list) else []

        # Construir lista de nombres para mostrar en el combobox
        nombres = []
        for cliente in self.clientes_data:
            # Priorizar claves: 'CLIENTE' > 'RAZÓN SOCIAL ' > 'RAZON SOCIAL' > RFC/CONTRATO
            nombre = None
            if isinstance(cliente, dict):
                nombre = cliente.get('CLIENTE') or cliente.get('RAZÓN SOCIAL ') or cliente.get('RAZON SOCIAL') or cliente.get('RAZON_SOCIAL')
                if not nombre:
                    # Fallbacks
                    nombre = cliente.get('RFC') or cliente.get('NÚMERO_DE_CONTRATO') or cliente.get('NOMBRE')
            if nombre and isinstance(nombre, str) and nombre.strip() != '':
                nombres.append(nombre.strip())

        # Remover duplicados manteniendo orden
        seen = set()
        nombres_unicos = []
        for n in nombres:
            if n not in seen:
                seen.add(n)
                nombres_unicos.append(n)

        # Preparar valores para el combo
        valores = ['Seleccione un cliente...'] + sorted(nombres_unicos, key=lambda s: s.lower())

        try:
            self.combo_cliente.configure(values=valores)
            self.combo_cliente.set('Seleccione un cliente...')
        except Exception:
            pass

    def _guardar_cliente_desde_form(self):
        """Lee los campos del formulario de Reportes y guarda un nuevo cliente en Clientes.json"""
        # Construir el registro con la estructura estándar requerida
        nuevo = {
            'RFC': '',
            'CLIENTE': '',
            'NÚMERO_DE_CONTRATO': '',
            'ACTIVIDAD': '',
            'CURP': '',
            'DIRECCIONES': []
        }
        try:
            # Campos principales vienen de self.cliente_campos
            def _get_field(name):
                ent = (self.cliente_campos or {}).get(name)
                if not ent:
                    return ''
                try:
                    return ent.get().strip()
                except Exception:
                    try:
                        return ent.get() or ''
                    except Exception:
                        return ''

            nuevo['RFC'] = _get_field('RFC')
            nuevo['CLIENTE'] = _get_field('CLIENTE')
            # Mapear desde varios posibles labels en el formulario hacia 'NÚMERO_DE_CONTRATO'
            nro_candidates = ['NÚMERO_DE_CONTRATO', 'NÚMERO DE CONTRATO', 'No. DE CONTRATO', 'No. CONTRATO', 'NUMERO_DE_CONTRATO', 'NUMERO DE CONTRATO']
            nro_val = ''
            for c in nro_candidates:
                val = _get_field(c)
                if val:
                    nro_val = val
                    break
            nuevo['NÚMERO_DE_CONTRATO'] = nro_val
            # Fecha de firma del contrato: mapear desde el campo agregado 'FECHA DE CONTRATO'
            nuevo['FECHA_DE_CONTRATO'] = _get_field('FECHA DE CONTRATO') or _get_field('FECHA_DE_CONTRATO')
            nuevo['ACTIVIDAD'] = _get_field('ACTIVIDAD')
            nuevo['CURP'] = _get_field('CURP')
        except Exception:
            pass

        # Recopilar domicilios y almacenarlos en 'Direcciones' con claves estándar
        try:
            direcciones = []
            for rec in (self.dom_fields or []):
                try:
                    fields = rec.get('fields') or {}
                    d = {
                        'CALLE Y NO': '',
                        'COLONIA O POBLACION': '',
                        'MUNICIPIO O ALCADIA': '',
                        'CIUDAD O ESTADO': '',
                        'CP': '',
                        'SERVICIO': ''
                    }
                    # Los campos internos usan claves como 'CALLE_Y_NO', 'COLONIA_O_POBLACION', etc.
                    try:
                        if 'CALLE_Y_NO' in fields:
                            d['CALLE Y NO'] = (fields['CALLE_Y_NO'].get() or '').strip()
                        elif 'CALLE Y No' in fields:
                            d['CALLE Y NO'] = (fields['CALLE Y No'].get() or '').strip()
                    except Exception:
                        pass
                    try:
                        if 'COLONIA_O_POBLACION' in fields:
                            d['COLONIA O POBLACION'] = (fields['COLONIA_O_POBLACION'].get() or '').strip()
                    except Exception:
                        pass
                    try:
                        if 'MUNICIPIO_O_ALCADIA' in fields:
                            d['MUNICIPIO O ALCADIA'] = (fields['MUNICIPIO_O_ALCADIA'].get() or '').strip()
                    except Exception:
                        pass
                    try:
                        if 'CIUDAD_O_ESTADO' in fields:
                            d['CIUDAD O ESTADO'] = (fields['CIUDAD_O_ESTADO'].get() or '').strip()
                    except Exception:
                        pass
                    try:
                        if 'CP' in fields:
                            d['CP'] = (fields['CP'].get() or '').strip()
                    except Exception:
                        pass
                    try:
                        if 'SERVICIO' in fields:
                            d['SERVICIO'] = fields['SERVICIO'].get() if hasattr(fields['SERVICIO'], 'get') else ''
                    except Exception:
                        pass
                    direcciones.append(d)
                except Exception:
                    continue
            if direcciones:
                nuevo['DIRECCIONES'] = direcciones
        except Exception:
            pass

        # Ruta objetivo
        ruta = os.path.join(DATA_DIR, 'Clientes.json')
        datos = []
        try:
            if os.path.exists(ruta):
                with open(ruta, 'r', encoding='utf-8') as f:
                    datos = json.load(f) or []
        except Exception:
            datos = []

        # Si estamos en modo edición, intentar reemplazar el registro existente
        actualizado = False
        try:
            if getattr(self, 'editing_cliente_rfc', None):
                busc_rfc = str(self.editing_cliente_rfc)
                for i, c in enumerate(datos):
                    try:
                        if str(c.get('RFC', '')) == busc_rfc:
                            datos[i] = nuevo
                            actualizado = True
                            break
                    except Exception:
                        continue
            if not actualizado:
                datos.append(nuevo)

            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(datos, f, ensure_ascii=False, indent=2)

            if actualizado:
                messagebox.showinfo('Cliente actualizado', 'El cliente se ha actualizado en Clientes.json')
            else:
                messagebox.showinfo('Cliente guardado', 'El cliente se ha guardado en Clientes.json')
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo guardar el cliente: {e}')
            return

        # Refrescar cache y UI
        try:
            self.cargar_clientes_desde_json()
        except Exception:
            pass
        try:
            self._refrescar_tabla_clientes()
        except Exception:
            pass

        # Limpiar formulario
        for ent in (self.cliente_campos or {}).values():
            try:
                ent.delete(0, 'end')
            except Exception:
                pass
        # Si veníamos editando, resetear estado del botón
        try:
            if getattr(self, 'editing_cliente_rfc', None):
                self.editing_cliente_rfc = None
                try:
                    self.btn_guardar_cliente.configure(text="Guardar cliente")
                except Exception:
                    pass
                # Después de editar un cliente, limpiar la selección en la UI
                try:
                    self.limpiar_cliente()
                except Exception:
                    pass
        except Exception:
            pass

    def _limpiar_formulario_cliente(self):
        """Limpia los campos del formulario de cliente y elimina subformularios de domicilios."""
        try:
            for ent in (self.cliente_campos or {}).values():
                try:
                    ent.delete(0, 'end')
                except Exception:
                    pass
        except Exception:
            pass
        # quitar subformularios de domicilios
        try:
            for idx, rec in enumerate(list(self.dom_fields)):
                frm = rec.get('frame')
                try:
                    frm.destroy()
                except Exception:
                    pass
            self.dom_fields = []
            try:
                self.lbl_dom_count.configure(text="0 domicilios")
            except Exception:
                pass
        except Exception:
            pass

    def _refrescar_tabla_clientes(self):
        """Carga `Clientes.json` y muestra los registros en la tabla Treeview."""
        ruta = os.path.join(DATA_DIR, 'Clientes.json')
        datos = []
        try:
            if os.path.exists(ruta):
                with open(ruta, 'r', encoding='utf-8') as f:
                    datos = json.load(f) or []
        except Exception:
            datos = []

        # Limpiar tabla
        try:
            for r in self.tree_clientes.get_children():
                self.tree_clientes.delete(r)
        except Exception:
            pass

        # Obtener filtro de servicio desde UI (si existe)
        filtro_servicio = ''
        try:
            if getattr(self, 'combo_filtrar_servicio_clientes', None):
                filtro_servicio = (self.combo_filtrar_servicio_clientes.get() or '').strip().lower()
        except Exception:
            filtro_servicio = ''

        # Insertar filas
        for c in datos:
            try:
                rfc = c.get('RFC') or c.get('R.F.C') or ''
                nombre = c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or ''
                contrato = c.get('NÚMERO_DE_CONTRATO') or c.get('No. DE CONTRATO') or ''
                actividad = c.get('ACTIVIDAD') or ''
                # Si no hay CP/SERVICIO a nivel cliente, intentar usar el primer domicilio
                # CP se omite para la vista; aún lo extraemos por compatibilidad
                cp = c.get('CP') or ''
                servicio = c.get('SERVICIO') or ''
                try:
                    # soportar tanto 'DIRECCIONES' (nuevo) como 'DOMICILIOS' (antiguo)
                    direc = c.get('DIRECCIONES') if c.get('DIRECCIONES') is not None else c.get('DOMICILIOS')
                    if (not cp or not servicio) and isinstance(direc, (list, tuple)) and len(direc) > 0:
                        primera = direc[0]
                        if not cp:
                            cp = primera.get('CP') or primera.get('cp') or ''
                        if not servicio:
                            servicio = primera.get('SERVICIO') or primera.get('servicio') or ''
                except Exception:
                    pass
                # Aplicar filtro de servicio si se especificó
                if filtro_servicio:
                    try:
                        if filtro_servicio not in str(servicio).lower():
                            continue
                    except Exception:
                        continue
                # Insertar sin la columna CP; añadir texto en ACCIONES para permitir interacción
                self.tree_clientes.insert('', 'end', values=(rfc, nombre, contrato, actividad, servicio, 'Ver domicilios'))
            except Exception:
                continue

        # Auto-ajustar columnas según el contenido recién cargado
        try:
            self._auto_resize_tree_columns(self.tree_clientes)
        except Exception:
            pass

        # Asegurar que la columna ACCIONES permanezca visible ajustando anchos
        try:
            self._adjust_clientes_columns()
        except Exception:
            pass
        # Actualizar contador de clientes
        try:
            total = len(datos) if isinstance(datos, (list, tuple)) else 0
            try:
                self.lbl_total_clientes.configure(text=f"Total: {total}")
            except Exception:
                pass
        except Exception:
            pass

    def _buscar_clientes(self):
        """Busca clientes por texto en RFC, CLIENTE, NÚMERO DE CONTRATO o SERVICIO."""
        q = ''
        try:
            q = (self.entry_buscar_cliente.get() or '').strip().lower()
        except Exception:
            q = ''

        ruta = os.path.join(DATA_DIR, 'Clientes.json')
        datos = []
        try:
            if os.path.exists(ruta):
                with open(ruta, 'r', encoding='utf-8') as f:
                    datos = json.load(f) or []
        except Exception:
            datos = []

        # Limpiar tabla
        try:
            for r in self.tree_clientes.get_children():
                self.tree_clientes.delete(r)
        except Exception:
            pass

        # Obtener filtro de servicio (si existe)
        filtro_servicio = ''
        try:
            if getattr(self, 'combo_filtrar_servicio_clientes', None):
                filtro_servicio = (self.combo_filtrar_servicio_clientes.get() or '').strip().lower()
        except Exception:
            filtro_servicio = ''

        if not q:
            # si no hay query, recargar todo (aplicará el filtro si está presente)
            try:
                self._refrescar_tabla_clientes()
            except Exception:
                pass
            return

        for c in datos:
            try:
                rfc = (c.get('RFC') or c.get('R.F.C') or '')
                nombre = (c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or '')
                contrato = (c.get('NÚMERO_DE_CONTRATO') or c.get('No. DE CONTRATO') or '')
                servicio = (c.get('SERVICIO') or '')
                actividad = c.get('ACTIVIDAD') or ''
                cp = c.get('CP') or ''
                hay = False
                # Buscar en campos relevantes (omitimos CP de la lista de búsqueda visible)
                for field in (rfc, nombre, contrato, servicio, actividad):
                    try:
                        if q in str(field).lower():
                            hay = True
                            break
                    except Exception:
                        continue
                if hay:
                    # aplicar filtro de servicio a los resultados de búsqueda
                    if filtro_servicio:
                        try:
                            if filtro_servicio not in str(servicio).lower():
                                continue
                        except Exception:
                            continue
                    self.tree_clientes.insert('', 'end', values=(rfc, nombre, contrato, actividad, servicio, 'Ver domicilios'))
            except Exception:
                continue

    def _limpiar_busqueda_clientes(self):
        """Limpia la caja de búsqueda de clientes y recarga la tabla."""
        try:
            self.entry_buscar_cliente.delete(0, tk.END)
        except Exception:
            try:
                self.entry_buscar_cliente.delete(0, 'end')
            except Exception:
                pass
        try:
            self._refrescar_tabla_clientes()
        except Exception:
            pass

    def _eliminar_cliente_seleccionado(self):
        sel = None
        try:
            sel = self.tree_clientes.selection()[0]
        except Exception:
            messagebox.showwarning('Eliminar', 'No hay ningún cliente seleccionado')
            return

        vals = self.tree_clientes.item(sel, 'values')
        if not vals:
            return

        rfc_sel = vals[0]
        ruta = os.path.join(DATA_DIR, 'Clientes.json')
        try:
            with open(ruta, 'r', encoding='utf-8') as f:
                datos = json.load(f) or []
        except Exception:
            datos = []

        # Eliminar por RFC (si está vacío, por nombre)
        nuevos = []
        for c in datos:
            try:
                if rfc_sel and c.get('RFC') and str(c.get('RFC')) == str(rfc_sel):
                    continue
                nombre = c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or ''
                if not rfc_sel and nombre == vals[1]:
                    continue
                nuevos.append(c)
            except Exception:
                nuevos.append(c)

        # Confirmación fuerte: primero confirmar intención
        try:
            confirmar = messagebox.askyesno('Confirmar eliminación', f"¿Está seguro que desea eliminar al cliente '{vals[1]}'? Esta acción no se puede deshacer.")
            if not confirmar:
                return
            # Solicitar que escriba la palabra ELIMINAR para evitar borrados accidentales
            respuesta = simpledialog.askstring('Autenticación', "Escriba ELIMINAR para confirmar la eliminación:", parent=self)
            if not respuesta or respuesta.strip().upper() != 'ELIMINAR':
                messagebox.showinfo('Cancelado', 'Eliminación cancelada. No se escribieron las credenciales correctas.')
                return
        except Exception:
            # Si falla el diálogo, cancelar por seguridad
            messagebox.showwarning('Eliminar', 'No se pudo completar la confirmación. Cancelando eliminación.')
            return

        try:
            with open(ruta, 'w', encoding='utf-8') as f:
                json.dump(nuevos, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo eliminar: {e}')
            return

        messagebox.showinfo('Eliminado', 'Cliente eliminado correctamente')
        try:
            self._refrescar_tabla_clientes()
        except Exception:
            pass

    def _editar_cliente_seleccionado(self):
        """Carga el cliente seleccionado en el formulario para su edición."""
        sel = None
        try:
            sel = self.tree_clientes.selection()[0]
        except Exception:
            messagebox.showwarning('Editar', 'No hay ningún cliente seleccionado')
            return

        vals = self.tree_clientes.item(sel, 'values')
        if not vals:
            return

        rfc_sel = vals[0]
        ruta = os.path.join(DATA_DIR, 'Clientes.json')
        datos = []
        try:
            if os.path.exists(ruta):
                with open(ruta, 'r', encoding='utf-8') as f:
                    datos = json.load(f) or []
        except Exception:
            datos = []

        # Buscar registro completo por RFC (si no hay RFC, por nombre)
        registro = None
        for c in datos:
            try:
                if rfc_sel and str(c.get('RFC', '')) == str(rfc_sel):
                    registro = c
                    break
                nombre = c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or ''
                if not rfc_sel and nombre == vals[1]:
                    registro = c
                    break
            except Exception:
                continue

        if not registro:
            messagebox.showwarning('Editar', 'No se encontró el registro completo para edición')
            return

        # Llenar formulario con datos del registro
        for k, ent in (self.cliente_campos or {}).items():
            try:
                # Construir lista de posibles claves en el JSON para este campo
                candidates = []
                candidates.append(k)
                candidates.append(k.replace(' ', '_'))
                candidates.append(k.upper())
                candidates.append(k.replace('.', ''))
                candidates.append(k.replace('.', '').replace(' ', '_').upper())
                # Casos específicos mapeados
                if k.strip().lower().startswith('no') or 'contrato' in k.lower():
                    candidates = ['NÚMERO_DE_CONTRATO', 'NÚMERO DE CONTRATO', 'No. DE CONTRATO', 'No. CONTRATO', 'NUMERO_DE_CONTRATO', 'NUMERO DE CONTRATO'] + candidates
                # Buscar el primer valor existente
                v = ''
                for key in candidates:
                    try:
                        if key in registro and registro.get(key) is not None:
                            v = registro.get(key)
                            break
                    except Exception:
                        continue
                # Fallback: intentar buscar por coincidencia ignorando mayúsculas/minúsculas y signos
                if (v is None or v == '') and isinstance(registro, dict):
                    lk = k.replace('_', ' ').strip().lower()
                    for rk in registro.keys():
                        try:
                            if str(rk).strip().lower() == lk or str(rk).strip().lower().replace('.', '') == lk.replace('.', ''):
                                v = registro.get(rk)
                                break
                        except Exception:
                            continue

                try:
                    ent.delete(0, 'end')
                except Exception:
                    pass
                if v is None:
                    v = ''
                try:
                    ent.insert(0, str(v))
                except Exception:
                    pass
            except Exception:
                try:
                    ent.delete(0, 'end')
                except Exception:
                    pass

        # Llenar subformularios dinámicos de domicilios si existen
        try:
            # Aceptar varias claves posibles para domicilios: DIRECCIONES, DOMICILIOS, DOMICILIO
            domicilios = registro.get('DIRECCIONES') or registro.get('DOMICILIOS') or registro.get('DOMICILIO') or []
            if not isinstance(domicilios, (list, tuple)):
                domicilios = []
            # eliminar subformularios existentes
            try:
                for rec in list(self.dom_fields):
                    try:
                        rec.get('frame').destroy()
                    except Exception:
                        pass
                self.dom_fields = []
            except Exception:
                pass
            # crear subformularios y rellenar
            for i, ddata in enumerate(domicilios):
                try:
                    if len(self.dom_fields) >= self.max_domicilios:
                        break
                    frm, fields = self._crear_domicilio_subform(self.dom_container, len(self.dom_fields))
                    self.dom_fields.append({'frame': frm, 'fields': fields})
                    
                    key_variants = {
                        'CALLE_Y_NO': ['CALLE Y NO', 'CALLE_Y_NO', 'CALLE Y No', 'CALLEYNO', 'CALLE'],
                        'COLONIA_O_POBLACION': ['COLONIA O POBLACION', 'COLONIA_O_POBLACION', 'COLONIA', 'POBLACION'],
                        'MUNICIPIO_O_ALCADIA': ['MUNICIPIO O ALCADIA', 'MUNICIPIO_O_ALCADIA', 'MUNICIPIO', 'ALCADIA'],
                        'CIUDAD_O_ESTADO': ['CIUDAD O ESTADO', 'CIUDAD_O_ESTADO', 'CIUDAD', 'ESTADO'],
                        'CP': ['CP', 'cp', 'Codigo Postal', 'C.P.'],
                        'SERVICIO': ['SERVICIO', 'servicio']
                    }
                    for k, widget in (fields or {}).items():
                        try:
                            val = ''
                            if isinstance(ddata, dict):
                                # buscar por variantes de llave
                                variants = key_variants.get(k, [k, k.replace('_', ' '), k.lower()])
                                for key in variants:
                                    if key in ddata and ddata.get(key) is not None:
                                        val = ddata.get(key)
                                        break
                                # también probar claves con distintos capitalizations
                                if val == '':
                                    for key in list(ddata.keys()):
                                        if key.strip().lower() == k.replace('_', ' ').strip().lower():
                                            val = ddata.get(key)
                                            break
                            # Asignar al widget: combo -> set(), entry -> insert
                            try:
                                if hasattr(widget, 'set'):
                                    try:
                                        widget.set(str(val))
                                    except Exception:
                                        pass
                                else:
                                    try:
                                        widget.delete(0, 'end')
                                    except Exception:
                                        pass
                                    if val is not None:
                                        try:
                                            widget.insert(0, str(val))
                                        except Exception:
                                            pass
                            except Exception:
                                pass
                        except Exception:
                            pass
                except Exception:
                    continue
            try:
                self.lbl_dom_count.configure(text=f"{len(self.dom_fields)} domicilios")
            except Exception:
                pass
        except Exception:
            pass

        # Marcar modo edición
        try:
            self.editing_cliente_rfc = registro.get('RFC') or ''
            try:
                self.btn_guardar_cliente.configure(text="Actualizar cliente")
            except Exception:
                pass
        except Exception:
            pass

    def _on_client_tree_click(self, event):
        """Detecta clicks en la columna de ACCIONES y muestra domicilios del cliente."""
        try:
            col = self.tree_clientes.identify_column(event.x)
            iid = self.tree_clientes.identify_row(event.y)
            if not iid:
                return
            last_col = f"#{len(self.tree_clientes['columns'])}"
            if col == last_col:
                vals = self.tree_clientes.item(iid, 'values')
                if not vals:
                    return
                rfc = vals[0]
                try:
                    self._mostrar_domicilios_cliente(rfc)
                except Exception:
                    pass
                return "break"
        except Exception:
            pass

    def _mostrar_domicilios_cliente(self, rfc_or_name):
        """Abre una ventana modal mostrando los domicilios del cliente identificado por RFC o nombre."""
        ruta = os.path.join(DATA_DIR, 'Clientes.json')
        datos = []
        try:
            if os.path.exists(ruta):
                with open(ruta, 'r', encoding='utf-8') as f:
                    datos = json.load(f) or []
        except Exception:
            datos = []

        registro = None
        for c in datos:
            try:
                if rfc_or_name and c.get('RFC') and str(c.get('RFC')) == str(rfc_or_name):
                    registro = c
                    break
                nombre = c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or ''
                if (not registro) and str(nombre) == str(rfc_or_name):
                    registro = c
                    break
            except Exception:
                continue

        if not registro:
            messagebox.showwarning('Domicilios', 'No se encontró el cliente o no tiene domicilios registrados.')
            return

        domicilios = registro.get('DIRECCIONES') or registro.get('DOMICILIOS') or registro.get('DOMICILIO') or []
        if not isinstance(domicilios, (list, tuple)):
            domicilios = []

        # Crear ventana modal
        win = ctk.CTkToplevel(self)
        win.title(f"Domicilios - {registro.get('CLIENTE') or registro.get('RFC')}")
        win.geometry('640x420')
        win.transient(self)
        win.grab_set()

        frame = ctk.CTkFrame(win, fg_color=STYLE['surface'])
        frame.pack(fill='both', expand=True, padx=12, pady=12)

        hdr = ctk.CTkLabel(frame, text=f"Domicilios de {registro.get('CLIENTE') or registro.get('RFC')}", font=FONT_SUBTITLE, text_color=STYLE['texto_oscuro'])
        hdr.pack(anchor='w', pady=(0,8))

        scroll = ctk.CTkScrollableFrame(frame, fg_color='transparent')
        scroll.pack(fill='both', expand=True)

        if not domicilios:
            ctk.CTkLabel(scroll, text='No hay domicilios registrados.', text_color=STYLE['texto_oscuro']).pack(anchor='w', pady=6)
        else:
            for i, d in enumerate(domicilios, start=1):
                box = ctk.CTkFrame(scroll, fg_color=STYLE['surface'], corner_radius=6)
                box.pack(fill='x', pady=6, padx=6)
                ctk.CTkLabel(box, text=f"Domicilio {i}", font=("Inter", 12, "bold"), text_color=STYLE['texto_oscuro']).pack(anchor='w', padx=8, pady=(6,2))
                # Mostrar campos relevantes y soportar claves con espacios/variantes
                field_variants = [
                    ('CALLE Y NO', ['CALLE Y NO', 'CALLE_Y_NO', 'CALLE Y No', 'CALLEYNO']),
                    ('COLONIA O POBLACION', ['COLONIA O POBLACION', 'COLONIA_O_POBLACION', 'COLONIA']),
                    ('MUNICIPIO O ALCADIA', ['MUNICIPIO O ALCADIA', 'MUNICIPIO_O_ALCADIA', 'MUNICIPIO']),
                    ('CIUDAD O ESTADO', ['CIUDAD O ESTADO', 'CIUDAD_O_ESTADO', 'ESTADO']),
                    ('CP', ['CP', 'cp']),
                    ('SERVICIO', ['SERVICIO', 'servicio'])
                ]
                for label, variants in field_variants:
                    try:
                        val = ''
                        if isinstance(d, dict):
                            for key in variants:
                                if key in d and d.get(key) is not None:
                                    val = d.get(key)
                                    break
                        ctk.CTkLabel(box, text=f"{label}: {val}", text_color=STYLE['texto_oscuro']).pack(anchor='w', padx=12, pady=2)
                    except Exception:
                        continue

        ctk.CTkButton(frame, text='Cerrar', command=win.destroy, fg_color=STYLE['secundario'], hover_color=STYLE['secundario'], text_color=STYLE['surface']).pack(pady=(8,0))

    def _export_catalogo_clientes(self):
        """Exporta `data/Clientes.json` a un archivo Excel seleccionado por el usuario.
        Cada domicilio del cliente se exporta como una fila separada con columnas:
        RFC, CLIENTE, CALLE Y NO, COLONIA O POBLACION, MUNICIPIO O ALCADIA, CIUDAD O ESTADO, CP, SERVICIO
        """
        ruta = os.path.join(DATA_DIR, 'Clientes.json')
        datos = []
        try:
            if os.path.exists(ruta):
                with open(ruta, 'r', encoding='utf-8') as f:
                    datos = json.load(f) or []
        except Exception as e:
            messagebox.showerror('Exportar', f'No se pudo leer Clientes.json: {e}')
            return

        if not datos:
            messagebox.showinfo('Exportar', 'No hay datos para exportar.')
            return

        rows = []
        for c in datos:
            try:
                rfc = c.get('RFC') or c.get('R.F.C') or ''
                cliente = c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or ''
                domicilios = c.get('DIRECCIONES') or c.get('DOMICILIOS') or c.get('DOMICILIO') or []
                if not isinstance(domicilios, (list, tuple)):
                    domicilios = []
                if domicilios:
                    for d in domicilios:
                        try:
                            # extraer campos con variantes de nombre
                            def g(obj, *keys):
                                for k in keys:
                                    if isinstance(obj, dict) and k in obj and obj.get(k) is not None:
                                        return obj.get(k)
                                return ''
                            calle = g(d, 'CALLE Y NO', 'CALLE_Y_NO', 'CALLE Y No', 'CALLEYNO')
                            colonia = g(d, 'COLONIA O POBLACION', 'COLONIA_O_POBLACION', 'COLONIA')
                            municipio = g(d, 'MUNICIPIO O ALCADIA', 'MUNICIPIO_O_ALCADIA', 'MUNICIPIO')
                            ciudad = g(d, 'CIUDAD O ESTADO', 'CIUDAD_O_ESTADO', 'ESTADO')
                            cp = g(d, 'CP', 'cp')
                            servicio = g(d, 'SERVICIO', 'servicio')
                            rows.append({
                                'RFC': rfc,
                                'CLIENTE': cliente,
                                'CALLE Y NO': calle,
                                'COLONIA O POBLACION': colonia,
                                'MUNICIPIO O ALCADIA': municipio,
                                'CIUDAD O ESTADO': ciudad,
                                'CP': cp,
                                'SERVICIO': servicio
                            })
                        except Exception:
                            continue
                else:
                    rows.append({
                        'RFC': rfc,
                        'CLIENTE': cliente,
                        'CALLE Y NO': '',
                        'COLONIA O POBLACION': '',
                        'MUNICIPIO O ALCADIA': '',
                        'CIUDAD O ESTADO': '',
                        'CP': '',
                        'SERVICIO': ''
                    })
            except Exception:
                continue

        # Preguntar ubicación de guardado
        try:
            save_path = filedialog.asksaveasfilename(defaultextension='.xlsx', filetypes=[('Excel','*.xlsx')], title='Guardar catálogo de clientes')
            if not save_path:
                return
            df = pd.DataFrame(rows, columns=['RFC','CLIENTE','CALLE Y NO','COLONIA O POBLACION','MUNICIPIO O ALCADIA','CIUDAD O ESTADO','CP','SERVICIO'])
            try:
                # Intentar ajustar anchos en Excel usando openpyxl si está disponible
                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Catalogo')
                    try:
                        ws = writer.sheets['Catalogo']
                        # Ajustar anchos de columna en base al contenido (simple heurística)
                        from openpyxl.utils import get_column_letter
                        for i, col in enumerate(df.columns, 1):
                            col_letter = get_column_letter(i)
                            try:
                                max_len = max(df[col].astype(str).map(len).max(), len(col)) + 2
                                ws.column_dimensions[col_letter].width = max_len
                            except Exception:
                                pass
                    except Exception:
                        pass
            except Exception:
                df.to_excel(save_path, index=False)

            messagebox.showinfo('Exportar', f'Catálogo exportado correctamente a:\n{save_path}')
        except Exception as e:
            messagebox.showerror('Exportar', f'Error al exportar: {e}')

    # Handlers para los subformularios de domicilios
    def _on_change_num_domicilios(self, valor):
        """Muestra u oculta los subformularios de domicilios según el número seleccionado."""
        # deprecated (now using dynamic add/remove). keep for safety
        return

    def _construir_tab_reportes_ejecutivo(self, parent):
        """Construye la UI mínima para que un ejecutivo descargue su reporte diario."""
        try:
            parent.pack_propagate(False)
            header = ctk.CTkFrame(parent, fg_color="transparent")
            header.pack(fill='x', padx=12, pady=(8,6))
            ctk.CTkLabel(header, text="Reportes de Productividad", font=FONT_SUBTITLE, text_color=STYLE['texto_oscuro']).pack(anchor='w')

            body = ctk.CTkFrame(parent, fg_color="transparent")
            body.pack(fill='both', expand=True, padx=12, pady=8)

            # Layout: calendario a la izquierda + controles (días disponibles / acciones) a la derecha
            main_controls = ctk.CTkFrame(body, fg_color='transparent')
            main_controls.pack(fill='x', pady=(6,12))

            # Calendar (usar tkcalendar si está disponible)
            cal_frame = ctk.CTkFrame(main_controls, fg_color='transparent')
            # Mostrar calendario más grande y aprovechar el alto disponible
            cal_frame.pack(side='left', padx=(0,12), fill='both')
            if Calendar is not None:
                try:
                    # Calendar devuelve fechas en formato dd/mm/YYYY con date_pattern
                    # Empujar un poco el tamaño y padding para que se vea más grande
                    self.reporte_calendar = Calendar(cal_frame, selectmode='day', date_pattern='dd/mm/yyyy')
                    self.reporte_calendar.pack(padx=8, pady=8, ipadx=6, ipady=6)
                except Exception:
                    self.reporte_calendar = None
            else:
                # Fallback: entrada de texto si tkcalendar no está disponible
                self.reporte_calendar = None
                ctk.CTkLabel(cal_frame, text='(Instale tkcalendar para ver calendario)', text_color=STYLE['texto_oscuro']).pack()
                # Crear una entrada de texto como alternativa
                self.entry_reporte_fecha = ctk.CTkEntry(cal_frame, width=140)
                try:
                    self.entry_reporte_fecha.insert(0, datetime.now().strftime('%d/%m/%Y'))
                except Exception:
                    pass
                self.entry_reporte_fecha.pack()

            # Right-side controls: lista de días con datos y botones
            right_frame = ctk.CTkFrame(main_controls, fg_color='transparent')
            right_frame.pack(side='left', fill='both', expand=True)

            days_label = ctk.CTkLabel(right_frame, text='Productividad:', text_color=STYLE['texto_oscuro'])
            days_label.pack(anchor='nw')

            # Lista con los días donde existen registros de producción para el usuario
            # Mostrar más filas y permitir que la lista crezca verticalmente
            self.lista_dias_produccion = tk.Listbox(right_frame, height=12)
            self.lista_dias_produccion.pack(fill='both', expand=True, pady=(6,6))
            # Bind selection to sync with calendar
            try:
                self.lista_dias_produccion.bind('<<ListboxSelect>>', lambda e: self._on_select_dia_produccion(e))
                self.lista_dias_produccion.bind('<Double-Button-1>', lambda e: self._on_doubleclick_dia(e))
            except Exception:
                pass

            btns_row = ctk.CTkFrame(right_frame, fg_color='transparent')
            btns_row.pack(anchor='s', pady=(6,0), fill='x')

            ctk.CTkButton(btns_row, text='Actualizar días', fg_color=STYLE['primario'], text_color=STYLE['secundario'], hover_color=STYLE['primario'], command=self._refresh_production_days).pack(side='left', padx=(0,8))
            ctk.CTkButton(btns_row, text='Descargar reporte', fg_color=STYLE['primario'], text_color=STYLE['secundario'], hover_color=STYLE['primario'], command=self._descargar_reporte_ejecutivo).pack(side='left')

            # Nota: se eliminó el cuadro de texto inferior (área de logs) —
            # los mensajes de estado se mantienen en `self.reporte_log` (etiqueta),
            # y las acciones de guardado muestran cuadros de diálogo.

            # Inicializar lista de días
            try:
                self._refresh_production_days()
            except Exception:
                pass
        except Exception:
            pass

    def _normalize_fecha_str(self, s):
        """Normaliza una cadena de fecha a formato 'dd/mm/YYYY'. Devuelve cadena vacía si no es válida."""
        try:
            if not s:
                return ''
            # eliminar caracteres invisibles
            s2 = ''.join(ch for ch in str(s) if ch.isprintable())
            s2 = s2.strip()
            # Reemplazar separadores comunes y limpiar caracteres no numéricos excepto '/'
            s2 = s2.replace('-', '/').replace('.', '/').replace('\\', '/')
            # Collapsar múltiples espacios
            s2 = re.sub(r"\s+", '', s2)
            # intentar parsear formatos esperados
            for fmt in ('%d/%m/%Y', '%d/%m/%y'):
                try:
                    dt = datetime.strptime(s2, fmt)
                    return dt.strftime('%d/%m/%Y')
                except Exception:
                    continue
            # intentar extraer con regex dd/mm/YYYY
            m = re.search(r"(\d{1,2})\D(\d{1,2})\D(\d{2,4})", s2)
            if m:
                day, mon, year = m.group(1), m.group(2), m.group(3)
                if len(year) == 2:
                    # asumir 20xx si empieza con '20' probablemente
                    year = '20' + year
                try:
                    dt = datetime.strptime(f"{int(day):02d}/{int(mon):02d}/{int(year)}", '%d/%m/%Y')
                    return dt.strftime('%d/%m/%Y')
                except Exception:
                    return ''
            return ''
        except Exception:
            return ''

    def _on_select_dia_produccion(self, event=None):
        try:
            sel = None
            try:
                sel = self.lista_dias_produccion.get(self.lista_dias_produccion.curselection())
            except Exception:
                sel = None
            if not sel:
                return
            if getattr(self, 'reporte_calendar', None) and Calendar is not None:
                try:
                    # seleccionar en el calendario
                    self.reporte_calendar.selection_set(datetime.strptime(sel, '%d/%m/%Y'))
                    self.reporte_calendar.see(datetime.strptime(sel, '%d/%m/%Y'))
                except Exception:
                    pass
            else:
                try:
                    # fallback: poner en entry_reporte_fecha
                    if getattr(self, 'entry_reporte_fecha', None):
                        self.entry_reporte_fecha.delete(0, tk.END)
                        self.entry_reporte_fecha.insert(0, sel)
                except Exception:
                    pass
        except Exception:
            pass

    def _on_doubleclick_dia(self, event=None):
        try:
            # al hacer doble click, iniciar descarga directamente
            self._on_select_dia_produccion(event)
            self._descargar_reporte_ejecutivo()
        except Exception:
            pass

    def _get_production_dates(self, owner_username=None):
        """Retorna un conjunto de fechas (dd/mm/YYYY) para las cuales hay registros de producción del usuario."""
        fechas = set()
        try:
            user = owner_username or getattr(self, 'current_user', None)
            role = getattr(self, 'current_role', None)
            prod_root = os.path.join(DATA_DIR, 'produccion')

            # Si es admin o no se especificó usuario, agregar fechas de todos los subdirectorios
            owners_to_scan = []
            if role == 'admin' or not user:
                try:
                    for name in os.listdir(prod_root):
                        p = os.path.join(prod_root, name)
                        if os.path.isdir(p):
                            owners_to_scan.append(p)
                except Exception:
                    owners_to_scan = []
            else:
                uname = ''
                if isinstance(user, dict):
                    uname = user.get('username') or ''
                else:
                    uname = str(user or '')
                owner_safe = re.sub(r"[^A-Za-z0-9_.-]", '_', str(uname))
                owners_to_scan = [os.path.join(prod_root, owner_safe)]

            for owner_dir in owners_to_scan:
                if not os.path.exists(owner_dir):
                    continue
                for fn in os.listdir(owner_dir):
                    if not fn.lower().endswith('.json'):
                        continue
                    fp = os.path.join(owner_dir, fn)
                    try:
                        with open(fp, 'r', encoding='utf-8') as jf:
                            obj = json.load(jf)
                        recs = []
                        if isinstance(obj, dict) and 'records' in obj:
                            recs = obj.get('records') or []
                        elif isinstance(obj, list):
                            recs = obj
                        else:
                            recs = [obj]

                        for rec in recs:
                            try:
                                raw = (rec.get('fecha_inicio') or '') if isinstance(rec, dict) else ''
                                fnorm = self._normalize_fecha_str(raw)
                                if fnorm:
                                    fechas.add(fnorm)
                            except Exception:
                                continue
                    except Exception:
                        continue
        except Exception:
            pass
        # devolver ordenadas cronológicamente
        try:
            lista = list(fechas)
            lista_dt = []
            for s in lista:
                try:
                    dt = datetime.strptime(s, '%d/%m/%Y')
                    lista_dt.append((dt, s))
                except Exception:
                    try:
                        # fallback, meter al final
                        lista_dt.append((datetime.max, s))
                    except Exception:
                        lista_dt.append((datetime.max, s))
            lista_dt.sort(key=lambda x: x[0])
            return [s for (_dt, s) in lista_dt]
        except Exception:
            return sorted(list(fechas))

    def _refresh_production_days(self):
        """Actualiza la lista de días con producción y marca el calendario (si existe)."""
        try:
            dias = []
            try:
                dias = self._get_production_dates()
            except Exception:
                dias = []

            # actualizar listbox
            try:
                self.lista_dias_produccion.delete(0, tk.END)
                for d in dias:
                    self.lista_dias_produccion.insert(tk.END, d)
            except Exception:
                pass

            # marcar calendario
            if getattr(self, 'reporte_calendar', None) and Calendar is not None:
                try:
                    # Limpiar eventos previos
                    for ev in list(self.reporte_calendar.get_calevents()):
                        try:
                            self.reporte_calendar.calevent_remove(ev)
                        except Exception:
                            pass
                    # Crear eventos para cada día
                    for d in dias:
                        try:
                            # Calendar espera formato dd/mm/yyyy
                            self.reporte_calendar.calevent_create(datetime.strptime(d, '%d/%m/%Y'), 'Produccion', 'prod')
                        except Exception:
                            continue
                    # configurar estilo para tag 'prod'
                    try:
                        self.reporte_calendar.tag_config('prod', background='lightblue')
                    except Exception:
                        pass
                except Exception:
                    pass
        except Exception:
            pass

    def _add_domicilio(self):
        """Añade un subformulario de domicilio si no supera el máximo."""
        try:
            if len(self.dom_fields) >= self.max_domicilios:
                messagebox.showwarning('Máximo', f'Solo se permiten hasta {self.max_domicilios} domicilios')
                return
            frm, fields = self._crear_domicilio_subform(self.dom_container, len(self.dom_fields))
            self.dom_fields.append({'frame': frm, 'fields': fields})
            try:
                self.lbl_dom_count.configure(text=f"{len(self.dom_fields)} domicilios")
            except Exception:
                pass
        except Exception:
            pass

    def _remove_domicilio_by_frame(self, frame):
        """Elimina el subform asociado al frame y actualiza el contador."""
        try:
            found = None
            for rec in self.dom_fields:
                if rec.get('frame') == frame:
                    found = rec
                    break
            if not found:
                return
            try:
                found.get('frame').destroy()
            except Exception:
                pass
            try:
                self.dom_fields.remove(found)
            except Exception:
                pass
            try:
                self.lbl_dom_count.configure(text=f"{len(self.dom_fields)} domicilios")
            except Exception:
                pass
        except Exception:
            pass

    def safe_forget(self, widget):
        """Evita errores al ocultar widgets ya olvidados."""
        try:
            if widget and widget.winfo_ismapped():
                widget.pack_forget()
        except Exception:
            pass

    def safe_pack(self, widget, **kwargs):
        """Evita errores de Tkinter al volver a empacar widgets."""
        try:
            if widget and not widget.winfo_ismapped():
                widget.pack(**kwargs)
        except Exception:
            pass

    def actualizar_cliente_seleccionado(self, cliente_nombre):

        # Reset si selecciona opción vacía
        if cliente_nombre == "Seleccione un cliente...":
            self.cliente_seleccionado = None
            self.info_cliente.configure(
                text="No se ha seleccionado ningún cliente",
                text_color=STYLE["texto_claro"]
            )
            self.boton_limpiar_cliente.configure(state="disabled")
            self.safe_forget(self.boton_subir_etiquetado)
            self.safe_forget(self.info_etiquetado)
            # Mantener los botones de pegado visibles incluso si no hay cliente seleccionado
            return

        # ─────────────────────────────────────────────
        #  1) CLIENTES QUE SE TRATAN COMO EVIDENCIA
        # ─────────────────────────────────────────────
        CLIENTES_EVIDENCIA = set()

        # ─────────────────────────────────────────────
        # 2) CLIENTES QUE PEGAN ETIQUETAS
        # ─────────────────────────────────────────────
        CLIENTES_ETIQUETA = {
            "ARTICULOS DEPORTIVOS DECATHLON SA DE CV",
            "FERRAGAMO MEXICO S. DE R.L. DE C.V.",
            "ULTA BEAUTY S.A.P.I. DE C.V.",  # Regla especial
        }

        # Buscar cliente en la lista; aceptar varias claves de nombre
        encontrado = None
        for cliente in self.clientes_data:
            if not isinstance(cliente, dict):
                continue
            nombre_cliente = cliente.get('CLIENTE') or cliente.get('RAZÓN SOCIAL ') or cliente.get('RAZON SOCIAL') or cliente.get('RAZON_SOCIAL') or cliente.get('NOMBRE')
            if nombre_cliente and isinstance(nombre_cliente, str) and nombre_cliente.strip() == cliente_nombre:
                encontrado = cliente
                break

        if encontrado is None:
            # No se encontró por claves comunes; intentar por RFC o contrato si el nombre coincide
            for cliente in self.clientes_data:
                if not isinstance(cliente, dict):
                    continue
                # construir un nombre de fallback mostrado en el combo (tal como se poblaron los valores)
                fallback = cliente.get('CLIENTE') or cliente.get('RAZÓN SOCIAL ') or cliente.get('RAZON SOCIAL') or cliente.get('RAZON_SOCIAL') or cliente.get('RFC') or cliente.get('NÚMERO_DE_CONTRATO')
                if fallback and isinstance(fallback, str) and fallback.strip() == cliente_nombre:
                    encontrado = cliente
                    break

        if not encontrado:
            # No encontrado; mostrar mensaje y salir
            try:
                self.info_cliente.configure(text="Cliente no encontrado", text_color=STYLE["advertencia"])
            except Exception:
                pass
            return

        cliente = encontrado
        self.cliente_seleccionado = cliente
        rfc = cliente.get("RFC", "No disponible")

        display_name = cliente.get('CLIENTE') or cliente.get('RAZÓN SOCIAL ') or cliente.get('RAZON SOCIAL') or cliente.get('RAZON_SOCIAL') or cliente.get('RFC') or cliente.get('NÚMERO_DE_CONTRATO')

        self.info_cliente.configure(
            text=f"✅ {display_name}\n📋 RFC: {rfc}",
            text_color=STYLE["exito"]
        )
        self.boton_limpiar_cliente.configure(state="normal")

        # Rellenar lista de domicilios para este cliente (si existen)
        domicilios = []
        try:
            direcciones = cliente.get('DIRECCIONES')
            if isinstance(direcciones, list) and direcciones:
                for d in direcciones:
                    if not isinstance(d, dict):
                        continue
                    parts = []
                    for k in ('CALLE Y NO', 'CALLE', 'CALLE_Y_NO', 'CALLE_Y_NRO', 'NUMERO'):
                        v = d.get(k) or d.get(k.upper()) if isinstance(d, dict) else None
                        if v:
                            parts.append(str(v))
                    for k in ('COLONIA O POBLACION', 'COLONIA'):
                        v = d.get(k)
                        if v:
                            parts.append(str(v))
                    for k in ('MUNICIPIO O ALCADIA', 'MUNICIPIO'):
                        v = d.get(k)
                        if v:
                            parts.append(str(v))
                    if d.get('CIUDAD O ESTADO'):
                        parts.append(str(d.get('CIUDAD O ESTADO')))
                    if d.get('CP'):
                        parts.append(str(d.get('CP')))
                    addr = ", ".join(parts).strip()
                    if addr:
                        domicilios.append(addr)

            # si no hay lista de direcciones, intentar con campos a nivel superior
            if not domicilios:
                parts = []
                for k in ('CALLE Y NO', 'CALLE', 'CALLE_Y_NO'):
                    v = cliente.get(k) or cliente.get(k.upper())
                    if v:
                        parts.append(str(v))
                for k in ('COLONIA O POBLACION', 'COLONIA'):
                    v = cliente.get(k)
                    if v:
                        parts.append(str(v))
                for k in ('MUNICIPIO O ALCADIA', 'MUNICIPIO'):
                    v = cliente.get(k)
                    if v:
                        parts.append(str(v))
                if cliente.get('CIUDAD O ESTADO'):
                    parts.append(str(cliente.get('CIUDAD O ESTADO')))
                if cliente.get('CP') is not None:
                    parts.append(str(cliente.get('CP')))
                addr = ", ".join(parts).strip()
                if addr:
                    domicilios.append(addr)
        except Exception:
            domicilios = []

        if not domicilios:
            domicilios = ["Domicilio no disponible"]

        # Configurar combo de domicilios
        try:
            vals = ['Seleccione un domicilio...'] + domicilios
            self.combo_domicilios.configure(values=vals, state='readonly')
            self.combo_domicilios.set('Seleccione un domicilio...')
            # almacenar lista para referencia y raw dicts alineados
            self._domicilios_list = domicilios
            # construir _domicilios_raw: si DIRECCIONES existían usamos dicts, else build one
            raw = []
            try:
                direcciones = cliente.get('DIRECCIONES')
                if isinstance(direcciones, list) and direcciones:
                    for d in direcciones:
                        if isinstance(d, dict):
                            raw.append(d)
                else:
                    # fallback: construir dict a partir de campos de cliente
                    d = {
                        'CALLE Y NO': cliente.get('CALLE Y NO') or cliente.get('CALLE') or cliente.get('CALLE_Y_NO') or '',
                        'COLONIA O POBLACION': cliente.get('COLONIA O POBLACION') or cliente.get('COLONIA') or '',
                        'MUNICIPIO O ALCADIA': cliente.get('MUNICIPIO O ALCADIA') or cliente.get('MUNICIPIO') or '',
                        'CIUDAD O ESTADO': cliente.get('CIUDAD O ESTADO') or cliente.get('CIUDAD') or '',
                        'CP': cliente.get('CP')
                    }
                    raw.append(d)
            except Exception:
                raw = []

            # Ensure lengths match: if not, pad with minimal dicts
            if len(raw) != len(self._domicilios_list):
                # try to align by creating dicts from the display strings
                aligned = []
                for s in self._domicilios_list:
                    aligned.append({'_display': s})
                raw = aligned

            self._domicilios_raw = raw
            self.domicilio_seleccionado = None
            # limpiar campos individuales
            self.direccion_seleccionada = None
            self.colonia_seleccionada = None
            self.municipio_seleccionado = None
            self.ciudad_seleccionada = None
            self.cp_seleccionado = None
        except Exception:
            pass

        # ─────────────────────────────────────────────
        # CASO 1: EVIDENCIA (por defecto para TODOS excepto `CLIENTES_ETIQUETA`)
        # ─────────────────────────────────────────────
        if cliente_nombre not in CLIENTES_ETIQUETA:
            self.tipo_operacion = "EVIDENCIA"
            self.safe_forget(self.boton_subir_etiquetado)
            self.safe_forget(self.info_etiquetado)
            # Mostrar botones de pegado (no persisten rutas)
            try:
                self.safe_pack(self.boton_pegado_simple, side="left", padx=(0, 12))
                self.safe_pack(self.boton_pegado_carpetas, side="left", padx=(0, 12))
                # Empacar el botón Limpiar (Reservar Folio está en la fila de carga)
                self.safe_pack(self.boton_limpiar_rutas_evidencias, side="left", padx=(0, 12))
            except Exception:
                pass

        # ─────────────────────────────────────────────
        # CASO 2: PEGADO DE ETIQUETAS
        # ─────────────────────────────────────────────
        if cliente_nombre in CLIENTES_ETIQUETA:
            self.tipo_operacion = "ETIQUETA"

            # ULTA BEAUTY — flujo mixto dependiendo de la NOM
            if cliente_nombre == "ULTA BEAUTY SAPI DE CV":
                self.tipo_operacion = "ULTA"

            # Mostrar botón de carga de etiquetado
            self.safe_pack(self.boton_subir_etiquetado, side="left", padx=(0, 8))

            if self.archivo_etiquetado_json:
                self.safe_pack(self.info_etiquetado, anchor="w", fill="x", pady=(5, 0))

            # En flujos de etiquetas no se ocultan los botones de pegado; se mantienen visibles
            pass

        # ─────────────────────────────────────────────
        # SI YA SE CARGÓ EL JSON → habilitar dictamen
        # ─────────────────────────────────────────────
        # Mostrar/ocultar el botón de configuración según el cliente seleccionado
        # Los botones de pegado se mantienen visibles por diseño; no se ocultan según cliente

        if self.archivo_json_generado:
            self.boton_generar_dictamen.configure(state="normal")

    def cargar_base_etiquetado(self):
        file_path = filedialog.askopenfilename(
            title="Seleccionar Base de Etiquetado DECATHLON",
            filetypes=[("Archivos Excel", "*.xlsx;*.xls")]
        )

        if not file_path:
            return

        try:
            df = pd.read_excel(file_path)

            if df.empty:
                messagebox.showwarning("Archivo vacío", "El archivo de etiquetado no contiene datos.")
                return

            for col in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    df[col] = df[col].astype(str)

            registros = df.to_dict(orient="records")

            data_dir = DATA_DIR
            os.makedirs(data_dir, exist_ok=True)

            output_json = os.path.join(data_dir, "base_etiquetado.json")

            with open(output_json, "w", encoding="utf-8") as f:
                json.dump(registros, f, ensure_ascii=False, indent=2)

            self.archivo_etiquetado_json = output_json

            self.info_etiquetado.configure(
                text=f"📄 Base de etiquetado cargada ({len(registros)} registros)",
                text_color=STYLE["exito"]
            )
            self.info_etiquetado.pack(anchor="w", fill="x", pady=(5, 0))

            messagebox.showinfo(
                "Base cargada",
                f"Base de etiquetado convertida exitosamente.\n\nGuardado en:\n{output_json}"
            )

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo procesar la base de etiquetado:\n{e}")

    def limpiar_cliente(self):
        self.combo_cliente.set("Seleccione un cliente...")
        self.cliente_seleccionado = None
        self.info_cliente.configure(
            text="No se ha seleccionado ningún cliente",
            text_color=STYLE["texto_claro"]
        )
        self.boton_limpiar_cliente.configure(state="disabled")
        self.boton_generar_dictamen.configure(state="disabled")
        self.boton_subir_etiquetado.pack_forget()
        # No ocultar los botones de pegado al limpiar cliente; deben permanecer visibles
        self.info_etiquetado.pack_forget()
        try:
            self.combo_domicilios.configure(values=["Seleccione un domicilio..."], state='disabled')
            self.combo_domicilios.set('Seleccione un domicilio...')
            self.domicilio_seleccionado = None
        except Exception:
            pass

    def desmarcar_cliente(self):
        """Desmarca cualquier cliente seleccionado y sale del modo edición."""
        try:
            # Limpiar el formulario de edición
            self._limpiar_formulario_cliente()
        except Exception:
            pass
        try:
            # Limpiar la selección visual y los datos de cliente seleccionado
            self.limpiar_cliente()
        except Exception:
            pass
        try:
            # Resetear flag de edición y texto del botón guardar
            self.editing_cliente_rfc = None
            try:
                self.btn_guardar_cliente.configure(text="Guardar cliente")
            except Exception:
                pass
        except Exception:
            pass
        try:
            # Quitar selección en la tabla si existe
            if hasattr(self, 'tree_clientes'):
                for iid in list(self.tree_clientes.selection()):
                    try:
                        self.tree_clientes.selection_remove(iid)
                    except Exception:
                        pass
        except Exception:
            pass

    def _seleccionar_domicilio(self, domicilio_text):
        """Handler para seleccionar domicilio del cliente."""
        try:
            if domicilio_text == 'Seleccione un domicilio...' or not domicilio_text:
                self.domicilio_seleccionado = None
                # reset component fields
                self.direccion_seleccionada = None
                self.colonia_seleccionada = None
                self.municipio_seleccionado = None
                self.ciudad_seleccionada = None
                self.cp_seleccionado = None
            else:
                # almacenar texto seleccionado y mapear a raw dict si existe
                self.domicilio_seleccionado = domicilio_text
                try:
                    idx = self._domicilios_list.index(domicilio_text)
                except Exception:
                    idx = None
                raw = None
                try:
                    if idx is not None and hasattr(self, '_domicilios_raw') and idx < len(self._domicilios_raw):
                        raw = self._domicilios_raw[idx]
                except Exception:
                    raw = None

                if raw and isinstance(raw, dict):
                    # prefer explicit keys
                    self.direccion_seleccionada = raw.get('CALLE Y NO') or raw.get('CALLE') or raw.get('calle_numero') or raw.get('CALLE_Y_NO') or raw.get('_display')
                    self.colonia_seleccionada = raw.get('COLONIA O POBLACION') or raw.get('COLONIA') or raw.get('colonia')
                    self.municipio_seleccionado = raw.get('MUNICIPIO O ALCADIA') or raw.get('MUNICIPIO') or raw.get('municipio')
                    self.ciudad_seleccionada = raw.get('CIUDAD O ESTADO') or raw.get('CIUDAD') or raw.get('ciudad_estado')
                    self.cp_seleccionado = raw.get('CP')
                else:
                    # fallback: store full text in direccion_seleccionada
                    self.direccion_seleccionada = domicilio_text
                    self.colonia_seleccionada = None
                    self.municipio_seleccionado = None
                    self.ciudad_seleccionada = None
                    self.cp_seleccionado = None

            # Actualizar la vista de info_cliente para mostrar domicilio elegido
            try:
                if self.cliente_seleccionado:
                    display_name = self.cliente_seleccionado.get('CLIENTE') or self.cliente_seleccionado.get('RAZÓN SOCIAL ') or self.cliente_seleccionado.get('RAZON SOCIAL') or self.cliente_seleccionado.get('RFC') or ''
                    rfc = self.cliente_seleccionado.get('RFC', 'No disponible')
                    if self.domicilio_seleccionado:
                        self.info_cliente.configure(text=f"✅ {display_name}\n📋 RFC: {rfc}\n🏠 {self.direccion_seleccionada}", text_color=STYLE['exito'])
                    else:
                        self.info_cliente.configure(text=f"✅ {display_name}\n📋 RFC: {rfc}", text_color=STYLE['exito'])
            except Exception:
                pass
        except Exception:
            pass

    # -----------------------------------------------------------
    # MÉTODOS MEJORADOS PARA GESTIÓN DE FOLIOS
    # -----------------------------------------------------------
    def cargar_ultimo_folio(self):
        """Carga el último folio utilizado y determina el siguiente disponible"""
        try:
            # Determinar el siguiente folio de visita (CP...) a partir del historial
            # de visitas (campo `folio_visita`). Esto es independiente del siguiente
            # folio de documento (usado para los dictámenes) que se calcula con
            # `_get_next_document_folio()` y se muestra en el footer.
            try:
                if os.path.exists(self.historial_path):
                    with open(self.historial_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    visitas = data.get("visitas", [])
                    if visitas:
                        maxv = 0
                        for visita in visitas:
                            folio_raw = visita.get("folio_visita", "")
                            # Extraer solo dígitos (soporta formatos como 'CP000012')
                            folio_digits = ''.join([c for c in str(folio_raw) if c.isdigit()])
                            if folio_digits:
                                try:
                                    n = int(folio_digits)
                                    if n > maxv:
                                        maxv = n
                                except Exception:
                                    pass
                        # siguiente visita = maxv + 1 (si maxv==0 -> 1)
                        self.current_folio = f"{(maxv + 1):06d}"
                    else:
                        self.current_folio = "000001"
                else:
                    self.current_folio = "000001"
            except Exception:
                self.current_folio = None
            # Mantener el comportamiento original: el folio de visita (CP/AC)
            # se determina únicamente a partir del historial de visitas en disco
            # (no se fuerza desde el contador de documentos central).

            # Actualizar el campo en la interfaz con prefijo CP (si existen widgets)
            try:
                if hasattr(self, 'entry_folio_visita') and hasattr(self, 'entry_folio_acta'):
                    self.entry_folio_visita.configure(state="normal")
                    self.entry_folio_visita.delete(0, "end")
                    folio_con_prefijo = f"CP{self.current_folio}"
                    self.entry_folio_visita.insert(0, folio_con_prefijo)
                    self.entry_folio_visita.configure(state="normal")

                    # Actualizar también el folio del acta
                    self.entry_folio_acta.configure(state="normal")
                    self.entry_folio_acta.delete(0, "end")
                    self.entry_folio_acta.insert(0, f"AC{self.current_folio}")
                    self.entry_folio_acta.configure(state="normal")
            except Exception:
                pass
            # Actualizar etiqueta visual del siguiente folio de documento
            try:
                self._update_siguiente_folio_label()
            except Exception:
                pass
                    
        except Exception as e:
            print(f"❌ Error cargando último folio: {e}")
            self.current_folio = "000001"

    def _folio_visita_exists(self, folio_visita, exclude_id=None):
        """Devuelve True si `folio_visita` ya existe en el historial en disco.
        Opción `exclude_id` permite omitir un registro por su _id al validar (útil en actualizaciones).
        Comparación es insensible a mayúsculas y espacios.
        """
        try:
            if not folio_visita:
                return False
            fv = str(folio_visita).strip().lower()
            hist_path = getattr(self, 'historial_path', None) or os.path.join(DATA_DIR, 'historial_visitas.json')
            visitas = []
            if os.path.exists(hist_path):
                try:
                    with open(hist_path, 'r', encoding='utf-8') as hf:
                        hobj = json.load(hf) or {}
                        visitas = hobj.get('visitas', []) if isinstance(hobj, dict) else (hobj or [])
                except Exception:
                    visitas = self.historial.get('visitas', []) or []
            else:
                visitas = self.historial.get('visitas', []) or []

            for rec in (visitas or []):
                try:
                    if exclude_id and (rec.get('_id') == exclude_id or rec.get('id') == exclude_id):
                        continue
                    other = str(rec.get('folio_visita','') or '').strip().lower()
                    if other and other == fv:
                        return True
                except Exception:
                    continue
            return False
        except Exception:
            return False

    def crear_nueva_visita(self):
        """Prepara el formulario para una nueva visita"""
        try:
            # No reservar ni avanzar el contador en disco aquí. Solo calcular/usar
            # el siguiente folio a partir del historial en memoria o recargarlo.
            try:
                # recargar desde historial para evitar usar folio_counter.json
                self.cargar_ultimo_folio()
            except Exception:
                pass

            # Actualizar campos con prefijo CP
            self.entry_folio_visita.configure(state="normal")
            self.entry_folio_visita.delete(0, "end")
            folio_con_prefijo = f"CP{self.current_folio}"
            self.entry_folio_visita.insert(0, folio_con_prefijo)
            self.entry_folio_visita.configure(state="normal")

            # Actualizar folio acta automáticamente
            self.entry_folio_acta.configure(state="normal")
            self.entry_folio_acta.delete(0, "end")
            self.entry_folio_acta.insert(0, f"AC{self.current_folio}")
            self.entry_folio_acta.configure(state="normal")
            # Limpiar/poner valores por defecto en otros campos de fecha/hora
            try:
                self.entry_fecha_inicio.delete(0, "end")
                self.entry_fecha_inicio.insert(0, datetime.now().strftime("%d/%m/%Y"))
            except Exception:
                pass
            try:
                self.entry_hora_inicio.delete(0, "end")
                self.entry_hora_inicio.insert(0, datetime.now().strftime("%H:%M"))
            except Exception:
                pass
            try:
                self.entry_fecha_termino.delete(0, "end")
            except Exception:
                pass
            try:
                self.entry_hora_termino.delete(0, "end")
            except Exception:
                pass

            # No forzamos el tipo de documento: respetar la selección actual
            try:
                seleccionado = self.combo_tipo_documento.get().strip() if hasattr(self, 'combo_tipo_documento') else None
            except Exception:
                seleccionado = None

            # No consumir automáticamente las reservas al crear un formulario nuevo.
            # Dejaremos las visitas en estado 'Pendiente' hasta que el usuario
            # seleccione explícitamente una de las reservas desde el combobox.
            try:
                # Asegurar que el combobox de folios pendientes esté actualizado
                if hasattr(self, '_refresh_pending_folios_dropdown'):
                    try:
                        self._refresh_pending_folios_dropdown()
                    except Exception:
                        pass
                # Indicar al usuario si existe alguna reserva para el tipo seleccionado
                try:
                    if seleccionado:
                        pendientes = [r for r in getattr(self, 'historial_data', []) if (r.get('tipo_documento') or '').strip() == seleccionado and (r.get('estatus','').lower() == 'pendiente')]
                        if pendientes:
                            # Mostrar mensaje no intrusivo en la etiqueta de info
                            try:
                                if hasattr(self, 'info_folio_pendiente'):
                                    p0 = pendientes[0]
                                    self.info_folio_pendiente.configure(text=f"Folio pendiente disponible: {p0.get('folio_visita','-')}")
                            except Exception:
                                pass
                except Exception:
                    pass
            except Exception:
                pass

            messagebox.showinfo("Nueva Visita", "Formulario listo para nueva visita")
        except Exception:
            pass

    # ----------------- Folio counter (archivo compartido, con lock) -----------------
    def _acquire_file_lock(self, lock_path, timeout=5.0, poll=0.08):
        """Intenta crear un archivo lock de manera atómica. Devuelve el file descriptor.
        Usa O_EXCL para evitar race conditions. Lanzará Exception si no pudo adquirir en timeout."""
        start = time.time()
        while True:
            try:
                fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                # escribir PID/ts
                try:
                    os.write(fd, f"{os.getpid()}\n{time.time()}".encode('utf-8'))
                except Exception:
                    pass
                return fd
            except FileExistsError:
                if (time.time() - start) >= timeout:
                    raise TimeoutError(f"No se pudo adquirir lock {lock_path}")
                time.sleep(poll)

    def _release_file_lock(self, fd, lock_path):
        try:
            os.close(fd)
        except Exception:
            pass
        try:
            if os.path.exists(lock_path):
                os.remove(lock_path)
        except Exception:
            pass

    def guardar_visita_desde_formulario(self):
        """Guarda una nueva visita desde el formulario principal"""
        try:
            if not self.cliente_seleccionado:
                messagebox.showwarning("Cliente requerido", "Por favor seleccione un cliente primero.")
                return

            if not getattr(self, 'domicilio_seleccionado', None):
                messagebox.showwarning("Domicilio requerido", "Por favor seleccione un domicilio para el cliente antes de guardar la visita.")
                return

            # Recoger datos del formulario
            folio_visita = self.entry_folio_visita.get().strip()
            folio_acta = self.entry_folio_acta.get().strip()
            fecha_inicio = self.entry_fecha_inicio.get().strip()
            fecha_termino = self.entry_fecha_termino.get().strip()
            hora_inicio = self.entry_hora_inicio.get().strip()
            hora_termino = self.entry_hora_termino.get().strip()
            # Leer supervisor de forma segura (puede no existir en algunos flujos)
            safe_supervisor_widget = getattr(self, 'entry_supervisor', None)
            try:
                supervisor = safe_supervisor_widget.get().strip() if safe_supervisor_widget and safe_supervisor_widget.winfo_exists() else ""
            except Exception:
                supervisor = ""

            # Leer tipo de documento (conservar la selección tal cual: título / mayúsculas según opciones)
            tipo_documento = (self.combo_tipo_documento.get().strip()
                               if hasattr(self, 'combo_tipo_documento') else "Dictamen")

            # Permitir guardar aunque no haya folio_acta si hay tipo_documento
            if not folio_acta:
                if tipo_documento:
                    # Guardar registro incompleto solo con tipo de documento
                    payload = {
                        "folio_visita": folio_visita,
                        "folio_acta": folio_acta,
                        "fecha_inicio": fecha_inicio,
                        "fecha_termino": fecha_termino,
                        "hora_inicio": hora_inicio,
                        "hora_termino": hora_termino,
                        "norma": "",
                        "cliente": self.cliente_seleccionado['CLIENTE'],
                        "nfirma1": supervisor,
                        "nfirma2": "",
                        "estatus": "En proceso",
                            "tipo_documento": tipo_documento,
                        "folios_utilizados": f"{folio_visita} - {folio_visita}"  # Guardar el folio como rango único
                    }
                    self.hist_create_visita(payload, show_notification=False)
                    self.crear_nueva_visita()
                    messagebox.showinfo("Registro guardado", "El folio se guardó como registro incompleto. Podrá completarlo más adelante.")
                    return
                else:
                    messagebox.showwarning("Datos incompletos", "Por favor ingrese el folio de acta o seleccione un tipo de documento.")
                    return

            # Validar que el folio acta tenga formato correcto
            if not folio_acta.startswith("AC") or len(folio_acta) != 6:
                messagebox.showwarning("Formato incorrecto", "El folio de acta debe tener formato ACXXXX (ej: AC0001).")
                return

            # Crear payload con todos los campos
            payload = {
                "folio_visita": folio_visita,
                "folio_acta": folio_acta,
                "fecha_inicio": fecha_inicio,
                "fecha_termino": fecha_termino,
                "hora_inicio": hora_inicio,
                "hora_termino": hora_termino,
                "norma": "",
                "cliente": self.cliente_seleccionado['CLIENTE'],
                "nfirma1": supervisor,  # Usamos supervisor como única firma
                "nfirma2": "",
                "estatus": "En proceso",
                "tipo_documento": tipo_documento
            }

            # Añadir datos de dirección seleccionada si existen
            try:
                payload['direccion'] = getattr(self, 'direccion_seleccionada', '') or getattr(self, 'domicilio_seleccionado', '')
                # también guardar alias `calle_numero` para compatibilidad con generadores
                payload['calle_numero'] = payload.get('direccion') or getattr(self, 'direccion_seleccionada', '')
                payload['colonia'] = getattr(self, 'colonia_seleccionada', '')
                payload['municipio'] = getattr(self, 'municipio_seleccionado', '')
                payload['ciudad_estado'] = getattr(self, 'ciudad_seleccionada', '')
                payload['cp'] = getattr(self, 'cp_seleccionado', '')
            except Exception:
                pass

            # Guardar visita: validar unicidad de folio de visita y folio de acta
            new_fv = str(payload.get('folio_visita','') or '').strip()
            new_fa = str(payload.get('folio_acta','') or '').strip()
            if new_fv and self._folio_visita_exists(new_fv):
                messagebox.showwarning("Folio duplicado", f"El folio de visita {new_fv} ya está registrado en el historial. Elimine el registro existente para volver a usarlo.")
                return
            # Para folio de acta reutilizar la validación previa (comprobar en disco)
            if new_fa:
                # Buscar AC duplicada en disco
                try:
                    hist_path = getattr(self, 'historial_path', None) or os.path.join(DATA_DIR, 'historial_visitas.json')
                    if os.path.exists(hist_path):
                        with open(hist_path, 'r', encoding='utf-8') as hf:
                            hobj = json.load(hf) or {}
                            latest_visitas = hobj.get('visitas', []) if isinstance(hobj, dict) else (hobj or [])
                    else:
                        latest_visitas = self.historial.get('visitas', []) or []
                except Exception:
                    latest_visitas = self.historial.get('visitas', []) or []
                for v in (latest_visitas or []):
                    try:
                        if new_fa and str(v.get('folio_acta','') or '').strip().lower() == new_fa.lower():
                            messagebox.showwarning("Folio duplicado", f"El folio de acta {new_fa} ya está en uso. No se puede duplicar AC.")
                            return
                    except Exception:
                        continue

            self.hist_create_visita(payload, show_notification=False)
            # Forzar actualización inmediata de la etiqueta de siguiente folio
            try:
                self._update_siguiente_folio_label()
            except Exception:
                pass

            # Limpiar formulario después de guardar
            self.crear_nueva_visita()
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar la visita:\n{e}")
            return

        # Si algo falla al guardar, ya se manejó arriba

    def cargar_excel(self):
        """Carga un archivo Excel y actualiza el UI con el nombre del archivo cargado."""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Archivos Excel", "*.xlsx;*.xls")]
        )
        if not file_path:
            return

        self.archivo_excel_cargado = file_path
        nombre_archivo = os.path.basename(file_path)
        
        try:
            self.info_archivo.configure(
                text=f"📄 {nombre_archivo}",
                text_color=STYLE["exito"]
            )
        except Exception:
            pass
        try:
            self.boton_cargar_excel.configure(state="disabled")
            self.boton_limpiar.configure(state="normal")
        except Exception:
            pass
        
        self.etiqueta_estado.configure(
            text="⏳ Convirtiendo a JSON...", 
            text_color=STYLE["advertencia"]
        )
        self.check_label.configure(text="")
        self.update_idletasks()

        thread = threading.Thread(target=self.convertir_a_json, args=(file_path,))
        thread.daemon = True
        thread.start()

    def convertir_a_json(self, file_path):
        try:
            # Detectar si el Excel tiene una hoja llamada CONCENTRADO (índice)
            try:
                xls = pd.ExcelFile(file_path)
                sheets = [s.upper() for s in (xls.sheet_names or [])]
            except Exception:
                sheets = []

            if 'CONCENTRADO' in sheets:
                # Advertir al usuario: este archivo parece ser un índice, no la tabla principal
                from tkinter import messagebox as _mb
                confirmar = _mb.askyesno(
                    "Archivo con hoja CONCENTRADO detectado",
                    "El archivo seleccionado contiene una hoja llamada 'CONCENTRADO', que parece ser un índice para Pegado por Índice.\n\n¿Desea importarlo AHORA como 'Tabla de Relación' y sobrescribir data/tabla_de_relacion.json?\n\n(Si NO está seguro, use la opción 'Pegado por Índice' en la UI en lugar de 'Cargar Tabla de Relación')"
                )
                if not confirmar:
                    # cancelar la conversión para evitar sobrescritura accidental
                    self.after(0, self.mostrar_error, "Importación cancelada: el archivo parece ser un índice. Use 'Pegado por Índice' para importar índices.")
                    return

            df = pd.read_excel(file_path)
            if df.empty:
                self.mostrar_error("El archivo seleccionado no contiene datos.")
                return

            # Convertir columnas de fecha a string
            for col in df.columns:
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    df[col] = df[col].astype(str)

            # Limpiar nombres de columnas (eliminar espacios extra)
            df.columns = df.columns.str.strip()

            # Buscar y renombrar la columna de solicitud para consistencia
            col_solicitud = self._obtener_columna_solicitud(df)
            if col_solicitud and col_solicitud != 'SOLICITUD':
                df.rename(columns={col_solicitud: 'SOLICITUD'}, inplace=True)

            # Reemplazar NaN por None para mantener claves presentes y serializables
            try:
                df = df.astype(object).where(pd.notnull(df), None)
            except Exception:
                pass

            # Construir lista de registros respetando el orden original de columnas
            cols = list(df.columns)
            records = []
            for idx, row in df.iterrows():
                rec = {}
                for c in cols:
                    # Asegurarse de que la clave exista incluso si el valor es None
                    rec[c] = row.get(c, None)
                # Guardar índice original del DataFrame y una estimación de fila Excel
                try:
                    rec['_orig_index'] = idx
                    if isinstance(idx, int):
                        # Excel típico: cabecera ocupa 1 fila -> pandas idx 0 corresponde a Excel fila 2
                        rec['_excel_row'] = idx + 2
                    else:
                        rec['_excel_row'] = str(idx)
                except Exception:
                    rec['_orig_index'] = None
                    rec['_excel_row'] = None
                records.append(rec)

            # ----------------- VALIDACIÓN DE CAMPOS REQUERIDOS -----------------
            # No permitir continuar si faltan datos críticos en columnas esperadas.
            try:
                # Definir aliases aceptables para cada campo requerido
                required_aliases = {
                    'SOLICITUD': ['SOLICITUD', 'Solicitud', 'solicitud'],
                    'LISTA': ['LISTA', 'Lista', 'lista'],
                    'FIRMA': ['FIRMA', 'Firma', 'firma', 'INSPECTOR', 'Inspector']
                }

                # Normalizar nombres de columnas del DataFrame para comparación
                cols_upper = [c.upper().strip() for c in df.columns]

                missing_cols = []
                # Detectar qué columnas (por campo requerido) no existen bajo ningún alias
                for canonical, aliases in required_aliases.items():
                    present = any(a.upper().strip() in cols_upper for a in aliases)
                    if not present:
                        missing_cols.append(canonical)

                if missing_cols:
                    message = (
                        "Columnas requeridas no encontradas:\n"
                        + "\n".join([f"- {c}" for c in missing_cols])
                        + "\n\nPor favor renombre o agregue esas columnas en el Excel antes de generar los documentos.\n"
                        "(La herramienta considera alias comunes; revise encabezados de columna.)"
                    )
                    try:
                        messagebox.showerror("Campos requeridos faltantes", message)
                    except Exception:
                        self.mostrar_error(message)
                    return

                # Ahora comprobar filas con valores faltantes en al menos una de las claves requeridas
                problematic = []
                for i, rec in enumerate(records, start=1):
                    miss = []
                    current_vals = {}
                    for canonical, aliases in required_aliases.items():
                        found_val = None
                        # Buscar la primera alias presente en el registro con valor no vacío
                        for a in aliases:
                            # Búsqueda de clave exacta (respetando mayúsculas/espacios tal como vienen en rec)
                            if a in rec and rec.get(a) not in (None, ''):
                                found_val = str(rec.get(a)).strip()
                                break
                            # También intentar con versión uppercase/strip de la clave en el registro
                            for rk in list(rec.keys()):
                                if rk and rk.upper().strip() == a.upper().strip():
                                    if rec.get(rk) not in (None, ''):
                                        found_val = str(rec.get(rk)).strip()
                                        break
                            if found_val:
                                break
                        current_vals[canonical] = found_val
                        if not found_val:
                            miss.append(canonical)

                    if miss:
                        problematic.append((i, miss, current_vals))

                if problematic:
                    # Resumen por columna: contar cuántas filas tienen el campo vacío
                    missing_counts = {k: 0 for k in required_aliases.keys()}
                    for rec in records:
                        for canonical, aliases in required_aliases.items():
                            found = False
                            for a in aliases:
                                if a in rec and rec.get(a) not in (None, ''):
                                    found = True
                                    break
                                for rk in list(rec.keys()):
                                    if rk and rk.upper().strip() == a.upper().strip():
                                        if rec.get(rk) not in (None, ''):
                                            found = True
                                            break
                                if found:
                                    break
                            if not found:
                                missing_counts[canonical] += 1

                    # Si alguna columna está vacía en TODAS las filas, informar específicamente
                    full_empty = [c for c, cnt in missing_counts.items() if cnt == len(records)]
                    if full_empty:
                        message = (
                            "Columnas detectadas pero VACÍAS en todos los registros:\n"
                            + "\n".join([f"- {c}" for c in full_empty])
                            + "\n\nPor favor rellene esa(s) columna(s) en el Excel antes de continuar."
                        )
                        try:
                            messagebox.showerror("Columnas vacías", message)
                        except Exception:
                            self.mostrar_error(message)
                        return

                    # Construir mensaje claro y compacto: resumen por columna + ejemplos de filas Excel
                    # Recolectar ejemplos por columna (usar fila Excel si está disponible)
                    missing_examples = {k: [] for k in required_aliases.keys()}
                    for idx, miss, vals in problematic:
                        try:
                            rec = records[idx-1]
                            excel_row = rec.get('_excel_row') or idx
                        except Exception:
                            excel_row = idx
                        for c in miss:
                            if excel_row not in missing_examples[c]:
                                missing_examples[c].append(excel_row)

                    resumen_lines = []
                    for c in required_aliases.keys():
                        cnt = missing_counts.get(c, 0)
                        examples = missing_examples.get(c, [])
                        # Mostrar hasta 30 filas para casos con muchas
                        ex_trim = examples[:30]
                        ex_text = ", ".join(str(x) for x in ex_trim) if ex_trim else "-"
                        resumen_lines.append(f"- {c}: {cnt} filas vacías. Ejemplos (Excel rows): {ex_text}")

                    # Si sólo una columna tiene problemas, mostrar mensaje directo y todas las filas afectadas
                    fields_with_issues = [c for c, cnt in missing_counts.items() if cnt > 0]
                    if len(fields_with_issues) == 1:
                        field = fields_with_issues[0]
                        all_rows = missing_examples.get(field, [])
                        rows_text = ", ".join(str(x) for x in all_rows) if all_rows else "(no disponible)"
                        message = (
                            f"La columna '{field}' contiene {missing_counts.get(field,0)} fila(s) vacías.\n"
                            f"Filas (Excel): {rows_text}\n\n"
                            "Por favor complete la columna '" + field + "' en esas filas y vuelva a importar la tabla de relación."
                        )
                    else:
                        message = (
                            f"Se detectaron {len(problematic)} fila(s) con datos faltantes.\n\n"
                            "Detalles por campo:\n"
                            + "\n".join(resumen_lines)
                            + "\n\nPor favor corrija las filas indicadas en el archivo Excel (use las columnas mostradas arriba) y vuelva a importar la tabla de relación."
                        )
                    try:
                        messagebox.showerror("Filas con datos incompletos", message)
                    except Exception:
                        self.mostrar_error(message)
                    return
            except Exception:
                # Si la validación falla por algún motivo, no bloquear pero registrar
                print('⚠️ Error en validación de columnas requeridas; se continuará con precaución')
            # ----------------- ASIGNAR FOLIOS USANDO FOLIO_MANAGER -----------------
            try:
                # Recolectar pares únicos (SOLICITUD, LISTA)
                pares_vistos = []
                for r in records:
                    sol_val = None
                    for sk in ('SOLICITUD', 'Solicitud', 'solicitud'):
                        if sk in r and r.get(sk) is not None and str(r.get(sk)).strip() != "":
                            sol_val = str(r.get(sk)).strip()
                            break
                    lista_val = None
                    for key in ('LISTA', 'Lista', 'lista'):
                        if key in r and r.get(key) is not None and str(r.get(key)).strip() != "":
                            lista_val = str(r.get(key)).strip()
                            break
                    if lista_val is None:
                        continue
                    pair = (sol_val or '', lista_val)
                    if pair not in pares_vistos:
                        pares_vistos.append(pair)

                total_necesarios = len(pares_vistos)
                pair_to_folio = {}

                if total_necesarios > 0:
                    try:
                        maxf = 0
                        visitas = []
                        if os.path.exists(self.historial_path):
                            with open(self.historial_path, 'r', encoding='utf-8') as hf:
                                hj = json.load(hf)
                                visitas = hj.get('visitas', [])
                        import re
                        for v in visitas:
                            # 1) Preferir leer archivo `data/folios_visitas/folios_<folio_visita>.json`
                            fid = v.get('folio_visita') or v.get('folio')
                            try:
                                if fid:
                                    archivo_f = os.path.join(self.folios_visita_path, f"folios_{fid}.json")
                                    if os.path.exists(archivo_f):
                                        with open(archivo_f, 'r', encoding='utf-8') as fh:
                                            arr = json.load(fh) or []
                                        for entry in arr:
                                            fol = entry.get('FOLIOS') or ''
                                            nums = re.findall(r"\d+", str(fol))
                                            for d in nums:
                                                try:
                                                    n = int(d)
                                                    if n > maxf:
                                                        maxf = n
                                                except Exception:
                                                    pass
                                        continue
                            except Exception:
                                pass

                            # 2) Fallback: usar campo `folios_utilizados` del historial (puede contener rango o lista)
                            try:
                                fu = v.get('folios_utilizados') or ''
                                if fu:
                                    nums = re.findall(r"\d+", str(fu))
                                    for d in nums:
                                        try:
                                            n = int(d)
                                            if n > maxf:
                                                maxf = n
                                        except Exception:
                                            pass
                            except Exception:
                                pass

                        next_local = maxf + 1
                        if int(next_local) == 1:
                            try:
                                import folio_manager
                                curr = folio_manager.get_last()
                                if curr and int(curr) > 0:
                                    next_local = int(curr) + 1
                            except Exception:
                                pass
                    except Exception:
                        next_local = 1
                    for pair in pares_vistos:
                        pair_to_folio[pair] = int(next_local)
                        print(f"      → SOL {pair[0]} LISTA {pair[1]} → Folio {int(next_local):06d} (in-memory)")
                        next_local += 1

                # Propagar folios a los registros
                asignados = 0
                for r in records:
                    sol_val = None
                    for sk in ('SOLICITUD', 'Solicitud', 'solicitud'):
                        if sk in r and r.get(sk) is not None and str(r.get(sk)).strip() != "":
                            sol_val = str(r.get(sk)).strip()
                            break
                    lista_val = None
                    for key in ('LISTA', 'Lista', 'lista'):
                        if key in r and r.get(key) is not None and str(r.get(key)).strip() != "":
                            lista_val = str(r.get(key)).strip()
                            break
                    if lista_val is None:
                        continue
                    fol_asig = pair_to_folio.get((sol_val or '', lista_val))
                    if fol_asig is not None:
                        try:
                            r['FOLIO'] = int(fol_asig)
                        except Exception:
                            r['FOLIO'] = str(fol_asig)
                        asignados += 1

                if asignados:
                    print(f"🔢 Asignados {asignados} registros a {len(pair_to_folio)} folios únicos")
            except Exception as e:
                print(f"⚠️ Error asignando folios automáticos secuenciales: {e}")


            data_folder = DATA_DIR
            os.makedirs(data_folder, exist_ok=True)

            self.json_filename = "tabla_de_relacion.json"
            output_path = os.path.join(data_folder, self.json_filename)

            # Normalizar campos CODIGO y SKU para que se guarden como strings sin '.0'
            def _norm_code(v):
                try:
                    import pandas as _pd
                    if _pd.isna(v):
                        return None
                except Exception:
                    pass
                if v is None:
                    return None
                if isinstance(v, float):
                    if v.is_integer():
                        return str(int(v))
                    return format(v, 'g')
                if isinstance(v, int):
                    return str(v)
                s = str(v).strip()
                if s.endswith('.0'):
                    s = s[:-2]
                if s.lower() == 'nan' or s == '':
                    return None
                return s

            for rec in records:
                # Normalizar claves comunes (mayúsculas esperadas en la tabla)
                if 'CODIGO' in rec:
                    rec['CODIGO'] = _norm_code(rec.get('CODIGO'))
                if 'SKU' in rec:
                    rec['SKU'] = _norm_code(rec.get('SKU'))

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(records, f, ensure_ascii=False, indent=2)

        
            try:
                print("   ℹ️ Conversión completada: no se crea backup PERSIST en esta etapa.")
            except Exception:
                pass

            # EXTRAER Y GUARDAR INFORMACIÓN DE FOLIOS
            self._extraer_informacion_folios(records)
            try:
                # actualizar indicador visual de siguiente folio
                self._update_siguiente_folio_label()
            except Exception:
                pass

            # Validar que existan firmas/inspectores que cubran las normas requeridas
            try:
                ok = self._validate_tabla_normas(records)
                if not ok:
                    # Bloquear la continuación hasta que el usuario agregue/seleccione firma(s)
                    self.after(0, self.mostrar_error, "La tabla de relación requiere firmas para todas las normas. Por favor agregue la(s) firma(s) necesarias desde el catálogo de supervisores.")
                    return
            except Exception:
                # Si falla la validación no evitar la conversión, pero informar en consola
                print("⚠️ Error validando firmas de la tabla (continuando):", sys.exc_info()[0])

        
            if hasattr(self, 'current_folio') and self.current_folio:
                # Regenerar cache exportable para Excel (persistente)
                try:
                    self._generar_datos_exportable()
                except Exception:
                    pass

            self.after(0, self._actualizar_ui_conversion_exitosa, output_path, len(records))

        except Exception as e:
            self.after(0, self.mostrar_error, f"Error al convertir el archivo:\n{e}")
    
    def _extraer_informacion_folios(self, datos_tabla):
        """Extrae y procesa la información de folios de la tabla de relación"""
        try:
            # Verificar si hay datos en la tabla
            if not datos_tabla:
                return {
                    "hay_folios": False,
                    "total_folios": 0,
                    "total_folios_numericos": 0,
                    "mensaje": "No hay datos en la tabla"
                }
            
            folios_encontrados = []
            folios_numericos = []
            hay_folios_asignados = False
            
            # Buscar la columna FOLIO en los datos
            for item in datos_tabla:
                if 'FOLIO' in item:
                    folio_valor = item['FOLIO']
                    
                    # Verificar si el folio tiene un valor asignado (no NaN, None o vacío)
                    if (folio_valor is not None and 
                        str(folio_valor).strip() != "" and 
                        str(folio_valor).lower() != "nan" and
                        str(folio_valor).lower() != "none"):
                        
                        hay_folios_asignados = True
                        folio_str = str(folio_valor).strip()
                        
                        # Intentar convertir a número y formatear a 6 dígitos
                        try:
                            # Manejar casos donde folio_str puede ser decimal
                            folio_num = int(float(folio_str))
                            folios_numericos.append(folio_num)
                            folios_encontrados.append(f"{folio_num:06d}")
                        except (ValueError, TypeError):
                            # Si no se puede convertir, usar el valor original
                            folios_encontrados.append(folio_str)
            
            # Procesar la información de folios
            info_folios = {
                "hay_folios": hay_folios_asignados,
                "total_folios": len(folios_encontrados),
                "total_registros": len(datos_tabla),
                "total_folios_numericos": len(folios_numericos),
                "rango_folios": "",
                "lista_folios": folios_encontrados,
                "folios_formateados": folios_encontrados,
                "mensaje": ""
            }
            
            # Calcular rango si hay folios numéricos (usar folios NUMÉRICOS ÚNICOS)
            if folios_numericos:
                try:
                    unique_nums = sorted(set(int(x) for x in folios_numericos))
                except Exception:
                    unique_nums = sorted(set([int(x) for x in folios_numericos if isinstance(x, int)])) if folios_numericos else []

                if unique_nums:
                    min_folio = unique_nums[0]
                    max_folio = unique_nums[-1]
                    info_folios["rango_folios"] = f"{min_folio:06d} - {max_folio:06d}"
                    info_folios["rango_numerico"] = f"{min_folio} - {max_folio}"

                    # Determinar mensaje usando folios únicos
                    if len(unique_nums) == 1:
                        info_folios["mensaje"] = f"Folio: {min_folio:06d}"
                    else:
                        es_consecutivo = all(
                            unique_nums[i] + 1 == unique_nums[i + 1]
                            for i in range(len(unique_nums) - 1)
                        )
                        if es_consecutivo:
                            info_folios["mensaje"] = f"Total: {len(unique_nums)} | Rango: {min_folio:06d} - {max_folio:06d}"
                        else:
                            info_folios["mensaje"] = f"Total: {len(unique_nums)} | Folios asignados"

                    # Actualizar listas formateadas y totales con folios únicos
                    info_folios["lista_folios"] = [f"{x:06d}" for x in unique_nums]
                    info_folios["folios_formateados"] = info_folios["lista_folios"]
                    info_folios["total_folios"] = len(info_folios["lista_folios"])
                    info_folios["total_folios_numericos"] = len(info_folios["lista_folios"])
                else:
                    info_folios["mensaje"] = f"Total: {len(folios_encontrados)} | Folios asignados"
            elif hay_folios_asignados:
                # Si hay folios pero no son numéricos
                info_folios["mensaje"] = f"Total: {len(folios_encontrados)} | Folios no numéricos"
            else:
                # Si no hay folios asignados
                info_folios["mensaje"] = f"Total: {len(datos_tabla)} | Sin folios asignados"
            
            # Guardar información de folios para usar después
            self.info_folios_actual = info_folios
            
            print(f"📊 Información de folios extraída:")
            print(f"   - ¿Hay folios asignados?: {'Sí' if hay_folios_asignados else 'No'}")
            print(f"   - Total registros: {info_folios['total_registros']}")
            print(f"   - Folios asignados: {info_folios['total_folios']}")
            print(f"   - Folios numéricos: {info_folios['total_folios_numericos']}")
            print(f"   - Mensaje: {info_folios['mensaje']}")
            if folios_numericos and len(folios_numericos) > 1:
                print(f"   - Rango: {info_folios['rango_folios']}")
            
            return info_folios
            
        except Exception as e:
            print(f"⚠️ Error extrayendo información de folios: {e}")
            return {
                "hay_folios": False,
                "total_folios": 0,
                "total_folios_numericos": 0,
                "mensaje": f"Error: {str(e)}"
            }

    def _extract_normas_from_records(self, records):
        """Intento heurístico para extraer el conjunto de normas requeridas desde la tabla de relación."""
        try:
            import re
            posibles = ['NORMA','Norma','norma','NORMAS','Normas','normas','REQUISITOS','REQUERIMIENTOS','REQUISITO','requisito']
            normas = set()
            if not records:
                return normas
            for r in records:
                if not isinstance(r, dict):
                    continue
                for k in posibles:
                    if k in r and r.get(k):
                        v = r.get(k)
                        if isinstance(v, (list, tuple, set)):
                            for it in v:
                                for part in re.split(r'[;,/\\|]', str(it)):
                                    p = part.strip()
                                    if p:
                                        normas.add(p)
                        else:
                            for part in re.split(r'[;,/\\|]', str(v)):
                                p = part.strip()
                                if p:
                                    normas.add(p)
                        break
            # Normalizar espacios
            normas = set([n.strip() for n in normas if n and str(n).strip()])
            return normas
        except Exception:
            return set()

    def _load_supervisores_catalog(self):
        """Carga el catálogo de supervisores/firmas desde `DATA_DIR/Firmas.json` si existe."""
        path = os.path.join(DATA_DIR, 'Firmas.json')
        if not os.path.exists(path):
            return []
        try:
            with open(path, 'r', encoding='utf-8') as f:
                arr = json.load(f) or []
            # Normalizar esquema mínimo
            cleaned = []
            for s in arr:
                if not isinstance(s, dict):
                    continue
                nombre = s.get('nombre') or s.get('Nombre') or s.get('nombre_supervisor') or ''
                normas_field = s.get('normas') or s.get('Normas') or s.get('NORMAS') or s.get('lista_normas') or ''
                normas_set = set()
                if isinstance(normas_field, (list, tuple, set)):
                    for n in normas_field:
                        if n:
                            normas_set.add(str(n).strip())
                else:
                    import re
                    for p in re.split(r'[;,/\\|]', str(normas_field)):
                        if p and p.strip():
                            normas_set.add(p.strip())
                cleaned.append({'nombre': nombre, 'normas': sorted(normas_set)})
            return cleaned
        except Exception:
            return []

    def _save_supervisores_catalog(self, arr):
        try:
            path = os.path.join(DATA_DIR, 'Firmas.json')
            with open(path, 'w', encoding='utf-8') as f:
                json.dump(arr, f, ensure_ascii=False, indent=2)
            return True
        except Exception:
            return False

    def _validate_tabla_normas(self, records):
        """Valida que el catálogo de supervisores cubra las normas encontradas en la tabla.
        Si faltan normas, solicita al usuario agregar un supervisor que las cubra.
        Devuelve True si la validación queda satisfecha (se encontró o se agregó supervisor).
        Devuelve False si el usuario cancela.
        """
        try:
            needed = self._extract_normas_from_records(records)
            if not needed:
                return True

            supervisors = self._load_supervisores_catalog()
            covered = set()
            for s in supervisors:
                for n in (s.get('normas') or []):
                    if n:
                        covered.add(str(n).strip())

            missing = set(n for n in needed if n not in covered)
            if not missing:
                return True

            # Informar al usuario y ofrecer agregar un supervisor que cubra las normas faltantes
            msg = "Se requieren firmas para las normas:\n\n" + "\n".join(sorted(missing))
            msg += "\n\n¿Desea agregar un supervisor que cubra estas normas ahora?"
            abrir = messagebox.askyesno("Faltan firmas", msg)
            if not abrir:
                return False

            # Pedir nombre y normas (prellenar con las faltantes)
            nombre = simpledialog.askstring("Añadir supervisor", "Ingrese el nombre del supervisor que cubrirá las normas:", initialvalue="")
            if not nombre or not str(nombre).strip():
                return False
            import re
            normas_input = simpledialog.askstring("Normas", "Ingrese las normas separadas por comas", initialvalue=", ".join(sorted(missing)))
            if not normas_input:
                return False

            normas_list = [p.strip() for p in re.split(r'[;,/\\|]', normas_input) if p and p.strip()]
            new = {'nombre': str(nombre).strip(), 'normas': normas_list}
            # Añadir al catálogo y guardar
            supervisors.append(new)
            self._save_supervisores_catalog(supervisors)
            messagebox.showinfo("Supervisor agregado", f"Se agregó el supervisor '{new['nombre']}' con normas: {', '.join(new['normas'])}.")
            return True
        except Exception as e:
            print("⚠️ Error en validación de normas:", e)
            return False

    def verificar_datos_folios_existentes(self):
        """Verifica y repara datos de folios existentes para asegurar consistencia"""
        try:
            print("🔍 Verificando datos de folios existentes...")
            
            if not os.path.exists(self.folios_visita_path):
                print("ℹ️ No hay carpeta de folios para verificar")
                return
            
            # Listar todos los archivos JSON de folios
            archivos_folios = [f for f in os.listdir(self.folios_visita_path) if f.endswith('.json')]
            
            archivos_reparados = 0
            for archivo in archivos_folios:
                archivo_path = os.path.join(self.folios_visita_path, archivo)
                
                try:
                    with open(archivo_path, 'r', encoding='utf-8') as f:
                        datos = json.load(f)
                    
                    datos_modificados = False
                    
                    # Verificar y reparar cada registro
                    for item in datos:
                        # Reparar formato de FOLIOS a 6 dígitos
                        if 'FOLIOS' in item:
                            folio_raw = item['FOLIOS']
                            if folio_raw:
                                try:
                                    # Intentar convertir a número y formatear
                                    folio_num = int(float(str(folio_raw)))
                                    folio_formateado = f"{folio_num:06d}"
                                    
                                    if folio_formateado != str(folio_raw):
                                        item['FOLIOS'] = folio_formateado
                                        datos_modificados = True
                                        print(f"   🔧 Reparado: {folio_raw} -> {folio_formateado}")
                                except (ValueError, TypeError):
                                    pass
                    
                    # Guardar si hubo modificaciones
                    if datos_modificados:
                        with open(archivo_path, 'w', encoding='utf-8') as f:
                            json.dump(datos, f, ensure_ascii=False, indent=2)
                        archivos_reparados += 1
                        print(f"✅ Archivo reparado: {archivo}")
                        
                except Exception as e:
                    print(f"⚠️ Error procesando archivo {archivo}: {e}")
            
            print(f"📊 Verificación completada. Archivos reparados: {archivos_reparados}/{len(archivos_folios)}")
            
        except Exception as e:
            print(f"❌ Error en verificación de datos: {e}")

    def _obtener_folios_de_tabla(self):
        """Obtiene la información de folios de la tabla de relación con formato mejorado"""
        try:
            if not hasattr(self, 'info_folios_actual') or not self.info_folios_actual:
                return "No disponible"
            
            info = self.info_folios_actual
            
            if info['rango_folios']:
                return f"Total: {info['total_folios']} | Rango: {info['rango_folios']}"
            else:
                return f"Total: {info['total_folios']} folios"
                
        except Exception as e:
            print(f"⚠️ Error obteniendo folios de tabla: {e}")
            return "Error al obtener folios"

    def _actualizar_ui_conversion_exitosa(self, output_path, num_registros):
        self.archivo_json_generado = output_path
        self.etiqueta_estado.configure(
            text=f"✅ Convertido - {num_registros} registros", 
            text_color=STYLE["exito"]
        )
        self.check_label.configure(text="✓")
        
        if self.cliente_seleccionado:
            self.boton_generar_dictamen.configure(state="normal")
        
        messagebox.showinfo(
            "Conversión exitosa",
            f"Archivo convertido correctamente.\n\n"
            f"Ubicación: {output_path}\n"
            f"Total de registros: {num_registros}"
        )

    # ----------------- Helpers para folio documento (visual) -----------------
    def _get_next_document_folio(self):
        """Calcula el siguiente folio de documento disponible a partir del historial.
        Esto es solo informativo y no persiste nada. Devuelve entero (1-based).
        """
        try:
            # Buscar el máximo folio numérico entre los archivos data/folios_visitas/folios_*.json
            maxf = 0
            try:
                dirp = self.folios_visita_path
                if os.path.exists(dirp):
                    for fn in os.listdir(dirp):
                        if fn.startswith('folios_') and fn.endswith('.json'):
                            pathf = os.path.join(dirp, fn)
                            try:
                                with open(pathf, 'r', encoding='utf-8') as fh:
                                    arr = json.load(fh) or []
                                    for entry in arr:
                                        fol = entry.get('FOLIOS') or entry.get('FOLIOS', '')
                                        if not fol:
                                            continue
                                        digits = ''.join([c for c in str(fol) if c.isdigit()])
                                        if digits:
                                            try:
                                                n = int(digits)
                                                if n > maxf:
                                                    maxf = n
                                            except Exception:
                                                pass
                            except Exception:
                                continue
            except Exception:
                pass

            # Considerar reservas pendientes y folios recién calculados en memoria
            try:
                # `self.pending_folios` contiene reservas no persistidas o recién añadidas
                for p in getattr(self, 'pending_folios', []) or []:
                    try:
                        fus = p.get('folios_utilizados') or p.get('folios') or []
                        if isinstance(fus, list):
                            for f in fus:
                                digits = ''.join([c for c in str(f) if c.isdigit()])
                                if digits:
                                    try:
                                        n = int(digits)
                                        if n > maxf:
                                            maxf = n
                                    except Exception:
                                        pass
                        else:
                            digits = ''.join([c for c in str(fus) if c.isdigit()])
                            if digits:
                                try:
                                    n = int(digits)
                                    if n > maxf:
                                        maxf = n
                                except Exception:
                                    pass
                    except Exception:
                        continue

                # También considerar `folios_utilizados_actual` (datos cargados desde la tabla)
                try:
                    for f in getattr(self, 'folios_utilizados_actual', []) or []:
                        digits = ''.join([c for c in str(f) if c.isdigit()])
                        if digits:
                            try:
                                n = int(digits)
                                if n > maxf:
                                    maxf = n
                            except Exception:
                                pass
                except Exception:
                    pass
            except Exception:
                pass

            # Además, consultar el contador maestro a través de `folio_manager`
            # (usa `FOLIO_DATA_DIR` cuando esté definido). Esto evita que la
            # UI lea un archivo distinto al que usa el proceso que reserva
            # folios (por ejemplo, al ejecutar en .exe con rutas embebidas).
            try:
                import folio_manager
                try:
                    last = int(folio_manager.get_last() or 0)
                    if last > maxf:
                        maxf = last
                except Exception:
                    pass
            except Exception:
                # Fallback: mantener comportamiento anterior si folio_manager
                # no está disponible por alguna razón.
                try:
                    data_dir = os.path.dirname(self.historial_path)
                    contador_path = os.path.join(data_dir, "folio_counter.json")
                    if os.path.exists(contador_path):
                        try:
                            with open(contador_path, 'r', encoding='utf-8') as cf:
                                j = json.load(cf) or {}
                                last = int(j.get('last', 0))
                                if last > maxf:
                                    maxf = last
                        except Exception:
                            pass
                except Exception:
                    pass

            return maxf + 1
        except Exception:
            return 1








    def _update_siguiente_folio_label(self):
        try:
            # Si hay un archivo cargado y se extrajo la lista de folios usados
            # mostramos el rango propuesto (visual, no persiste el contador)
            nxt = self._get_next_document_folio()
            txt = None
            try:
                if hasattr(self, 'folios_utilizados_actual') and getattr(self, 'archivo_json_generado', None):
                    total = len(self.folios_utilizados_actual or [])
                    if total > 0:
                        start = nxt
                        end = nxt + total - 1
                        if start == end:
                            txt = f"Siguiente folio documento: {start:06d}"
                        else:
                            txt = f"Siguiente folio documento: {start:06d} - {end:06d} (propuesta)"
            except Exception:
                txt = None

            if not txt:
                txt = f"Siguiente folio documento: {nxt:06d}"

            if hasattr(self, 'lbl_siguiente_folio_doc'):
                try:
                    self.lbl_siguiente_folio_doc.configure(text=txt)
                    # Forzar refresco inmediato del GUI para que el cambio se vea al instante
                    try:
                        self.update_idletasks()
                    except Exception:
                        try:
                            self.update()
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception:
            pass












    def limpiar_archivo(self):
        self.archivo_excel_cargado = None
        self.archivo_json_generado = None
        self.json_filename = None

        self.info_archivo.configure(
            text="No se ha cargado ningún archivo",
            text_color=STYLE["texto_claro"]
        )

        self.boton_cargar_excel.configure(state="normal")
        self.boton_limpiar.configure(state="disabled")
        self.boton_generar_dictamen.configure(state="disabled")

        self.etiqueta_estado.configure(text="", text_color=STYLE["texto_claro"])
        self.check_label.configure(text="")
        self.barra_progreso.set(0)
        self.etiqueta_progreso.configure(text="")

        try:
            data_dir = DATA_DIR
            os.makedirs(data_dir, exist_ok=True)

            archivos_a_eliminar = [
                "base_etiquetado.json",
                "tabla_de_relacion.json"
            ]

            archivos_eliminados = []
            
            for archivo in archivos_a_eliminar:
                ruta_archivo = os.path.join(data_dir, archivo)
                if os.path.exists(ruta_archivo):
                    os.remove(ruta_archivo)
                    archivos_eliminados.append(archivo)
                    print(f"🗑️ {archivo} eliminado correctamente.")
            
            if archivos_eliminados:
                print(f"✅ Se eliminaron {len(archivos_eliminados)} archivos: {', '.join(archivos_eliminados)}")
            else:
                print("ℹ️ No se encontraron archivos para eliminar.")

            self.archivo_etiquetado_json = None
            self.info_etiquetado.configure(text="")
            self.info_etiquetado.pack_forget()

            try:
                if hasattr(self, 'current_folio') and self.current_folio:
                    import re
                    backup_dir = os.path.join(data_dir, 'tabla_relacion_backups')
                    if os.path.exists(backup_dir):
                        deleted = 0
                        # Buscar secuencia de dígitos en current_folio
                        matches = re.findall(r"\d+", str(self.current_folio))
                        key = matches[-1] if matches else str(self.current_folio)
                        for fn in os.listdir(backup_dir):
                            try:
                                # No eliminar backups marcados como persistentes
                                if key and key in fn and 'PERSIST' not in fn.upper():
                                    os.remove(os.path.join(backup_dir, fn))
                                    deleted += 1
                            except Exception:
                                continue
                        if deleted:
                            print(f"🗑️ Eliminados {deleted} backups relacionados en tabla_relacion_backups")
            except Exception as e:
                print(f"⚠️ Error eliminando backups de tabla_relacion_backups: {e}")
        except Exception:
            pass

        except Exception as e:
            print(f"⚠️ Error al eliminar archivos: {e}")

        messagebox.showinfo("Limpieza completa", "Los datos del archivo y el etiquetado han sido limpiados.")

    def generar_dictamenes(self):
        if not self.archivo_json_generado:
            messagebox.showwarning("Sin datos", "No hay archivo JSON disponible para generar dictámenes.")
            return

        if not self.cliente_seleccionado:
            messagebox.showwarning("Cliente no seleccionado", "Por favor seleccione un cliente antes de generar los dictámenes.")
            return

        if not getattr(self, 'domicilio_seleccionado', None):
            messagebox.showwarning("Domicilio no seleccionado", "Por favor seleccione un domicilio para el cliente antes de generar los dictámenes.")
            return

        try:
            # Leer el archivo JSON y extraer los folios
            with open(self.archivo_json_generado, 'r', encoding='utf-8') as f:
                datos = json.load(f)

            # Extraer folios únicos y ordenados
            folios = set()
            for item in datos:
                if 'FOLIO' in item and item['FOLIO']:
                    try:
                        folio = int(item['FOLIO'])
                        folios.add(folio)
                    except (ValueError, TypeError):
                        # Si no se puede convertir a entero, ignorar
                        pass

            # Convertir a lista ordenada
            folios_ordenados = sorted(folios)
            self.folios_utilizados_actual = folios_ordenados

            # Continuar con la generación...
            # Si hay un folio reservado seleccionado, advertir al usuario
            if getattr(self, 'usando_folio_reservado', False) and getattr(self, 'selected_pending_id', None):
                sel_msg = f"Hay un folio reservado seleccionado (ID: {self.selected_pending_id}). Si confirma, se usará ese folio para la visita.\n\n"
                sel_msg += "¿Desea continuar y usar el folio reservado?"
                if not messagebox.askyesno("Folio reservado seleccionado", sel_msg):
                    return

            # Antes de confirmar, validar que los inspectores asignados estén acreditados
            try:
                # Determinar visit actual en el historial por folio
                visit_folio_key = f"CP{self.current_folio}" if hasattr(self, 'current_folio') and self.current_folio else None
                visit_actual = None
                try:
                    visitas_hist = self.historial.get('visitas', []) if isinstance(self.historial, dict) else []
                    for v in (visitas_hist or []):
                        try:
                            if visit_folio_key and (v.get('folio_visita') or v.get('folio') ) == visit_folio_key:
                                visit_actual = v
                                break
                        except Exception:
                            continue
                except Exception:
                    visit_actual = None

                # Extraer normas desde los datos (buscar claves comunes)
                normas_en_datos = set()
                for item in (datos or []):
                    try:
                        # buscar claves que contengan 'norm' o exactamente 'nom'
                        for k, val in (item or {}).items():
                            if not k or val is None:
                                continue
                            kn = str(k).lower()
                            if 'norm' in kn or kn == 'nom':
                                # soportar listas o strings
                                if isinstance(val, (list, tuple)):
                                    for s in val:
                                        try:
                                            normas_en_datos.add(str(s).strip())
                                        except Exception:
                                            continue
                                else:
                                    for s in str(val).split(','):
                                        s2 = s.strip()
                                        if s2:
                                            normas_en_datos.add(s2)
                    except Exception:
                        continue

                problemas = []
                sugerencias = {}
                if visit_actual and normas_en_datos:
                    # cargar Firmas.json para mapa nombre->normas
                    try:
                        firmas_path = os.path.join(DATA_DIR, 'Firmas.json')
                        with open(firmas_path, 'r', encoding='utf-8') as ff:
                            firmas_data = json.load(ff)
                    except Exception:
                        firmas_data = []

                    firma_map = {}
                    for f in (firmas_data or []):
                        try:
                            name = f.get('NOMBRE DE INSPECTOR') or f.get('NOMBRE') or ''
                            normas_ac = f.get('Normas acreditadas') or f.get('Normas') or []
                            firma_map[name] = [str(n).strip() for n in (normas_ac or [])]
                        except Exception:
                            continue

                    # inspectores asignados en la visita
                    assigned_raw = (visit_actual.get('supervisores_tabla') or visit_actual.get('nfirma1') or '')
                    assigned = [s.strip() for s in str(assigned_raw).split(',') if s.strip()]

                    for norma in normas_en_datos:
                        # verificar si al menos un assigned tiene la norma
                        ok = False
                        missing_inspectores = []
                        for a in assigned:
                            try:
                                acc = firma_map.get(a, [])
                                if any(str(n).strip() == str(norma).strip() for n in (acc or [])):
                                    ok = True
                                    break
                                else:
                                    missing_inspectores.append(a)
                            except Exception:
                                missing_inspectores.append(a)
                        if not ok:
                            problemas.append((norma, missing_inspectores))
                            # sugerir inspectores que sí tienen la norma
                            sugeridas = []
                            for name, acc in firma_map.items():
                                try:
                                    if any(str(n).strip() == str(norma).strip() for n in (acc or [])):
                                        sugeridas.append(name)
                                except Exception:
                                    continue
                            sugerencias[norma] = sugeridas

                if problemas:
                    # Construir mensaje resumido
                    msg_lines = ["Se detectaron posibles conflictos de acreditación para las siguientes normas:"]
                    for norma, miss in problemas:
                        msg_lines.append(f"- {norma}: inspectores asignados no acreditados -> {', '.join(miss) or 'N/A'}")
                        sug = sugerencias.get(norma) or []
                        if sug:
                            msg_lines.append(f"  Sugeridos: {', '.join(sug[:5])}{'...' if len(sug)>5 else ''}")
                        else:
                            msg_lines.append(f"  Sugeridos: (ninguno encontrado)")

                    msg_lines.append("")
                    # Preparar lista única de sugerencias
                    sugeridos_unicos = []
                    try:
                        for norma, lst in sugerencias.items():
                            for s in (lst or []):
                                if s and s not in sugeridos_unicos:
                                    sugeridos_unicos.append(s)
                    except Exception:
                        sugeridos_unicos = []

                    # No abrir interfaz modal: registrar sugerencias en log y continuar.
                    if sugeridos_unicos:
                        try:
                            print("[ADVERTENCIA] Inspectores sugeridos detectados:", ", ".join(sugeridos_unicos))
                        except Exception:
                            pass

            except Exception:
                # en caso de errores en la validación, continuar con confirmación normal
                pass

            tipo_sel = (self.combo_tipo_documento.get().strip() if hasattr(self, 'combo_tipo_documento') else 'Dictamen')
            titulo_confirm = f"Generar {tipo_sel}s" if tipo_sel else "Generar Documentos"
            confirmacion = messagebox.askyesno(
                titulo_confirm,
                f"¿Está seguro de que desea generar {tipo_sel.lower()}s?\n\n"
                f"📄 Archivo: {os.path.basename(self.archivo_json_generado) if getattr(self, 'archivo_json_generado', None) else 'N/A'}\n"
                f"👤 Cliente: {self.cliente_seleccionado['CLIENTE']}\n"
                f"📋 RFC: {self.cliente_seleccionado.get('RFC', 'No disponible')}\n"
                f"📊 Total de folios: {len(folios_ordenados)}"
            )
            
            if not confirmacion:
                return

            self.generando_dictamenes = True
            self.boton_generar_dictamen.configure(state="disabled")
            self.barra_progreso.set(0)
            self.etiqueta_progreso.configure(
                text="⏳ Iniciando generación de documentos...",
                text_color=STYLE["advertencia"]
            )
            self.update_idletasks()

            thread = threading.Thread(target=self._ejecutar_generador_con_progreso)
            thread.daemon = True
            thread.start()
            try:
                # iniciar watcher que mantiene la barra moviéndose si el generador no informa progreso
                self._start_progress_watcher()
            except Exception:
                pass

        except Exception as e:
            self.mostrar_error(f"No se pudo iniciar el generador:\n{e}")

    def _actualizar_ui_conversion_exitosa(self, output_path, num_registros):
        self.archivo_json_generado = output_path
        info_folios_text = ""
        try:
            if hasattr(self, 'info_folios_actual') and self.info_folios_actual:
                info = self.info_folios_actual
                total = int(info.get('total_folios') or info.get('total_folios_numericos') or 0)
                if total > 0:
                    try:
                        start = int(self._get_next_document_folio())
                        end = start + total - 1
                        info_folios_text = f" | 📋 Folios: {start:06d} - {end:06d}"
                    except Exception:
                        if info.get('rango_folios'):
                            info_folios_text = f" | 📋 Folios: {info['rango_folios']}"
                        else:
                            info_folios_text = f" | 📋 Folios: {info['total_folios']} encontrados"
                else:
                    if info.get('rango_folios'):
                        info_folios_text = f" | 📋 Folios: {info['rango_folios']}"
                    else:
                        info_folios_text = f" | 📋 Folios: {info['total_folios']} encontrados"
        except Exception:
            info_folios_text = ""

        self.etiqueta_estado.configure(text=f" {info_folios_text}", text_color=STYLE["exito"]) 
        self.check_label.configure(text="✓")
        
        if self.cliente_seleccionado:
            self.boton_generar_dictamen.configure(state="normal")

    def _ejecutar_generador_con_progreso(self):
        try:
            if not self.winfo_exists():
                return
                
            tipo_sel = (self.combo_tipo_documento.get().strip() if hasattr(self, 'combo_tipo_documento') else 'Dictamen')
            tipo_upper = tipo_sel.upper() if tipo_sel else 'DICTAMEN'

            if 'CONSTANCIA' in tipo_upper:
                try:
                    import importlib.util, time
                    const_file = os.path.join(BASE_DIR, 'Documentos Inspeccion', 'Constancia.py')
                    if os.path.exists(const_file):
                        spec = importlib.util.spec_from_file_location('Constancia', const_file)
                        if spec and getattr(spec, 'loader', None):
                            mod = importlib.util.module_from_spec(spec)
                            sys.modules['Constancia'] = mod
                            spec.loader.exec_module(mod)
                        else:
                            mod = None
                    else:
                        mod = None
                except Exception:
                    mod = None

                created = []
                errores = []
                try:
                    tabla_path = os.path.join(DATA_DIR, 'tabla_de_relacion.json')
                    tabla_data = []
                    if os.path.exists(tabla_path):
                        with open(tabla_path, 'r', encoding='utf-8') as tf:
                            tabla_data = json.load(tf) or []

                    folios_sel = set(getattr(self, 'folios_utilizados_actual', []) or [])
                    filas = []
                    if tabla_data:
                        for r in tabla_data:
                            try:
                                f = r.get('FOLIO')
                                if f is None:
                                    filas.append(r)
                                    continue
                                if folios_sel:
                                    if (str(f).isdigit() and int(f) in folios_sel) or (str(f) in map(str, folios_sel)):
                                        filas.append(r)
                                else:
                                    filas.append(r)
                            except Exception:
                                continue

                    grupos = {}
                    for r in filas:
                        lista = r.get('LISTA') or r.get('Lista') or r.get('lista') or '0'
                        grupos.setdefault(str(lista), []).append(r)

                    # Preparar seguimiento de progreso para Constancias
                    try:
                        total_listas = len(grupos) if grupos else 0
                    except Exception:
                        total_listas = 0
                    listas_procesadas = 0

                    try:
                        sys.path.append(BASE_DIR)
                        import generador_dictamen as gd
                        limpiar_nombre = getattr(gd, 'limpiar_nombre_archivo', lambda n: n.replace('/', '_'))
                    except Exception:
                        limpiar_nombre = lambda n: n.replace('/', '_')

                    ConstGen = getattr(mod, 'ConstanciaPDFGenerator', None) if mod is not None else None

                    # Directorio para JSONs (siempre dentro de data/Constancias)
                    out_base_json = os.path.join(DATA_DIR, 'Constancias')
                    os.makedirs(out_base_json, exist_ok=True)

                    # Solicitar al usuario la carpeta destino para los PDFs (se ejecuta en hilo UI)
                    selected = {'done': False, 'path': None}
                    def _askdir():
                        try:
                            p = filedialog.askdirectory(initialdir=DATA_DIR, title='Seleccionar carpeta destino para Constancias')
                        except Exception:
                            p = ''
                        selected['path'] = p or ''
                        selected['done'] = True

                    # Llamar en thread principal
                    if self.winfo_exists():
                        self.after(0, _askdir)

                    # esperar hasta que el usuario seleccione o cancele (timeout razonable)
                    timeout = 60.0
                    waited = 0.0
                    while not selected['done'] and waited < timeout:
                        time.sleep(0.05)
                        waited += 0.05

                    out_base_pdf = selected['path'] or None

                    # Crear carpeta raíz timestamp una sola vez para todas las listas
                    carpeta_raiz = None
                    if out_base_pdf:
                        dt_now_root = datetime.now()
                        carpeta_raiz = os.path.join(out_base_pdf, f"Constancia_{dt_now_root.strftime('%Y%m%d_%H%M%S')}")
                        os.makedirs(carpeta_raiz, exist_ok=True)

                    for lista, filas_grp in grupos.items():
                        try:
                            # Actualizar progreso por lista (evitar bloquear UI)
                            try:
                                listas_procesadas += 1
                                pct = 0
                                if total_listas:
                                    pct = min(95, int((listas_procesadas / float(total_listas)) * 90))
                                else:
                                    pct = 10
                                # usar el callback seguro si existe
                                try:
                                    if self.winfo_exists():
                                        self.actualizar_progreso(pct, f"Generando constancias: lista {lista} ({listas_procesadas}/{total_listas})")
                                except Exception:
                                    pass
                            except Exception:
                                pass
                            solicitud_raw = ''
                            if filas_grp:
                                solicitud_raw = str(filas_grp[0].get('SOLICITUD') or '').strip()
                            sol_no = ''
                            sol_year = ''
                            if '/' in solicitud_raw:
                                parts = solicitud_raw.split('/')
                                sol_no = parts[0].strip()
                                sol_year = parts[1].strip() if len(parts) > 1 else ''
                            else:
                                sol_no = solicitud_raw

                            folio_vis = (self.entry_folio_visita.get().strip() if hasattr(self, 'entry_folio_visita') else '') or str(filas_grp[0].get('FOLIO',''))
                            fecha_em = (self.entry_fecha_termino.get().strip() if hasattr(self, 'entry_fecha_termino') else '') or datetime.now().strftime('%d/%m/%Y')
                            cliente_name = self.cliente_seleccionado['CLIENTE'] if getattr(self, 'cliente_seleccionado', None) else ''
                            rfc = (self.cliente_seleccionado.get('RFC','') if getattr(self, 'cliente_seleccionado', None) else '')
                            no_contrato = (self.cliente_seleccionado.get('NÚMERO_DE_CONTRATO','') if getattr(self, 'cliente_seleccionado', None) else '')

                            # Resolver norma y nombre de la norma consultando data/Normas.json cuando sea posible
                            norma_raw = filas_grp[0].get('NORMA','') or filas_grp[0].get('NORMA UVA','') or filas_grp[0].get('NORMA_UVA','') or ''
                            norma_val = str(norma_raw).strip()
                            nombre_norma = ''
                            try:
                                normas_path = os.path.join(DATA_DIR, 'Normas.json')
                                if os.path.exists(normas_path):
                                    with open(normas_path, 'r', encoding='utf-8') as nf:
                                        normas_raw = json.load(nf) or []
                                        # soportar dos formatos: lista de objetos [{"NOM":..., "NOMBRE":...}] o dict {"NOM-...": {...}}
                                        if isinstance(normas_raw, list):
                                            for entry in normas_raw:
                                                try:
                                                    nom_code = (entry.get('NOM') or entry.get('NOMINA') or '')
                                                    nom_name = entry.get('NOMBRE') or entry.get('NOMBRE DE NORMA') or ''
                                                    if not nom_code:
                                                        continue
                                                    if norma_val and norma_val.isdigit():
                                                        if norma_val.zfill(3) in nom_code or nom_code in norma_val:
                                                            norma_val = nom_code
                                                            nombre_norma = nom_name
                                                            break
                                                    else:
                                                        if norma_val and (norma_val in nom_code or nom_code in norma_val):
                                                            norma_val = nom_code
                                                            nombre_norma = nom_name
                                                            break
                                                except Exception:
                                                    # continuar con la siguiente entrada si hay datos inesperados
                                                    continue
                                            # fallback: si no se encontró y norma_val coincide exactamente con algún NOM
                                            if not nombre_norma:
                                                for entry in normas_raw:
                                                    nom_code = entry.get('NOM') or ''
                                                    if nom_code == norma_val:
                                                        nombre_norma = entry.get('NOMBRE') or ''
                                                        break
                                        elif isinstance(normas_raw, dict):
                                            normas_map = normas_raw
                                            if norma_val and norma_val.isdigit():
                                                search = norma_val.zfill(3)
                                                for k, v in normas_map.items():
                                                    if search in k or k in norma_val:
                                                        if isinstance(v, dict):
                                                            nombre_norma = v.get('NOMBRE', '') or v.get('NOMBRE DE NORMA', '') or ''
                                                        else:
                                                            nombre_norma = str(v)
                                                        norma_val = k
                                                        break
                                            else:
                                                for k, v in normas_map.items():
                                                    if norma_val and (norma_val in k or k in norma_val):
                                                        if isinstance(v, dict):
                                                            nombre_norma = v.get('NOMBRE', '') or v.get('NOMBRE DE NORMA', '') or ''
                                                        else:
                                                            nombre_norma = str(v)
                                                        norma_val = k
                                                        break
                                            if not nombre_norma and norma_val in normas_map:
                                                v = normas_map.get(norma_val)
                                                if isinstance(v, dict):
                                                    nombre_norma = v.get('NOMBRE', '') or v.get('NOMBRE DE NORMA', '') or ''
                                                else:
                                                    nombre_norma = str(v)
                            except Exception:
                                pass

                            datos_const = {
                                'folio_constancia': folio_vis,
                                'fecha_emision': fecha_em,
                                'cliente': cliente_name,
                                'rfc': rfc,
                                'no_contrato': no_contrato,
                                'fecha_contrato': '',
                                'norma': norma_val,
                                'normades': nombre_norma,
                                'nombre_norma': nombre_norma,
                                'producto': filas_grp[0].get('DESCRIPCION',''),
                                'marca': filas_grp[0].get('MARCA',''),
                                'modelo': filas_grp[0].get('MODELO',''),
                                'tabla_relacion': filas_grp,
                                'lista': lista
                            }

                            folio_num = str(filas_grp[0].get('FOLIO','') or '').strip()
                            folio_for_name = folio_num

                            # Crear una carpeta raíz `Constancias` dentro de la carpeta elegida por el usuario
                            nombre_pdf = f"Constancia_Lista_{lista}_{folio_for_name}_{sol_no}_{sol_year}.pdf"
                            nombre_pdf = limpiar_nombre(nombre_pdf)

                            ruta_pdf = None
                            if out_base_pdf and carpeta_raiz:
                                # Crear subcarpeta por solicitud (usar solicitud_raw saneado)
                                sol_folder_name = solicitud_raw or f"Solicitud_{lista}"
                                try:
                                    sol_folder_name = limpiar_nombre(sol_folder_name)
                                except Exception:
                                    sol_folder_name = f"Solicitud_{lista}"
                                carpeta_solicitud = os.path.join(carpeta_raiz, sol_folder_name)
                                os.makedirs(carpeta_solicitud, exist_ok=True)
                                ruta_pdf = os.path.join(carpeta_solicitud, nombre_pdf)

                            pdf_ok = False
                            gen_exception = None
                            # Persistir folios para la visita ANTES de generar el PDF
                            try:
                                if filas_grp:
                                    try:
                                        self.guardar_folios_visita(folio_vis, filas_grp, persist_counter=True)
                                    except Exception:
                                        # no interrumpir generación si la persistencia falla
                                        pass
                            except Exception:
                                pass
                            try:
                                if ruta_pdf and ConstGen:
                                    gen_inst = ConstGen(datos_const, base_dir=BASE_DIR)
                                    ruta_generada = gen_inst.generar(ruta_pdf)
                                    pdf_ok = True if ruta_generada and os.path.exists(ruta_generada) else False
                                elif ruta_pdf and mod is not None:
                                    try:
                                        ruta_generada = getattr(mod, 'generar_constancia_desde_visita')(folio_vis, salida=ruta_pdf)
                                        pdf_ok = True if ruta_generada and os.path.exists(ruta_generada) else False
                                    except Exception as e:
                                        pdf_ok = False
                                        import traceback
                                        gen_exception = traceback.format_exc()
                                else:
                                    # Usuario canceló selección de carpeta para PDFs -> no generar PDF
                                    pdf_ok = False
                            except Exception as e:
                                pdf_ok = False
                                import traceback
                                gen_exception = traceback.format_exc()

                            try:
                                json_name = f"Constancia_Lista_{lista}_{folio_for_name}_{sol_no}_{sol_year}.json"
                                json_name = limpiar_nombre(json_name)
                                json_path = os.path.join(out_base_json, json_name)
                                try:
                                    if mod is not None and hasattr(mod, 'convertir_constancia_a_json'):
                                        json_data = getattr(mod, 'convertir_constancia_a_json')(datos_const)
                                    else:
                                        # convertir de forma básica local
                                        def _basic_convert(d):
                                            return d
                                        json_data = _basic_convert(datos_const)
                                except Exception:
                                    json_data = datos_const
                                with open(json_path, 'w', encoding='utf-8') as jf:
                                    json.dump(json_data, jf, ensure_ascii=False, indent=2)
                            except Exception:
                                pass

                            if pdf_ok:
                                created.append(ruta_pdf)
                                try:
                                    try:
                                        if filas_grp:
                                            self.guardar_folios_visita(folio_vis, filas_grp, persist_counter=True)
                                    except Exception:
                                        pass

                                    try:
                                        tabla_relacion_path = os.path.join(DATA_DIR, 'tabla_de_relacion.json')
                                        if os.path.exists(tabla_relacion_path):
                                            backup_dir = os.path.join(DATA_DIR, 'tabla_relacion_backups')
                                            os.makedirs(backup_dir, exist_ok=True)
                                            ts = datetime.now().strftime('%Y%m%d%H%M%S')
                                            dest_name = f"tabla_de_relacion_{folio_vis}_PERSIST_{ts}.json"
                                            try:
                                                shutil.copyfile(tabla_relacion_path, os.path.join(backup_dir, dest_name))
                                                print(f"📦 Respaldo persistente creado para constancia: {dest_name}")
                                            except Exception as e:
                                                print(f"⚠️ No se pudo crear respaldo persistente de tabla_de_relacion para constancia: {e}")
                                    except Exception:
                                        pass
                                except Exception:
                                    pass
                            else:
                                # sólo registrar error si intentó generar y falló
                                if ruta_pdf:
                                    errores.append(lista)
                                    # registrar traza detallada en el log de constancias
                                    try:
                                        import traceback
                                        dbg_lines = []
                                        dbg_lines.append(f"[{datetime.now().isoformat()}] GENERACION FAILED: lista={lista} ruta_pdf={ruta_pdf} folio_vis={folio_vis}")
                                        if gen_exception:
                                            dbg_lines.append("-- exception trace --")
                                            dbg_lines.append(gen_exception)
                                        else:
                                            dbg_lines.append("-- no exception trace captured --")
                                        log_path = os.path.join(DATA_DIR, 'constancia_debug.log')
                                        os.makedirs(DATA_DIR, exist_ok=True)
                                        with open(log_path, 'a', encoding='utf-8') as lf:
                                            for L in dbg_lines:
                                                lf.write(L + '\n')
                                            lf.write('\n')
                                    except Exception:
                                        pass
                            # Actualizar progreso tras procesar la lista
                            try:
                                if self.winfo_exists():
                                    # Incrementar un poco el progreso localmente
                                    pct2 = min(98, int((listas_procesadas / float(total_listas)) * 95) if total_listas else 10)
                                    try:
                                        self.actualizar_progreso(pct2, f"Procesadas {listas_procesadas}/{total_listas} listas")
                                    except Exception:
                                        pass
                            except Exception:
                                pass
                        except Exception as e:
                            errores.append((lista, str(e)))

                except Exception as e:
                    created = []
                    errores.append(str(e))

                # Preparar resultado y notificar
                try:
                    mensaje_final = f"✅ Se generaron {len(created)} constancias (PDF).\n"
                    if out_base_pdf:
                        mensaje_final += f"\n📁 Ubicación PDFs: {out_base_pdf}"
                    else:
                        mensaje_final += "\nℹ️ No se guardaron PDFs (usuario no seleccionó carpeta)."

                    mensaje_final += f"\n\n📁 JSONs guardados en: {out_base_json}"

                    if errores:
                        mensaje_final += f"\n\n⚠️ Errores en listas: {errores}"

                    if self.winfo_exists():
                        self.after(0, lambda: messagebox.showinfo("Generación Completada", mensaje_final) if self.winfo_exists() else None)

                    resultado = {
                        'directorio': out_base_pdf or out_base_json,
                        'total_generados': len(created),
                        'total_familias': len(grupos),
                        'dictamenes_fallidos': len(errores),
                        'folios_utilizados': getattr(self, 'folios_utilizados_actual', [])
                    }
                    try:
                        resultado['folios_utilizados_info'] = getattr(self, 'folios_utilizados_actual', [])
                        self.registrar_visita_automatica(resultado)
                        # Forzar recarga y refresco del historial en el hilo principal
                        try:
                            def _refresh_hist_ui():
                                try:
                                    self._cargar_historial()
                                except Exception:
                                    pass
                                try:
                                    self._poblar_historial_ui()
                                except Exception:
                                    pass
                            if self.winfo_exists():
                                self.after(150, _refresh_hist_ui)
                        except Exception:
                            pass
                        # Marcar progreso final
                        try:
                            if self.winfo_exists():
                                self.after(200, lambda: self.actualizar_progreso(100, "Completado") if self.winfo_exists() else None)
                        except Exception:
                            pass
                    except Exception:
                        pass

                    # Abrir carpeta seleccionada por el usuario (si existe)
                    try:
                        if out_base_pdf and os.path.exists(out_base_pdf) and self.winfo_exists():
                            self.after(1000, lambda: self._abrir_carpeta(out_base_pdf) if self.winfo_exists() else None)
                    except Exception:
                        pass
                except Exception:
                    pass
                return

            # flujo por defecto: dictámenes
            sys.path.append(BASE_DIR)
            from generador_dictamen import generar_dictamenes_gui
            
            def actualizar_progreso(porcentaje, mensaje):
                # VERIFICACIÓN EN CALLBACK
                if self.winfo_exists():
                    self.actualizar_progreso(porcentaje, mensaje)
            
            def finalizado(exito, mensaje, resultado):
                # VERIFICACIÓN EN CALLBACK FINAL
                if not self.winfo_exists():
                    return
                    
                if exito and resultado:
                    directorio = resultado['directorio']
                    total_gen = resultado['total_generados']
                    total_fam = resultado['total_familias']
                    
                    dictamenes_fallidos = resultado.get('dictamenes_fallidos', 0)
                    folios_fallidos = resultado.get('folios_fallidos', [])
                    folios_utilizados = resultado.get('folios_utilizados', "No disponible")
                    
                    archivos_existentes = []
                    if os.path.exists(directorio):
                        archivos_existentes = [f for f in os.listdir(directorio) if f.endswith('.pdf')]
                    
                    mensaje_final = f"✅ {mensaje}\n\n📁 Ubicación: {directorio}"
                    
                    if archivos_existentes:
                        mensaje_final += f"\n📄 Archivos creados: {len(archivos_existentes)}"
                    
                    
                    if dictamenes_fallidos > 0:
                        mensaje_final += f"\n❌ Dictámenes no generados: {dictamenes_fallidos}"
                        if folios_fallidos:
                            mensaje_final += f"\n📋 Folios fallidos: {', '.join(map(str, folios_fallidos))}"
                    
                    # VERIFICAR ANTES DE MOSTRAR MESSAGEBOX
                    if self.winfo_exists():
                        self.after(0, lambda: messagebox.showinfo("Generación Completada", mensaje_final) if self.winfo_exists() else None)
                        
                        resultado['folios_utilizados_info'] = folios_utilizados
                        self.registrar_visita_automatica(resultado)

                        # Si se generó usando un folio reservado seleccionado, marcarlo como completado
                        try:
                            sel_id = getattr(self, 'selected_pending_id', None)
                            if sel_id:
                                try:
                                    self.hist_update_visita(sel_id, {'estatus': 'Completada'})
                                except Exception:
                                    # fallback: buscar y modificar manualmente
                                    for v in self.historial.get('visitas', []):
                                        if v.get('_id') == sel_id or v.get('id') == sel_id:
                                            v['estatus'] = 'Completada'
                                    try:
                                        self._guardar_historial()
                                    except Exception:
                                        pass

                                # Eliminar de archivo de reservas
                                try:
                                    pf = os.path.join(DATA_DIR, 'pending_folios.json')
                                    if os.path.exists(pf):
                                        with open(pf, 'r', encoding='utf-8') as f:
                                            arr = json.load(f) or []
                                        # Eliminar por _id / id si coincide con sel_id
                                        try:
                                            arr = [p for p in arr if ((p.get('_id') or p.get('id')) != sel_id)]
                                        except Exception:
                                            arr = [p for p in arr if p.get('folio_visita') != (getattr(self, 'entry_folio_visita', None).get() if hasattr(self, 'entry_folio_visita') else None)]
                                        with open(pf, 'w', encoding='utf-8') as f:
                                  
                                            json.dump(arr, f, ensure_ascii=False, indent=2)
                                        self.pending_folios = arr
                                except Exception:
                                    pass
                                # limpiar selección
                                try:
                                    self.selected_pending_id = None
                                    self.usando_folio_reservado = False
                                except Exception:
                                    pass
                        except Exception:
                            pass
                        
                        if archivos_existentes and self.winfo_exists():
                            self.after(1000, lambda: self._abrir_carpeta(directorio) if self.winfo_exists() else None)
                    
                else:
                    if self.winfo_exists():
                        self.after(0, lambda: self.mostrar_error(mensaje) if self.winfo_exists() else None)
            
            # LLAMADA CORREGIDA - sin folios_info
            generar_dictamenes_gui(
                cliente_manual=self.cliente_seleccionado['CLIENTE'],
                rfc_manual=self.cliente_seleccionado.get('RFC', ''),
                callback_progreso=actualizar_progreso,
                callback_finalizado=finalizado
            )
            
        except Exception as e:
            error_msg = f"Error iniciando generador: {str(e)}"
            if self.winfo_exists():
                self.after(0, lambda: self.mostrar_error(error_msg) if self.winfo_exists() else None)
        finally:
            if self.winfo_exists():
                self.after(0, self._finalizar_generacion)

    def _abrir_carpeta(self, directorio):
        try:
            if os.path.exists(directorio):
                if os.name == 'nt':
                    os.startfile(directorio)
                elif os.name == 'posix':
                    os.system(f'open "{directorio}"' if sys.platform == 'darwin' else f'xdg-open "{directorio}"')
        except Exception as e:
            print(f"Error abriendo carpeta: {e}")

    def actualizar_progreso(self, porcentaje, mensaje):
        def _actualizar():
            if self.winfo_exists():  # Verificar si la ventana aún existe
                # Asegurar porcentaje entre 0 y 100
                try:
                    pct = float(porcentaje)
                except Exception:
                    pct = 0.0
                pct = max(0.0, min(100.0, pct))
                self.barra_progreso.set(pct / 100.0)
                # Mostrar porcentaje y mensaje breve
                if mensaje:
                    self.etiqueta_progreso.configure(text=f"{int(pct)}% - {mensaje}")
                else:
                    self.etiqueta_progreso.configure(text=f"{int(pct)}%")
                self.update_idletasks()
                # Registrar marca de tiempo y valor para el watcher
                try:
                    self._last_progress_value = float(pct)
                    self._last_progress_ts = time.time()
                except Exception:
                    pass
        
        self.after(0, _actualizar)

    def _start_progress_watcher(self):
        """Inicia un hilo que anima ligeramente la barra de progreso mientras
        `self.generando_dictamenes` es True y el generador no actualiza con frecuencia.
        """
        try:
            # Si ya existe, no crear otro
            if getattr(self, '_progress_watcher_thread', None) and self._progress_watcher_thread.is_alive():
                return
            stop_event = threading.Event()
            self._progress_watcher_stop_event = stop_event

            def _watcher():
                try:
                    while not stop_event.is_set() and getattr(self, 'generando_dictamenes', False):
                        try:
                            now = time.time()
                            last_ts = getattr(self, '_last_progress_ts', 0) or 0
                            last_val = getattr(self, '_last_progress_value', 0.0) or 0.0
                            # Si no hubo actualización en 1.2s, animar un pequeño paso
                            if now - last_ts > 1.2 and last_val < 95:
                                # avanzar un paso pequeño entre 1-4%
                                step = 1 + int((now - last_ts) % 4)
                                new_val = min(95.0, last_val + step)
                                try:
                                    if self.winfo_exists():
                                        self.after(0, lambda nv=new_val: self.actualizar_progreso(nv, "Procesando...") )
                                except Exception:
                                    pass
                        except Exception:
                            pass
                        time.sleep(0.8)
                except Exception:
                    pass

            t = threading.Thread(target=_watcher, daemon=True)
            self._progress_watcher_thread = t
            t.start()
        except Exception:
            pass

    def _stop_progress_watcher(self):
        try:
            ev = getattr(self, '_progress_watcher_stop_event', None)
            if ev:
                ev.set()
            th = getattr(self, '_progress_watcher_thread', None)
            try:
                if th and th.is_alive():
                    # give it a moment to exit
                    th.join(timeout=0.5)
            except Exception:
                pass
        except Exception:
            pass

    def actualizar_tipo_documento(self, valor=None):
        """Actualiza la UI del panel Generador según el tipo de documento seleccionado."""
        try:
            # Obtener selección (si valor pasado por callback, usarlo)
            seleccionado = valor if valor else (self.combo_tipo_documento.get() if hasattr(self, 'combo_tipo_documento') else 'Dictamen')
            seleccionado = seleccionado.strip()

            # Actualizar título del panel de generación
            title_map = {
                'Dictamen': 'Generar Dictámenes',
                'Negación de Dictamen': 'Generar Negación de Dictamen',
                'Constancia': 'Generar Constancias',
                'Negación de Constancia': 'Generar Negación de Constancia'
            }
            nuevo_titulo = title_map.get(seleccionado, 'Generador')
            if hasattr(self, 'generacion_title'):
                self.generacion_title.configure(text=f"🚀 {nuevo_titulo}")

            # Cambiar texto del botón principal de generación
            if hasattr(self, 'boton_generar_dictamen'):
                self.boton_generar_dictamen.configure(text=f"{nuevo_titulo}")

            if hasattr(self, 'boton_guardar_folio'):
                self.boton_guardar_folio.configure(state='normal')

            # Buscar en historial si hay folio pendiente para este tipo
            pendiente_msg = ''
            try:
                pendientes = [r for r in getattr(self, 'historial_data', []) if (r.get('tipo_documento') or '').strip() == seleccionado and r.get('estatus','').lower() in ('en proceso','pendiente')]
                if pendientes:
                    first = pendientes[0]
                    pendiente_msg = f"Folio pendiente: {first.get('folio_visita','-')} / {first.get('folio_acta','-')} (puede usarlo al iniciar una {seleccionado})"
            except Exception:
                pendiente_msg = ''

            if hasattr(self, 'info_folio_pendiente'):
                self.info_folio_pendiente.configure(text=pendiente_msg)

            # Ajustes visuales adicionales: si no hay cliente seleccionado, deshabilitar generación
            if not getattr(self, 'cliente_seleccionado', None):
                if hasattr(self, 'boton_generar_dictamen'):
                    self.boton_generar_dictamen.configure(state='disabled')
            else:
                if hasattr(self, 'boton_generar_dictamen'):
                    self.boton_generar_dictamen.configure(state='normal')

            # Ocultar / deshabilitar el campo Folio Acta para las Constancias (no requerido)
            try:
                folio_widget = getattr(self, 'entry_folio_acta', None)
                if folio_widget is not None:
                    parent = getattr(folio_widget, 'master', None)
                    is_constancia = 'constancia' in seleccionado.lower()
                    if is_constancia:
                        # borrar valor y ocultar
                        try:
                            folio_widget.delete(0, 'end')
                        except Exception:
                            pass
                        try:
                            folio_widget.configure(state='disabled')
                        except Exception:
                            pass
                        try:
                            if parent and getattr(parent, 'winfo_ismapped', lambda: False)():
                                parent.pack_forget()
                        except Exception:
                            pass
                    else:
                        # volver a mostrar y activar
                        try:
                            if parent and not getattr(parent, 'winfo_ismapped', lambda: False)():
                                parent.pack(fill='x', pady=(0, 10))
                        except Exception:
                            pass
                        try:
                            folio_widget.configure(state='normal')
                        except Exception:
                            pass
                        # si está vacío, rellenar con valor por defecto
                        try:
                            if folio_widget.get().strip() == '':
                                folio_widget.delete(0, 'end')
                                folio_widget.insert(0, f"AC{self.current_folio}")
                        except Exception:
                            pass
            except Exception:
                pass
        except Exception as e:
            # No bloquear la aplicación por errores en esta actualización
            print(f"Error actualizando tipo de documento UI: {e}")

        

        # Refrescar lista de folios pendientes en el combobox (si existe)
        try:
            if hasattr(self, '_refresh_pending_folios_dropdown'):
                self._refresh_pending_folios_dropdown()
            elif hasattr(self, 'combo_folios_pendientes'):
                self._refresh_pending_folios_dropdown()
        except Exception:
            pass

    def guardar_folio_historial(self):
        """Guarda el folio actual en el historial como registro incompleto para retomarlo después."""
        try:
            tipo_documento = (self.combo_tipo_documento.get().strip() if hasattr(self, 'combo_tipo_documento') else 'Dictamen')
            folio_visita = (self.entry_folio_visita.get().strip() if hasattr(self, 'entry_folio_visita') else '')
            folio_acta = (self.entry_folio_acta.get().strip() if hasattr(self, 'entry_folio_acta') else '')

            if not folio_visita:
                messagebox.showwarning("Folio requerido", "No hay folio de visita disponible para guardar.")
                return

            # Requerir cliente y domicilio para guardar/registrar visita
            if not getattr(self, 'cliente_seleccionado', None):
                messagebox.showwarning("Cliente requerido", "Por favor seleccione un cliente antes de guardar la visita.")
                return
            if not getattr(self, 'domicilio_seleccionado', None):
                messagebox.showwarning("Domicilio requerido", "Por favor seleccione un domicilio para el cliente antes de guardar la visita.")
                return

            # Garantizar que si no hay fecha/hora/cliente en el formulario, se registre la marca temporal y el cliente actual
            fecha_inicio_val = (self.entry_fecha_inicio.get().strip() if hasattr(self, 'entry_fecha_inicio') else '')
            hora_inicio_val = (self.entry_hora_inicio.get().strip() if hasattr(self, 'entry_hora_inicio') else '')
            if not fecha_inicio_val:
                fecha_inicio_val = datetime.now().strftime("%d/%m/%Y")
            if not hora_inicio_val:
                hora_inicio_val = datetime.now().strftime("%H:%M")

            cliente_val = ""
            try:
                # soportar dict con clave 'CLIENTE' o 'cliente'
                if getattr(self, 'cliente_seleccionado', None):
                    cliente_val = self.cliente_seleccionado.get('CLIENTE') or self.cliente_seleccionado.get('cliente') or str(self.cliente_seleccionado)
            except Exception:
                cliente_val = ""
            # Determinar los folios reales usados para generación (tomados de la tabla cargada)
            folios_utilizados_val = ""
            try:
                info = getattr(self, 'info_folios_actual', None)
                if info:
                    if info.get('rango_folios'):
                        folios_utilizados_val = info.get('rango_folios')
                    elif info.get('lista_folios'):
                        # unir una lista corta para mostrar
                        lf = info.get('lista_folios')
                        if isinstance(lf, (list, tuple)) and lf:
                            folios_utilizados_val = ','.join(lf[:20])
                        else:
                            folios_utilizados_val = str(lf)
            except Exception:
                folios_utilizados_val = ""

            # Normalizar tipo de documento a valores esperados
            def _normalizar_td(raw):
                if not raw:
                    return 'Dictamen'
                s = str(raw).strip()
                low = s.lower()
                if 'dictamen' in low and ('neg' in low or 'negación' in low or 'negacion' in low):
                    return 'Negación de Dictamen'
                if 'dictamen' in low:
                    return 'Dictamen'
                if 'constancia' in low and ('neg' in low or 'negación' in low or 'negacion' in low):
                    return 'Negación de Constancia'
                if 'constancia' in low:
                    return 'Constancia'
                return s

            tipo_documento_norm = _normalizar_td(tipo_documento)

            payload = {
                "folio_visita": folio_visita,
                "folio_acta": folio_acta,
                "fecha_inicio": fecha_inicio_val,
                "fecha_termino": self.entry_fecha_termino.get().strip() if hasattr(self, 'entry_fecha_termino') and self.entry_fecha_termino.get().strip() else "",
                "hora_inicio": hora_inicio_val,
                "hora_termino": self.entry_hora_termino.get().strip() if hasattr(self, 'entry_hora_termino') and self.entry_hora_termino.get().strip() else "",
                "norma": "",
                "cliente": cliente_val,
                "nfirma1": "",
                "nfirma2": "",
                "estatus": "Pendiente",
                "tipo_documento": tipo_documento_norm,
                "folios_utilizados": folios_utilizados_val,
                # Dirección seleccionada (si existe)
                "direccion": getattr(self, 'direccion_seleccionada', '') or getattr(self, 'domicilio_seleccionado', ''),
                "colonia": getattr(self, 'colonia_seleccionada', ''),
                "municipio": getattr(self, 'municipio_seleccionado', ''),
                "ciudad_estado": getattr(self, 'ciudad_seleccionada', ''),
                "cp": getattr(self, 'cp_seleccionado', '')
            }

            # Log breve: indicar folio guardado (evitar volcar todo el payload)
            try:
                print(f"[INFO] Visita pendiente guardada: {payload.get('folio_visita','-')} tipo={payload.get('tipo_documento','-')}")
            except Exception:
                try:
                    print(f"[INFO] Visita pendiente guardada")
                except Exception:
                    pass

            # Guardar usando la función existente (suprimir notificación interna
            # porque este llamador mostrará su propio messagebox)
            self.hist_create_visita(payload, show_notification=False)
            try:
                # actualizar indicador visual del siguiente folio
                self._update_siguiente_folio_label()
            except Exception:
                pass
            # Persistir también en archivo de reservas (pending_folios.json)
            try:
                pf_path = os.path.join(DATA_DIR, 'pending_folios.json')
                arr = []
                if os.path.exists(pf_path):
                    try:
                        with open(pf_path, 'r', encoding='utf-8') as f:
                            arr = json.load(f) or []
                    except Exception:
                        arr = []
                # evitar duplicados por folio_visita
                if not any(p.get('folio_visita') == payload.get('folio_visita') for p in arr):
                    arr.append(payload)
                    with open(pf_path, 'w', encoding='utf-8') as f:
                        json.dump(arr, f, ensure_ascii=False, indent=2)
                    self.pending_folios = arr
                    # Forzar refresco inmediato de la UI de folios pendientes
                    try:
                        if hasattr(self, '_refresh_pending_folios_dropdown'):
                            self._refresh_pending_folios_dropdown()
                    except Exception:
                        pass
                    try:
                        if hasattr(self, 'cliente_folios_frame') and not self.cliente_folios_frame.winfo_ismapped():
                            self.cliente_folios_frame.pack(fill="x", pady=(4,4))
                    except Exception:
                        pass
            except Exception as e:
                print(f"[WARN] No se pudo persistir reserva en pending_folios.json: {e}")
            # DEBUG: leer historial inmediatamente y confirmar último registro
            try:
                self._cargar_historial()
                print(f"[DEBUG] after guardar_folio_historial -> total historial: {len(self.historial_data)}")
                if self.historial_data:
                    print(f"[DEBUG] ultimo registro: {self.historial_data[-1]}")
            except Exception:
                pass
            messagebox.showinfo("Folio guardado", f"El folio {folio_visita} ha sido guardado como {tipo_documento} pendiente.")
            # Preparar siguiente folio
            try:
                self.crear_nueva_visita()
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar el folio: {e}")
        finally:
            # Refrescar lista de folios pendientes
            try:
                # Programar el refresco para que ocurra después de que `hist_create_visita`
                # haya tenido oportunidad de aplicar la visita en el hilo principal.
                try:
                    self.after(250, self._refresh_pending_folios_dropdown)
                except Exception:
                    # Fallback síncrono si after no está disponible
                    self._refresh_pending_folios_dropdown()
            except Exception:
                pass

    def reservar_folios_tabla(self):
        """Reserva un bloque de folios y sobrescribe la columna 'FOLIO' en
        `data/tabla_de_relacion.json` asignando un folio por familia (LISTA).
        """
        try:
            try:
                self.guardar_folio_historial()
            except Exception as e:
                print(f"[ERROR] reservar_folios_tabla -> guardar_folio_historial: {e}")
        except Exception as e:
            # No interrumpimos la UX: mostrar un error genérico si ocurre
            try:
                messagebox.showerror("Error", f"Error guardando visita: {e}")
            except Exception:
                print(f"Error guardando visita: {e}")

    def _finalizar_generacion(self):
        if self.winfo_exists():  # Verificar si la ventana aún existe
            self.generando_dictamenes = False
            self.boton_generar_dictamen.configure(state="normal")
        # Detener watcher si existe
        try:
            self._stop_progress_watcher()
        except Exception:
            pass
        # Asegurar que la barra muestre completado cuando terminemos
        try:
            if self.winfo_exists():
                self.after(50, lambda: self.actualizar_progreso(100, "Completado"))
        except Exception:
            pass

    def mostrar_error(self, mensaje):
        if self.winfo_exists():  # Verificar si la ventana aún existe
            self.etiqueta_estado.configure(
                text="❌ Error en el proceso", 
                text_color=STYLE["peligro"]
            )
            self.check_label.configure(text="")
            messagebox.showerror("Error", mensaje)

    # -----------------------------------------------------------
    # MÉTODOS DEL HISTORIAL
    # -----------------------------------------------------------
    def _cargar_historial(self):
        """Carga los datos del historial desde el archivo JSON con validación"""
        try:
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(self.historial_path), exist_ok=True)
            
            if os.path.exists(self.historial_path):
                with open(self.historial_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    # Extraer solo las visitas
                    self.historial_data = data.get("visitas", [])
                    self.historial = data  # CARGAR EL DICCIONARIO COMPLETO
                    
                    # Validar que los datos sean consistentes
                    if not isinstance(self.historial_data, list):
                        self.historial_data = []
                    
                    # Log de carga exitosa
                    print(f"✅ Historial cargado: {len(self.historial_data)} registros desde {self.historial_path}")
            else:
                self.historial_data = []
                self.historial = {"visitas": []}
                print(f"📝 Archivo de historial no existe, se creará uno nuevo")
                
            # Inicializar también historial_data_original
            self.historial_data_original = self.historial_data.copy()
            # Normalizar campo 'tipo_documento' en los registros existentes
            try:
                cambios = False
                for reg in self.historial_data:
                    if not isinstance(reg, dict):
                        continue
                    td = reg.get('tipo_documento')
                    # Normalizar a las formas amigables: Dictamen, Negación de Dictamen, Constancia, Negación de Constancia
                    def _normalizar_td(raw):
                        if not raw:
                            return 'Dictamen'
                        s = str(raw).strip().lower()
                        if 'dictamen' in s and ('neg' in s or 'negación' in s or 'negacion' in s):
                            return 'Negación de Dictamen'
                        if 'dictamen' in s:
                            return 'Dictamen'
                        if 'constancia' in s and ('neg' in s or 'negación' in s or 'negacion' in s):
                            return 'Negación de Constancia'
                        if 'constancia' in s:
                            return 'Constancia'
                        # Default: capitalize first letter
                        return str(raw).strip()

                    nuevo_td = _normalizar_td(td)
                    if nuevo_td != td:
                        reg['tipo_documento'] = nuevo_td
                        cambios = True
                if cambios:
                    # Actualizar estructura y persistir cambios
                    self.historial['visitas'] = self.historial_data
                    self._guardar_historial()
            except Exception:
                pass
                
        except json.JSONDecodeError as e:
            print(f"❌ Error decodificando JSON: {e}")
            # Intentar recuperar del backup
            backup_path = self.historial_path + ".backup"
            if os.path.exists(backup_path):
                try:
                    print(f"🔄 Recuperando desde backup...")
                    with open(backup_path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        self.historial_data = data.get("visitas", [])
                        self.historial = data
                except Exception:
                    self.historial_data = []
                    self.historial = {"visitas": []}
            else:
                self.historial_data = []
                self.historial = {"visitas": []}
        except Exception as e:
            print(f"❌ Error cargando historial: {e}")
            self.historial_data = []
            self.historial_data_original = []
            self.historial = {"visitas": []}

    def _sincronizar_historial(self):
        """Sincroniza los datos en memoria con el archivo JSON para asegurar persistencia"""
        try:
            # Actualizar self.historial con los datos actuales de historial_data
            self.historial["visitas"] = self.historial_data
            
            # Guardar el archivo
            self._guardar_historial()
            
            # Actualizar original
            self.historial_data_original = self.historial_data.copy()
            # Regenerar cache exportable para Excel (persistente)
            try:
                self._generar_datos_exportable()
            except Exception:
                pass
            
            return True
        except Exception as e:
            print(f"❌ Error sincronizando historial: {e}")
            return False

    def hist_borrar_visita(self, id_):
        """Elimina una visita y recalcula el folio actual si es necesario"""
        if not messagebox.askyesno("Confirmar borrado", "¿Eliminar este registro?"):
            return
        
        try:
            # Encontrar la visita a borrar
            visita_a_borrar = next((v for v in self.historial.get("visitas",[]) if v["_id"] == id_), None)
            if not visita_a_borrar:
                messagebox.showerror("Error", "No se encontró la visita para borrar")
                return

            folio_borrado = visita_a_borrar.get("folio_visita", "")
            try:
                print(f"[DEBUG] borrar_visita id_={id_} visita={visita_a_borrar}")
                print(f"[DEBUG] borrar_visita folio_borrado={folio_borrado}")
            except Exception:
                pass
            
            # Borrar la visita
            self.historial["visitas"] = [v for v in self.historial.get("visitas",[]) if v["_id"] != id_]
            self._guardar_historial()
            self._poblar_historial_ui()

            # Leer previamente el archivo `data/folios_visitas/folios_<visita>.json`
            # para obtener la lista autoritativa de folios asignados a esta visita.
            try:
                assigned_nums = set()
                fv_file_early = os.path.join(self.folios_visita_path, f"folios_{folio_borrado}.json")
                if os.path.exists(fv_file_early):
                    try:
                        with open(fv_file_early, 'r', encoding='utf-8') as _ff_early:
                            _arr_early = json.load(_ff_early) or []
                        for _entry_early in _arr_early:
                            folv = _entry_early.get('FOLIOS') or _entry_early.get('FOLIO') or ''
                            digs = ''.join([c for c in str(folv) if c.isdigit()])
                            if digs:
                                assigned_nums.add(int(digs.lstrip('0') or '0'))
                    except Exception:
                        assigned_nums = set()
                    # Log early read result
                    try:
                        dbg_dir_early = DATA_DIR
                        os.makedirs(dbg_dir_early, exist_ok=True)
                        dbg_path_early = os.path.join(dbg_dir_early, 'hist_borrar_debug.log')
                        from datetime import datetime as _dt
                        with open(dbg_path_early, 'a', encoding='utf-8') as _dbg_early:
                            _dbg_early.write(f"[{_dt.now().isoformat()}] early_read visita={folio_borrado} file_exists={os.path.exists(fv_file_early)} assigned_early={sorted(list(assigned_nums))}\n")
                    except Exception:
                        pass
                    try:
                        print(f"[DEBUG] early_read visita={folio_borrado} file_exists={os.path.exists(fv_file_early)} assigned_early={sorted(list(assigned_nums))[:200]}")
                    except Exception:
                        pass
            except Exception:
                assigned_nums = set()

            # Recalcular el folio actual (buscar el siguiente disponible)
            self.cargar_ultimo_folio()
            try:
                self._update_siguiente_folio_label()
            except Exception:
                pass
            # Además de borrar el registro en el historial, eliminar archivos JSON
            # de dictámenes que correspondan a los folios asociados a esta visita.
            try:
                # Intentar cargar folios y solicitudes desde el archivo folios_{folio}.json (si existe)
                folios_a_eliminar = set()
                solicitudes_a_eliminar = set()
                try:
                    folios_visita_file = os.path.join(self.folios_visita_path, f"folios_{folio_borrado}.json")
                    if os.path.exists(folios_visita_file):
                        with open(folios_visita_file, 'r', encoding='utf-8') as ff:
                            fv = json.load(ff) or []
                            for entry in fv:
                                try:
                                    fval = str(entry.get('FOLIOS') or entry.get('FOLIO') or '').strip()
                                    if not fval:
                                        continue
                                    # Extraer todos los números y expandir rangos (ej. 849-857)
                                    import re
                                    # First, expand explicit ranges like 849-857
                                    ranges = re.findall(r"(\d{1,6})\s*-\s*(\d{1,6})", fval)
                                    for a,b in ranges:
                                        try:
                                            start = int(a.lstrip('0') or '0')
                                            end = int(b.lstrip('0') or '0')
                                            if start <= end:
                                                # limit expansion to reasonable size
                                                span = end - start + 1
                                                if span <= 2000:
                                                    for n in range(start, end+1):
                                                        folios_a_eliminar.add(str(n))
                                                        folios_a_eliminar.add(n)
                                                        folios_a_eliminar.add(str(n).zfill(6))
                                        except Exception:
                                            continue
                                    # Then add any standalone numbers
                                    nums = re.findall(r"\d{1,6}", fval)
                                    if nums:
                                        for num in nums:
                                            try:
                                                n = int(num.lstrip('0') or '0')
                                                folios_a_eliminar.add(str(n))
                                                folios_a_eliminar.add(n)
                                                folios_a_eliminar.add(str(n).zfill(6))
                                            except Exception:
                                                continue
                                    else:
                                        folios_a_eliminar.add(fval)
                                    # También recoger solicitud si está presente
                                    try:
                                        sol = entry.get('SOLICITUDES') or entry.get('SOLICITUD') or entry.get('SOLICITUDE') or ''
                                        sol = str(sol).strip()
                                        if sol:
                                            # extraer dígitos de solicitud también
                                            s_nums = re.findall(r"\d{1,8}", sol)
                                            if s_nums:
                                                for sd in s_nums:
                                                    try:
                                                        sni = sd.lstrip('0')
                                                        solicitudes_a_eliminar.add(sni)
                                                        solicitudes_a_eliminar.add(sd)
                                                        solicitudes_a_eliminar.add(sd.zfill(6))
                                                    except Exception:
                                                        continue
                                            else:
                                                solicitudes_a_eliminar.add(sol)
                                    except Exception:
                                        pass
                                except Exception:
                                    continue
                except Exception:
                    folios_a_eliminar = set()
                    solicitudes_a_eliminar = set()

                # Fallback: intentar extraer folios desde el campo del registro en historial
                # Fallback: intentar extraer folios y solicitudes desde el campo del registro en historial
                if not folios_a_eliminar or not solicitudes_a_eliminar:
                    raw = visita_a_borrar.get('folios_utilizados') or visita_a_borrar.get('folios') or visita_a_borrar.get('folios_usados') or ''
                    try:
                        import re
                        raw_str = str(raw)
                        # expand ranges like 849-857
                        ranges = re.findall(r"(\d{1,6})\s*-\s*(\d{1,6})", raw_str)
                        for a,b in ranges:
                            try:
                                start = int(a.lstrip('0') or '0')
                                end = int(b.lstrip('0') or '0')
                                if start <= end:
                                    span = end - start + 1
                                    if span <= 2000:
                                        for n in range(start, end+1):
                                            folios_a_eliminar.add(str(n))
                                            folios_a_eliminar.add(n)
                                            folios_a_eliminar.add(str(n).zfill(6))
                            except Exception:
                                continue
                        posibles = re.findall(r"\d{1,6}", raw_str)
                        for p in posibles:
                            try:
                                n = int(p.lstrip('0') or '0')
                                folios_a_eliminar.add(str(n))
                                folios_a_eliminar.add(p.zfill(6))
                                # algunas solicitudes pueden estar incluidas; añadir también como solicitud
                                solicitudes_a_eliminar.add(p)
                                solicitudes_a_eliminar.add(p.zfill(6))
                            except Exception:
                                continue
                        # también intentar extraer solicitudes formateadas (ej '004227')
                        posibles_sol = re.findall(r"\b\d{4,8}\b", str(raw))
                        for s in posibles_sol:
                            solicitudes_a_eliminar.add(s)
                            solicitudes_a_eliminar.add(s.lstrip('0'))
                            solicitudes_a_eliminar.add(s.zfill(6))
                    except Exception:
                        pass

                # Ruta a data/Dictamenes (compatible con exe y desarrollo)
                dicts_dir = os.path.join(DATA_DIR, 'Dictamenes')
                deleted_files = []
                # Siempre escanear `data/Dictamenes` para intentar localizar archivos
                # relacionados con la visita aunque no tengamos `folios_a_eliminar`.
                if os.path.exists(dicts_dir):
                    for fn in os.listdir(dicts_dir):
                        if not fn.lower().endswith('.json'):
                            continue
                        fp = os.path.join(dicts_dir, fn)
                        try:
                            with open(fp, 'r', encoding='utf-8') as jf:
                                d = json.load(jf)
                        except Exception:
                            continue
                        ident = d.get('identificacion') or {}
                        fol_file = str(ident.get('folio') or '').strip()
                        sol_file = str(ident.get('solicitud') or '').strip()
                        cadena = str(ident.get('cadena_identificacion') or '').strip()
                        # Normalizar folio del archivo a dígitos si es posible
                        fol_file_digits = ''.join([c for c in fol_file if c.isdigit()])
                        fol_file_norm = (fol_file_digits.lstrip('0') or '0') if fol_file_digits else fol_file
                        fol_file_z6 = fol_file_digits.zfill(6) if fol_file_digits else fol_file
                        candidates = {fol_file, fol_file_digits, fol_file_norm, fol_file_z6}

                        # Normalizar solicitud del archivo a dígitos sin ceros a la izquierda
                        sol_digits = ''.join([c for c in sol_file if c.isdigit()])
                        sol_norm = sol_digits.lstrip('0') if sol_digits else sol_file

                        # Preparar solicitudes a eliminar normalizadas (strip after '/')
                        normalized_solicitudes = set()
                        for s in solicitudes_a_eliminar:
                            try:
                                s_str = str(s).split('/')[0].strip()
                                s_digits = ''.join([c for c in s_str if c.isdigit()])
                                if not s_digits:
                                    continue
                                s_norm = s_digits.lstrip('0') or '0'
                                # evitar coincidencias con números muy cortos (ej. '1','25')
                                if len(s_norm) < 3:
                                    continue
                                normalized_solicitudes.add(s_norm)
                                normalized_solicitudes.add(s_norm.zfill(6))
                            except Exception:
                                continue

                        match = False
                        # Comprobar coincidencia por folio (comparación de formas normalizadas)
                        for f_candidate in folios_a_eliminar:
                            f_c = ''.join([c for c in str(f_candidate) if c.isdigit()]) or str(f_candidate)
                            f_c_norm = f_c.lstrip('0') if f_c else f_c
                            try:
                                if f_c_norm and (f_c_norm == fol_file_norm or f_c_norm == fol_file_digits or f_c_norm == fol_file_z6 or f_c_norm == fol_file):
                                    match = True
                                    break
                            except Exception:
                                continue

                        # Comprobar coincidencia por solicitud (exigir igualdad tras normalizar)
                        if not match and normalized_solicitudes:
                            import re
                            digit_groups = re.findall(r"\d+", cadena)
                            for s_norm in normalized_solicitudes:
                                # comparar con el campo solicitud del archivo
                                if sol_digits and (s_norm == sol_digits.lstrip('0') or s_norm == sol_digits or s_norm == sol_digits.zfill(6)):
                                    match = True
                                    break
                                # buscar en la cadena de identificacion grupos de dígitos y comparar normalizados
                                for dg in digit_groups:
                                    dg_norm = dg.lstrip('0')
                                    if dg_norm == s_norm or dg == s_norm or dg.zfill(6) == s_norm:
                                        match = True
                                        break
                                if match:
                                    break
                        if match:
                            # Si encontramos coincidencia, asegurar que el folio del archivo
                            # queda registrado en `folios_a_eliminar` para posible recuperación
                            try:
                                if fol_file_digits:
                                    n = int(fol_file_digits.lstrip('0') or '0')
                                    folios_a_eliminar.add(str(n))
                                    folios_a_eliminar.add(n)
                                    folios_a_eliminar.add(str(n).zfill(6))
                            except Exception:
                                pass
                            # Evitar borrar archivos claramente no relacionados (ej. nombres que contengan 'style')
                            if 'style' in fn.lower():
                                continue
                            try:
                                os.remove(fp)
                                deleted_files.append(fn)
                            except Exception:
                                continue
                        else:
                            # Si no hubo match pero tenemos el folio de visita, intentar
                            # emparejar por presencia de ese valor en el nombre o cadena.
                            try:
                                if folio_borrado:
                                    # Evitar confundirse con el identificador de visita (ej. 'CP000001')
                                    fb_digits = ''.join([c for c in str(folio_borrado) if c.isdigit()])
                                    fb_norm = fb_digits.lstrip('0') or '0'
                                    # Solo usar la heurística de subcadena numérica si tiene al menos 3 dígitos
                                    if fb_norm and len(fb_norm) >= 3 and (fb_digits in fn or fb_digits in cadena):
                                        # registrar este folio también (solo si el folio del archivo parece un folio de documento)
                                        try:
                                            if fol_file_digits and len(fol_file_digits.lstrip('0') or '') >= 3:
                                                n = int(fol_file_digits.lstrip('0') or '0')
                                                folios_a_eliminar.add(str(n))
                                                folios_a_eliminar.add(n)
                                                folios_a_eliminar.add(str(n).zfill(6))
                                        except Exception:
                                            pass
                                        try:
                                            os.remove(fp)
                                            deleted_files.append(fn)
                                        except Exception:
                                            pass
                                        # mark match to avoid double-handling
                                        match = True
                                    else:
                                        # Si fb_digits es corto (p.ej. '1' de CP000001), buscar el prefijo completo
                                        try:
                                            if isinstance(folio_borrado, str) and folio_borrado.lower() in fn.lower():
                                                if fol_file_digits and len(fol_file_digits.lstrip('0') or '') >= 3:
                                                    try:
                                                        n = int(fol_file_digits.lstrip('0') or '0')
                                                        folios_a_eliminar.add(str(n))
                                                        folios_a_eliminar.add(n)
                                                        folios_a_eliminar.add(str(n).zfill(6))
                                                    except Exception:
                                                        pass
                                                try:
                                                    os.remove(fp)
                                                    deleted_files.append(fn)
                                                except Exception:
                                                    pass
                                                match = True
                                        except Exception:
                                            pass
                            except Exception:
                                pass

                # NOTA: no eliminar aquí el archivo folios_{folio}.json; se eliminará
                # después de intentar recalcular y persistir el contador para
                # que podamos usar su contenido como fuente autoritativa.

                # Intentar devolver folios al contador: calcular el máximo folio
                # existente tras el borrado y ajustar el contador a ese valor
                try:
                    import re
                    try:
                        import folio_manager
                    except Exception:
                        folio_manager = None

                    existing = set()
                    # Escanear data/Dictamenes
                    try:
                        dicts_dir = os.path.join(DATA_DIR, 'Dictamenes')
                        if os.path.exists(dicts_dir):
                            for fn in os.listdir(dicts_dir):
                                if not fn.lower().endswith('.json'):
                                    continue
                                fp = os.path.join(dicts_dir, fn)
                                try:
                                    with open(fp, 'r', encoding='utf-8') as jf:
                                        d = json.load(jf)
                                    ident = d.get('identificacion') or {}
                                    fol_file = str(ident.get('folio') or '').strip()
                                    digits = ''.join([c for c in fol_file if c.isdigit()])
                                    if digits:
                                        existing.add(int(digits.lstrip('0') or '0'))
                                except Exception:
                                    # fallback: extraer dígitos del nombre de archivo
                                    for g in re.findall(r"\d{3,7}", fn):
                                        try:
                                            existing.add(int(g.lstrip('0') or '0'))
                                        except Exception:
                                            pass
                    except Exception:
                        pass

                    # Escanear data/folios_visitas
                    try:
                        fv_dir = os.path.join(DATA_DIR, 'folios_visitas')
                        if os.path.exists(fv_dir):
                            for ffn in os.listdir(fv_dir):
                                if not ffn.lower().endswith('.json'):
                                    continue
                                try:
                                    with open(os.path.join(fv_dir, ffn), 'r', encoding='utf-8') as fjs:
                                        arr = json.load(fjs) or []
                                except Exception:
                                    continue
                                for entry in arr:
                                    try:
                                        fol = entry.get('FOLIOS') or entry.get('FOLIO') or ''
                                        s = ''.join([c for c in str(fol) if c.isdigit()])
                                        if s:
                                            existing.add(int(s.lstrip('0') or '0'))
                                    except Exception:
                                        continue
                    except Exception:
                        pass

                    # NOTA: no añadir aquí `folios_a_eliminar` a `existing` porque
                    # esos folios pertenecen a la visita que estamos borrando;
                    # en su lugar los excluiremos explícitamente más abajo tras
                    # leer `assigned_nums` desde el archivo correspondiente.

                    try:
                        curr = int(folio_manager.get_last() or 0) if folio_manager is not None else 0
                    except Exception:
                        curr = 0

                    # Construir lista ordenada de folios detectados para esta visita.
                    # Priorizar el archivo `data/folios_visitas/folios_<visita>.json` si existe.
                    # Si ya leímos `assigned_nums` al inicio, no sobrescribirlo.
                    try:
                        if 'assigned_nums' not in locals() or not assigned_nums:
                            assigned_nums = set()
                            fv_file = os.path.join(self.folios_visita_path, f"folios_{folio_borrado}.json")
                            if os.path.exists(fv_file):
                                try:
                                    with open(fv_file, 'r', encoding='utf-8') as _ff:
                                        _arr = json.load(_ff) or []
                                    for _rec in _arr:
                                        fol = _rec.get('FOLIOS') or _rec.get('FOLIO') or ''
                                        digs = ''.join([c for c in str(fol) if c.isdigit()])
                                        if digs:
                                            assigned_nums.add(int(digs.lstrip('0') or '0'))
                                except Exception:
                                    pass
                    except Exception:
                        assigned_nums = set()
                    # Si no hubo archivo o no pudo leerse, usar folios_a_eliminar recopilados antes
                    if not assigned_nums:
                        try:
                            for f in folios_a_eliminar:
                                s = ''.join([c for c in str(f) if c.isdigit()])
                                if s:
                                    assigned_nums.add(int(s.lstrip('0') or '0'))
                        except Exception:
                            assigned_nums = set()

                    # Si no se detectaron folios desde archivos, intentar parsear el campo
                    # `folios_utilizados` del registro (ej. "000857 - 000865").
                    if not assigned_nums:
                        try:
                            raw = visita_a_borrar.get('folios_utilizados') or visita_a_borrar.get('folios') or ''
                            if raw:
                                import re as _re
                                rngs = _re.findall(r"(\d{1,6})\s*-\s*(\d{1,6})", str(raw))
                                for a,b in rngs:
                                    a_i = int(a.lstrip('0') or '0')
                                    b_i = int(b.lstrip('0') or '0')
                                    if a_i <= b_i:
                                        for n in range(a_i, b_i+1):
                                            assigned_nums.add(n)
                                # also collect standalone numbers
                                for m in _re.findall(r"\d{1,6}", str(raw)):
                                    try:
                                        assigned_nums.add(int(m.lstrip('0') or '0'))
                                    except Exception:
                                        pass
                        except Exception:
                            pass

                    set_ok = False
                    dbg_messages = []

                    # Strategy 0: forced restore from the visit's file (automatic)
                    # If we have assigned folios from the visit file, prefer restoring
                    # the counter to min(assigned)-1 so those folios return to pool.
                    try:
                        if assigned_nums:
                            candidate = max(0, min(assigned_nums) - 1)
                            try:
                                if folio_manager is not None:
                                    # Only lower the counter (never increase)
                                    curr_check = int(folio_manager.get_last() or 0)
                                else:
                                    curr_check = None
                            except Exception:
                                curr_check = None
                            if curr_check is None or candidate < curr_check:
                                try:
                                    if folio_manager is not None:
                                        folio_manager.set_last(candidate)
                                        set_ok = True
                                        dbg_messages.append(f"forced_restore_from_file: candidate={candidate}")
                                except Exception as e:
                                    dbg_messages.append(f"forced_restore failed: {e}")
                    except Exception:
                        pass

                    # Strategy 1: if the deleted visit used the top-most folios (contiguous ending at current),
                    # decrement by the contiguous count.
                    if assigned_nums and max(assigned_nums) == curr:
                        try:
                            cnt = 0
                            n = curr
                            while n in assigned_nums:
                                cnt += 1
                                n -= 1
                            new_last = curr - cnt
                            if folio_manager is not None:
                                folio_manager.set_last(new_last)
                                set_ok = True
                                dbg_messages.append(f"top_contiguous rollback: curr={curr} cnt={cnt} new_last={new_last}")
                        except Exception as e:
                            dbg_messages.append(f"top_contiguous failed: {e}")

                    # Strategy 2: fallback to scanning remaining files and set to max existing
                    if not set_ok:
                        try:
                            # Excluir folios asignados a la visita borrada
                            try:
                                if assigned_nums:
                                    existing = {x for x in existing if x not in assigned_nums}
                            except Exception:
                                pass

                            max_existing = max(existing) if existing else 0
                            # If there are no remaining files but we have assigned folios,
                            # fallback to set the counter to just before the smallest
                            # assigned folio (min(assigned)-1). This restores those folios
                            # to the pool when the visit consumed the top block.
                            if (not existing) and assigned_nums:
                                try:
                                    new_last_candidate = max(0, min(assigned_nums) - 1)
                                    max_existing = new_last_candidate
                                except Exception:
                                    pass
                            else:
                                # ensure assigned numbers are considered (in case files were removed)
                                if assigned_nums:
                                    max_assigned = max(assigned_nums)
                                    if max_assigned > max_existing:
                                        # if files were removed but assigned contained higher folios,
                                        # treat those as candidates to reduce the counter
                                        max_existing = max_assigned
                            if max_existing < curr:
                                new_last = max_existing
                                try:
                                    if folio_manager is not None:
                                        folio_manager.set_last(new_last)
                                        set_ok = True
                                        dbg_messages.append(f"scan_max rollback: curr={curr} max_existing={max_existing} new_last={new_last}")
                                except Exception:
                                    set_ok = False
                                if not set_ok:
                                    try:
                                        fc_dir = DATA_DIR
                                        os.makedirs(fc_dir, exist_ok=True)
                                        tmp_path = os.path.join(fc_dir, 'folio_counter.json.tmp')
                                        real_path = os.path.join(fc_dir, 'folio_counter.json')
                                        with open(tmp_path, 'w', encoding='utf-8') as tf:
                                            json.dump({"last": int(new_last)}, tf)
                                        try:
                                            os.replace(tmp_path, real_path)
                                            set_ok = True
                                            dbg_messages.append(f"atomic write fallback: set to {new_last}")
                                        except Exception:
                                            try:
                                                if os.path.exists(tmp_path):
                                                    os.remove(tmp_path)
                                            except Exception:
                                                pass
                                    except Exception as e:
                                        dbg_messages.append(f"atomic write failed: {e}")
                        except Exception as e:
                            dbg_messages.append(f"scan_max failed: {e}")

                    # Write debug log entries
                    try:
                        dbg_dir = DATA_DIR
                        os.makedirs(dbg_dir, exist_ok=True)
                        dbg_path = os.path.join(dbg_dir, 'hist_borrar_debug.log')
                        with open(dbg_path, 'a', encoding='utf-8') as dbgf:
                            from datetime import datetime as _dt
                            # Compute current persisted last for accurate reporting
                            try:
                                persisted_last = int(folio_manager.get_last() or 0) if folio_manager is not None else None
                            except Exception:
                                persisted_last = None
                            dbgf.write(f"[{_dt.now().isoformat()}] hist_borrar id={folio_borrado} curr={curr} assigned={sorted(list(assigned_nums))[:200]} existing_sample={sorted(list(existing))[:200]} deleted_files={deleted_files} set_ok={set_ok} persisted_last={persisted_last} msgs={dbg_messages}\n")
                            try:
                                print(f"[DEBUG] hist_borrar id={folio_borrado} curr={curr} assigned={sorted(list(assigned_nums))[:200]} existing_sample={sorted(list(existing))[:200]} deleted_files={deleted_files} set_ok={set_ok} persisted_last={persisted_last} msgs={dbg_messages}")
                            except Exception:
                                pass
                    except Exception:
                        pass

                    # Intentar eliminar ahora el archivo folios_{folio}.json asociado
                    # sólo después de haber usado su contenido para la recuperación.
                    try:
                        fv_file = os.path.join(self.folios_visita_path, f"folios_{folio_borrado}.json")
                        if os.path.exists(fv_file):
                            try:
                                os.remove(fv_file)
                                deleted_files.append(os.path.basename(fv_file))
                            except Exception:
                                pass
                    except Exception:
                        pass

                    if set_ok:
                        try:
                            resumen = (resumen if 'resumen' in locals() else '') + f"\nSe recuperaron folios (contador actualizado)."
                        except Exception:
                            pass
                except Exception:
                    pass

                resumen = f"Se borró el folio {folio_borrado}. Folio actual recalculado: {self.current_folio}."
                if deleted_files:
                    resumen += f"\nSe eliminaron {len(deleted_files)} archivos JSON de dictámenes asociados: {', '.join(deleted_files)}"
                messagebox.showinfo("Folio actualizado", resumen)
            except Exception as e:
                try:
                    messagebox.showinfo("Folio actualizado", f"Se borró el folio {folio_borrado}. Folio actual recalculado: {self.current_folio}")
                except Exception:
                    pass

        except Exception as e:
            messagebox.showerror("Error", str(e))
  
    def _guardar_historial(self):
        """Guarda el historial en un único archivo con validación de persistencia"""
        try:
            # ACTUALIZAR self.historial_data DESDE self.historial
            self.historial_data = self.historial.get("visitas", [])
            self.historial_data_original = self.historial_data.copy()

            # Ordenar historial por `folio_visita` (CP) de menor a mayor cuando sea posible.
            # Intentamos extraer el primer número presente en el campo `folio_visita` y
            # ordenar numéricamente; si no es posible, caer al comparador de texto.
            try:
                import re
                def _folio_key(rec):
                    try:
                        fv = str(rec.get('folio_visita') or '')
                    except Exception:
                        fv = ''
                    # Buscar primer bloque de dígitos
                    m = re.search(r'(\d+)', fv)
                    if m:
                        try:
                            return int(m.group(1))
                        except Exception:
                            pass
                    # Intentar convertir entero directo
                    try:
                        return int(fv)
                    except Exception:
                        return fv.lower() if isinstance(fv, str) else fv

                visitas_list = self.historial.get('visitas')
                if isinstance(visitas_list, list) and visitas_list:
                    visitas_list.sort(key=_folio_key)
                    # Asegurar que self.historial_data refleje el orden actual
                    self.historial_data = visitas_list
            except Exception:
                pass
            
            # Determinar ruta de guardado (soporte para .exe congelado y rutas no escribibles)
            target_path = self.historial_path
            try:
                base_dir = os.path.dirname(self.historial_path)
                os.makedirs(base_dir, exist_ok=True)
            except Exception:
                base_dir = None

            try:
                # Si estamos en un ejecutable congelado (PyInstaller), preferir escribir
                # en APP_DIR/data (carpeta junto al exe) para que los archivos sean visibles
                # al usuario. Solo si eso falla, redirigir a APPDATA.
                if getattr(sys, 'frozen', False):
                    try:
                        app_data_dir = DATA_DIR
                        os.makedirs(app_data_dir, exist_ok=True)
                        if os.access(app_data_dir, os.W_OK):
                            target_path = os.path.join(app_data_dir, os.path.basename(self.historial_path))
                            self.historial_path = target_path
                        else:
                            raise Exception("No write access to APP_DIR/data")
                    except Exception:
                        alt_base = os.path.join(os.environ.get('APPDATA', os.path.expanduser('~')), 'GeneradorDictamenes')
                        os.makedirs(alt_base, exist_ok=True)
                        target_path = os.path.join(alt_base, os.path.basename(self.historial_path))
                        self.historial_path = target_path
                elif base_dir and not os.access(base_dir, os.W_OK):
                    alt_base = os.path.join(os.environ.get('APPDATA', os.path.expanduser('~')), 'GeneradorDictamenes')
                    os.makedirs(alt_base, exist_ok=True)
                    target_path = os.path.join(alt_base, os.path.basename(self.historial_path))
                    self.historial_path = target_path
            except Exception:
                target_path = self.historial_path

            # Guardar con respaldo (backup)
            backup_path = target_path + ".backup"
            if os.path.exists(target_path):
                try:
                    shutil.copy2(target_path, backup_path)
                except Exception:
                    pass

            # Escribir archivo principal
            # Escritura atómica: volcar a tmp y luego reemplazar
            tmp_path = target_path + '.tmp'
            try:
                with open(tmp_path, 'w', encoding='utf-8') as f:
                    json.dump(self.historial, f, ensure_ascii=False, indent=2)
                    f.flush()
                    try:
                        os.fsync(f.fileno())
                    except Exception:
                        pass

                # Reemplazar de forma atómica
                try:
                    os.replace(tmp_path, target_path)
                except Exception:
                    # Fallback a copy if replace falla
                    shutil.copy2(tmp_path, target_path)
            finally:
                try:
                    if os.path.exists(tmp_path):
                        os.remove(tmp_path)
                except Exception:
                    pass

            # Verificar que se escribió correctamente
            if os.path.exists(target_path):
                with open(target_path, 'r', encoding='utf-8') as f:
                    verificacion = json.load(f)
                    if verificacion.get('visitas'):
                        lbl = getattr(self, 'hist_info_label', None)
                        if lbl and hasattr(lbl, 'winfo_exists') and lbl.winfo_exists():
                            try:
                                lbl.configure(text=f"✅ Guardado — {len(self.historial_data)} registros")
                            except Exception:
                                pass
            else:
                print("⚠️ Error: No se pudo verificar el archivo guardado")
            print(f"✅ Historial guardado: {len(self.historial_data)} registros (ruta: {target_path})")
            
        except Exception as e:
            print(f"❌ Error guardando historial: {e}")
            lbl = getattr(self, 'hist_info_label', None)
            if lbl and hasattr(lbl, 'configure'):
                try:
                    lbl.configure(text=f"Error guardando: {e}")
                except Exception:
                    pass
    
    def hist_hacer_backup(self):
        """Crea un respaldo manual del historial"""
        try:
            if os.path.exists(self.historial_path):
                backup_dir = os.path.join(os.path.dirname(self.historial_path), "backups")
                os.makedirs(backup_dir, exist_ok=True)
                
                backup_name = f"historial_visitas_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                backup_path = os.path.join(backup_dir, backup_name)
                
                shutil.copy2(self.historial_path, backup_path)
                messagebox.showinfo("Backup", f"Backup creado:\n{backup_path}")
            else:
                messagebox.showinfo("Backup", "No existe historial para respaldar.")
        except Exception as e:
            messagebox.showerror("Backup error", str(e))

    def _limpiar_scroll_hist(self):
        # Si el Treeview está presente, vaciarlo (operación rápida)
        try:
            if hasattr(self, 'hist_tree') and self.hist_tree is not None:
                try:
                    self.hist_tree.delete(*self.hist_tree.get_children())
                except Exception:
                    pass
                return
        except Exception:
            pass

        # Fallback: recrear el scrollable frame para versiones antiguas
        old = getattr(self, 'hist_scroll', None)
        parent = old.master if old is not None else None
        if parent is None:
            return

        new_scroll = ctk.CTkScrollableFrame(
            parent,
            fg_color=STYLE["fondo"],
            scrollbar_button_color=STYLE["primario"],
            scrollbar_button_hover_color=STYLE["primario"]
        )
        new_scroll.pack(fill="both", expand=True, padx=0, pady=0)
        self.hist_scroll = new_scroll

        if old is not None and isinstance(old, ctk.CTkScrollableFrame):
            try:
                self.after(250, old.destroy)
            except Exception:
                try:
                    old.destroy()
                except Exception:
                    pass

    # -- BOTONES DE ACCION PARA CADA VISITA -- #
    def _poblar_historial_ui(self):
        """Poblar historial usando Treeview virtualizado (más eficiente)."""
        # Cargar datos solo si no existen o si se solicita recarga (permite que búsquedas filtradas persistan)
        if (not hasattr(self, 'historial_data') or not self.historial_data) or getattr(self, '_force_reload_hist', False):
            self._cargar_historial()
            self._force_reload_hist = False
        regs = getattr(self, 'historial_data', []) or []
        role = getattr(self, 'current_role', None)
        user = getattr(self, 'current_user', None)
        encontrados = []
        # Si no es admin, leer los JSON individuales guardados en data/produccion/<usuario>
        if role != 'admin':
            try:
                # obtener nombre de usuario
                uname = ''
                if isinstance(user, dict):
                    uname = user.get('username') or ''
                else:
                    uname = str(user or '')
                owner_safe = re.sub(r"[^A-Za-z0-9_.-]", '_', str(uname))
                owner_dir = os.path.join(DATA_DIR, 'produccion', owner_safe)
                if not os.path.exists(owner_dir):
                    try:
                        self.reporte_log.configure(text='No existen registros guardados para su usuario.')
                    except Exception:
                        pass
                    # Evitar mostrar un popup modal al iniciar la aplicación.
                    # Mostrar el estado en la UI si el widget existe y devolver.
                    return

                for fn in os.listdir(owner_dir):
                    if not fn.lower().endswith('.json'):
                        continue
                    fp = os.path.join(owner_dir, fn)
                    try:
                        with open(fp, 'r', encoding='utf-8') as jf:
                            obj = json.load(jf)
                        # Añadir todas las visitas guardadas del usuario (sin filtrado por fecha aquí)
                        encontrados.append(obj)
                    except Exception:
                        continue
            except Exception:
                encontrados = []
        else:
            # admin: comportamiento global sobre historial en memoria
            fuente = getattr(self, 'historial_data', []) or []
            for r in fuente:
                try:
                    encontrados.append(r)
                except Exception:
                    continue
        # Actualizar opciones del combo de supervisores con valores únicos encontrados
        try:
            if getattr(self, 'combo_filtrar_supervisor', None):
                sups = []
                for r in (regs or []):
                    try:
                        s = (r.get('supervisor') or r.get('nfirma1') or '')
                        s = str(s).strip()
                        if s and s not in sups:
                            sups.append(s)
                    except Exception:
                        continue
                try:
                    vals = [""] + sorted(sups, key=lambda x: x.lower())
                    self.combo_filtrar_supervisor.configure(values=vals)
                except Exception:
                    pass
        except Exception:
            pass
        # Actualizar opciones del combo de clientes con valores únicos encontrados
        try:
            if getattr(self, 'combo_filtrar_cliente', None):
                clis = []
                for r in (regs or []):
                    try:
                        c = (r.get('cliente') or '')
                        c = str(c).strip()
                        if c and c not in clis:
                            clis.append(c)
                    except Exception:
                        continue
                try:
                    vals = [""] + sorted(clis, key=lambda x: x.lower())
                    self.combo_filtrar_cliente.configure(values=vals)
                except Exception:
                    pass
        except Exception:
            pass
        # Ordenar visitas por folio_visita (CP) ascendente, y luego por folio_acta (AC)
        try:
            def _folio_key(r):
                def digits_of(s):
                    try:
                        s = str(s) or ''
                        digs = ''.join([c for c in s if c.isdigit()])
                        return int(digs) if digs else 0
                    except Exception:
                        return 0

                return (digits_of(r.get('folio_visita') or r.get('folio') or ''), digits_of(r.get('folio_acta') or ''))

            regs = sorted(regs, key=_folio_key)
        except Exception:
            pass

        total_registros = len(regs)
        regs_pagina = self.HISTORIAL_REGS_POR_PAGINA
        pagina_actual = getattr(self, 'HISTORIAL_PAGINA_ACTUAL', 1)
        total_paginas = max(1, (total_registros + regs_pagina - 1) // regs_pagina)
        inicio = (pagina_actual - 1) * regs_pagina
        fin = min(inicio + regs_pagina, total_registros)

        # actualizar controles de paginación si existen
        try:
            if hasattr(self, 'hist_pagina_label'):
                self.hist_pagina_label.configure(text=f"Página {pagina_actual} de {total_paginas}")
            if hasattr(self, 'btn_hist_prev'):
                self.btn_hist_prev.configure(state="normal" if pagina_actual > 1 else "disabled")
            if hasattr(self, 'btn_hist_next'):
                self.btn_hist_next.configure(state="normal" if pagina_actual < total_paginas else "disabled")
        except Exception:
            pass

        # Vaciar Treeview actual
        try:
            self.hist_tree.delete(*self.hist_tree.get_children())
        except Exception:
            pass

        # Configurar tags de colores para estatus (cancelada=rojo, pendiente=amarillo, completado=verde)
        try:
            # usar colores suaves para legibilidad
            self.hist_tree.tag_configure('cancelado', background='#FFC7CE')
            self.hist_tree.tag_configure('pendiente', background='#FFF2CC')
            self.hist_tree.tag_configure('completado', background='#C6EFCE')
        except Exception:
            pass

        # Insertar registros de la página actual
        for idx in range(inicio, fin):
            registro = regs[idx]
            hora_inicio = self._formatear_hora_12h(registro.get('hora_inicio', ''))
            hora_termino = self._formatear_hora_12h(registro.get('hora_termino', ''))

            # Preferir calcular rango real desde el archivo `folios_<folio_visita>.json`
            folios_display = ''
            try:
                fid = registro.get('folio_visita') or registro.get('folio')
                if fid:
                    archivo_f = os.path.join(self.folios_visita_path, f"folios_{fid}.json")
                    if os.path.exists(archivo_f):
                        try:
                            with open(archivo_f, 'r', encoding='utf-8') as ff:
                                arr = json.load(ff) or []
                            nums = []
                            for it in arr:
                                fol = it.get('FOLIOS') or it.get('FOLIOS', '') or ''
                                # extraer dígitos
                                digs = ''.join([c for c in str(fol) if c.isdigit()])
                                if digs:
                                    try:
                                        nums.append(int(digs))
                                    except Exception:
                                        continue
                            nums = sorted(set(nums))
                            if nums:
                                if len(nums) == 1:
                                    folios_display = f"{nums[0]:06d}"
                                else:
                                    folios_display = f"{nums[0]:06d} - {nums[-1]:06d}"
                        except Exception:
                            folios_display = ''

            except Exception:
                folios_display = ''

            # Si no hubo archivo de folios, usar el campo guardado en el registro
            if not folios_display:
                folios_str = registro.get('folios_utilizados', '') or ''
                if not folios_str or folios_str in ('0', '-'):
                    folios_display = ''
                else:
                    folios_display = self._formatear_folios_rango(folios_str)

            cliente_short = self._acortar_texto(registro.get('cliente', '-'), 20)
            nfirma1_short = self._acortar_texto(registro.get('nfirma1', 'N/A'), 12)

            datos = [
                registro.get('folio_visita', '-') or '-',
                registro.get('folio_acta', '-') or '-',
                registro.get('fecha_inicio', '-') or '-',
                registro.get('fecha_termino', '-') or '-',
                hora_inicio or '-',
                hora_termino or '-',
                cliente_short,
                nfirma1_short,
                registro.get('tipo_documento', '-') or '-',
                registro.get('estatus', 'Completado') or 'Completado',
                folios_display,
                "📁 Folios  •  📎 Archivos  •  ✏️ Editar  •  🗑️ Borrar"
            ]

            # Insertar en tree
            iid = f"h_{idx}"
            try:
                # determinar tag según estatus
                est_raw = str(registro.get('estatus', '') or '').strip().lower()
                tag = None
                if 'cancel' in est_raw:
                    tag = 'cancelado'
                elif 'pend' in est_raw:
                    tag = 'pendiente'
                elif 'complet' in est_raw or 'finaliz' in est_raw:
                    tag = 'completado'

                if tag:
                    self.hist_tree.insert('', 'end', iid=iid, values=datos, tags=(tag,))
                else:
                    self.hist_tree.insert('', 'end', iid=iid, values=datos)
                self._hist_map[iid] = registro
            except Exception:
                pass

        # actualizar info pie
        try:
            if hasattr(self, 'hist_info_label'):
                self.hist_info_label.configure(
                    text=f"Registros: {total_registros} | Sistema V&C - Generador de Dictámenes de Comprimiento"
                )
        except Exception:
            pass

    def _hist_show_context_menu(self, event):
        try:
            iid = self.hist_tree.identify_row(event.y)
            if not iid:
                return
            self.hist_tree.selection_set(iid)
            self._hist_context_selected = iid
            self.hist_context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            try:
                self.hist_context_menu.grab_release()
            except Exception:
                pass

    def _hist_on_double_click(self, event):
        iid = self.hist_tree.identify_row(event.y)
        if not iid:
            return
        reg = self._hist_map.get(iid)
        if reg:
            try:
                self.mostrar_opciones_documentos(reg)
            except Exception:
                pass

    def _hist_on_left_click(self, event):
        # Si el usuario hizo click en la columna de Acciones, abrir menú contextual
        try:
            col = self.hist_tree.identify_column(event.x)
            # columnas vienen como '#1', '#2', ...
            cols = list(self.hist_tree['columns'])
            if not cols:
                return
            last_index = len(cols)
            if col == f"#{last_index}":
                iid = self.hist_tree.identify_row(event.y)
                if not iid:
                    return
                self.hist_tree.selection_set(iid)
                self._hist_context_selected = iid
                try:
                    self.hist_context_menu.tk_popup(event.x_root, event.y_root)
                finally:
                    try:
                        self.hist_context_menu.grab_release()
                    except Exception:
                        pass
        except Exception:
            pass

    def _hist_menu_action(self, action):
        iid = getattr(self, '_hist_context_selected', None)
        if not iid:
            sel = self.hist_tree.selection()
            iid = sel[0] if sel else None
        if not iid:
            return
        reg = self._hist_map.get(iid)
        if not reg:
            return
        try:
            if action == 'folios':
                self.descargar_folios_visita(reg)
            elif action == 'archivos':
                self.mostrar_opciones_documentos(reg)
            elif action == 'editar':
                self.hist_editar_registro(reg)
            elif action == 'borrar':
                self.hist_eliminar_registro(reg)
        except Exception:
            pass

    def hist_pagina_anterior(self):
        try:
            # debug log to help diagnose exe issues
            try:
                dbg = os.path.join(DATA_DIR, 'pagination_debug.log')
                with open(dbg, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now().isoformat()}] hist_pagina_anterior called. current={self.HISTORIAL_PAGINA_ACTUAL}\n")
            except Exception:
                pass

            if getattr(self, 'HISTORIAL_PAGINA_ACTUAL', 1) > 1:
                self.HISTORIAL_PAGINA_ACTUAL -= 1
                self._poblar_historial_ui()

            try:
                with open(dbg, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now().isoformat()}] hist_pagina_anterior finished. new={self.HISTORIAL_PAGINA_ACTUAL}\n")
            except Exception:
                pass
        except Exception as e:
            try:
                with open(os.path.join(DATA_DIR, 'pagination_debug.log'), 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now().isoformat()}] hist_pagina_anterior ERROR: {e}\n")
            except Exception:
                pass

    def hist_pagina_siguiente(self):
        try:
            try:
                dbg = os.path.join(DATA_DIR, 'pagination_debug.log')
                with open(dbg, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now().isoformat()}] hist_pagina_siguiente called. current={getattr(self, 'HISTORIAL_PAGINA_ACTUAL', None)}\n")
            except Exception:
                pass

            total_registros = len(getattr(self, 'historial_data', []) or [])
            total_paginas = max(1, (total_registros + self.HISTORIAL_REGS_POR_PAGINA - 1) // self.HISTORIAL_REGS_POR_PAGINA)
            if getattr(self, 'HISTORIAL_PAGINA_ACTUAL', 1) < total_paginas:
                self.HISTORIAL_PAGINA_ACTUAL += 1
                self._poblar_historial_ui()

            try:
                with open(dbg, 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now().isoformat()}] hist_pagina_siguiente finished. new={self.HISTORIAL_PAGINA_ACTUAL}\n")
            except Exception:
                pass
        except Exception as e:
            try:
                with open(os.path.join(DATA_DIR, 'pagination_debug.log'), 'a', encoding='utf-8') as f:
                    f.write(f"[{datetime.now().isoformat()}] hist_pagina_siguiente ERROR: {e}\n")
            except Exception:
                pass

    def _formatear_folios_rango(self, folios_str):
        """Formatea los folios para mostrar solo el rango (inicio-fin)"""
        if not folios_str or folios_str == '0' or folios_str == '-':
            return '-'
        
        # Si ya es un rango simple
        if ' - ' in folios_str and not folios_str.startswith('Total:'):
            # Extraer solo el rango (puede venir como "000001 - 000010")
            return folios_str
        
        # Si es un solo folio
        if folios_str.startswith('Folio: '):
            return folios_str.replace('Folio: ', '')
        
        # Si tiene formato de total con lista
        if folios_str.startswith('Total: '):
            # Intentar extraer folios de la lista
            if '| Folios:' in folios_str:
                try:
                    partes = folios_str.split('| Folios:')
                    if len(partes) > 1:
                        folios_lista = partes[1].strip().split(', ')
                        if folios_lista:
                            # Obtener primer y último folio
                            primer = folios_lista[0].strip()
                            ultimo = folios_lista[-1].strip().replace('...', '').strip()
                            if primer and ultimo and primer != ultimo:
                                return f"{primer} - {ultimo}"
                            elif primer:
                                return primer
                except:
                    pass
            
            # Si no se pudo extraer, mostrar solo el total
            try:
                total_part = folios_str.split('|')[0].strip()
                return total_part.replace('Total: ', '') + ' folios'
            except:
                return folios_str
        
        # Si tiene muchos folios separados por comas
        if ',' in folios_str:
            folios_list = [f.strip() for f in folios_str.split(',') if f.strip()]
            if len(folios_list) > 1:
                return f"{folios_list[0]} - {folios_list[-1]}"
            elif folios_list:
                return folios_list[0]
        
        return folios_str[:20] + ('...' if len(folios_str) > 20 else '')

    def _acortar_texto(self, texto, max_caracteres=20):
        """Acorta el texto si es muy largo, agregando '...' al final"""
        if not texto:
            return ""
        
        texto_str = str(texto)
        if len(texto_str) <= max_caracteres:
            return texto_str
        
        return texto_str[:max_caracteres-3] + "..."

    def _formatear_hora_12h(self, hora_str):
        """Formatea una hora a formato 12h (AM/PM)"""
        try:
            if ":" in hora_str:
                partes = hora_str.split(":")
                if len(partes) >= 2:
                    horas = int(partes[0])
                    minutos = int(partes[1])
                    
                    periodo = "AM" if horas < 12 else "PM"
                    horas_12 = horas if horas <= 12 else horas - 12
                    if horas_12 == 0:
                        horas_12 = 12
                    
                    return f"{horas_12}:{minutos:02d} {periodo}"
        except:
            pass
        
        return hora_str

    def mostrar_opciones_documentos(self, registro):
        """Muestra una ventana con opciones para descargar documentos"""
        # Crear ventana modal
        modal = ctk.CTkToplevel(self)
        modal.title("Descargar Documentos")
        modal.geometry("750x600")
        modal.transient(self)
        modal.grab_set()
        
        # Centrar ventana
        modal.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - modal.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - modal.winfo_height()) // 2
        modal.geometry(f"+{x}+{y}")
        
        # Frame principal
        main_frame = ctk.CTkFrame(modal, fg_color=STYLE["surface"], corner_radius=0)
        main_frame.pack(fill="both", expand=True, padx=0, pady=0)
        
        # Título
        ctk.CTkLabel(
            main_frame,
            text="📄 Documentos de la Visita",
            font=("Inter", 20, "bold"),
            text_color=STYLE["texto_oscuro"]
        ).pack(pady=(15, 10))
        
        # Información de la visita
        info_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        info_frame.pack(fill="x", padx=0, pady=(0, 0))
        
        ctk.CTkLabel(
            info_frame,
            text=f"Folio Visita: {registro.get('folio_visita', 'N/A')} | Cliente: {registro.get('cliente', 'N/A')}",
            font=("Inter", 13),
            text_color=STYLE["texto_oscuro"]
        ).pack()
        
        # Línea separadora
        separador = ctk.CTkFrame(main_frame, fg_color=STYLE["borde"], height=1)
        separador.pack(fill="x", padx=0, pady=(0, 0))
        
        # Frame para las opciones de documentos en horizontal
        documentos_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        documentos_frame.pack(fill="both", expand=True, padx=0, pady=0)
        
        # Configurar grid para 3 columnas
        documentos_frame.grid_columnconfigure(0, weight=1)
        documentos_frame.grid_columnconfigure(1, weight=1)
        documentos_frame.grid_columnconfigure(2, weight=1)
        documentos_frame.grid_rowconfigure(0, weight=1)
        
        # Función para manejar la descarga de documentos
        def descargar_documento(tipo, nombre):
            modal.destroy()
            try:
                if tipo == "acta":
                    folio = registro.get('folio_visita', '')
                    if not folio:
                        messagebox.showwarning("Error", "No se encontró el folio de la visita para generar el acta.")
                        return

                    # Preferir el backup más reciente en data/tabla_relacion_backups si existe
                    data_dir_local = DATA_DIR
                    backups_dir = os.path.join(data_dir_local, 'tabla_relacion_backups')
                    tabla_dest = os.path.join(data_dir_local, 'tabla_de_relacion.json')

                    if os.path.exists(backups_dir):
                        archivos = [os.path.join(backups_dir, f) for f in os.listdir(backups_dir) if f.lower().endswith('.json')]
                        if archivos:
                            latest = max(archivos, key=os.path.getmtime)
                            try:
                                shutil.copy2(latest, tabla_dest)
                                print(f"📁 Usando backup de tabla de relación: {latest}")
                            except Exception as e:
                                print(f"⚠️ No se pudo copiar backup de tabla de relación: {e}")

                    # Pedir al usuario dónde guardar el PDF (Explorador de archivos)
                    default_name = f"Acta_{folio}.pdf"
                    save_path = filedialog.asksaveasfilename(
                        title="Guardar Acta de Inspección",
                        defaultextension=".pdf",
                        filetypes=[("PDF", "*.pdf")],
                        initialfile=default_name
                    )

                    if not save_path:
                        return

                    # Importar dinámicamente el generador de actas y generar
                    try:
                        import importlib.util
                        acta_file = os.path.join(BASE_DIR, 'Documentos Inspeccion', 'Acta_inspeccion.py')
                        if not os.path.exists(acta_file):
                            messagebox.showerror("Error", f"No se encontró el generador de actas: {acta_file}")
                            return

                        spec = importlib.util.spec_from_file_location("Acta_inspeccion", acta_file)
                        if spec is None or getattr(spec, 'loader', None) is None:
                            messagebox.showerror("Error", f"No se pudo cargar el módulo Acta_inspeccion (spec inválido): {acta_file}")
                            return
                        acta_mod = importlib.util.module_from_spec(spec)
                        import sys
                        # Registrar el módulo temporalmente para permitir reload() desde el código
                        sys.modules["Acta_inspeccion"] = acta_mod
                        spec.loader.exec_module(acta_mod)

                        # Generar acta para el folio y guardarla en la ruta indicada
                        ruta_generada = acta_mod.generar_acta_desde_visita(folio_visita=folio, ruta_salida=save_path)

                        # Persistir la ruta del acta en el historial (si corresponde)
                        try:
                            for v in self.historial.get('visitas', []):
                                if v.get('folio_visita') == folio:
                                    v['ruta_acta'] = save_path
                                    break
                            # Guardar historial
                            self._guardar_historial()
                        except Exception as e:
                            print(f"⚠️ Error guardando ruta de acta en historial: {e}")

                        messagebox.showinfo("Acta generada", f"Acta guardada en:\n{ruta_generada}")
                        return
                    except Exception as e:
                        messagebox.showerror("Error", f"Error generando acta:\n{e}")
                        return

                # Otros documentos: Oficio y Formato -> generar desde módulos correspondientes
                folio = registro.get('folio_visita', '')
                if not folio:
                    messagebox.showwarning("Error", "No se encontró el folio de la visita para generar el documento.")
                    return

                # Asegurar que usamos el backup más reciente para tabla_de_relacion
                data_dir_local = DATA_DIR
                backups_dir = os.path.join(data_dir_local, 'tabla_relacion_backups')
                tabla_dest = os.path.join(data_dir_local, 'tabla_de_relacion.json')

                if os.path.exists(backups_dir):
                    archivos = [os.path.join(backups_dir, f) for f in os.listdir(backups_dir) if f.lower().endswith('.json')]
                    if archivos:
                        latest = max(archivos, key=os.path.getmtime)
                        try:
                            shutil.copy2(latest, tabla_dest)
                            print(f"📁 Usando backup de tabla de relación: {latest}")
                        except Exception as e:
                            print(f"⚠️ No se pudo copiar backup de tabla de relación: {e}")

                # Calcular lista de folios asociados a la visita (int)
                folios_list = []
                try:
                    fol_num = ''.join([c for c in folio if c.isdigit()])
                    folios_file = os.path.join(data_dir_local, 'folios_visitas', f'folios_{fol_num}.json')
                    if os.path.exists(folios_file):
                        with open(folios_file, 'r', encoding='utf-8') as ff:
                            data_f = json.load(ff)
                            if isinstance(data_f, list):
                                folios_list = [int(x) for x in data_f if str(x).isdigit()]
                            elif isinstance(data_f, dict) and 'folios' in data_f:
                                folios_list = [int(x) for x in data_f.get('folios', []) if str(x).isdigit()]
                except Exception:
                    folios_list = []

                # fallback: parsear rango en registro
                if not folios_list:
                    fu = registro.get('folios_utilizados') or ''
                    if fu and isinstance(fu, str):
                        if '-' in fu:
                            parts = [p.strip() for p in fu.split('-')]
                            try:
                                start = int(parts[0])
                                end = int(parts[1]) if len(parts) > 1 else start
                                folios_list = list(range(start, end+1))
                            except Exception:
                                folios_list = []
                        elif ',' in fu:
                            vals = [p.strip() for p in fu.split(',')]
                            for v in vals:
                                if v.isdigit():
                                    folios_list.append(int(v))

                # Cargar tabla_de_relacion y extraer solicitudes únicas y FECHA DE VERIFICACION
                solicitudes = set()
                fecha_verificacion = None
                try:
                    if os.path.exists(tabla_dest):
                        with open(tabla_dest, 'r', encoding='utf-8') as tf:
                            tabla = json.load(tf)
                            for rec in tabla:
                                try:
                                    fol = rec.get('FOLIO')
                                    fol_int = int(fol) if fol is not None and str(fol).isdigit() else None
                                except Exception:
                                    fol_int = None
                                if folios_list and fol_int in folios_list:
                                    sol = rec.get('SOLICITUD') or rec.get('SOLICITUDES')
                                    if sol:
                                        solicitudes.add(str(sol).strip())
                                    if not fecha_verificacion and rec.get('FECHA DE VERIFICACION'):
                                        fecha_verificacion = rec.get('FECHA DE VERIFICACION')
                            # si no filtró por folios, intentar colectar solicitudes globales
                            if not solicitudes and isinstance(tabla, list):
                                for rec in tabla:
                                    sol = rec.get('SOLICITUD') or rec.get('SOLICITUDES')
                                    if sol:
                                        solicitudes.add(str(sol).strip())
                except Exception as e:
                    print(f"⚠️ Error leyendo tabla de relación para generar documento: {e}")

                # Formatear fecha_verificacion si está en ISO
                fecha_formateada = None
                if fecha_verificacion:
                    try:
                        if '-' in fecha_verificacion:
                            dt = datetime.strptime(fecha_verificacion[:10], '%Y-%m-%d')
                        else:
                            dt = datetime.strptime(fecha_verificacion[:10], '%d/%m/%Y')
                        fecha_formateada = dt.strftime('%d/%m/%Y')
                    except Exception:
                        fecha_formateada = fecha_verificacion

                # Preparar nombre por defecto y pedir ruta
                default_name = f"{nombre.replace(' ', '_')}_{folio}.pdf"
                save_path = filedialog.asksaveasfilename(
                    title=f"Guardar {nombre}",
                    defaultextension=".pdf",
                    filetypes=[("PDF", "*.pdf")],
                    initialfile=default_name
                )
                if not save_path:
                    return

                # Importar dinámicamente y generar según tipo
                try:
                    import importlib.util
                    # --- Generador especial: Reporte Decathlon ---
                    if tipo == 'decathlon':
                        # Generar y guardar automáticamente en data/Reportes_Decathlon
                        folio = registro.get('folio_visita', '')
                        if not folio:
                            messagebox.showwarning('Error', 'No se encontró el folio de la visita para generar el reporte Decathlon.')
                            return

                        # Buscar backup de tabla_relacion correspondiente al folio
                        tabla_relacion_data = None
                        try:
                            if os.path.exists(backups_dir):
                                archivos = [os.path.join(backups_dir, f) for f in os.listdir(backups_dir) if f.lower().endswith('.json')]
                                # Prefer files that contain the folio identifier (CP...)
                                matched = [a for a in archivos if folio in os.path.basename(a)]
                                src = None
                                if matched:
                                    src = max(matched, key=os.path.getmtime)
                                elif archivos:
                                    src = max(archivos, key=os.path.getmtime)
                                if src:
                                    with open(src, 'r', encoding='utf-8') as tf:
                                        try:
                                            tabla_relacion_data = json.load(tf)
                                        except Exception:
                                            tabla_relacion_data = None
                        except Exception:
                            tabla_relacion_data = None

                        # Ensure reports dir exists
                        reports_dir = os.path.join(data_dir_local, 'Reportes_Decathlon')
                        try:
                            os.makedirs(reports_dir, exist_ok=True)
                        except Exception:
                            pass

                        output_name = f"Reporte_Decathlon_{folio}.pdf"
                        output_path = os.path.join(reports_dir, output_name)

                        # Cargar módulo Reporte_Decathlon
                        rep_file = os.path.join(BASE_DIR, 'Documentos Inspeccion', 'Reporte_Decathlon.py')
                        if not os.path.exists(rep_file):
                            messagebox.showerror('Error', f'No se encontró el generador de reportes: {rep_file}')
                            return
                        spec_rep = importlib.util.spec_from_file_location('Reporte_Decathlon', rep_file)
                        if spec_rep is None or getattr(spec_rep, 'loader', None) is None:
                            messagebox.showerror('Error', f'No se pudo cargar el módulo Reporte_Decathlon (spec inválido): {rep_file}')
                            return
                        rep_mod = importlib.util.module_from_spec(spec_rep)
                        import sys as _sys
                        _sys.modules['Reporte_Decathlon'] = rep_mod
                        spec_rep.loader.exec_module(rep_mod)

                        # Llamar al generador
                        try:
                            # Si el usuario seleccionó una ruta en el diálogo, usaremos esa ruta como salida final
                            ruta_usuario = save_path if save_path else output_path

                            rep_mod.generar_reporte_decathlon(
                                nombre_cliente=registro.get('cliente',''),
                                folio_inspeccion=registro.get('folio_visita',''),
                                fecha_inspeccion=registro.get('fecha_termino') or registro.get('fecha_inicio') or datetime.now().strftime('%Y-%m-%d'),
                                tabla_relacion=tabla_relacion_data,
                                lista_equipos=None,
                                observaciones_generales=registro.get('observaciones',''),
                                ruta_guardado=ruta_usuario,
                                folio_visita=registro.get('folio_visita',''),
                                destinatario=registro.get('destinatario') or 'Diana Cumpean'
                            )
                            # No crear copia PDF en carpeta local (solo guardamos el JSON con metadatos)
                        except Exception as e:
                            messagebox.showerror('Error', f'Error generando Reporte Decathlon:\n{e}')
                            return

                        # Guardar JSON meta en carpeta local de Reportes_Decathlon
                        try:
                            # Construir JSON reducido con solo los datos impresos
                            fecha_verif = None
                            printed_rows = []
                            if isinstance(tabla_relacion_data, list):
                                seen = {}
                                order = []
                                for item in tabla_relacion_data:
                                    if fecha_verif is None:
                                        fecha_verif = item.get('FECHA DE VERIFICACION') or item.get('Fecha de Verificacion') or item.get('FECHA_VERIFICACION')
                                    sol = item.get('SOLICITUD') or item.get('Solicitud') or item.get('solicitud') or ''
                                    fol = item.get('FOLIO') or item.get('folio') or item.get('Folio') or ''
                                    ped = item.get('PEDIMENTO') or item.get('pedimento') or item.get('Referencia') or item.get('referencia') or ''
                                    key = str(fol)
                                    if key not in seen:
                                        seen[key] = {'SOLICITUD': sol, 'FOLIO': fol, 'PEDIMENTO': ped}
                                        order.append(key)
                                for k in order:
                                    printed_rows.append(seen[k])

                            meta = {
                                'generado_en': datetime.now().isoformat(),
                                'archivo_usuario': ruta_usuario,
                                'folio_visita': folio,
                                'cliente': registro.get('cliente',''),
                                'destinatario': registro.get('destinatario') or 'Diana Cumpean',
                                'fecha_verificacion': fecha_verif,
                                'observaciones': registro.get('observaciones',''),
                                'dictamenes_impresos': printed_rows
                            }
                            json_path = os.path.join(reports_dir, f"Reporte_Decathlon_{folio}.json")
                            with open(json_path, 'w', encoding='utf-8') as jf:
                                json.dump(meta, jf, ensure_ascii=False, indent=2)
                        except Exception:
                            json_path = None

                        # Persistir ruta en historial (usar ruta de usuario si existe)
                        try:
                            for v in self.historial.get('visitas', []):
                                if v.get('folio_visita') == folio:
                                    v['ruta_reporte_decathlon'] = ruta_usuario
                                    if json_path:
                                        v['ruta_reporte_decathlon_meta'] = json_path
                                    break
                            self._guardar_historial()
                        except Exception:
                            pass

                        # Informar al usuario rutas generadas
                        info_msg = f'Reporte guardado en:\n{ruta_usuario}'
                        if json_path:
                            info_msg += f"\nCopia y metadatos locales: {json_path}"
                        messagebox.showinfo('Reporte Decathlon generado', info_msg)
                        return
                    if tipo == 'formato':
                        formato_file = os.path.join(BASE_DIR, 'Documentos Inspeccion', 'Formato_supervision.py')
                        if not os.path.exists(formato_file):
                            messagebox.showerror('Error', f'No se encontró el módulo: {formato_file}')
                            return
                        spec = importlib.util.spec_from_file_location('Formato_supervision', formato_file)
                        if spec is None or getattr(spec, 'loader', None) is None:
                            messagebox.showerror('Error', f'No se pudo cargar el módulo Formato_supervision (spec inválido): {formato_file}')
                            return
                        mod = importlib.util.module_from_spec(spec)
                        import sys
                        sys.modules['Formato_supervision'] = mod
                        spec.loader.exec_module(mod)

                        datos = {
                            'solicitud': ', '.join(sorted(list(solicitudes))) if solicitudes else registro.get('folio_visita',''),
                            'servicio': None,
                            'fecha': registro.get('fecha_inicio') or registro.get('fecha_termino') or datetime.now().strftime('%d/%m/%Y'),
                            'cliente': registro.get('cliente',''),
                            'supervisor': 'Mario Terrez Gonzalez'
                        }
                        # Determinar servicio a partir de tabla (si hay alguna entrada con TIPO DE DOCUMENTO)
                        try:
                            # buscar primer registro en tabla_dest que corresponda a folios_list
                            if os.path.exists(tabla_dest):
                                with open(tabla_dest, 'r', encoding='utf-8') as tf:
                                    tabla = json.load(tf)
                                    for rec in tabla:
                                        try:
                                            fol = rec.get('FOLIO')
                                            fol_int = int(fol) if fol is not None and str(fol).isdigit() else None
                                        except Exception:
                                            fol_int = None
                                        if folios_list and fol_int in folios_list:
                                            tipo_doc = (rec.get('TIPO DE DOCUMENTO') or rec.get('TIPO_DE_DOCUMENTO') or '')
                                            tipo_code = str(tipo_doc).strip().upper()
                                            mapping = {
                                                'D': 'Dictamen',
                                                'C': 'Constancia',
                                                'ND': 'Negacion de Dictamen',
                                                'NC': 'Negacion de Constancia'
                                            }
                                            datos['servicio'] = mapping.get(tipo_code, str(tipo_doc) if tipo_doc else 'Dictamen')
                                            break
                        except Exception:
                            datos['servicio'] = datos.get('servicio') or 'Dictamen'

                        # Si no se determinó por folios, intentar inferir servicio desde cualquier registro de la tabla
                        if not datos.get('servicio'):
                            try:
                                if os.path.exists(tabla_dest):
                                    with open(tabla_dest, 'r', encoding='utf-8') as tf2:
                                        tabla_all = json.load(tf2)
                                        if isinstance(tabla_all, list):
                                            for rec in tabla_all:
                                                tipo_doc = rec.get('TIPO DE DOCUMENTO') or rec.get('TIPO_DE_DOCUMENTO') or ''
                                                if tipo_doc:
                                                    tipo_code = str(tipo_doc).strip().upper()
                                                    mapping = {
                                                        'D': 'Dictamen',
                                                        'C': 'Constancia',
                                                        'ND': 'Negacion de Dictamen',
                                                        'NC': 'Negacion de Constancia'
                                                    }
                                                    datos['servicio'] = mapping.get(tipo_code, str(tipo_doc))
                                                    break
                            except Exception:
                                pass

                        # Valor por defecto si no se logró inferir
                        if not datos.get('servicio'):
                            datos['servicio'] = 'Dictamen'

                        # Llamar al generador
                        try:
                            mod.generar_supervision(datos, save_path)
                        except Exception as e:
                            messagebox.showerror('Error', f'Error generando Formato de Supervisión:\n{e}')
                            return

                    elif tipo == 'oficio':
                        # Prefer a fixed fallback module if present to avoid importing a corrupted original
                        oficio_file = os.path.join(BASE_DIR, 'Documentos Inspeccion', 'Oficio_comision.py')
                        oficio_fixed = os.path.join(BASE_DIR, 'Documentos Inspeccion', 'Oficio_comision_fixed.py')
                        if os.path.exists(oficio_fixed):
                            oficio_file = oficio_fixed
                        if not os.path.exists(oficio_file):
                            messagebox.showerror('Error', f'No se encontró el módulo: {oficio_file}')
                            return
                        spec = importlib.util.spec_from_file_location('Oficio_comision', oficio_file)
                        if spec is None or getattr(spec, 'loader', None) is None:
                            messagebox.showerror('Error', f'No se pudo cargar el módulo Oficio_comision (spec inválido): {oficio_file}')
                            return
                        mod = importlib.util.module_from_spec(spec)
                        import sys
                        sys.modules['Oficio_comision'] = mod
                        spec.loader.exec_module(mod)

                        # Preferir usar la función de preparación del propio módulo si existe
                        datos_oficio = None
                        try:
                            if hasattr(mod, 'preparar_datos_desde_visita'):
                                datos_oficio = mod.preparar_datos_desde_visita(registro)
                            else:
                                # Heurística local: priorizar 'calle_numero' y anexar CP a colonia
                                calle = registro.get('calle_numero') or registro.get('direccion','') or ''
                                colonia = registro.get('colonia','') or ''
                                cp = registro.get('cp') or registro.get('CP') or ''
                                if cp and colonia:
                                    colonia = f"{colonia}, {cp}"
                                datos_oficio = {
                                    'no_oficio': registro.get('folio_visita',''),
                                    'fecha_inspeccion': fecha_formateada or registro.get('fecha_termino') or datetime.now().strftime('%d/%m/%Y'),
                                    'normas': registro.get('norma','').split(', ') if registro.get('norma') else [],
                                    'empresa_visitada': registro.get('cliente',''),
                                    'calle_numero': calle,
                                    'colonia': colonia,
                                    'municipio': registro.get('municipio',''),
                                    'ciudad_estado': registro.get('ciudad_estado',''),
                                    'fecha_confirmacion': registro.get('fecha_inicio') or datetime.now().strftime('%d/%m/%Y'),
                                    'medio_confirmacion': 'correo electrónico',
                                    'inspectores': [s.strip() for s in (registro.get('supervisores_tabla') or registro.get('nfirma1') or '').split(',') if s.strip()],
                                    'observaciones': registro.get('observaciones',''),
                                    'num_solicitudes': ', '.join(sorted(list(solicitudes))) if solicitudes else ''
                                }
                        except Exception:
                            datos_oficio = {
                                'no_oficio': registro.get('folio_visita',''),
                                'fecha_inspeccion': fecha_formateada or registro.get('fecha_termino') or datetime.now().strftime('%d/%m/%Y'),
                                'normas': registro.get('norma','').split(', ') if registro.get('norma') else [],
                                'empresa_visitada': registro.get('cliente',''),
                                'calle_numero': registro.get('calle_numero') or registro.get('direccion',''),
                                'colonia': registro.get('colonia',''),
                                'municipio': registro.get('municipio',''),
                                'ciudad_estado': registro.get('ciudad_estado',''),
                                'fecha_confirmacion': registro.get('fecha_inicio') or datetime.now().strftime('%d/%m/%Y'),
                                'medio_confirmacion': 'correo electrónico',
                                'inspectores': [s.strip() for s in (registro.get('supervisores_tabla') or registro.get('nfirma1') or '').split(',') if s.strip()],
                                'observaciones': registro.get('observaciones',''),
                                'num_solicitudes': ', '.join(sorted(list(solicitudes))) if solicitudes else ''
                            }

                        try:
                            mod.generar_oficio_pdf(datos_oficio, save_path)
                        except TypeError as e:
                            # Intentar recargar el módulo y reintentar: puede ocurrir si el archivo fue editado
                            try:
                                import importlib
                                importlib.reload(mod)
                                mod.generar_oficio_pdf(datos_oficio, save_path)
                            except Exception as e2:
                                messagebox.showerror('Error', f'Error generando Oficio de Comisión:\n{e2}')
                                return
                        except Exception as e:
                            messagebox.showerror('Error', f'Error generando Oficio de Comisión:\n{e}')
                            return

                    # Persistir la ruta en historial
                    try:
                        for v in self.historial.get('visitas', []):
                            if v.get('folio_visita') == folio:
                                key = 'ruta_' + nombre.replace(' ', '_').lower()
                                v[key] = save_path
                                break
                        self._guardar_historial()
                    except Exception as e:
                        print(f"⚠️ Error guardando ruta en historial: {e}")

                    messagebox.showinfo(f"{nombre} generado", f"{nombre} guardado en:\n{save_path}")
                    return
                except Exception as e:
                    messagebox.showerror("Error", f"Error generando documento {nombre}:\n{e}")
                    return
            except Exception as e:
                messagebox.showerror("Error", f"Error al procesar descarga:\n{e}")
        
        # Botón 1: Oficio de Comisión
        oficio_frame = ctk.CTkFrame(documentos_frame, fg_color=STYLE["surface"], 
                        border_width=0, border_color=STYLE["borde"], 
                        corner_radius=10)
        oficio_frame.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")
        
        # Icono grande
        ctk.CTkLabel(
            oficio_frame,
            text="📝",
            font=("Inter", 48),
            text_color=STYLE["primario"]
        ).pack(pady=(25, 15))
        
        # Nombre del documento
        ctk.CTkLabel(
            oficio_frame,
            text="OFICIO DE COMISIÓN",
            font=("Inter", 14, "bold"),
            text_color=STYLE["texto_oscuro"]
        ).pack(pady=(0, 10))
        
        # Descripción
        ctk.CTkLabel(
            oficio_frame,
            text="Documento que autoriza la comisión de inspección",
            font=("Inter", 10),
            text_color=STYLE["texto_oscuro"],
            wraplength=180,
            justify="center"
        ).pack(pady=(0, 15), padx=15)
        
        # Botón de descarga - CAMBIADO: Color secundario con texto claro
        btn_oficio = ctk.CTkButton(
            oficio_frame,
            text="Descargar",
            command=lambda: descargar_documento("oficio", "Oficio de Comisión"),
            font=("Inter", 12, "bold"),
            fg_color=STYLE["secundario"],  # Cambiado a color secundario
            hover_color="#1a1a1a",  # Hover más oscuro
            text_color=STYLE["texto_claro"],  # Cambiado a texto claro
            height=35,
            corner_radius=6
        )
        btn_oficio.pack(pady=(0, 20), padx=15, fill="x")
        
        # Botón 2: Formato de Supervisión
        formato_frame = ctk.CTkFrame(documentos_frame, fg_color=STYLE["surface"], 
                        border_width=0, border_color=STYLE["borde"], 
                        corner_radius=10)
        formato_frame.grid(row=0, column=1, padx=10, pady=5, sticky="nsew")
        
        # Icono grande
        ctk.CTkLabel(
            formato_frame,
            text="📊",
            font=("Inter", 48),
            text_color=STYLE["primario"]
        ).pack(pady=(25, 15))
        
        # Nombre del documento
        ctk.CTkLabel(
            formato_frame,
            text="FORMATO DE SUPERVISIÓN",
            font=("Inter", 14, "bold"),
            text_color=STYLE["texto_oscuro"]
        ).pack(pady=(0, 10))
        
        # Descripción
        ctk.CTkLabel(
            formato_frame,
            text="Formato para registrar observaciones de supervisión",
            font=("Inter", 10),
            text_color=STYLE["texto_oscuro"],
            wraplength=180,
            justify="center"
        ).pack(pady=(0, 15), padx=15)
        
        # Botón de descarga - CAMBIADO: Color secundario con texto claro
        btn_formato = ctk.CTkButton(
            formato_frame,
            text="Descargar",
            command=lambda: descargar_documento("formato", "Formato de Supervisión"),
            font=("Inter", 12, "bold"),
            fg_color=STYLE["secundario"],  # Cambiado a color secundario
            hover_color="#1a1a1a",  # Hover más oscuro
            text_color=STYLE["texto_claro"],  # Cambiado a texto claro
            height=35,
            corner_radius=6
        )
        btn_formato.pack(pady=(0, 20), padx=15, fill="x")
        
        # Botón 3: Acta de Inspección
        acta_frame = ctk.CTkFrame(documentos_frame, fg_color=STYLE["surface"], 
                    border_width=0, border_color=STYLE["borde"], 
                    corner_radius=10)
        acta_frame.grid(row=0, column=2, padx=10, pady=5, sticky="nsew")
        
        # Icono grande
        ctk.CTkLabel(
            acta_frame,
            text="📋",
            font=("Inter", 48),
            text_color=STYLE["primario"]
        ).pack(pady=(25, 15))
        
        # Nombre del documento
        ctk.CTkLabel(
            acta_frame,
            text="ACTA DE INSPECCIÓN",
            font=("Inter", 14, "bold"),
            text_color=STYLE["texto_oscuro"]
        ).pack(pady=(0, 10))
        
        # Descripción
        ctk.CTkLabel(
            acta_frame,
            text="Documento oficial de la visita de inspección",
            font=("Inter", 10),
            text_color=STYLE["texto_oscuro"],
            wraplength=180,
            justify="center"
        ).pack(pady=(0, 15), padx=15)
        
        # Botón de descarga - CAMBIADO: Color secundario con texto claro
        btn_acta = ctk.CTkButton(
            acta_frame,
            text="Descargar",
            command=lambda: descargar_documento("acta", "Acta de Inspección"),
            font=("Inter", 12, "bold"),
            fg_color=STYLE["secundario"],  # Cambiado a color secundario
            hover_color="#1a1a1a",  # Hover más oscuro
            text_color=STYLE["texto_claro"],  # Cambiado a texto claro
            height=35,
            corner_radius=6
        )
        btn_acta.pack(pady=(0, 20), padx=15, fill="x")
        
        # Frame para botón cerrar
        footer_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        footer_frame.pack(fill="x", pady=(20, 0))

        # --- Tile especial: Reporte Decathlon (solo para cliente Decathlon) ---
        try:
            if registro.get('cliente','').strip().upper() == 'ARTICULOS DEPORTIVOS DECATHLON SA DE CV'.upper():
                # Colocar en una fila nueva dentro de documentos_frame para que sea visible debajo de los 3 tiles
                try:
                    dec_frame = ctk.CTkFrame(documentos_frame, fg_color=STYLE["surface"], border_width=0, border_color=STYLE["borde"], corner_radius=10)
                    # Grid en fila 1 y ocupar las 3 columnas
                    dec_frame.grid(row=1, column=0, columnspan=3, padx=10, pady=(12,0), sticky='nsew')
                    # Contenido centrado
                    ctk.CTkLabel(dec_frame, text="📄", font=("Inter", 36), text_color=STYLE["primario"]).pack(pady=(8,4))
                    ctk.CTkLabel(dec_frame, text="REPORTE DECATHLON", font=("Inter", 14, "bold"), text_color=STYLE["texto_oscuro"]).pack()
                    ctk.CTkLabel(dec_frame, text="Reporte específico para Decathlon", font=("Inter", 10), text_color=STYLE["texto_oscuro"], wraplength=520, justify='center').pack(pady=(6,8), padx=10)
                    btn_dec = ctk.CTkButton(dec_frame, text="Descargar", command=lambda: descargar_documento('decathlon', 'Reporte Decathlon'), font=("Inter",12,"bold"), fg_color=STYLE['secundario'], hover_color="#1a1a1a", text_color=STYLE['texto_claro'], height=36, corner_radius=6)
                    btn_dec.pack(pady=(0,12), padx=120)
                except Exception:
                    pass
        except Exception:
            pass
        
    def guardar_folios_visita(self, folio_visita, datos_tabla, persist_counter=True):
        """Guarda los folios de una visita en un archivo JSON con formato 6 dígitos.

        Args:
            folio_visita: identificador de la visita (string).
            datos_tabla: lista de registros con campo 'FOLIO'.
            persist_counter: si True, actualizar `data/folio_counter.json` al
                mayor folio encontrado (no decrementa el contador).
        """
        try:
            if not datos_tabla:
                print(f"⚠️ No hay datos de folios para guardar en la visita {folio_visita}")

            
                return False
            
            # Preparar datos para el archivo JSON
            folios_data = []
            
            # Cargar mapeos de firmas -> nombre completo y de normas
            firmas_map = {}
            try:
                firmas_path = os.path.join(DATA_DIR, 'Firmas.json')
                if os.path.exists(firmas_path):
                    with open(firmas_path, 'r', encoding='utf-8') as ff:
                        fdata = json.load(ff) or []
                        for ent in fdata:
                            try:
                                key = (str(ent.get('FIRMA') or '').strip()).upper()
                                val = str(ent.get('NOMBRE DE INSPECTOR') or '').strip()
                                if key:
                                    firmas_map[key] = val
                            except Exception:
                                continue
            except Exception:
                firmas_map = {}

            normas_map = {}
            try:
                normas_path = os.path.join(DATA_DIR, 'Normas.json')
                if os.path.exists(normas_path):
                    with open(normas_path, 'r', encoding='utf-8') as nf:
                        ndata = json.load(nf) or []
                        import re
                        for n in ndata:
                            try:
                                nom_code = str(n.get('NOM') or '').strip()
                                m = re.search(r'NOM-(\d+)-', nom_code)
                                if m:
                                    num = int(m.group(1))
                                    normas_map[num] = nom_code
                            except Exception:
                                continue
            except Exception:
                normas_map = {}

            for item in datos_tabla:
                # Obtener y formatear el folio a 6 dígitos
                folio_raw = item.get('FOLIO', '')
                folio_formateado = ""
                
                if folio_raw is not None:
                    try:
                        # Convertir a entero y formatear a 6 dígitos
                        folio_num = int(float(str(folio_raw).strip()))
                        folio_formateado = f"{folio_num:06d}"
                    except (ValueError, TypeError):
                        # Si no se puede convertir, usar el valor original
                        folio_formateado = str(folio_raw).strip()
                
                # Obtener solicitud - buscar en varias posibles columnas
                solicitud = ""
                posibles_columnas_solicitud = ['SOLICITUD', 'SOLICITUDES', 'NO. SOLICITUD']
                for col in posibles_columnas_solicitud:
                    if col in item and item[col] is not None:
                        solicitud = str(item[col]).strip()
                        break
                
                # Extraer los campos necesarios
                folio_data = {
                    "FOLIOS": folio_formateado,
                    "MARCA": str(item.get('MARCA', '')).strip() if item.get('MARCA') else "",
                    "SOLICITUDES": solicitud,
                    "INSPECTOR": str(item.get('INSPECTOR', '') or item.get('inspector', '') or '').strip(),
                    # Supervisor: si la tabla trae 'FIRMA' (codificado), mapear a nombre completo desde Firmas.json
                    "SUPERVISOR": "",
                    "NORMA": "",
                    "LISTA": str(item.get('LISTA', '') or item.get('lista', '') or '').strip(),
                    "CODIGO": str(item.get('CODIGO', '') or item.get('codigo', '') or '').strip(),
                    "FECHA DE IMPRESION": self.entry_fecha_termino.get().strip() or datetime.now().strftime("%d/%m/%Y"),
                    "FECHA DE VERIFICACION": str(item.get('FECHA DE VERIFICACION', '')).strip() if item.get('FECHA DE VERIFICACION') else "",
                    "TIPO DE DOCUMENTO": str(item.get('TIPO DE DOCUMENTO', 'D')).strip()
                }

                # Resolver SUPERVISOR: preferir mapeo por 'FIRMA', luego 'INSPECTOR'
                try:
                    firma_code = (str(item.get('FIRMA') or '')).strip()
                    if not firma_code:
                        # some tables may store signature code under 'FIRMA' uppercase/lower
                        for k in ('FIRMA', 'firma'):
                            if k in item and item.get(k):
                                firma_code = str(item.get(k)).strip()
                                break
                    if firma_code:
                        nombre = firmas_map.get(firma_code.upper()) or firmas_map.get(firma_code)
                        if nombre:
                            folio_data['SUPERVISOR'] = nombre
                        else:
                            # fallback: if item contains inspector full name already
                            folio_data['SUPERVISOR'] = str(item.get('INSPECTOR') or item.get('inspector') or '')
                except Exception:
                    pass

                # Resolver NORMA desde 'CLASIF UVA' o 'NORMA UVA' (numérico)
                try:
                    val = None
                    for k in ('CLASIF UVA', 'CLASIF_UVA', 'NORMA UVA', 'NORMA_UVA', 'NORMA'):
                        if k in item and item.get(k) is not None:
                            val = item.get(k)
                            break
                    if val is not None and str(val).strip() != '':
                        try:
                            num = int(float(val))
                            nom = normas_map.get(int(num))
                            if nom:
                                folio_data['NORMA'] = nom
                        except Exception:
                            pass
                except Exception:
                    pass
                
                # intentar resolver LISTA desde posibles ubicaciones si no viene en item
                try:
                    lista_val = ''
                    if 'LISTA' in item and item.get('LISTA') is not None and str(item.get('LISTA')).strip() != '':
                        lista_val = str(item.get('LISTA')).strip()
                    elif 'lista' in item and item.get('lista') is not None and str(item.get('lista')).strip() != '':
                        lista_val = str(item.get('lista')).strip()
                    else:
                        # buscar en estructuras anidadas comunes (p.ej. 'identificacion':{'lista': '2'})
                        try:
                            ident = item.get('identificacion') if isinstance(item, dict) else None
                            if isinstance(ident, dict):
                                lv = ident.get('lista') or ident.get('LISTA')
                                if lv is not None and str(lv).strip() != '':
                                    lista_val = str(lv).strip()
                        except Exception:
                            lista_val = lista_val
                except Exception:
                    lista_val = ''

                folio_data['LISTA'] = lista_val

                # Agregar solo si tiene folio
                if folio_data["FOLIOS"]:
                    folios_data.append(folio_data)

            # Deduplicar por número de folio y por LISTA (mantener primer registro para cada par)
            # Algunos documentos (p.ej. Constancias) pueden compartir el mismo número de folio
            # pero pertenecer a listas diferentes; en ese caso queremos mantener ambos registros.
            seen = set()
            deduped = []
            for f in folios_data:
                fol_val = f.get('FOLIOS') or ''
                lista_val = f.get('LISTA') or ''
                if not fol_val:
                    continue
                key = (str(fol_val).strip(), str(lista_val).strip())
                if key in seen:
                    continue
                seen.add(key)
                deduped.append(f)
            folios_data = deduped
            
            if not folios_data:
                print(f"⚠️ No se encontraron folios válidos para guardar en la visita {folio_visita}")
                return False
            
            # Crear archivo JSON
            archivo_folios = os.path.join(self.folios_visita_path, f"folios_{folio_visita}.json")

            # Si ya existe un archivo de folios para esta visita, cargarlo y fusionar
            try:
                if os.path.exists(archivo_folios):
                    with open(archivo_folios, 'r', encoding='utf-8') as ef:
                        existing = json.load(ef) or {}
                        existing_list = existing.get('folios') if isinstance(existing, dict) else existing
                        if isinstance(existing_list, list) and existing_list:
                            # Prepend existing so their entries keep priority on dedupe
                            folios_data = existing_list + folios_data
            except Exception:
                # Si falla la carga, continuar con los datos nuevos
                pass

            # Capturar contador actual antes de persistir (meta) para poder
            # restaurarlo si la visita se elimina posteriormente.
            try:
                try:
                    current_counter = int(folio_manager.get_last() or 0)
                except Exception:
                    current_counter = 0
            except Exception:
                current_counter = 0

            data_to_write = {
                "_meta": {"counter_before": int(current_counter)},
                "folios": folios_data
            }

            with open(archivo_folios, 'w', encoding='utf-8') as f:
                json.dump(data_to_write, f, ensure_ascii=False, indent=2)
            
            print(f"✅ Folios guardados para visita {folio_visita}: {len(folios_data)} registros")
            # Actualizar contador legacy `folio_counter.json` con el mayor folio
            # Usamos el módulo `folio_manager` para escritura atómica y lockado.
            if persist_counter:
                try:
                    max_num = 0
                    for fdata in folios_data:
                        fol = fdata.get('FOLIOS') or ''
                        digits = ''.join([c for c in str(fol) if c.isdigit()])
                        if digits:
                            try:
                                n = int(digits)
                                if n > max_num:
                                    max_num = n
                            except Exception:
                                pass
                    if max_num > 0:
                        try:
                            try:
                                current_counter = int(folio_manager.get_last() or 0)
                            except Exception:
                                current_counter = 0
                            safe_max = max(current_counter, int(max_num or 0))
                            if safe_max > 0 and safe_max != current_counter:
                                try:
                                    folio_manager.set_last(safe_max)
                                    print(f"🔁 folio_counter.json actualizado a {safe_max:06d}")
                                except Exception:
                                    # Fallback atómico directo al archivo si folio_manager falla
                                    try:
                                        counter_path = os.path.join(DATA_DIR, 'folio_counter.json')
                                        tmp = counter_path + '.tmp'
                                        with open(tmp, 'w', encoding='utf-8') as tf:
                                            json.dump({'last': int(safe_max)}, tf)
                                        try:
                                            os.replace(tmp, counter_path)
                                        except Exception:
                                            if os.path.exists(counter_path):
                                                os.remove(counter_path)
                                            os.replace(tmp, counter_path)
                                        print(f"🔁 folio_counter.json actualizado (fallback) a {safe_max:06d}")
                                    except Exception:
                                        pass
                        except Exception:
                            pass
                except Exception:
                    pass
            return True
            
        except Exception as e:
            print(f"❌ Error guardando folios para visita {folio_visita}: {e}")
            return False

    def descargar_folios_visita(self, registro):
        """Descarga los folios de una visita en formato Excel con columnas personalizadas"""
        try:
            folio_visita = registro.get('folio_visita', '')
            if not folio_visita:
                messagebox.showwarning("Error", "No se pudo obtener el folio de la visita.")
                return
            
            # Buscar el archivo JSON de folios
            archivo_folios = os.path.join(self.folios_visita_path, f"folios_{folio_visita}.json")
            
            if not os.path.exists(archivo_folios):
                messagebox.showinfo("Sin datos", f"No se encontró archivo de folios para la visita {folio_visita}.")
                return
            
            # Cargar los datos (aceptar ambos formatos: lista antigua o dict nuevo)
            with open(archivo_folios, 'r', encoding='utf-8') as f:
                obj = json.load(f)
                if isinstance(obj, dict) and 'folios' in obj:
                    folios_data = obj.get('folios') or []
                else:
                    folios_data = obj or []
            
            if not folios_data:
                messagebox.showinfo("Sin datos", f"No hay datos de folios para la visita {folio_visita}.")
                return
            
            # Crear DataFrame con el orden de columnas específico
            df = pd.DataFrame(folios_data)
            # Garantizar columna CLIENTE incluso si no se enriquece desde la tabla de relación
            if 'CLIENTE' not in df.columns:
                df['CLIENTE'] = ''

            # Rellenar columna CLIENTE a partir del registro de la visita (o historial)
            cliente_default = ''
            try:
                cliente_default = registro.get('cliente') or registro.get('cliente_nombre') or ''
            except Exception:
                cliente_default = ''

            # Si no viene en el registro, intentar cargar desde data/historial_visitas.json
            if not cliente_default:
                try:
                    hist_path = os.path.join(DATA_DIR, 'historial_visitas.json')
                    if os.path.exists(hist_path):
                        with open(hist_path, 'r', encoding='utf-8') as hf:
                            hist = json.load(hf) or {}
                            visitas = hist.get('visitas') or []
                            for v in visitas:
                                if str(v.get('folio_visita', '')).strip() == str(folio_visita).strip():
                                    cliente_default = v.get('cliente', '') or v.get('cliente_nombre', '') or ''
                                    break
                except Exception:
                    cliente_default = ''

            # Aplicar cliente por defecto a las filas que no tengan CLIENTE
            try:
                if cliente_default:
                    # Asegurar no sobrescribir valores ya extraídos
                    df['CLIENTE'] = df['CLIENTE'].fillna('').astype(str)
                    df.loc[df['CLIENTE'].str(strip=True) == '', 'CLIENTE'] = cliente_default
            except Exception:
                # Fallback simple
                try:
                    df['CLIENTE'] = df['CLIENTE'].fillna('').replace('', cliente_default)
                except Exception:
                    pass

            # Intentar enriquecer el reporte con `LISTA` y `CODIGO` usando el backup
            try:
                # Buscar backups específicos para esta visita en data/tabla_relacion_backups
                try:
                    backups_dir = os.path.join(DATA_DIR, 'tabla_relacion_backups')
                except Exception:
                    backups_dir = os.path.join(os.path.dirname(__file__), 'data', 'tabla_relacion_backups')

                rel = None
                # Preferir archivo cuyo nombre contenga el folio de visita
                folio_key = folio_visita or ''
                folio_digits = ''.join([c for c in folio_key if c.isdigit()])

                if os.path.exists(backups_dir):
                    candidates = [os.path.join(backups_dir, f) for f in os.listdir(backups_dir) if f.lower().endswith('.json')]
                    # buscar coincidencias por nombre
                    matches_files = []
                    for p in candidates:
                        name = os.path.basename(p)
                        if folio_key and folio_key in name:
                            matches_files.append(p)
                        elif folio_digits and folio_digits in name:
                            matches_files.append(p)

                    chosen = None
                    if matches_files:
                        # elegir el más reciente entre matches
                        chosen = max(matches_files, key=os.path.getmtime)
                    elif candidates:
                        # fallback: el backup más reciente general
                        chosen = max(candidates, key=os.path.getmtime)

                    if chosen:
                        try:
                            with open(chosen, 'r', encoding='utf-8') as rf:
                                data_rel = json.load(rf)
                            rel = pd.DataFrame(data_rel)
                            print(f"📁 Usando backup de tabla de relación para reporte: {chosen}")
                        except Exception as e:
                            print(f"⚠️ Error cargando backup seleccionado: {e}")

                # Fallback a tabla principal si no se cargó backup
                if rel is None or rel.empty:
                    try:
                        df_rel = cargar_tabla_relacion()
                        if not df_rel.empty:
                            rel = df_rel.copy()
                    except Exception:
                        rel = pd.DataFrame()

                if rel is not None and not rel.empty:
                    # Normalizar nombres de columnas
                    rel_columns = [str(c).strip() for c in rel.columns]
                    rel.columns = rel_columns

                    listas_col = None
                    for c in ('LISTA', 'Lista', 'lista'):
                        if c in rel.columns:
                            listas_col = c
                            break

                    codigo_col = None
                    for c in ('CODIGO', 'Codigo', 'codigo'):
                        if c in rel.columns:
                            codigo_col = c
                            break

                        cliente_col = None
                        for c in ('CLIENTE', 'Cliente', 'cliente', 'NOMBRE CLIENTE'):
                            if c in rel.columns:
                                cliente_col = c
                                break

                    folio_col = None
                    for c in ('FOLIO', 'FOLIOS', 'folio'):
                        if c in rel.columns:
                            folio_col = c
                            break

                    solicitud_col = None
                    for c in ('SOLICITUD', 'SOLICITUDES', 'Solicitud', 'NO. SOLICITUD', 'NUMERO_SOLICITUD'):
                        if c in rel.columns:
                            solicitud_col = c
                            break

                    listas_vals = []
                    codigos_vals = []
                    clientes_vals = []

                    for _, r in df.iterrows():
                        fol = str(r.get('FOLIOS', '')).strip()
                        fol_digits = ''.join([ch for ch in fol if ch.isdigit()])
                        sol = str(r.get('SOLICITUDES', '')).strip()

                        matches = rel.copy()
                        # Filtrar por folio si existe columna (comparar como enteros para evitar ceros a la izquierda)
                        if folio_col and fol_digits:
                            try:
                                fol_int = int(fol_digits)
                                def extract_int(x):
                                    s = ''.join([ch for ch in str(x) if ch.isdigit()])
                                    return int(s) if s not in (None, '') else None
                                # Keep rows where folio_col numeric value matches fol_int
                                matches = matches[matches[folio_col].notna()]
                                matches = matches[matches[folio_col].astype(str).apply(lambda x: ''.join([ch for ch in str(x) if ch.isdigit()]) ) != '']
                                matches = matches[matches[folio_col].astype(str).apply(lambda x: extract_int(x) == fol_int if extract_int(x) is not None else False)]
                            except Exception:
                                # Fallback a comparación por dígitos como string
                                matches = matches[matches[folio_col].notna()]
                                matches = matches[matches[folio_col].astype(str).apply(lambda x: ''.join([ch for ch in str(x) if ch.isdigit()])) == fol_digits]

                        # Filtrar por solicitud si se proporcionó
                        if solicitud_col and sol:
                            matches = matches[matches[solicitud_col].astype(str).str.strip() == sol]

                        if not matches.empty:
                            listas_set = sorted({str(v).strip() for v in matches[listas_col].tolist() if listas_col and v is not None and str(v).strip() != ''}) if listas_col else []
                            # CODIGO puede contener listas separadas por comas; extraer todos
                            codigos_set = []
                            if codigo_col:
                                raw_codigos = []
                                for cv in matches[codigo_col].tolist():
                                    if cv is None:
                                        continue
                                    for part in str(cv).split(','):
                                        p = part.strip()
                                        if p:
                                            raw_codigos.append(p)
                                codigos_set = sorted(set(raw_codigos))

                            listas_vals.append(','.join(map(str, listas_set)) if listas_set else '')
                            codigos_vals.append(','.join(codigos_set) if codigos_set else '')
                            # Extraer nombre(s) de cliente desde la tabla de relación si existe
                            if cliente_col:
                                cliente_set = sorted({str(v).strip() for v in matches[cliente_col].tolist() if v is not None and str(v).strip() != ''})
                                clientes_vals.append(','.join(cliente_set) if cliente_set else '')
                            else:
                                clientes_vals.append('')
                            print(f"→ Folio {fol} | Solicitud '{sol}' → LISTA: {','.join(map(str, listas_set))} | CODIGO: {','.join(codigos_set)}")
                        else:
                            listas_vals.append('')
                            codigos_vals.append('')
                            clientes_vals.append('')
                            print(f"→ Folio {fol} | Solicitud '{sol}' → No se encontraron coincidencias en backups de tabla_de_relacion")

                    df['LISTA'] = listas_vals
                    df['CODIGO'] = codigos_vals
                    df['CLIENTE'] = clientes_vals
            except Exception as e:
                print(f"⚠️ No se pudo enriquecer reporte con tabla de relación: {e}")
            
            # Definir el orden de columnas deseado (incluir LISTA y CODIGO si existen)
            column_order = [
                "LISTA",
                "CODIGO",
                "FOLIOS",
                "CLIENTE",
                "MARCA",
                "SOLICITUDES",
                "FECHA DE IMPRESION",
                "FECHA DE VERIFICACION",
                "TIPO DE DOCUMENTO",
            ]
            
            # Reordenar columnas si existen
            existing_columns = [col for col in column_order if col in df.columns]
            df = df[existing_columns]
            
            # Preguntar donde guardar el archivo Excel
            file_path = filedialog.asksaveasfilename(
                title="Guardar archivo de folios",
                defaultextension=".xlsx",
                filetypes=[
                    ("Archivos Excel", "*.xlsx"),
                    ("Archivos Excel 97-2003", "*.xls"),
                    ("Archivos CSV", "*.csv"),
                    ("Todos los archivos", "*.*")
                ],
                initialfile=f"Folios_Visita_{folio_visita}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            
            if not file_path:
                return
            
            # Guardar en Excel con formato
            if file_path.endswith('.csv'):
                df.to_csv(file_path, index=False, encoding='utf-8-sig')
            else:
                # Usar ExcelWriter para aplicar formato
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name='Folios')
                    
                    # Obtener el libro y la hoja para aplicar formato
                    workbook = writer.book
                    worksheet = writer.sheets['Folios']
                    
                    # Ajustar ancho de columnas automáticamente
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Verificar persistencia - mantener una copia en la carpeta de respaldo
            backup_dir = os.path.join(self.folios_visita_path, "backups")
            os.makedirs(backup_dir, exist_ok=True)
            
            backup_file = os.path.join(
                backup_dir, 
                f"Folios_Visita_{folio_visita}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            # Crear copia de respaldo
            try:
                if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                    shutil.copy2(file_path, backup_file)
                    print(f"📁 Copia de respaldo creada: {backup_file}")
            except Exception as backup_error:
                print(f"⚠️ No se pudo crear copia de respaldo: {backup_error}")
            
            # Mostrar información detallada
            info_mensaje = f"""
                                ✅ Folios descargados exitosamente:

                                📁 Archivo: {os.path.basename(file_path)}
                                📋 Total de folios: {len(folios_data)}
                                📅 Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
                                📍 Ubicación: {file_path}

                                📊 Columnas incluidas:
                                • FOLIOS (formato 6 dígitos: 000001)
                                • MARCA
                                • SOLICITUDES
                                • FECHA DE IMPRESION
                                • FECHA DE VERIFICACION
                                • TIPO DE DOCUMENTO
                            """
            
            messagebox.showinfo("Descarga completada", info_mensaje)
            
            # Opcional: Abrir el archivo
            respuesta = messagebox.askyesno("Abrir archivo", "¿Desea abrir el archivo descargado?")
            if respuesta:
                self._abrir_archivo(file_path)
                
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron descargar los folios:\n{str(e)}")
    
    # ----------------- FOLIOS PENDIENTES (UI helpers) -----------------
    def _get_folios_pendientes(self):
        """Retorna lista de registros pendientes.

        Los pendientes se cargan preferentemente desde `data/pending_folios.json` si existe
        (persistencia explícita de reservas). Si no existe, se hace fallback a `historial_data`.
        """
        try:
            # Si cargamos reservas persistentes, trabajar con esa lista
            if hasattr(self, 'pending_folios') and isinstance(self.pending_folios, list):
                return list(self.pending_folios)

            pendientes = []
            fuente = getattr(self, 'historial_data', []) or []
            for r in fuente:
                est = (r.get('estatus','') or '').strip().lower()
                if est != 'pendiente':
                    continue
                pendientes.append(r)
            return pendientes
        except Exception:
            return []

    def _refresh_pending_folios_dropdown(self):
        """Actualiza los valores del combobox de folios pendientes."""
        try:
            # Intentar filtrar por el tipo de documento actualmente seleccionado
            tipo_sel = None
            try:
                tipo_sel = self.combo_tipo_documento.get().strip() if hasattr(self, 'combo_tipo_documento') else None
            except Exception:
                tipo_sel = None

            # Cargar pendientes persistentes si existen
            pendientes_source = []
            try:
                # cargar desde memoria si ya leido
                if hasattr(self, 'pending_folios') and isinstance(self.pending_folios, list):
                    pendientes_source = list(self.pending_folios)
                else:
                    # intentar leer archivo
                    pf = os.path.join(DATA_DIR, 'pending_folios.json')
                    if os.path.exists(pf):
                        with open(pf, 'r', encoding='utf-8') as f:
                            pendientes_source = json.load(f) or []
                    else:
                        pendientes_source = list(getattr(self, 'historial_data', []) or [])
            except Exception:
                pendientes_source = list(getattr(self, 'historial_data', []) or [])

            pendientes = []
            for r in pendientes_source:
                try:
                    est = (r.get('estatus','') or '').strip().lower()
                    td = (r.get('tipo_documento','') or '').strip()
                    # Mostrar todas las reservas pendientes independientemente del tipo
                    if est != 'pendiente':
                        continue
                    pendientes.append(r)
                except Exception:
                    continue
            vals = []
            pendientes_map = {}
            # Construir valores legibles e índices alternativos por folio/acta
            for i, r in enumerate(pendientes):
                fol = r.get('folio_visita','')
                act = r.get('folio_acta','')
                cliente = r.get('cliente','')
                fecha = r.get('fecha_inicio','')
                tipo = r.get('tipo_documento','') or ''
                display = f"{fol} — {tipo} — {cliente} ({fecha})"
                vals.append(display)
                # mapear la etiqueta completa
                pendientes_map[display] = r
                # mapear por folio exacto (con y sin prefijo)
                try:
                    if fol:
                        pendientes_map[fol] = r
                        # sin prefijo CP/AC
                        no_pref = fol
                        if fol.upper().startswith('CP') or fol.upper().startswith('AC'):
                            no_pref = fol[2:]
                        pendientes_map[no_pref] = r
                except Exception:
                    pass
                # mapear por folio de acta
                try:
                    if act:
                        pendientes_map[act] = r
                except Exception:
                    pass

            self._pendientes_map = pendientes_map

            if hasattr(self, 'combo_folios_pendientes'):
                try:
                    self.combo_folios_pendientes.configure(values=vals)
                    # Mostrar/ocultar la sección de folios según existan reservas
                    if vals:
                        # Asegurar que el contenedor de folios está visible
                        try:
                            if not self.cliente_folios_frame.winfo_ismapped():
                                self.cliente_folios_frame.pack(fill="x", pady=(4,4))
                        except Exception:
                            pass
                        
                        # Asegurar que los widgets están visibles
                        try:
                            if not self.lbl_folios_pendientes.winfo_ismapped():
                                self.lbl_folios_pendientes.pack(side="left", padx=(0,8))
                        except Exception:
                            pass
                        try:
                            if not self.combo_folios_pendientes.winfo_ismapped():
                                self.combo_folios_pendientes.pack(side="left", fill="x", expand=True, padx=(0, 8))
                        except Exception:
                            pass
                        try:
                            if not self.btn_desmarcar_folio.winfo_ismapped():
                                self.btn_desmarcar_folio.pack(side="left", padx=(0, 6))
                        except Exception:
                            pass
                        try:
                            if not self.btn_eliminar_folio_pendiente.winfo_ismapped():
                                self.btn_eliminar_folio_pendiente.pack(side="left")
                        except Exception:
                            pass

                        # Dejar sin seleccionar para evitar consumir automáticamente
                        try:
                            sel_id = getattr(self, 'selected_pending_id', None)
                            if sel_id:
                                display_to_set = None
                                for d in vals:
                                    r = pendientes_map.get(d)
                                    try:
                                        if r and (r.get('_id') == sel_id or r.get('id') == sel_id):
                                            display_to_set = d
                                            break
                                    except Exception:
                                        continue
                                if display_to_set:
                                    self.combo_folios_pendientes.set(display_to_set)
                                else:
                                    self.combo_folios_pendientes.set("")
                            else:
                                self.combo_folios_pendientes.set("")
                        except Exception:
                            self.combo_folios_pendientes.set(vals[0])
                        try:
                            # Intentar abrir el desplegable para que el usuario vea las opciones
                            try:
                                self.combo_folios_pendientes.focus_set()
                                self.combo_folios_pendientes.event_generate('<Button-1>')
                                self.combo_folios_pendientes.event_generate('<Down>')
                            except Exception:
                                pass
                        except Exception:
                            pass
                    else:
                        # No hay reservas: ocultar los widgets relacionados
                        try:
                            self.lbl_folios_pendientes.pack_forget()
                        except Exception:
                            pass
                        try:
                            self.combo_folios_pendientes.pack_forget()
                        except Exception:
                            pass
                        try:
                            self.btn_desmarcar_folio.pack_forget()
                        except Exception:
                            pass
                        try:
                            self.btn_eliminar_folio_pendiente.pack_forget()
                        except Exception:
                            pass
                        try:
                            self.combo_folios_pendientes.set("")
                        except Exception:
                            pass
                except Exception:
                    pass
        except Exception as e:
            print(f"Error refrescando folios pendientes: {e}")

    def _seleccionar_folio_pendiente(self, seleccionado_text):
        """Al seleccionar un folio pendiente, cargar sus datos en el formulario y marcar como En proceso."""
        try:
            if not seleccionado_text:
                return
            registro = getattr(self, '_pendientes_map', {}).get(seleccionado_text)

            # DEBUG
            try:
                print(f"[DEBUG] _seleccionar_folio_pendiente seleccionado_text='{seleccionado_text}' registro_found={bool(registro)}")
            except Exception:
                pass

            # Si el usuario escribió solo el folio (p. ej. "CP0001") o el texto no coincide
            # intentar hacer una búsqueda por folio_visita, folio_acta o substring en los valores
            if not registro:
                try:
                    pendientes = self._get_folios_pendientes()
                    # búsqueda exacta por folio
                    for r in pendientes:
                        if seleccionado_text == r.get('folio_visita') or seleccionado_text == r.get('folio_acta'):
                            registro = r
                            break
                    # búsqueda por substring en la representación mostrada
                    if not registro:
                        for k, r in getattr(self, '_pendientes_map', {}).items():
                            try:
                                if seleccionado_text.strip() and seleccionado_text.strip().lower() in str(k).lower():
                                    registro = r
                                    break
                            except Exception:
                                continue
                except Exception:
                    registro = None

            if not registro:
                messagebox.showwarning("No encontrado", "No se encontró el folio pendiente seleccionado.")
                return

            # Cargar datos en la sección Información de Visita
            try:
                self.entry_folio_visita.configure(state='normal')
                self.entry_folio_visita.delete(0, 'end')
                folio_to_set = registro.get('folio_visita','')
                # Asegurar formato CP/AC si viene sin prefijo
                if folio_to_set and not (folio_to_set.upper().startswith('CP') or folio_to_set.upper().startswith('AC')):
                    # intentar deducir si corresponde a CP
                    if folio_to_set.isdigit():
                        folio_to_set = f"CP{folio_to_set.zfill(6)[-6:]}"
                self.entry_folio_visita.insert(0, folio_to_set)
                self.entry_folio_visita.configure(state='readonly')

                self.entry_folio_acta.configure(state='normal')
                self.entry_folio_acta.delete(0, 'end')
                self.entry_folio_acta.insert(0, registro.get('folio_acta',''))
                self.entry_folio_acta.configure(state='readonly')

                self.entry_fecha_inicio.delete(0, 'end')
                self.entry_fecha_inicio.insert(0, registro.get('fecha_inicio',''))
                self.entry_fecha_termino.delete(0, 'end')
                self.entry_fecha_termino.insert(0, registro.get('fecha_termino',''))

                try:
                    self.entry_hora_inicio.configure(state='normal')
                    self.entry_hora_inicio.delete(0, 'end')
                    self.entry_hora_inicio.insert(0, registro.get('hora_inicio',''))
                    self.entry_hora_inicio.configure(state='readonly')
                except Exception:
                    pass

                try:
                    self.entry_hora_termino.configure(state='normal')
                    self.entry_hora_termino.delete(0, 'end')
                    self.entry_hora_termino.insert(0, registro.get('hora_termino',''))
                    self.entry_hora_termino.configure(state='readonly')
                except Exception:
                    pass

                cliente = registro.get('cliente')
                if cliente and hasattr(self, 'combo_cliente'):
                    try:
                        self.combo_cliente.set(cliente)
                        try:
                            self.actualizar_cliente_seleccionado(cliente)
                        except Exception:
                            pass
                    except Exception:
                        pass

            except Exception as e:
                print(f"Error cargando registro pendiente en formulario: {e}")

            # No cambiar estatus en disco todavía: solo marcar en memoria que el usuario
            # seleccionó este folio pendiente. La visita permanecerá como 'Pendiente'
            # hasta que el usuario genere los documentos (o confirme su uso).
            try:
                rid = registro.get('_id') or registro.get('id')
                # Marcar folio como seleccionado para uso posterior (no persistir estatus)
                try:
                    fv = registro.get('folio_visita','') or ''
                    num = fv
                    if fv.upper().startswith('CP') or fv.upper().startswith('AC'):
                        num = fv[2:]
                    # Mantener formato de `current_folio` como 6 dígitos sin prefijo
                    num_only = ''.join([c for c in str(num) if c.isdigit()]) or ''
                    if num_only:
                        self.current_folio = num_only.zfill(6)
                    else:
                        # si no contiene dígitos, mantener el valor existente
                        pass
                    self.selected_pending_id = rid
                    self.usando_folio_reservado = True
                    try:
                        print(f"[DEBUG] seleccionado pending id={rid} folio={fv} current_folio set to {self.current_folio}")
                    except Exception:
                        pass
                except Exception:
                    pass
            except Exception as e:
                print(f"Error preparando registro seleccionado: {e}")

            try:
                self._refresh_pending_folios_dropdown()
            except Exception:
                pass

        except Exception as e:
            print(f"Error al seleccionar folio pendiente: {e}")

    def _desmarcar_folio_seleccionado(self):
        """Desmarca la selección actual sin eliminar la reserva persistente."""
        try:
            self.selected_pending_id = None
            self.usando_folio_reservado = False
            # Restaurar current_folio al cálculo normal (recalcular)
            try:
                self.cargar_ultimo_folio()
            except Exception:
                pass
            # refrescar UI
            try:
                self._refresh_pending_folios_dropdown()
            except Exception:
                pass

            # Actualizar etiqueta de siguiente folio en la UI sin mostrar popup
            try:
                self._update_siguiente_folio_label()
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo desmarcar la selección: {e}")

        except Exception as e:
            print(f"Error al seleccionar folio pendiente: {e}")

    def _eliminar_folio_pendiente(self):
        """Eliminar el folio pendiente seleccionado (con confirmación)."""
        try:
            seleccionado_text = None
            try:
                seleccionado_text = self.combo_folios_pendientes.get()
            except Exception:
                pass
            if not seleccionado_text:
                messagebox.showwarning("Seleccionar folio", "Seleccione un folio pendiente para eliminar.")
                return

            registro = getattr(self, '_pendientes_map', {}).get(seleccionado_text)
            # Intento fallback si el usuario solo escribió el folio
            if not registro:
                try:
                    pendientes = self._get_folios_pendientes()
                    for r in pendientes:
                        if seleccionado_text == r.get('folio_visita') or seleccionado_text == r.get('folio_acta'):
                            registro = r
                            break
                    if not registro:
                        for k, r in getattr(self, '_pendientes_map', {}).items():
                            if seleccionado_text.strip() and seleccionado_text.strip() in k:
                                registro = r
                                break
                except Exception:
                    registro = None

            if not registro:
                messagebox.showwarning("No encontrado", "No se encontró el folio seleccionado.")
                return

            if not messagebox.askyesno("Confirmar eliminación", f"¿Eliminar el folio pendiente {registro.get('folio_visita')}? Esta acción no se puede deshacer."):
                return

            try:
                self.hist_eliminar_registro(registro)
            except Exception:
                try:
                    self.historial['visitas'] = [v for v in self.historial.get('visitas', []) if v.get('folio_visita') != registro.get('folio_visita')]
                    self._guardar_historial()
                except Exception as e:
                    print(f"Error eliminando folio pendiente manualmente: {e}")

            # También eliminar de archivo de reservas si existe
            try:
                pf_path = os.path.join(DATA_DIR, 'pending_folios.json')
                if os.path.exists(pf_path):
                    with open(pf_path, 'r', encoding='utf-8') as f:
                        arr = json.load(f) or []
                    arr = [p for p in arr if p.get('folio_visita') != registro.get('folio_visita')]
                    with open(pf_path, 'w', encoding='utf-8') as f:
                        json.dump(arr, f, ensure_ascii=False, indent=2)
                    # actualizar memoria
                    self.pending_folios = arr
            except Exception:
                pass

            try:
                self._refresh_pending_folios_dropdown()
            except Exception:
                pass

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo eliminar el folio pendiente: {e}")

    # Agregar este método para buscar la columna correcta de solicitud
    def _obtener_columna_solicitud(self, df):
        """Busca la columna correcta que contiene las solicitudes"""
        posibles_nombres = ['SOLICITUD', 'SOLICITUDES', 'NO. SOLICITUD', 'NO SOLICITUD', 'SOLICITUD NO.', 'NÚMERO DE SOLICITUD']
        
        for nombre in posibles_nombres:
            if nombre in df.columns:
                return nombre
        
        # Si no encuentra ninguna, buscar columnas que contengan "solicitud" (case insensitive)
        for col in df.columns:
            if isinstance(col, str) and 'solicitud' in col.lower():
                return col
        
        return None

    def _abrir_archivo(self, file_path):
        """Abre un archivo en el sistema operativo correspondiente"""
        try:
            if os.path.exists(file_path):
                if os.name == 'nt':  # Windows
                    os.startfile(file_path)
                elif os.name == 'posix':  # macOS o Linux
                    if sys.platform == 'darwin':  # macOS
                        subprocess.Popen(['open', file_path])
                    else:  # Linux
                        subprocess.Popen(['xdg-open', file_path])
        except Exception as e:
            print(f"Error abriendo archivo: {e}")

    # ----------------- PERSISTENCIA DE RESERVAS -----------------
    def _load_pending_folios(self):
        """Carga las reservas desde data/pending_folios.json en `self.pending_folios`."""
        try:
            arr = self._read_pending_folios_disk()
            self.pending_folios = arr
        except Exception as e:
            print(f"[WARN] Error cargando pending_folios.json: {e}")
            self.pending_folios = []

    def _save_pending_folios(self):
        """Guarda `self.pending_folios` en data/pending_folios.json con lock y escritura atómica."""
        pf = os.path.join(DATA_DIR, 'pending_folios.json')
        lock_path = pf + '.lock'
        fd = None
        try:
            fd = self._acquire_file_lock(lock_path, timeout=3.0)
            tmp_path = pf + '.tmp'
            with open(tmp_path, 'w', encoding='utf-8') as f:
                json.dump(self.pending_folios or [], f, ensure_ascii=False, indent=2)
                f.flush()
                try:
                    os.fsync(f.fileno())
                except Exception:
                    pass
            try:
                os.replace(tmp_path, pf)
            except Exception:
                try:
                    shutil.copy2(tmp_path, pf)
                except Exception:
                    pass
            try:
                self._pending_folios_last = json.dumps(self.pending_folios or [], sort_keys=True)
            except Exception:
                pass
        except Exception as e:
            print(f"[WARN] Error guardando pending_folios.json: {e}")
        finally:
            try:
                if fd:
                    self._release_file_lock(fd, lock_path)
            except Exception:
                pass

    def _read_pending_folios_disk(self, timeout=2.0):
        """Lee `pending_folios.json` desde disco usando lock. Devuelve lista de dicts."""
        pf = os.path.join(DATA_DIR, 'pending_folios.json')
        lock_path = pf + '.lock'
        try:
            fd = self._acquire_file_lock(lock_path, timeout=timeout)
            try:
                if os.path.exists(pf):
                    with open(pf, 'r', encoding='utf-8') as f:
                        arr = json.load(f) or []
                        return [r for r in arr if isinstance(r, dict)]
                else:
                    return []
            finally:
                try:
                    self._release_file_lock(fd, lock_path)
                except Exception:
                    pass
        except Exception:
            # Fallback: leer sin lock si no se pudo adquirir
            try:
                if os.path.exists(pf):
                    with open(pf, 'r', encoding='utf-8') as f:
                        arr = json.load(f) or []
                        return [r for r in arr if isinstance(r, dict)]
            except Exception:
                pass
            return []

    def _start_pending_folios_watcher(self):
        """Inicia un polling cada 5s para detectar cambios externos en pending_folios.json."""
        try:
            self._pending_folios_last = json.dumps(self.pending_folios or [], sort_keys=True)
        except Exception:
            self._pending_folios_last = None
        try:
            self.after(5000, self._pending_folios_watcher_tick)
        except Exception:
            pass

    def _pending_folios_watcher_tick(self):
        try:
            new = self._read_pending_folios_disk(timeout=1.0)
            try:
                new_s = json.dumps(new or [], sort_keys=True)
            except Exception:
                new_s = None
            if new_s != getattr(self, '_pending_folios_last', None):
                self.pending_folios = new
                try:
                    self._refresh_pending_folios_dropdown()
                except Exception:
                    pass
                self._pending_folios_last = new_s
        except Exception:
            pass
        finally:
            try:
                self.after(5000, self._pending_folios_watcher_tick)
            except Exception:
                pass

    # ----------------- Config y export persistente -----------------
    def _cargar_config_exportacion(self):
        """Carga o crea la configuración persistente para las exportaciones Excel."""
        try:
            data_folder = DATA_DIR
            os.makedirs(data_folder, exist_ok=True)
            cfg_path = os.path.join(data_folder, 'excel_export_config.json')
            if not os.path.exists(cfg_path):
                # Contenido por defecto
                default = {
                    "tabla_de_relacion": os.path.join(data_folder, 'tabla_de_relacion.json'),
                    "tabla_backups_dir": os.path.join(data_folder, 'tabla_relacion_backups'),
                    "clientes": os.path.join(data_folder, 'Clientes.json'),
                    "export_cache": os.path.join(data_folder, 'excel_export_data.json')
                }
                with open(cfg_path, 'w', encoding='utf-8') as f:
                    json.dump(default, f, ensure_ascii=False, indent=2)
                self.excel_export_config = default
            else:
                with open(cfg_path, 'r', encoding='utf-8') as f:
                    self.excel_export_config = json.load(f)
            # Normalizar rutas: si alguna apunta a una ubicación inexistente,
            # buscarla en APP_DIR/data y reemplazarla para usar la carpeta externa.
            try:
                for key in ('tabla_de_relacion', 'clientes', 'export_cache', 'tabla_backups_dir'):
                    val = self.excel_export_config.get(key)
                    if val and not os.path.exists(val):
                        candidate = os.path.join(DATA_DIR, os.path.basename(val))
                        if os.path.exists(candidate):
                            self.excel_export_config[key] = candidate
                        else:
                            # special case: backups dir should be a directory
                            if key == 'tabla_backups_dir':
                                candidate_dir = os.path.join(DATA_DIR, os.path.basename(val) if os.path.basename(val) else 'tabla_relacion_backups')
                                self.excel_export_config[key] = candidate_dir
            except Exception:
                pass
            # Ensure directories exist
            os.makedirs(os.path.dirname(self.excel_export_config.get('tabla_de_relacion','') or data_folder), exist_ok=True)
            os.makedirs(self.excel_export_config.get('tabla_backups_dir', data_folder), exist_ok=True)
        except Exception as e:
            print(f"Error cargando config exportacion: {e}")
            self.excel_export_config = {}

    def _guardar_config_exportacion(self):
        try:
            data_folder = DATA_DIR
            os.makedirs(data_folder, exist_ok=True)
            cfg_path = os.path.join(data_folder, 'excel_export_config.json')
            with open(cfg_path, 'w', encoding='utf-8') as f:
                json.dump(self.excel_export_config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"Error guardando config exportacion: {e}")

    def _generar_datos_exportable(self):
        """Genera y persiste un JSON consolidado que será la fuente para las exportaciones EMA y anual."""
        try:
            data_folder = DATA_DIR
            os.makedirs(data_folder, exist_ok=True)
            tabla_path = self.excel_export_config.get('tabla_de_relacion') or os.path.join(data_folder, 'tabla_de_relacion.json')
            clientes_path = self.excel_export_config.get('clientes') or os.path.join(data_folder, 'Clientes.json')
            export_cache = self.excel_export_config.get('export_cache') or os.path.join(data_folder, 'excel_export_data.json')

            # Cargar tabla de relación
            tabla = []
            if os.path.exists(tabla_path):
                try:
                    with open(tabla_path, 'r', encoding='utf-8') as f:
                        tabla = json.load(f)
                except Exception:
                    tabla = []

            # Cargar historial (ya en self.historial_data)
            visitas = getattr(self, 'historial_data', [])

            # Cargar clientes para enriquecer
            clientes = {}
            if os.path.exists(clientes_path):
                try:
                    with open(clientes_path, 'r', encoding='utf-8') as f:
                        cl = json.load(f)
                        if isinstance(cl, list):
                            for c in cl:
                                clientes[c.get('CLIENTE','').upper()] = c
                except Exception:
                    pass

            # Preparar estructura
            ema_rows = []
            for r in tabla:
                try:
                    cliente = r.get('EMPRESA','') or r.get('EMPRESA_VISITADA', r.get('CLIENTE',''))
                    cliente_key = (cliente or '').upper()
                    cliente_info = clientes.get(cliente_key, {})
                    # Enriquecer como en generar_reporte_ema
                    solicitud_full = r.get('ENCABEZADO', '') or r.get('SOLICITUD_ENCABEZADO', '') or r.get('SOLICITUD','')
                    sol_parts = str(solicitud_full).split()[-1] if solicitud_full else ''
                    ema_rows.append({
                        'NUMERO_SOLICITUD': sol_parts,
                        'CLIENTE': cliente,
                        'NUMERO_CONTRATO': cliente_info.get('NÚMERO_DE_CONTRATO',''),
                        'RFC': cliente_info.get('RFC',''),
                        'CURP': cliente_info.get('CURP','N/A') or 'N/A',
                        'PRODUCTO_VERIFICADO': r.get('DESCRIPCION',''),
                        'MARCAS': r.get('MARCA',''),
                        'NOM': r.get('CLASIF UVA') or r.get('CLASIF_UVA') or r.get('NOM',''),
                        'TIPO_DOCUMENTO': r.get('TIPO DE DOCUMENTO') or r.get('TIPO_DE_DOCUMENTO',''),
                        'DOCUMENTO_EMITIDO': solicitud_full,
                        'FECHA_DOCUMENTO_EMITIDO': r.get('FECHA DE VERIFICACION') or r.get('FECHA_DE_VERIFICACION') or '',
                        'VERIFICADOR': r.get('VERIFICADOR') or r.get('INSPECTOR',''),
                        'PEDIMENTO_IMPORTACION': r.get('PEDIMENTO',''),
                        'FECHA_DESADUANAMIENTO': r.get('FECHA DE ENTRADA') or r.get('FECHA_ENTRADA',''),
                        'MODELOS': r.get('CODIGO',''),
                        'FOLIO_EMA': str(r.get('FOLIO','')).zfill(6) if str(r.get('FOLIO','')).strip() else ''
                    })
                except Exception:
                    continue

            anual_rows = []
            for v in visitas:
                try:
                    anual_rows.append({
                        'FECHA_VISITA': v.get('fecha_termino') or v.get('fecha_inicio'),
                        'FOLIO_VISITA': v.get('folio_visita',''),
                        'CLIENTE': v.get('cliente',''),
                        'SOLICITUD': v.get('solicitud',''),
                        'FOLIOS_USADOS': v.get('folios_utilizados',''),
                        'NUM_SOLICITUDES': v.get('num_solicitudes',''),
                        'NORMAS': v.get('norma','')
                    })
                except Exception:
                    continue

            export_data = {
                'ema': ema_rows,
                'anual': anual_rows,
                'generated_at': datetime.now().isoformat()
            }

            # Guardar cache exportable
            try:
                with open(export_cache, 'w', encoding='utf-8') as f:
                    json.dump(export_data, f, ensure_ascii=False, indent=2)
            except Exception as e:
                print(f"Error guardando export cache: {e}")

            return export_data
        except Exception as e:
            print(f"Error generando datos exportable: {e}")
            return {}
    
    def descargar_excel_ema(self, registro=None):
        """Descarga el reporte EMA en Excel"""
        try:
            # Cargar el módulo control_folios_anual dinámicamente
            import importlib.util
            
            excel_gen_file = os.path.join(self.documentos_dir, 'control_folios_anual.py')
            
            if not os.path.exists(excel_gen_file):
                messagebox.showerror("Error", f"No se encontró el archivo generador de Excel: {excel_gen_file}")
                return
            
            spec = importlib.util.spec_from_file_location('control_folios_anual', excel_gen_file)
            excel_mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(excel_mod)
            
            # Preparar rutas (usar config persistente si existe)
            tabla_de_relacion_path = self.excel_export_config.get('tabla_de_relacion') if hasattr(self, 'excel_export_config') else os.path.join(self.documentos_dir, 'tabla_de_relacion.json')
            
            # Pedir ruta de guardado
            file_path = filedialog.asksaveasfilename(
                title="Guardar Reporte EMA",
                defaultextension=".xlsx",
                filetypes=[
                    ("Archivos Excel", "*.xlsx"),
                    ("Archivos Excel 97-2003", "*.xls"),
                    ("Todos los archivos", "*.*")
                ],
                initialfile=f"Reporte_EMA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not file_path:
                return
            
            # Si existe cache exportable, usar su sección 'ema' para generar el archivo
            export_cache = None
            if hasattr(self, 'excel_export_config'):
                export_cache = self.excel_export_config.get('export_cache')

            if export_cache and os.path.exists(export_cache):
                try:
                    with open(export_cache, 'r', encoding='utf-8') as f:
                        ec = json.load(f)
                    ema_list = ec.get('ema') if isinstance(ec, dict) else None
                    if ema_list is not None:
                        tmp_path = os.path.join(os.path.dirname(export_cache), f"_tmp_ema_{int(datetime.now().timestamp())}.json")
                        with open(tmp_path, 'w', encoding='utf-8') as tf:
                            json.dump(ema_list, tf, ensure_ascii=False, indent=2)
                        tabla_de_relacion_path_to_use = tmp_path
                    else:
                        tabla_de_relacion_path_to_use = tabla_de_relacion_path
                except Exception:
                    tabla_de_relacion_path_to_use = tabla_de_relacion_path
            else:
                tabla_de_relacion_path_to_use = tabla_de_relacion_path

            excel_mod.generar_reporte_ema(
                tabla_de_relacion_path_to_use,
                self.historial_path,
                file_path,
                export_cache=export_cache if hasattr(self, 'excel_export_config') else None
            )
            
            messagebox.showinfo("Éxito", f"Reporte EMA generado exitosamente:\n{file_path}")
            
            # Preguntar si abrir el archivo
            if messagebox.askyesno("Abrir archivo", "¿Desea abrir el archivo descargado?"):
                self._abrir_archivo(file_path)
                
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el reporte EMA:\n{str(e)}")
    
    def descargar_excel_anual(self, registro=None):
        """Descarga el reporte de control de folios anual en Excel"""
        try:
            # Cargar el módulo control_folios_anual dinámicamente
            import importlib.util
            
            excel_gen_file = os.path.join(self.documentos_dir, 'control_folios_anual.py')
            
            if not os.path.exists(excel_gen_file):
                messagebox.showerror("Error", f"No se encontró el archivo generador de Excel: {excel_gen_file}")
                return
            
            spec = importlib.util.spec_from_file_location('control_folios_anual', excel_gen_file)
            excel_mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(excel_mod)
            
            # Obtener el año actual
            year = datetime.now().year
            # Mostrar diálogo para seleccionar rango de fechas (opcional)
            def _pedir_rango_fechas_default():
                # Modal para pedir start/end date
                modal = ctk.CTkToplevel(self)
                modal.title("Rango para Control Anual")
                modal.geometry("420x180")
                modal.transient(self)
                modal.grab_set()

                ctk.CTkLabel(modal, text="Seleccione el rango de fechas (dd/mm/YYYY)\nDejar vacío para año completo:", anchor="w").pack(pady=(12,6), padx=12)

                frame = ctk.CTkFrame(modal, fg_color="transparent")
                frame.pack(fill="x", padx=12)

                ctk.CTkLabel(frame, text="Fecha inicio:").grid(row=0, column=0, sticky="w", padx=(0,6))
                ent_start = ctk.CTkEntry(frame, width=180)
                ent_start.grid(row=0, column=1, pady=6)

                ctk.CTkLabel(frame, text="Fecha fin:").grid(row=1, column=0, sticky="w", padx=(0,6))
                ent_end = ctk.CTkEntry(frame, width=180)
                ent_end.grid(row=1, column=1, pady=6)

                # Pre-fill with year bounds
                ent_start.insert(0, f"01/01/{year}")
                ent_end.insert(0, f"31/12/{year}")

                result = {"start": None, "end": None}

                def _aceptar():
                    s = ent_start.get().strip()
                    e = ent_end.get().strip()
                    result['start'] = s if s else None
                    result['end'] = e if e else None
                    modal.destroy()

                def _cancelar():
                    result['start'] = None
                    result['end'] = None
                    modal.destroy()

                btn_frame = ctk.CTkFrame(modal, fg_color="transparent")
                btn_frame.pack(fill="x", pady=8, padx=12)
                ctk.CTkButton(btn_frame, text="Aceptar", command=_aceptar, width=100).pack(side="right", padx=6)
                ctk.CTkButton(btn_frame, text="Cancelar", command=_cancelar, width=100).pack(side="right", padx=6)

                self.wait_window(modal)
                return result['start'], result['end']

            start_date, end_date = _pedir_rango_fechas_default()

            # Si el usuario no modificó los valores (usó los defaults del año),
            # tratarlos como "sin filtro" para mantener el comportamiento previo
            # (antes se generaba todo el año; ahora evitamos filtrar por defecto).
            try:
                default_start = f"01/01/{year}"
                default_end = f"31/12/{year}"
                if start_date == default_start and end_date == default_end:
                    start_date = None
                    end_date = None
            except Exception:
                pass

            # Pedir ruta de guardado
            file_path = filedialog.asksaveasfilename(
                title="Guardar Control de Folios Anual",
                defaultextension=".xlsx",
                filetypes=[
                    ("Archivos Excel", "*.xlsx"),
                    ("Archivos Excel 97-2003", "*.xls"),
                    ("Todos los archivos", "*.*")
                ],
                initialfile=f"Control_Folios_Anual_{year}_{datetime.now().strftime('%H%M%S')}.xlsx"
            )
            
            if not file_path:
                return
            
            # Generar el reporte anual (usar backups configurados si existen)
            tabla_backups_dir = self.excel_export_config.get('tabla_backups_dir') if hasattr(self, 'excel_export_config') else os.path.join(self.documentos_dir, 'tabla_relacion_backups')

            # Si existe cache exportable, usar su sección 'anual' para alimentar el generador
            export_cache = None
            if hasattr(self, 'excel_export_config'):
                export_cache = self.excel_export_config.get('export_cache')

            if export_cache and os.path.exists(export_cache):
                try:
                    with open(export_cache, 'r', encoding='utf-8') as f:
                        ec = json.load(f)
                    anual_list = ec.get('anual') if isinstance(ec, dict) else None
                    if anual_list is not None:
                        # No escribir archivo temporal: pasamos la lista directamente
                        historial_path_to_use = self.historial_path
                        historial_list_to_pass = anual_list
                    else:
                        historial_path_to_use = self.historial_path
                        historial_list_to_pass = None
                except Exception:
                    historial_path_to_use = self.historial_path
                    historial_list_to_pass = None
            else:
                historial_path_to_use = self.historial_path
                historial_list_to_pass = None
            excel_mod.generar_control_folios_anual(
                historial_path_to_use,
                tabla_backups_dir,
                file_path,
                year,
                start_date=start_date,
                end_date=end_date,
                export_cache=export_cache,
                historial_list=historial_list_to_pass
            )
            
            messagebox.showinfo("Éxito", f"Control de Folios Anual generado exitosamente:\n{file_path}")
            
            # Preguntar si abrir el archivo
            if messagebox.askyesno("Abrir archivo", "¿Desea abrir el archivo descargado?"):
                self._abrir_archivo(file_path)
                
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el control de folios anual:\n{str(e)}")

    def hist_editar_registro(self, registro):
        """Abre el formulario para editar un registro del historial"""
        self._crear_formulario_visita(registro)

    def hist_buscar_general(self, event=None):
        """Buscar en el historial por cualquier campo"""
        try:
            # resetear paginado al buscar
            self.HISTORIAL_PAGINA_ACTUAL = 1
            # Asegurarse de que los datos estén cargados
            if not hasattr(self, 'historial_data') or not self.historial_data:
                self._cargar_historial()
                
            # Guardar copia original si no existe
            if not hasattr(self, 'historial_data_original') or not self.historial_data_original:
                self.historial_data_original = self.historial_data.copy()
            
            # Obtener filtro por cliente desde el combo (reemplaza búsqueda general)
            busqueda_raw = ''
            try:
                if getattr(self, 'combo_filtrar_cliente', None):
                    busqueda_raw = (self.combo_filtrar_cliente.get() or '').strip()
                elif getattr(self, 'entry_buscar_general', None):
                    busqueda_raw = (self.entry_buscar_general.get() or '').strip()
            except Exception:
                busqueda_raw = ''
            # Normalizar (quitar acentos) y bajar a minúsculas para comparaciones
            def _norm(s):
                try:
                    s2 = str(s)
                    s2 = unicodedata.normalize('NFKD', s2).encode('ASCII', 'ignore').decode('ASCII')
                    return s2.lower()
                except Exception:
                    return str(s).lower()

            busqueda = _norm(busqueda_raw)

            if not busqueda_raw:
                # Si no hay búsqueda, mostrar todos los datos
                self.historial_data = self.historial_data_original.copy()
            else:
                # Filtrar datos
                resultados = []
                for registro in self.historial_data_original:
                    # Buscar en todos los campos relevantes (añadir supervisor y tipo de documento)
                    campos_busqueda = [
                        registro.get('folio_visita', ''),
                        registro.get('folio_acta', ''),
                        registro.get('fecha_inicio', ''),
                        registro.get('fecha_termino', ''),
                        registro.get('cliente', ''),
                        registro.get('estatus', ''),
                        registro.get('folios_utilizados', ''),
                        registro.get('nfirma1', ''),
                        registro.get('nfirma2', ''),
                        registro.get('supervisor', ''),
                        registro.get('tipo_documento', '')
                    ]

                    matched = False
                    # búsqueda tradicional (substring en texto)
                    for campo in campos_busqueda:
                        try:
                            if busqueda in _norm(campo):
                                matched = True
                                break
                        except Exception:
                            continue

                    # Si no coincidió, intentar comparar solo dígitos (útil para folios con padding)
                    if not matched:
                        digits_search = ''.join([c for c in busqueda_raw if c.isdigit()])
                        if digits_search:
                            for campo in campos_busqueda:
                                campo_digits = ''.join([c for c in str(campo) if c.isdigit()])
                                if campo_digits and digits_search in campo_digits:
                                    matched = True
                                    break

                    if matched:
                        # aplicar filtros adicionales: supervisor, tipo, estatus, rango de fecha
                        acepta = True
                        # Si hay filtro por cliente, exigir que coincida
                        try:
                            if getattr(self, 'combo_filtrar_cliente', None):
                                cliente_sel = (self.combo_filtrar_cliente.get() or '').strip().lower()
                                if cliente_sel:
                                    reg_cli = str(registro.get('cliente','') or '')
                                    if cliente_sel not in _norm(reg_cli):
                                        acepta = False
                        except Exception:
                            pass
                        # supervisor
                        try:
                            sup_f = ''
                            if getattr(self, 'combo_filtrar_supervisor', None):
                                sup_f = (self.combo_filtrar_supervisor.get() or '').strip().lower()
                            if sup_f:
                                reg_sup = str(registro.get('supervisor','') or '')
                                if sup_f not in _norm(reg_sup):
                                    aceita_tmp = False
                                    aceita_tmp = True
                                if sup_f and sup_f not in _norm(reg_sup):
                                    acepta = False
                        except Exception:
                            pass
                        # tipo de documento
                        try:
                            tipo_f = ''
                            if getattr(self, 'combo_filtrar_tipo', None):
                                tipo_f = (self.combo_filtrar_tipo.get() or '').strip().lower()
                            if tipo_f:
                                reg_tipo = str(registro.get('tipo_documento','') or '')
                                if tipo_f not in _norm(reg_tipo):
                                    acepta = False
                        except Exception:
                            pass
                        # estatus
                        try:
                            est_f = ''
                            if getattr(self, 'combo_filtrar_estatus', None):
                                est_f = (self.combo_filtrar_estatus.get() or '').strip().lower()
                            if est_f:
                                reg_est = str(registro.get('estatus','') or '')
                                if est_f not in _norm(reg_est):
                                    acepta = False
                        except Exception:
                            pass
                        # rango de fechas
                        try:
                            fd_raw = ''
                            fh_raw = ''
                            if getattr(self, 'entry_hist_fecha_desde', None):
                                fd_raw = (self.entry_hist_fecha_desde.get() or '').strip()
                            if getattr(self, 'entry_hist_fecha_hasta', None):
                                fh_raw = (self.entry_hist_fecha_hasta.get() or '').strip()
                            if fd_raw or fh_raw:
                                # intentar parsear fecha_inicio del registro
                                reg_fecha_s = registro.get('fecha_inicio') or registro.get('fecha_creacion') or ''
                                reg_fecha_norm = self._normalize_fecha_str(reg_fecha_s) if hasattr(self, '_normalize_fecha_str') else reg_fecha_s
                                try:
                                    reg_dt = datetime.strptime(reg_fecha_norm, "%d/%m/%Y")
                                except Exception:
                                    reg_dt = None
                                fd_dt = None
                                fh_dt = None
                                try:
                                    if fd_raw:
                                        fd_norm = self._normalize_fecha_str(fd_raw) if hasattr(self, '_normalize_fecha_str') else fd_raw
                                        fd_dt = datetime.strptime(fd_norm, "%d/%m/%Y")
                                except Exception:
                                    fd_dt = None
                                try:
                                    if fh_raw:
                                        fh_norm = self._normalize_fecha_str(fh_raw) if hasattr(self, '_normalize_fecha_str') else fh_raw
                                        fh_dt = datetime.strptime(fh_norm, "%d/%m/%Y")
                                except Exception:
                                    fh_dt = None
                                if reg_dt is not None:
                                    if fd_dt and reg_dt < fd_dt:
                                        acepta = False
                                    if fh_dt and reg_dt > fh_dt:
                                        acepta = False
                        except Exception:
                            pass

                        if acepta:
                            resultados.append(registro)
                
                self.historial_data = resultados
            
            self._poblar_historial_ui()
            
        except Exception as e:
            print(f"Error en búsqueda general: {e}")

    def hist_limpiar_busqueda(self):
        """Limpiar todas las búsquedas y mostrar todos los registros"""
        # limpiar filtro de cliente (combo) o entrada general si existe
        try:
            if getattr(self, 'combo_filtrar_cliente', None):
                try:
                    self.combo_filtrar_cliente.set("")
                except Exception:
                    pass
            elif getattr(self, 'entry_buscar_general', None):
                try:
                    self.entry_buscar_general.delete(0, 'end')
                except Exception:
                    try:
                        self.entry_buscar_general.delete(0, 'end')
                    except Exception:
                        pass
        except Exception:
            pass
        self.entry_buscar_folio.delete(0, 'end')
        # limpiar filtros adicionales si existen
        try:
            if getattr(self, 'combo_filtrar_supervisor', None):
                try:
                    self.combo_filtrar_supervisor.set("")
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'combo_filtrar_tipo', None):
                try:
                    self.combo_filtrar_tipo.set("")
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'combo_filtrar_estatus', None):
                try:
                    self.combo_filtrar_estatus.set("")
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'entry_hist_fecha_desde', None):
                try:
                    self.entry_hist_fecha_desde.delete(0, 'end')
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'entry_hist_fecha_hasta', None):
                try:
                    self.entry_hist_fecha_hasta.delete(0, 'end')
                except Exception:
                    pass
        except Exception:
            pass
        
        # Recargar datos originales y resetear paginado
        self.HISTORIAL_PAGINA_ACTUAL = 1
        if hasattr(self, 'historial_data_original'):
            self.historial_data = self.historial_data_original.copy()
        else:
            self._cargar_historial()
            
        self._poblar_historial_ui()

    def hist_limpiar_filtros(self):
        """Limpiar solo los filtros (no borra el campo 'Folio visita')."""
        try:
            if getattr(self, 'combo_filtrar_supervisor', None):
                try:
                    self.combo_filtrar_supervisor.set("")
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'combo_filtrar_tipo', None):
                try:
                    self.combo_filtrar_tipo.set("")
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'combo_filtrar_estatus', None):
                try:
                    self.combo_filtrar_estatus.set("")
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'entry_hist_fecha_desde', None):
                try:
                    self.entry_hist_fecha_desde.delete(0, 'end')
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'entry_hist_fecha_hasta', None):
                try:
                    self.entry_hist_fecha_hasta.delete(0, 'end')
                except Exception:
                    pass
        except Exception:
            pass
        try:
            if getattr(self, 'combo_filtrar_cliente', None):
                try:
                    self.combo_filtrar_cliente.set("")
                except Exception:
                    pass
            elif getattr(self, 'entry_buscar_general', None):
                try:
                    self.entry_buscar_general.delete(0, 'end')
                except Exception:
                    pass
        except Exception:
            pass

        # Resetear paginado y refrescar búsqueda con filtros limpias
        self.HISTORIAL_PAGINA_ACTUAL = 1
        try:
            self.hist_buscar_general()
        except Exception:
            try:
                # fallback: repoblar con los datos originales
                if hasattr(self, 'historial_data_original'):
                    self.historial_data = self.historial_data_original.copy()
                self._poblar_historial_ui()
            except Exception:
                pass

    def _eliminar_archivos_asociados_folio(self, folio, registro=None):
        """Elimina todos los archivos asociados a un folio de forma segura.
        Si se proporciona `registro`, se usa como fallback para extraer folios o solicitudes.
        Devuelve un dict con claves `eliminados` y `errores`.
        """
        resultados = {"eliminados": [], "errores": []}

        folios_a_eliminar = set()
        solicitudes_a_eliminar = set()

        try:
            # 1) Extraer de folios_{folio}.json si existe
            folios_visita_dir = getattr(self, 'folios_visita_path', os.path.join(DATA_DIR, 'folios_visitas'))
            folio_file = os.path.join(folios_visita_dir, f"folios_{folio}.json")
            if os.path.exists(folio_file):
                try:
                    with open(folio_file, 'r', encoding='utf-8') as ff:
                        fv = json.load(ff) or []
                        for entry in fv:
                            try:
                                fval = str(entry.get('FOLIOS') or entry.get('FOLIO') or '').strip()
                                if fval:
                                    digits = ''.join([c for c in fval if c.isdigit()])
                                    if digits:
                                        folios_a_eliminar.add(digits.lstrip('0') or '0')
                                        folios_a_eliminar.add(digits.zfill(6))
                                    else:
                                        folios_a_eliminar.add(fval)
                                sol = entry.get('SOLICITUDES') or entry.get('SOLICITUD') or entry.get('SOLICITUDE') or ''
                                sol = str(sol).strip()
                                if sol:
                                    solicitudes_a_eliminar.add(sol)
                                    sd = ''.join([c for c in sol if c.isdigit()])
                                    if sd:
                                        solicitudes_a_eliminar.add(sd)
                                        solicitudes_a_eliminar.add(sd.lstrip('0'))
                                        solicitudes_a_eliminar.add(sd.zfill(6))
                            except Exception:
                                continue
                except Exception as e:
                    resultados['errores'].append(f"Error leyendo {folio_file}: {e}")
                # intentar eliminar el archivo de folios
                try:
                    os.remove(folio_file)
                    resultados['eliminados'].append(f"Archivo de folios: {os.path.basename(folio_file)}")
                except Exception as e:
                    resultados['errores'].append(f"Error eliminando {folio_file}: {e}")

            # 2) Revisar backups locales en folios_visita/backups
            try:
                backup_dir = os.path.join(folios_visita_dir, 'backups')
                if os.path.exists(backup_dir):
                    for archivo in os.listdir(backup_dir):
                        if folio in archivo:
                            ruta = os.path.join(backup_dir, archivo)
                            # intentar extraer solicitudes desde el JSON antes de borrar
                            try:
                                if os.path.isfile(ruta):
                                    with open(ruta, 'r', encoding='utf-8') as bf:
                                        contenido = json.load(bf)
                                        if isinstance(contenido, list):
                                            for row in contenido:
                                                try:
                                                    sol = row.get('SOLICITUD') or row.get('SOLICITUDES') or ''
                                                    if sol:
                                                        s = str(sol).strip()
                                                        if s:
                                                            solicitudes_a_eliminar.add(s)
                                                except Exception:
                                                    continue
                            except Exception:
                                pass
                            try:
                                os.remove(ruta)
                                resultados['eliminados'].append(f"Backup: {archivo}")
                            except Exception as e:
                                resultados['errores'].append(f"Error eliminando backup {archivo}: {e}")
            except Exception as e:
                resultados['errores'].append(f"Error accediendo a backups locales: {e}")

            # 3) Revisar tabla_relacion_backups
            try:
                tabla_relacion_backup_dir = os.path.join(DATA_DIR, 'tabla_relacion_backups')
                if os.path.exists(tabla_relacion_backup_dir):
                    for archivo in os.listdir(tabla_relacion_backup_dir):
                        if folio in archivo:
                            ruta_archivo = os.path.join(tabla_relacion_backup_dir, archivo)
                            try:
                                if os.path.isfile(ruta_archivo):
                                    with open(ruta_archivo, 'r', encoding='utf-8') as bf:
                                        contenido = json.load(bf)
                                        if isinstance(contenido, list):
                                            for row in contenido:
                                                try:
                                                    sol = row.get('SOLICITUD') or row.get('SOLICITUDES') or row.get('SOLICITUDE') or ''
                                                    if sol:
                                                        s = str(sol).strip()
                                                        if s:
                                                            solicitudes_a_eliminar.add(s)
                                                except Exception:
                                                    continue
                            except Exception:
                                pass
                            try:
                                os.remove(ruta_archivo)
                                resultados['eliminados'].append(f"Tabla backup: {archivo}")
                            except Exception as e:
                                resultados['errores'].append(f"Error eliminando {archivo}: {e}")
            except Exception as e:
                resultados['errores'].append(f"Error accediendo a tabla_relacion_backups: {e}")

            # 4) Si aún no hay solicitudes, intentar leer tabla_de_relacion.json para filas asociadas al folio
            try:
                if not solicitudes_a_eliminar:
                    tabla_relacion_path = os.path.join(DATA_DIR, 'tabla_de_relacion.json')
                    if os.path.exists(tabla_relacion_path):
                        try:
                            with open(tabla_relacion_path, 'r', encoding='utf-8') as tf:
                                tbl = json.load(tf) or []
                            folio_digits = ''.join([c for c in str(folio) if c.isdigit()])
                            for row in tbl:
                                try:
                                    row_folio = row.get('FOLIO') or row.get('FOLIOS') or ''
                                    row_digits = ''.join([c for c in str(row_folio) if c.isdigit()])
                                    if folio_digits and folio_digits in row_digits:
                                        sol = row.get('SOLICITUD') or row.get('SOLICITUDES') or ''
                                        if sol:
                                            s = str(sol).strip()
                                            solicitudes_a_eliminar.add(s)
                                            sd = ''.join([c for c in s if c.isdigit()])
                                            if sd:
                                                solicitudes_a_eliminar.add(sd)
                                                solicitudes_a_eliminar.add(sd.lstrip('0'))
                                                solicitudes_a_eliminar.add(sd.zfill(6))
                                except Exception:
                                    continue
                        except Exception:
                            pass
            except Exception as e:
                resultados['errores'].append(f"Error leyendo tabla_de_relacion.json: {e}")

            # 5) Fallback: extraer números del propio folio string
            try:
                if not folios_a_eliminar and not solicitudes_a_eliminar:
                    import re
                    posibles = re.findall(r"\d{1,6}", str(folio))
                    for p in posibles:
                        folios_a_eliminar.add(str(int(p)))
                        folios_a_eliminar.add(p.zfill(6))
                        solicitudes_a_eliminar.add(p)
                        solicitudes_a_eliminar.add(p.zfill(6))
            except Exception:
                pass

            # 5b) Extraer candidatos directamente desde el registro (útil para Constancias)
            try:
                if registro and isinstance(registro, dict):
                    keys = ['folio_constancia', 'folio_acta', 'folio_acta_visita', 'folio', 'folio_visita', 'folio_acta_visita']
                    for k in keys:
                        try:
                            v = registro.get(k) or ''
                            if not v:
                                continue
                            s = str(v).strip()
                            if not s:
                                continue
                            digits = ''.join([c for c in s if c.isdigit()])
                            if digits:
                                folios_a_eliminar.add(digits.lstrip('0') or '0')
                                folios_a_eliminar.add(digits.zfill(6))
                            else:
                                folios_a_eliminar.add(s)
                        except Exception:
                            continue

                    # también intentar extraer solicitudes/solicitud
                    for sk in ('solicitud', 'solicitudes', 'SOLICITUD', 'SOLICITUDES', 'folios_utilizados'):
                        try:
                            sv = registro.get(sk) or ''
                            if not sv:
                                continue
                            # puede ser lista o string
                            if isinstance(sv, list):
                                for it in sv:
                                    try:
                                        ss = str(it).strip()
                                        if not ss:
                                            continue
                                        sd = ''.join([c for c in ss if c.isdigit()])
                                        if sd:
                                            solicitudes_a_eliminar.add(sd)
                                            solicitudes_a_eliminar.add(sd.zfill(6))
                                        else:
                                            solicitudes_a_eliminar.add(ss)
                                    except Exception:
                                        continue
                            else:
                                ss = str(sv).strip()
                                if ss:
                                    sd = ''.join([c for c in ss if c.isdigit()])
                                    if sd:
                                        solicitudes_a_eliminar.add(sd)
                                        solicitudes_a_eliminar.add(sd.zfill(6))
                                    else:
                                        solicitudes_a_eliminar.add(ss)
                        except Exception:
                            continue
            except Exception:
                pass

            # 6) Eliminar dictámenes en data/Dictamenes que coincidan
            try:
                dicts_dir = os.path.join(DATA_DIR, 'Dictamenes')
                if os.path.exists(dicts_dir) and (folios_a_eliminar or solicitudes_a_eliminar):
                    import re
                    for fn in os.listdir(dicts_dir):
                        if not fn.lower().endswith('.json'):
                            continue
                        if 'style' in fn.lower():
                            continue
                        fp = os.path.join(dicts_dir, fn)
                        try:
                            with open(fp, 'r', encoding='utf-8') as jf:
                                d = json.load(jf)
                        except Exception:
                            continue

                        ident = d.get('identificacion') or {}
                        fol_file = str(ident.get('folio') or '').strip()
                        sol_file = str(ident.get('solicitud') or '').strip()
                        cadena = str(ident.get('cadena_identificacion') or '').strip()

                        # Normalizar
                        fol_file_digits = ''.join([c for c in fol_file if c.isdigit()])
                        fol_file_norm = (fol_file_digits.lstrip('0') or '0') if fol_file_digits else fol_file
                        fol_file_z6 = fol_file_digits.zfill(6) if fol_file_digits else fol_file

                        sol_digits = ''.join([c for c in sol_file if c.isdigit()])
                        sol_norm = sol_digits.lstrip('0') if sol_digits else sol_file

                        match = False
                        # comparar folios
                        for f_candidate in folios_a_eliminar:
                            f_c = ''.join([c for c in str(f_candidate) if c.isdigit()]) or str(f_candidate)
                            f_c_norm = f_c.lstrip('0') if f_c else f_c
                            try:
                                if f_c_norm and (f_c_norm == fol_file_norm or f_c_norm == fol_file_digits or f_c_norm == fol_file_z6 or f_c_norm == fol_file):
                                    match = True
                                    break
                            except Exception:
                                continue

                        # comparar solicitudes (normalizar /xx)
                        if not match and solicitudes_a_eliminar:
                            normalized_solicitudes = set()
                            for s in solicitudes_a_eliminar:
                                try:
                                    s_str = str(s).split('/')[0].strip()
                                    s_digits = ''.join([c for c in s_str if c.isdigit()])
                                    if not s_digits:
                                        continue
                                    s_norm = s_digits.lstrip('0') or '0'
                                    normalized_solicitudes.add(s_norm)
                                    normalized_solicitudes.add(s_norm.zfill(6))
                                except Exception:
                                    continue

                            if normalized_solicitudes:
                                digit_groups = re.findall(r"\d+", cadena)
                                for s_norm in normalized_solicitudes:
                                    if sol_digits and (s_norm == sol_digits.lstrip('0') or s_norm == sol_digits or s_norm == sol_digits.zfill(6)):
                                        match = True
                                        break
                                    for dg in digit_groups:
                                        dg_norm = dg.lstrip('0')
                                        if dg_norm == s_norm or dg == s_norm or dg.zfill(6) == s_norm:
                                            match = True
                                            break
                                    if match:
                                        break

                        if match:
                            try:
                                os.remove(fp)
                                resultados['eliminados'].append(f"Dictamen eliminado: {fn}")
                            except Exception as e:
                                resultados['errores'].append(f"Error eliminando dictamen {fn}: {e}")
            except Exception as e:
                resultados['errores'].append(f"Error eliminando dictámenes asociados: {e}")

            # 7) Eliminar constancias en data/Constancias que coincidan
            try:
                const_dir = os.path.join(DATA_DIR, 'Constancias')
                if os.path.exists(const_dir) and (folios_a_eliminar or solicitudes_a_eliminar or folio):
                    for fn in os.listdir(const_dir):
                        if not fn.lower().endswith('.json'):
                            continue
                        if 'style' in fn.lower():
                            continue
                        fp = os.path.join(const_dir, fn)
                        try:
                            with open(fp, 'r', encoding='utf-8') as jf:
                                cdata = json.load(jf)
                        except Exception:
                            continue

                        match = False
                        # Revisar campo 'folio_constancia'
                        try:
                            fol_file = str(cdata.get('folio_constancia') or '').strip()
                            if fol_file:
                                fol_file_digits = ''.join([c for c in fol_file if c.isdigit()])
                                # comparar con folios_a_eliminar
                                for f_candidate in folios_a_eliminar:
                                    f_c = ''.join([c for c in str(f_candidate) if c.isdigit()]) or str(f_candidate)
                                    if f_c and (f_c == fol_file_digits.lstrip('0') or f_c == fol_file_digits or f_c == fol_file_digits.zfill(6) or f_c == fol_file):
                                        match = True
                                        break
                                if not match:
                                    if fol_file == str(folio):
                                        match = True
                        except Exception:
                            pass

                        # Revisar campo 'origen_visita' para folio_visita u otros identificadores
                        if not match:
                            try:
                                ov = cdata.get('origen_visita') or {}
                                if isinstance(ov, dict):
                                    ov_candidates = [ov.get('folio_visita'), ov.get('folio'), ov.get('folio_acta'), ov.get('folio_acta_visita')]
                                else:
                                    ov_candidates = [ov]
                                for oc in ov_candidates:
                                    try:
                                        if not oc:
                                            continue
                                        oc_str = str(oc).strip()
                                        oc_digits = ''.join([c for c in oc_str if c.isdigit()])
                                        for f_candidate in folios_a_eliminar:
                                            f_c = ''.join([c for c in str(f_candidate) if c.isdigit()]) or str(f_candidate)
                                            if f_c and (f_c == oc_digits.lstrip('0') or f_c == oc_digits or f_c == oc_digits.zfill(6) or f_c == oc_str):
                                                match = True
                                                break
                                        if match:
                                            break
                                    except Exception:
                                        continue
                            except Exception:
                                pass

                        if match:
                            try:
                                os.remove(fp)
                                resultados['eliminados'].append(f"Constancia eliminada: {fn}")
                            except Exception as e:
                                resultados['errores'].append(f"Error eliminando constancia {fn}: {e}")
                        else:
                            # Si no hubo coincidencia por contenido, intentar eliminar
                            # por nombre de archivo si el folio/solicitud aparece en el nombre.
                            try:
                                lower_fn = fn.lower()
                                filename_matchers = set()
                                try:
                                    cp_digits = ''.join([c for c in str(folio) if c.isdigit()])
                                except Exception:
                                    cp_digits = ''
                                if cp_digits:
                                    filename_matchers.add(cp_digits)
                                    filename_matchers.add(cp_digits.zfill(6))
                                    filename_matchers.add(f"cp{cp_digits.zfill(6)}")
                                # agregar candidatos desde folios_a_eliminar y solicitudes_a_eliminar
                                for fcan in list(folios_a_eliminar) + list(solicitudes_a_eliminar):
                                    try:
                                        s = str(fcan).lower()
                                        if s:
                                            filename_matchers.add(s)
                                            digits = ''.join([c for c in s if c.isdigit()])
                                            if digits:
                                                filename_matchers.add(digits)
                                                filename_matchers.add(digits.zfill(6))
                                    except Exception:
                                        continue

                                should_remove = False
                                for m in filename_matchers:
                                    try:
                                        if not m:
                                            continue
                                        if m in lower_fn:
                                            should_remove = True
                                            break
                                    except Exception:
                                        continue

                                if should_remove:
                                    try:
                                        os.remove(fp)
                                        resultados['eliminados'].append(f"Constancia eliminada (por nombre): {fn}")
                                    except Exception as e:
                                        resultados['errores'].append(f"Error eliminando constancia {fn}: {e}")
                            except Exception:
                                pass
            except Exception as e:
                resultados['errores'].append(f"Error eliminando constancias asociadas: {e}")

        except Exception as e:
            resultados['errores'].append(f"Error general en eliminación: {str(e)}")

        return resultados

    def _validar_integridad_historial(self):
        """Valida la integridad del historial y repara si es necesario"""
        try:
            # Verificar que historial_data_original esté sincronizado
            if len(self.historial_data) != len(self.historial_data_original):
                print("⚠️ Resincronizando historial_data_original...")
                self.historial_data_original = self.historial_data.copy()
            
            # Verificar que el archivo JSON existe y es válido
            if os.path.exists(self.historial_path):
                with open(self.historial_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                    json_visitas = json_data.get('visitas', [])
                    
                    # Si hay desincronización, sincronizar
                    if len(json_visitas) != len(self.historial_data):
                        print(f"⚠️ Desincronización detectada. JSON: {len(json_visitas)}, Memoria: {len(self.historial_data)}")
                        self._sincronizar_historial()
            
            return True
        except Exception as e:
            print(f"❌ Error en validación: {e}")
            return False

    def _garantizar_persistencia(self, folio):
        """Garantiza que un folio no exista en ninguna parte del sistema después de eliminación"""
        try:
            # Verificar JSON
            if os.path.exists(self.historial_path):
                with open(self.historial_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                folio_existe = any(v.get('folio_visita') == folio for v in data.get('visitas', []))
                
                if folio_existe:
                    # Eliminar y guardar de nuevo
                    data['visitas'] = [v for v in data.get('visitas', []) if v.get('folio_visita') != folio]
                    with open(self.historial_path, 'w', encoding='utf-8') as f:
                        json.dump(data, f, ensure_ascii=False, indent=2)
                    print(f"✅ Folio {folio} eliminado del JSON")
            
            # Verificar carpetas
            carpetas = [
                os.path.join(self.folios_visita_path, f"folios_{folio}.json"),
                os.path.join(APP_DIR, "data", "tabla_relacion_backups")
            ]
            
            for carpeta in carpetas:
                if os.path.exists(carpeta):
                    if os.path.isfile(carpeta):
                        os.remove(carpeta)
                        print(f"✅ Archivo eliminado: {carpeta}")
                    elif os.path.isdir(carpeta):
                        for archivo in os.listdir(carpeta):
                            if folio in archivo:
                                ruta = os.path.join(carpeta, archivo)
                                os.remove(ruta)
                                print(f"✅ Archivo de backup eliminado: {archivo}")
            
            return True
        except Exception as e:
            print(f"⚠️ Error en garantía de persistencia: {e}")
            return False

    def _registrar_operacion(self, tipo_operacion, folio, status, detalles=""):
        """Registra todas las operaciones para auditoría y persistencia"""
        try:
            log_path = os.path.join(APP_DIR, "data", "operaciones_log.json")
            
            # Cargar log existente o crear uno nuevo
            if os.path.exists(log_path):
                with open(log_path, 'r', encoding='utf-8') as f:
                    log_data = json.load(f)
            else:
                log_data = {"operaciones": []}
            
            # Agregar nueva operación
            operacion = {
                "timestamp": datetime.now().isoformat(),
                "tipo": tipo_operacion,
                "folio": folio,
                "status": status,
                "detalles": detalles
            }
            
            log_data["operaciones"].append(operacion)
            
            # Guardar log
            with open(log_path, 'w', encoding='utf-8') as f:
                json.dump(log_data, f, ensure_ascii=False, indent=2)
            
            return True
        except Exception as e:
            print(f"⚠️ Error registrando operación: {e}")
            return False

    def hist_eliminar_registro(self, registro):
        """Eliminar un registro del historial con persistencia completa"""
        folio = registro.get('folio_visita', '')
        confirmacion = messagebox.askyesno(
            "Confirmar eliminación",
            f"¿Está seguro de que desea eliminar el registro del folio {folio}?\n\nSe eliminarán todos los archivos asociados."
        )

        if not confirmacion:
            return

        # Mostrar diálogo de progreso modal (compacto)
        try:
            prog_win = tk.Toplevel(self)
            prog_win.title("Eliminando...")
            prog_win.geometry("420x110")
            prog_win.transient(self)
            prog_win.grab_set()
            prog_win.resizable(False, False)

            lbl = tk.Label(prog_win, text=f"Eliminando archivos del folio {folio}...", anchor='w')
            lbl.pack(fill='x', padx=12, pady=(12, 6))

            pb = ttk.Progressbar(prog_win, mode='indeterminate')
            pb.pack(fill='x', padx=12, pady=(0, 8))
            pb.start(10)

            info_lbl = tk.Label(prog_win, text="Esto puede tardar un momento. El detalle se guardará en un archivo de registro.")
            info_lbl.pack(fill='x', padx=12, pady=(0, 8))
        except Exception:
            prog_win = None
            pb = None

        # Ejecutar eliminación en hilo para no bloquear la UI
        def _worker():
            resultados = {'eliminados': [], 'errores': []}

            # Intentar leer meta de folios antes de eliminar archivos
            meta_counter_before = None
            try:
                folios_visita_file = os.path.join(self.folios_visita_path, f"folios_{folio}.json")
                if os.path.exists(folios_visita_file):
                    try:
                        with open(folios_visita_file, 'r', encoding='utf-8') as rf:
                            obj = json.load(rf)
                            if isinstance(obj, dict):
                                meta = obj.get('_meta') or {}
                                if isinstance(meta, dict) and 'counter_before' in meta:
                                    try:
                                        meta_counter_before = int(meta.get('counter_before') or 0)
                                    except Exception:
                                        meta_counter_before = None
                    except Exception:
                        meta_counter_before = None
            except Exception:
                meta_counter_before = None

            try:
                # 1) Borrar archivos asociados (método existente)
                try:
                    res = self._eliminar_archivos_asociados_folio(folio, registro)
                    # merge resultados
                    for k in ('eliminados', 'errores'):
                        if k in res:
                            resultados[k].extend(res[k])
                except Exception as e:
                    resultados['errores'].append(f"Error eliminando archivos asociados: {e}")

                # 2) Eliminar fila en memoria
                registro_id = registro.get('_id')
                if registro_id:
                    original_len = len(self.historial_data)
                    self.historial_data = [r for r in self.historial_data if r.get('_id') != registro_id]
                    self.historial_data_original = [r for r in self.historial_data_original if r.get('_id') != registro_id]
                    if len(self.historial_data) < original_len:
                        resultados['eliminados'].append(f"Entrada en historial (id={registro_id})")
                else:
                    self.historial_data = [r for r in self.historial_data if r.get('folio_visita') != folio]
                    self.historial_data_original = [r for r in self.historial_data_original if r.get('folio_visita') != folio]
                    resultados['eliminados'].append("Entrada en historial (por folio)")

                # 3) Sincronizar historial
                try:
                    sincronizacion_exitosa = self._sincronizar_historial()
                    if sincronizacion_exitosa:
                        resultados['eliminados'].append('✅ Entrada en historial visitas')
                    else:
                        resultados['errores'].append('⚠️ Error al sincronizar historial')
                except Exception as e:
                    resultados['errores'].append(f'Error sincronizando historial: {e}')

                # 4) Limpiar tabla_de_relacion (se deja como antes, pero en hilo)
                try:
                    # extraer folios asociados (int)
                    folios_asociados = set()
                    folios_visita_file = os.path.join(self.folios_visita_path, f"folios_{folio}.json")
                    if os.path.exists(folios_visita_file):
                        try:
                            with open(folios_visita_file, 'r', encoding='utf-8') as f:
                                fv = json.load(f)
                                if isinstance(fv, list):
                                    for v in fv:
                                        try:
                                            if isinstance(v, dict):
                                                fol = v.get('FOLIOS') or v.get('FOLIO') or ''
                                            else:
                                                fol = v
                                            digits = ''.join([c for c in str(fol) if c.isdigit()])
                                            if digits:
                                                folios_asociados.add(int(digits))
                                        except Exception:
                                            pass
                                elif isinstance(fv, dict):
                                    fl = fv.get('folios') or []
                                    for entry in fl:
                                        try:
                                            fol = entry.get('FOLIOS') or entry.get('FOLIO') or entry if isinstance(entry, dict) else entry
                                            digits = ''.join([c for c in str(fol) if c.isdigit()])
                                            if digits:
                                                folios_asociados.add(int(digits))
                                        except Exception:
                                            pass
                        except Exception:
                            pass

                    if not folios_asociados:
                        posibles = re.findall(r"\d{1,6}", str(registro.get('folios_utilizados', '') or ''))
                        cp_digits = ''.join([c for c in str(folio) if c.isdigit()])
                        acta = registro.get('folio_acta') or registro.get('folio_acta_visita') or ''
                        acta_digits = ''.join([c for c in str(acta) if c.isdigit()])
                        for p in posibles:
                            try:
                                if cp_digits and str(p).lstrip('0') == str(int(cp_digits)).lstrip('0'):
                                    continue
                                if acta_digits and str(p).lstrip('0') == str(int(acta_digits)).lstrip('0'):
                                    continue
                                folios_asociados.add(int(p))
                            except Exception:
                                pass

                    tabla_relacion_path = os.path.join(DATA_DIR, 'tabla_de_relacion.json')
                    if os.path.exists(tabla_relacion_path):
                        try:
                            backup_dir = os.path.join(DATA_DIR, 'tabla_relacion_backups')
                            os.makedirs(backup_dir, exist_ok=True)
                            ts = datetime.now().strftime('%Y%m%d%H%M%S')
                            cp_digits = ''.join([c for c in str(folio) if c.isdigit()]) or ''
                            cp = f"CP{int(cp_digits):06d}" if cp_digits else f"CP{str(folio)}"
                            dest_name = f"tabla_de_relacion_{cp}_BACKUP_{ts}.json"
                            shutil.copyfile(tabla_relacion_path, os.path.join(backup_dir, dest_name))
                        except Exception as e:
                            resultados['errores'].append(f"No se pudo crear backup de tabla_de_relacion: {e}")

                        try:
                            with open(tabla_relacion_path, 'r', encoding='utf-8') as f:
                                tabla = json.load(f)
                            nueva_tabla = []
                            for row in tabla:
                                try:
                                    val = row.get('FOLIO', None)
                                    if val is None:
                                        nueva_tabla.append(row)
                                        continue
                                    try:
                                        val_int = int(float(val))
                                    except Exception:
                                        val_int = None
                                    if val_int is not None and val_int in folios_asociados:
                                        continue
                                    if str(val).strip() in {str(f).zfill(6) for f in folios_asociados}:
                                        continue
                                    nueva_tabla.append(row)
                                except Exception:
                                    nueva_tabla.append(row)
                            with open(tabla_relacion_path, 'w', encoding='utf-8') as f:
                                json.dump(nueva_tabla, f, ensure_ascii=False, indent=2)
                            resultados['eliminados'].append('Entradas en tabla_de_relacion.json')
                        except Exception as e:
                            resultados['errores'].append(f"Error modificando tabla_de_relacion.json: {e}")
                except Exception as e:
                    resultados['errores'].append(f"Error eliminando entradas de tabla de relación: {e}")

                # Garantizar persistencia completa
                try:
                    persistencia_garantizada = self._garantizar_persistencia(folio)
                    if persistencia_garantizada:
                        resultados['eliminados'].append('✅ Persistencia verificada y garantizada')
                except Exception:
                    pass

                # Intentar recomputar y ajustar contador
                try:
                    try:
                        last_counter = int(folio_manager.get_last() or 0)
                    except Exception:
                        last_counter = 0
                    max_remain = 0
                    dirp = self.folios_visita_path
                    if os.path.exists(dirp):
                        for fn in os.listdir(dirp):
                            if fn.startswith('folios_') and fn.endswith('.json'):
                                pathf = os.path.join(dirp, fn)
                                try:
                                    with open(pathf, 'r', encoding='utf-8') as fh:
                                        obj = json.load(fh) or []
                                        arr = obj.get('folios') if isinstance(obj, dict) and 'folios' in obj else obj
                                        for entry in arr:
                                            try:
                                                fol = entry.get('FOLIOS') or entry.get('FOLIO') or entry if isinstance(entry, dict) else entry
                                                digits = ''.join([c for c in str(fol) if c.isdigit()])
                                                if digits:
                                                    n = int(digits)
                                                    if n > max_remain:
                                                        max_remain = n
                                            except Exception:
                                                continue
                                except Exception:
                                    continue
                    tabla_relacion_path = os.path.join(DATA_DIR, 'tabla_de_relacion.json')
                    if os.path.exists(tabla_relacion_path):
                        try:
                            with open(tabla_relacion_path, 'r', encoding='utf-8') as tf:
                                tabla = json.load(tf) or []
                            for row in tabla:
                                try:
                                    v = row.get('FOLIO') or row.get('FOLIOS') or ''
                                    digits = ''.join([c for c in str(v) if c.isdigit()])
                                    if digits:
                                        n = int(digits)
                                        if n > max_remain:
                                            max_remain = n
                                except Exception:
                                    continue
                        except Exception:
                            pass
                    desired = int(meta_counter_before) if meta_counter_before is not None else max_remain
                    if desired < last_counter:
                        try:
                            folio_manager.set_last(int(desired))
                            resultados['eliminados'].append(f"folio_counter.json ajustado a {int(desired):06d}")
                        except Exception as e:
                            resultados['errores'].append(f"No se pudo ajustar folio_counter: {e}")
                except Exception:
                    pass

                # Registrar operación
                detalles_eliminacion = f"Archivos: {len(resultados.get('eliminados', []))}, Errores: {len(resultados.get('errores', []))}"
                try:
                    self._registrar_operacion("eliminar_registro", folio, "exitosa", detalles_eliminacion)
                except Exception:
                    pass

            except Exception as e:
                resultados['errores'].append(f"Error general en proceso de eliminación: {e}")

            # Guardar log detallado en archivo para revisiones (no mostrar lista completa en mensaje)
            try:
                logs_dir = os.path.join(DATA_DIR, 'eliminacion_logs')
                os.makedirs(logs_dir, exist_ok=True)
                ts = datetime.now().strftime('%Y%m%d%H%M%S')
                log_path = os.path.join(logs_dir, f'eliminacion_{folio}_{ts}.txt')
                with open(log_path, 'w', encoding='utf-8') as lf:
                    lf.write(f"Eliminación folio: {folio}\n")
                    lf.write(f"Timestamp: {datetime.now().isoformat()}\n\n")
                    lf.write(f"Elementos eliminados ({len(resultados.get('eliminados', []))}):\n")
                    for it in resultados.get('eliminados', []):
                        lf.write(f" - {it}\n")
                    lf.write('\n')
                    lf.write(f"Errores ({len(resultados.get('errores', []))}):\n")
                    for er in resultados.get('errores', []):
                        lf.write(f" - {er}\n")
            except Exception:
                log_path = None

            # Actualizar UI en hilo principal cuando termine
            def _on_done():
                try:
                    if pb:
                        try:
                            pb.stop()
                        except Exception:
                            pass
                    if prog_win:
                        try:
                            prog_win.grab_release()
                        except Exception:
                            pass
                        try:
                            prog_win.destroy()
                        except Exception:
                            pass

                except Exception:
                    pass

                # Actualizar interfaz y mostrar resumen compacto
                try:
                    self._poblar_historial_ui()
                except Exception:
                    pass
                try:
                    self.cargar_ultimo_folio()
                except Exception:
                    pass
                try:
                    self._update_siguiente_folio_label()
                except Exception:
                    pass

                resumen = f"✅ Registro del folio {folio} eliminado correctamente.\nElementos eliminados: {len(resultados.get('eliminados', []))}\nErrores: {len(resultados.get('errores', []))}"
                if log_path:
                    resumen += f"\n\nDetalle guardado en: {log_path}"

                try:
                    messagebox.showinfo("Eliminación completada", resumen)
                except Exception:
                    print(resumen)

            try:
                self.after(50, _on_done)
            except Exception:
                _on_done()

        # lanzar hilo
        try:
            th = threading.Thread(target=_worker, daemon=True)
            th.start()
        except Exception:
            # fallback: ejecutar sin hilo
            _worker()

    def hist_create_visita(self, payload, es_automatica=False, show_notification=True):
        """Crea una nueva visita en el historial"""
        try:
            # Generar ID único
            payload["_id"] = str(uuid.uuid4())
            
            # Asegurar que estatus tenga valor
            payload.setdefault("estatus", "Completada" if es_automatica else "En proceso")
            
            # Asegurar que las fechas y horas estén presentes
            payload.setdefault("fecha_inicio", "")
            payload.setdefault("fecha_termino", "")
            payload.setdefault("hora_inicio", "")
            payload.setdefault("hora_termino", "")
            
            # Para evitar problemas al actualizar widgets desde hilos en background,
            # realizamos la mutación del historial y la actualización de la UI en el
            # hilo principal usando `self.after(0, ...)`.
            def _apply_and_refresh():
                try:
                    if "visitas" not in self.historial:
                        self.historial["visitas"] = []

                    # Buscar registro existente por folio_visita
                    existing_idx = None
                    try:
                        for idx, v in enumerate(self.historial.get('visitas', [])):
                            if str(v.get('folio_visita', '')).strip().lower() == str(payload.get('folio_visita', '')).strip().lower():
                                existing_idx = idx
                                break
                    except Exception:
                        existing_idx = None

                    # Normalizar campos de dirección en el payload antes de mezclar/añadir
                    for k in ('direccion','calle_numero','colonia','municipio','ciudad_estado','cp'):
                        if k not in payload:
                            payload[k] = ''
                    # ---------- Validación de unicidad (CP/AC) ----------
                    try:
                        new_fv = str(payload.get('folio_visita', '') or '').strip()
                        new_fa = str(payload.get('folio_acta', '') or '').strip()
                        # Leer versión en disco para evitar duplicados entre procesos
                        try:
                            hist_path = getattr(self, 'historial_path', None) or os.path.join(DATA_DIR, 'historial_visitas.json')
                            if os.path.exists(hist_path):
                                with open(hist_path, 'r', encoding='utf-8') as hf:
                                    hobj = json.load(hf) or {}
                                    disk_visitas = hobj.get('visitas', []) if isinstance(hobj, dict) else (hobj or [])
                            else:
                                disk_visitas = self.historial.get('visitas', []) or []
                        except Exception:
                            disk_visitas = self.historial.get('visitas', []) or []

                        if new_fv:
                            # Utilizar helper centralizado que chequea historial en disco
                            if self._folio_visita_exists(new_fv, exclude_id=None if existing_idx is None else self.historial.get('visitas', [])[existing_idx].get('_id')):
                                messagebox.showwarning("Folio duplicado", f"El folio de visita {new_fv} ya está en uso. No se puede duplicar CP.")
                                return
                        if new_fa:
                            for idx2, vv in enumerate(disk_visitas or []):
                                try:
                                    if idx2 == existing_idx:
                                        continue
                                    other_fa = str(vv.get('folio_acta', '') or '').strip()
                                    if other_fa and other_fa.lower() == new_fa.lower():
                                        messagebox.showwarning("Folio duplicado", f"El folio de acta {new_fa} ya está en uso. No se puede duplicar AC.")
                                        return
                                except Exception:
                                    continue
                    except Exception:
                        pass
                    # Asegurar que cp sea string (preservar ceros a la izquierda si los hay)
                    try:
                        if payload.get('cp') is not None and payload.get('cp') != '':
                            payload['cp'] = str(payload.get('cp'))
                    except Exception:
                        payload['cp'] = str(payload.get('cp') or '')

                    # Si faltan datos de dirección, intentar poblar desde data/Clientes.json
                    try:
                        need_addr = not payload.get('direccion') or not payload.get('calle_numero')
                        cliente_nombre = payload.get('cliente') or ''
                        if need_addr and cliente_nombre:
                            clientes_path = os.path.join(DATA_DIR, 'Clientes.json')
                            if os.path.exists(clientes_path):
                                try:
                                    with open(clientes_path, 'r', encoding='utf-8') as cf:
                                        clientes = json.load(cf)
                                    needle = str(cliente_nombre).strip().upper()
                                    for c in (clientes or []):
                                        try:
                                            name = (c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or c.get('RAZON_SOCIAL') or '')
                                            if not name:
                                                continue
                                            if str(name).strip().upper() == needle or needle in str(name).strip().upper() or str(name).strip().upper() in needle:
                                                direcciones = c.get('DIRECCIONES') or []
                                                first = None
                                                if isinstance(direcciones, list) and direcciones:
                                                    first = direcciones[0]
                                                if first and isinstance(first, dict):
                                                    payload['calle_numero'] = payload.get('calle_numero') or (first.get('CALLE Y NO') or first.get('CALLE') or '')
                                                    payload['colonia'] = payload.get('colonia') or (first.get('COLONIA O POBLACION') or first.get('COLONIA') or '')
                                                    payload['municipio'] = payload.get('municipio') or (first.get('MUNICIPIO O ALCADIA') or first.get('MUNICIPIO') or '')
                                                    payload['ciudad_estado'] = payload.get('ciudad_estado') or (first.get('CIUDAD O ESTADO') or first.get('CIUDAD') or '')
                                                    cpval = first.get('CP') or first.get('cp')
                                                    if cpval is not None and cpval != '':
                                                        payload['cp'] = str(cpval)
                                                else:
                                                    payload['calle_numero'] = payload.get('calle_numero') or (c.get('CALLE Y NO') or c.get('CALLE') or '')
                                                    payload['colonia'] = payload.get('colonia') or (c.get('COLONIA O POBLACION') or c.get('COLONIA') or '')
                                                    payload['municipio'] = payload.get('municipio') or (c.get('MUNICIPIO O ALCADIA') or c.get('MUNICIPIO') or '')
                                                    payload['ciudad_estado'] = payload.get('ciudad_estado') or (c.get('CIUDAD O ESTADO') or c.get('CIUDAD') or '')
                                                    cpval = c.get('CP') or c.get('cp')
                                                    if cpval is not None and cpval != '':
                                                        payload['cp'] = str(cpval)
                                                break
                                        except Exception:
                                            continue
                                except Exception:
                                    pass
                    except Exception:
                        pass

                    # Construir campo 'direccion' canónico a partir de componentes de dirección
                    try:
                        calle_val = (payload.get('calle_numero') or '').strip()
                        colonia_val = (payload.get('colonia') or '').strip()
                        municipio_val = (payload.get('municipio') or '').strip()
                        ciudad_estado_val = (payload.get('ciudad_estado') or '').strip()
                        cp_val = (payload.get('cp') or '').strip()
                        partes = [p for p in [calle_val, colonia_val, municipio_val, ciudad_estado_val] if p]
                        direccion_comp = ', '.join(partes)
                        if cp_val:
                            direccion_comp = (f"{direccion_comp}, C.P. {cp_val}" if direccion_comp else f"C.P. {cp_val}")
                        if direccion_comp:
                            payload['direccion'] = direccion_comp
                    except Exception:
                        pass

                    if existing_idx is not None:
                        # Mergear campos (no sobrescribir metadatos existentes innecesariamente)
                        existing = self.historial['visitas'][existing_idx]
                        for k, val in payload.items():
                            if k == '_id':
                                continue
                            if val is not None and val != '':
                                existing[k] = val
                        existing.setdefault('estatus', payload.get('estatus', 'En proceso'))
                    else:
                        # Añadir metadatos de creador si no están presentes
                        try:
                            if getattr(self, 'current_user', None):
                                payload.setdefault('creado_por', self.current_user.get('username') or '')
                                payload.setdefault('creado_nombre', self.current_user.get('name') or '')
                        except Exception:
                            pass

                        # Append payload ensuring address fields exist
                        self.historial["visitas"].append(payload)

                    # Actualizar datos en memoria
                    self.historial_data = self.historial.get("visitas", [])

                    # Asegurar que exista la carpeta de producción para el creador del registro
                    try:
                        creador = (payload.get('creado_por') or '')
                        if creador:
                            owner_safe = re.sub(r"[^A-Za-z0-9_.-]", '_', str(creador))
                            prod_dir = os.path.join(DATA_DIR, 'produccion', owner_safe)
                            os.makedirs(prod_dir, exist_ok=True)
                    except Exception:
                        pass

                    # Guardar copia individual de la visita en la carpeta del usuario (visita_<folio>_<ts>.json)
                    try:
                        try:
                            folio = str(payload.get('folio_visita') or payload.get('folio_acta') or payload.get('_id') or 'sinfolio')
                        except Exception:
                            folio = 'sinfolio'
                        tsfn = datetime.now().strftime('%Y%m%d_%H%M%S')
                        safe_folio = re.sub(r"[^A-Za-z0-9_.-]", '_', folio)
                        filename = f"visita_{safe_folio}_{tsfn}.json"
                        if creador:
                            file_path = os.path.join(prod_dir, filename)
                            with open(file_path, 'w', encoding='utf-8') as jf:
                                json.dump(payload, jf, ensure_ascii=False, indent=2)
                    except Exception:
                        pass

                    # Guardar y refrescar UI
                    self._guardar_historial()
                    self._poblar_historial_ui()

                    # Recalcular folio actual inmediatamente
                    try:
                        self.cargar_ultimo_folio()
                    except Exception:
                        pass
                    try:
                        if hasattr(self, '_update_siguiente_folio_label'):
                            self._update_siguiente_folio_label()
                    except Exception:
                        pass

                    # Forzar actualización visual de la etiqueta de siguiente folio
                    try:
                        if hasattr(self, '_update_siguiente_folio_label'):
                            self._update_siguiente_folio_label()
                    except Exception:
                        pass

                    # Refrescar dropdown de folios pendientes al añadir una visita
                    try:
                        if hasattr(self, '_refresh_pending_folios_dropdown'):
                            self._refresh_pending_folios_dropdown()
                    except Exception:
                        pass

                    if not es_automatica and show_notification:
                        messagebox.showinfo("OK", f"Visita {payload.get('folio_visita','-')} guardada correctamente")

                    # DEBUG: mostrar resumen mínimo del historial después de añadir
                    try:
                        print(f"[DEBUG] hist_create_visita: total registros = {len(self.historial.get('visitas', []))}")
                    except Exception:
                        pass
                except Exception as e:
                    print(f"❌ Error aplicando visita en hilo principal: {e}")
                    # Refrescar dropdown de folios pendientes cuando se aplica una modificación
                    try:
                        if hasattr(self, '_refresh_pending_folios_dropdown'):
                            self._refresh_pending_folios_dropdown()
                    except Exception:
                        pass

            # Programar la aplicación en el hilo principal
            try:
                self.after(0, _apply_and_refresh)
            except Exception:
                # Como fallback, intentar aplicar inmediatamente
                _apply_and_refresh()
                
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def hist_buscar_por_folio(self):
        """Buscar en el historial por folio de visita"""
        try:
            # resetear paginado al buscar
            self.HISTORIAL_PAGINA_ACTUAL = 1
            folio_busqueda_raw = self.entry_buscar_folio.get().strip()
            folio_busqueda = folio_busqueda_raw.lower()

            if not folio_busqueda_raw:
                # Si no hay búsqueda, mostrar todos los datos
                self.historial_data = self.historial_data_original.copy() if hasattr(self, 'historial_data_original') else self.historial_data
            else:
                # Filtrar datos por folio con normalización de dígitos
                resultados = []
                fuente = (self.historial_data_original if hasattr(self, 'historial_data_original') else self.historial_data)
                digits_search = ''.join([c for c in folio_busqueda_raw if c.isdigit()])
                for registro in fuente:
                    folio_actual = str(registro.get('folio_visita', '') or '')
                    # coincidencia directa (substring)
                    if folio_busqueda in folio_actual.lower():
                        resultados.append(registro)
                        continue
                    # coincidencia por dígitos (ignora padding)
                    if digits_search:
                        folio_digits = ''.join([c for c in folio_actual if c.isdigit()])
                        if folio_digits and digits_search in folio_digits:
                            resultados.append(registro)
                            continue

                self.historial_data = resultados
            
            self._poblar_historial_ui()
            
        except Exception as e:
            print(f"Error en búsqueda por folio: {e}")

    def hist_borrar_por_folio(self):
        """Borrar una visita usando el folio ingresado en la barra de búsqueda.
        Busca el registro por folio (coincidencia exacta primero, luego parcial)
        y delega en `hist_eliminar_registro` para la eliminación con confirmación.
        """
        try:
            folio = self.entry_buscar_folio.get().strip()
            if not folio:
                messagebox.showwarning("Advertencia", "Ingrese el folio en la barra de búsqueda para eliminar una visita.")
                return

            fuente = self.historial_data_original if hasattr(self, 'historial_data_original') else self.historial_data

            # Buscar coincidencia exacta (case-insensitive)
            matches = [r for r in fuente if str(r.get('folio_visita', '')).strip().lower() == folio.lower()]

            # Si no hay exactas, buscar por contains
            if not matches:
                matches = [r for r in fuente if folio.lower() in str(r.get('folio_visita', '')).lower()]

            if not matches:
                messagebox.showinfo("No encontrado", f"No se encontró ningún registro con folio '{folio}'.")
                return

            if len(matches) > 1:
                # Informar que se encontró más de una coincidencia y proceder con la primera
                confirmar = messagebox.askyesno(
                    "Confirmar eliminación",
                    f"Se encontraron {len(matches)} registros que coinciden con '{folio}'.\n\nSe eliminará el primer registro encontrado: {matches[0].get('folio_visita')}\n\n¿Desea continuar?"
                )
                if not confirmar:
                    return

            # Delegar en la función existente para eliminar (esta función pedirá su propia confirmación también)
            # Llamamos a hist_eliminar_registro con el primer match
            self.hist_eliminar_registro(matches[0])

        except Exception as e:
            messagebox.showerror("Error", f"Error al intentar eliminar por folio:\n{e}")

    def hist_update_visita(self, id_, nuevos):
        """Actualiza una visita existente"""
        try:
            # Buscar la visita a actualizar y mezclar (merge) los campos nuevos
            visitas = self.historial.get("visitas", [])
            encontrado = False
            for i, v in enumerate(visitas):
                try:
                    if v.get("_id") == id_ or v.get("id") == id_:
                        encontrado = True
                    else:
                        # permitir búsquedas por folio_visita o folio_acta si se pasó un folio
                        if id_ and isinstance(id_, str):
                            if id_.strip() and (id_.strip() == (v.get('folio_visita','') or '').strip() or id_.strip() == (v.get('folio_acta','') or '').strip()):
                                encontrado = True
                except Exception:
                    continue

                if encontrado:
                    actualizado = v.copy()
                    # Mezclar claves de 'nuevos' sobre el registro existente
                    for k, val in (nuevos or {}).items():
                        if k == "_id":
                            continue
                        actualizado[k] = val

                    # Normalizar y asegurar campos de dirección persisten
                    for k in ('direccion','calle_numero','colonia','municipio','ciudad_estado','cp'):
                        if k not in actualizado:
                            actualizado[k] = ''
                    # Si 'calle_numero' no existe pero 'direccion' sí, sincronizar
                    if not actualizado.get('calle_numero') and actualizado.get('direccion'):
                        actualizado['calle_numero'] = actualizado.get('direccion')
                    # Forzar cp como string
                    try:
                        if actualizado.get('cp') is not None and actualizado.get('cp') != '':
                            actualizado['cp'] = str(actualizado.get('cp'))
                    except Exception:
                        actualizado['cp'] = str(actualizado.get('cp') or '')

                    # Construir campo 'direccion' canónico a partir de componentes de dirección
                    try:
                        calle_val = (actualizado.get('calle_numero') or '').strip()
                        colonia_val = (actualizado.get('colonia') or '').strip()
                        municipio_val = (actualizado.get('municipio') or '').strip()
                        ciudad_estado_val = (actualizado.get('ciudad_estado') or '').strip()
                        cp_val = (actualizado.get('cp') or '').strip()
                        partes = [p for p in [calle_val, colonia_val, municipio_val, ciudad_estado_val] if p]
                        direccion_comp = ', '.join(partes)
                        if cp_val:
                            direccion_comp = (f"{direccion_comp}, C.P. {cp_val}" if direccion_comp else f"C.P. {cp_val}")
                        if direccion_comp:
                            actualizado['direccion'] = direccion_comp
                    except Exception:
                        pass
                    # Si faltan campos de dirección, intentar poblar desde data/Clientes.json
                    try:
                        need_addr = not actualizado.get('direccion') or not actualizado.get('calle_numero')
                        cliente_nombre = actualizado.get('cliente') or ''
                        if need_addr and cliente_nombre:
                            clientes_path = os.path.join(DATA_DIR, 'Clientes.json')
                            if os.path.exists(clientes_path):
                                try:
                                    with open(clientes_path, 'r', encoding='utf-8') as cf:
                                        clientes = json.load(cf)
                                    needle = str(cliente_nombre).strip().upper()
                                    for c in (clientes or []):
                                        try:
                                            name = (c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or c.get('RAZON_SOCIAL') or '')
                                            if not name:
                                                continue
                                            if str(name).strip().upper() == needle or needle in str(name).strip().upper() or str(name).strip().upper() in needle:
                                                # try DIRECCIONES first
                                                direcciones = c.get('DIRECCIONES') or []
                                                first = None
                                                if isinstance(direcciones, list) and direcciones:
                                                    first = direcciones[0]
                                                if first and isinstance(first, dict):
                                                    actualizado['calle_numero'] = actualizado.get('calle_numero') or (first.get('CALLE Y NO') or first.get('CALLE') or '')
                                                    actualizado['colonia'] = actualizado.get('colonia') or (first.get('COLONIA O POBLACION') or first.get('COLONIA') or '')
                                                    actualizado['municipio'] = actualizado.get('municipio') or (first.get('MUNICIPIO O ALCADIA') or first.get('MUNICIPIO') or '')
                                                    actualizado['ciudad_estado'] = actualizado.get('ciudad_estado') or (first.get('CIUDAD O ESTADO') or first.get('CIUDAD') or '')
                                                    cpval = first.get('CP') or first.get('cp')
                                                    if cpval is not None and cpval != '':
                                                        actualizado['cp'] = str(cpval)
                                                else:
                                                    # try top-level keys
                                                    actualizado['calle_numero'] = actualizado.get('calle_numero') or (c.get('CALLE Y NO') or c.get('CALLE') or '')
                                                    actualizado['colonia'] = actualizado.get('colonia') or (c.get('COLONIA O POBLACION') or c.get('COLONIA') or '')
                                                    actualizado['municipio'] = actualizado.get('municipio') or (c.get('MUNICIPIO O ALCADIA') or c.get('MUNICIPIO') or '')
                                                    actualizado['ciudad_estado'] = actualizado.get('ciudad_estado') or (c.get('CIUDAD O ESTADO') or c.get('CIUDAD') or '')
                                                    cpval = c.get('CP') or c.get('cp')
                                                    if cpval is not None and cpval != '':
                                                        actualizado['cp'] = str(cpval)
                                                break
                                        except Exception:
                                            continue
                                except Exception:
                                    pass
                    except Exception:
                        pass

                    # Antes de reemplazar, validar unicidad CP/AC frente a otros registros
                    try:
                        new_fv = str(actualizado.get('folio_visita','') or '').strip()
                        new_fa = str(actualizado.get('folio_acta','') or '').strip()
                        # Leer versión en disco para evitar duplicados entre procesos
                        try:
                            hist_path = getattr(self, 'historial_path', None) or os.path.join(DATA_DIR, 'historial_visitas.json')
                            if os.path.exists(hist_path):
                                with open(hist_path, 'r', encoding='utf-8') as hf:
                                    hobj = json.load(hf) or {}
                                    disk_visitas = hobj.get('visitas', []) if isinstance(hobj, dict) else (hobj or [])
                            else:
                                disk_visitas = visitas
                        except Exception:
                            disk_visitas = visitas

                        # Usar helper para validar folio de visita (excluir el registro actual por _id)
                        exclude_id = actualizado.get('_id') or None
                        if new_fv and self._folio_visita_exists(new_fv, exclude_id=exclude_id):
                            messagebox.showwarning("Folio duplicado", f"El folio de visita {new_fv} ya está en uso por otro registro. No se puede duplicar CP.")
                            return

                        for j, other in enumerate(disk_visitas or []):
                            try:
                                # si es el mismo índice ignorar
                                if j == i:
                                    continue
                                ofa = str(other.get('folio_acta','') or '').strip()
                                if new_fa and ofa and ofa.lower() == new_fa.lower():
                                    messagebox.showwarning("Folio duplicado", f"El folio de acta {new_fa} ya está en uso por otro registro. No se puede duplicar AC.")
                                    return
                            except Exception:
                                continue
                    except Exception:
                        pass

                    # Reemplazar el registro en la lista
                    try:
                        self.historial['visitas'][i] = actualizado
                    except Exception:
                        pass

                    # Actualizar vistas en memoria y persistir
                    self.historial_data = self.historial.get("visitas", [])
                    self._guardar_historial()
                    self._poblar_historial_ui()
                    try:
                        self.cargar_ultimo_folio()
                    except Exception:
                        pass
                    # Debug: confirmar en consola los campos de dirección guardados
                    try:
                        print(f"[DEBUG] visita actualizada _id={actualizado.get('_id')} folio={actualizado.get('folio_visita')} direccion={actualizado.get('direccion')} calle_numero={actualizado.get('calle_numero')} colonia={actualizado.get('colonia')} municipio={actualizado.get('municipio')} cp={actualizado.get('cp')}")
                    except Exception:
                        pass
                    messagebox.showinfo("OK", f"Visita {actualizado.get('folio_visita','-')} actualizada")
                    # Refrescar dropdown de pendientes por si cambió estatus/tipo
                    try:
                        if hasattr(self, '_refresh_pending_folios_dropdown'):
                            self._refresh_pending_folios_dropdown()
                    except Exception:
                        pass

                    # También actualizar en memoria/pending_folios si existe
                    try:
                        if hasattr(self, 'pending_folios') and isinstance(self.pending_folios, list):
                            for j, p in enumerate(self.pending_folios):
                                try:
                                    pid = p.get('_id') or p.get('id')
                                    if pid == id_ or p.get('folio_visita') == id_ or p.get('folio_acta') == id_:
                                        self.pending_folios[j].update(nuevos or {})
                                except Exception:
                                    continue
                            # persistir cambios
                            try:
                                self._save_pending_folios()
                            except Exception:
                                pass
                    except Exception:
                        pass

                    return

            # Si no encontramos coincidencias, mostrar advertencia (no lanzar excepción)
            messagebox.showerror("Error", "No se encontró la visita para actualizar")
        except Exception as e:
            messagebox.showerror("Error actualizando", str(e))

    def registrar_visita_automatica(self, resultado_dictamenes):
        """Registra automáticamente una visita al generar dictámenes con información de folios"""
        try:
            if not self.cliente_seleccionado:
                return

            # Obtener datos del formulario
            folio_visita = self.entry_folio_visita.get().strip()
            folio_acta = self.entry_folio_acta.get().strip()
            fecha_inicio = self.entry_fecha_inicio.get().strip()
            fecha_termino = self.entry_fecha_termino.get().strip()
            hora_inicio = self.entry_hora_inicio.get().strip()
            hora_termino = self.entry_hora_termino.get().strip()
            # Leer supervisor de forma segura (proviene de la tabla de relación; el campo UI fue removido)
            safe_supervisor_widget = getattr(self, 'entry_supervisor', None)
            try:
                supervisor = safe_supervisor_widget.get().strip() if safe_supervisor_widget and safe_supervisor_widget.winfo_exists() else ""
            except Exception:
                supervisor = ""

            # Convertir horas a formato consistente (24h para almacenamiento)
            def estandarizar_hora_24h(hora_str):
                """Estandariza hora a formato 24h HH:MM"""
                if not hora_str or hora_str.strip() == "":
                    return ""
                
                try:
                    hora_str = str(hora_str).strip()
                    # Reemplazar punto por dos puntos
                    hora_str = hora_str.replace(".", ":")
                    
                    if ":" in hora_str:
                        partes = hora_str.split(":")
                        hora = int(partes[0].strip())
                        minutos = partes[1].strip()[:2]
                        
                        # Asegurar rango válido
                        if hora < 0 or hora > 23:
                            hora = 0
                        
                        # Formatear a 2 dígitos
                        return f"{hora:02d}:{minutos}"
                    else:
                        return hora_str
                except:
                    return hora_str
            
            # Estandarizar horas a 24h
            hora_inicio_24h = estandarizar_hora_24h(hora_inicio)
            hora_termino_24h = estandarizar_hora_24h(hora_termino)
            
            # Formatear horas a 12h para visualización
            hora_inicio_formateada = self._formatear_hora_12h(hora_inicio_24h) if hora_inicio_24h else ""
            hora_termino_formateada = self._formatear_hora_12h(hora_termino_24h) if hora_termino_24h else ""

            # Si no hay fecha/hora de término, usar la actual
            if not fecha_termino:
                fecha_termino = datetime.now().strftime("%d/%m/%Y")
            if not hora_termino_24h:
                hora_termino_24h = datetime.now().strftime("%H:%M")
                hora_termino_formateada = self._formatear_hora_12h(hora_termino_24h)

            # CARGAR DATOS DE TABLA DE RELACIÓN SI EXISTEN
            datos_tabla = []
            # Si el generador devolvió una tabla actualizada con folios asignados, preferirla
            tabla_actualizada_path = None
            try:
                if isinstance(resultado_dictamenes, dict):
                    tabla_actualizada_path = resultado_dictamenes.get('tabla_relacion_actualizada')
            except Exception:
                tabla_actualizada_path = None

            if tabla_actualizada_path and os.path.exists(tabla_actualizada_path):
                try:
                    with open(tabla_actualizada_path, 'r', encoding='utf-8') as f:
                        datos_tabla = json.load(f)
                except Exception:
                    datos_tabla = []
            else:
                if self.archivo_json_generado and os.path.exists(self.archivo_json_generado):
                    try:
                        with open(self.archivo_json_generado, 'r', encoding='utf-8') as f:
                            datos_tabla = json.load(f)
                    except Exception:
                        datos_tabla = []

            # Determinar si debemos persistir el contador global: solo si el generador
            # realmente creó archivos (resultado_dictamenes contiene 'archivos')
            persist_flag = False
            try:
                if isinstance(resultado_dictamenes, dict) and resultado_dictamenes.get('archivos'):
                    persist_flag = True
            except Exception:
                persist_flag = False

            # Guardar folios específicos para esta visita solo SI el generador
            # realmente creó documentos (evita escribir en `folios_visitas` en
            # cargas/previas que luego se limpian).
            if persist_flag and datos_tabla:
                self.guardar_folios_visita(folio_visita, datos_tabla, persist_counter=persist_flag)
                # Crear respaldo persistente de la tabla_de_relacion para visitas generadas
                try:
                    tabla_relacion_path = os.path.join(DATA_DIR, 'tabla_de_relacion.json')
                    if os.path.exists(tabla_relacion_path):
                        backup_dir = os.path.join(DATA_DIR, 'tabla_relacion_backups')
                        os.makedirs(backup_dir, exist_ok=True)
                        ts = datetime.now().strftime('%Y%m%d%H%M%S')
                        # Marcar como PERSIST para que no sea eliminado por limpiar
                        dest_name = f"tabla_de_relacion_{folio_visita}_PERSIST_{ts}.json"
                        try:
                            shutil.copyfile(tabla_relacion_path, os.path.join(backup_dir, dest_name))
                            print(f"📦 Respaldo persistente creado: {dest_name}")
                        except Exception as e:
                            print(f"⚠️ No se pudo crear respaldo persistente: {e}")
                except Exception:
                    pass
            else:
                print(f"ℹ️ No se almacenaron folios para visita {folio_visita} porque no se generaron documentos o no hay tabla.")

            # ===== EXTRACCIÓN DE FOLIOS NUMÉRICOS ÚNICOS DE LA TABLA DE RELACIÓN =====
            folios_numericos_set = set()
            total_registros = 0

            if datos_tabla:
                for registro in datos_tabla:
                    total_registros += 1
                    if "FOLIO" in registro:
                        folio_valor = registro["FOLIO"]
                        if (folio_valor is not None and str(folio_valor).strip() != "" and
                            str(folio_valor).lower() != "nan" and str(folio_valor).lower() != "none"):
                            try:
                                folio_int = int(float(folio_valor))
                                folios_numericos_set.add(folio_int)
                            except (ValueError, TypeError):
                                pass

            # Ordenar folios únicos
            folios_numericos_ordenados = sorted(folios_numericos_set)

            # Formatear información de folios para mostrar y persistir
            if folios_numericos_ordenados:
                if len(folios_numericos_ordenados) == 1:
                    folios_str = f"Folio: {folios_numericos_ordenados[0]:06d}"
                else:
                    folios_str = f"{folios_numericos_ordenados[0]:06d} - {folios_numericos_ordenados[-1]:06d}"
            else:
                if total_registros > 0:
                    folios_str = f"Total registros: {total_registros} (sin folios numéricos)"
                else:
                    folios_str = "No se encontraron folios"

            # ===== EXTRACCIÓN DE NORMAS DE LA TABLA DE RELACIÓN =====
            normas_encontradas = set()  # Usamos set para evitar duplicados
            
            if datos_tabla:
                # Cargar el archivo de normas
                normas_path = os.path.join(APP_DIR, "data", "Normas.json")
                
                if os.path.exists(normas_path):
                    with open(normas_path, 'r', encoding='utf-8') as f:
                        normas_data = json.load(f)
                    
                    # Crear un diccionario para mapear números de norma a códigos NOM completos
                    normas_mapeadas = {}
                    for norma_obj in normas_data:
                        if isinstance(norma_obj, dict) and "NOM" in norma_obj:
                            nom_code = norma_obj["NOM"]
                            # Extraer el número del código NOM
                            try:
                                import re
                                match = re.search(r'NOM-(\d+)-', nom_code)
                                if match:
                                    num_norma = int(match.group(1))
                                    normas_mapeadas[num_norma] = nom_code
                            except (ValueError, AttributeError):
                                pass
                
                # Buscar normas UVA en la tabla de relación
                for registro in datos_tabla:
                    if "NORMA UVA" in registro:
                        norma_uva = registro["NORMA UVA"]
                        # Verificar que no sea NaN o vacío
                        if norma_uva is not None and str(norma_uva).strip() != "" and str(norma_uva).lower() != "nan":
                            try:
                                # Convertir a entero (puede venir como string "4" o float 4.0)
                                norma_num = int(float(norma_uva))
                                
                                # Buscar la NOM correspondiente en el mapeo
                                if norma_num in normas_mapeadas:
                                    normas_encontradas.add(normas_mapeadas[norma_num])
                                else:
                                    # Si no encontramos mapeo, agregar como "NORMA UVA X"
                                    normas_encontradas.add(f"NORMA UVA {norma_num}")
                            except (ValueError, TypeError):
                                # Si no se puede convertir a número, agregar el valor tal cual
                                if str(norma_uva).strip():
                                    normas_encontradas.add(str(norma_uva).strip())
            
            # Crear cadena de normas (ordenar alfabéticamente para consistencia)
            normas_str = ", ".join(sorted(normas_encontradas)) if normas_encontradas else ""

            # ===== EXTRACCIÓN DE FIRMAS (SUPERVISORES) DE LA TABLA DE RELACIÓN =====
            supervisores_encontrados = set()  # Usamos set para evitar duplicados
            firmas_originales = set()  # Para guardar las firmas originales también
            
            if datos_tabla:
                # Cargar el archivo de firmas
                firmas_path = os.path.join(APP_DIR, "data", "Firmas.json")
                
                # Prepare mapping dict even if file missing
                firmas_mapeadas = {}
                # collector for suggestions for unrecognized firmas
                firma_sugerencias = {}
                if os.path.exists(firmas_path):
                    with open(firmas_path, 'r', encoding='utf-8') as f:
                        firmas_data = json.load(f)

                    # Crear un diccionario para mapear firmas (normalizadas) a nombres completos
                    for inspector_obj in firmas_data:
                        if isinstance(inspector_obj, dict) and "FIRMA" in inspector_obj and "NOMBRE DE INSPECTOR" in inspector_obj:
                            raw_firma = inspector_obj["FIRMA"]
                            nombre_completo = inspector_obj["NOMBRE DE INSPECTOR"]
                            if raw_firma is None:
                                continue
                            key = str(raw_firma).strip().upper()
                            firmas_mapeadas[key] = nombre_completo
                
                # Buscar firmas en la tabla de relación
                for registro in datos_tabla:
                    if "FIRMA" in registro:
                        firma = registro["FIRMA"]
                        # Verificar que no sea NaN o vacío
                        if firma is not None and str(firma).strip() != "" and str(firma).lower() != "nan":
                            firma_str = str(firma).strip()
                            # Permitir múltiples firmas en la misma celda separadas por , ; / |
                            import re, difflib
                            parts = [p.strip() for p in re.split(r"[;,/|]+", firma_str) if p.strip()]
                            for part in parts:
                                firmas_originales.add(part)
                                # Normalizar para búsqueda
                                buscar_clave = part.upper()
                                if buscar_clave in firmas_mapeadas:
                                    supervisores_encontrados.add(firmas_mapeadas[buscar_clave])
                                else:
                                    # intentar búsqueda por similitud en claves conocidas
                                    posibles = difflib.get_close_matches(buscar_clave, list(firmas_mapeadas.keys()), n=3, cutoff=0.6)
                                    if posibles:
                                        # sugerir la primera coincidencia (sin cambiar mapeo)
                                        sugerencias = [firmas_mapeadas[p] for p in posibles if p in firmas_mapeadas]
                                        # anotar como no reconocida pero con sugerencia
                                        supervisores_encontrados.add(part)
                                        # guardar sugerencia en un dict auxiliar
                                        try:
                                            if 'firma_sugerencias' not in locals():
                                                firma_sugerencias = {}
                                            firma_sugerencias[part] = sugerencias
                                        except Exception:
                                            pass
                                    else:
                                        # ninguna sugerencia, marcar como no reconocida
                                        supervisores_encontrados.add(part)
            
            # Crear cadena de supervisores (ordenar alfabéticamente)
            supervisores_str = ", ".join(sorted(supervisores_encontrados)) if supervisores_encontrados else ""

            # Construir advertencias para firmas no reconocidas (si las hay)
            firma_warnings = ""
            try:
                # firmas_originales contiene las entradas tal cual del archivo
                desconocidas = [s for s in sorted(firmas_originales) if s and s.upper() not in firmas_mapeadas]
                if desconocidas:
                    parts = []
                    for d in desconocidas:
                        sug = firma_sugerencias.get(d)
                        if sug:
                            parts.append(f"{d} (sugerencia: {', '.join(sug)})")
                        else:
                            parts.append(f"{d}")
                    firma_warnings = "Firmas no reconocidas: " + "; ".join(parts)
            except Exception:
                firma_warnings = ""
            
            # Determinar qué supervisor mostrar en el campo principal
            # Prioridad: 1. Supervisores de la tabla, 2. Supervisor del formulario
            supervisor_mostrar = supervisores_str if supervisores_str else supervisor

            # Determinar tipo de documento para visitas automáticas (usar selección si existe)
            tipo_documento = (self.combo_tipo_documento.get().strip()
                               if hasattr(self, 'combo_tipo_documento') else "Dictamen")

            # Crear payload para visita automática con información de folios
            payload = {
                "folio_visita": folio_visita,
                "folio_acta": folio_acta or f"AC{self.current_folio}",
                "fecha_inicio": fecha_inicio or datetime.now().strftime("%d/%m/%Y"),
                "fecha_termino": fecha_termino,
                "hora_inicio_24h": hora_inicio_24h or datetime.now().strftime("%H:%M"),
                "hora_termino_24h": hora_termino_24h or datetime.now().strftime("%H:%M"),
                "hora_inicio": hora_inicio_formateada or self._formatear_hora_12h(datetime.now().strftime("%H:%M")),
                "hora_termino": hora_termino_formateada,
                "norma": normas_str,  # Normas encontradas
                "cliente": self.cliente_seleccionado['CLIENTE'],
                "nfirma1": supervisor_mostrar or " ",  # Supervisor principal (prioridad a los de la tabla)
                "nfirma2": "",
                "estatus": "Completada",
                "tipo_documento": tipo_documento,
                "folios_utilizados": folios_str,  # Información formateada de folios (únicos)
                "total_folios": len(folios_numericos_ordenados),
                "total_folios_numericos": len(folios_numericos_ordenados),
                "supervisores_tabla": supervisores_str,  # Todos los supervisores de la tabla
                "supervisor_formulario": supervisor  # Supervisor del formulario (por si se necesita)
            }
            # Añadir advertencias de firmas no reconocidas al payload si existen
            if firma_warnings:
                payload['firma_warnings'] = firma_warnings

            # Guardar visita automática
            self.hist_create_visita(payload, es_automatica=True)
            
            # Preparar nueva visita después de guardar
            self.crear_nueva_visita()
            
        except Exception as e:
            print(f"⚠️ Error registrando visita automática: {e}")

    def limpiar_archivo(self):
        self.archivo_excel_cargado = None
        self.archivo_json_generado = None
        self.json_filename = None
        
        # Limpiar también la información de folios
        if hasattr(self, 'info_folios_actual'):
            del self.info_folios_actual

        self.info_archivo.configure(
            text="No se ha cargado ningún archivo",
            text_color=STYLE["texto_claro"]
        )

        self.boton_cargar_excel.configure(state="normal")
        self.boton_limpiar.configure(state="disabled")
        self.boton_generar_dictamen.configure(state="disabled")

        self.etiqueta_estado.configure(text="", text_color=STYLE["texto_claro"])
        self.check_label.configure(text="")
        self.barra_progreso.set(0)
        self.etiqueta_progreso.configure(text="")

        try:
            data_dir = DATA_DIR
            os.makedirs(data_dir, exist_ok=True)

            # Archivos a eliminar (pero NO los de folios_visitas)
            archivos_a_eliminar = [
                "base_etiquetado.json",
                "tabla_de_relacion.json"
            ]
            
            archivos_eliminados = []
            
            for archivo in archivos_a_eliminar:
                ruta_archivo = os.path.join(data_dir, archivo)
                if os.path.exists(ruta_archivo):
                    os.remove(ruta_archivo)
                    archivos_eliminados.append(archivo)
                    print(f"🗑️ {archivo} eliminado correctamente.")
            
            if archivos_eliminados:
                print(f"✅ Se eliminaron {len(archivos_eliminados)} archivos: {', '.join(archivos_eliminados)}")
            else:
                print("ℹ️ No se encontraron archivos para eliminar.")

            self.archivo_etiquetado_json = None
            self.info_etiquetado.configure(text="")
            self.info_etiquetado.pack_forget()

        except Exception as e:
            print(f"⚠️ Error al eliminar archivos: {e}")

        messagebox.showinfo("Limpieza completa", "Los datos del archivo y el etiquetado han sido limpiados.\n\nNota: Los archivos de folios por visita se conservan en la carpeta 'folios_visitas'.")

    def _crear_formulario_visita(self, datos=None):
        """Crea un formulario modal para editar visitas con disposición organizada
        Añade un combobox de domicilios dependiente del cliente para permitir
        seleccionar la dirección registrada y guardarla en la visita.
        """
        datos = datos or {}
        modal = ctk.CTkToplevel(self)
        modal.title("Editar Visita")
        modal.geometry("1200x600")  # Aumentado altura para mejor visibilidad
        modal.transient(self)
        modal.grab_set()

        # Centrar ventana
        modal.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - modal.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - modal.winfo_height()) // 2
        modal.geometry(f"+{x}+{y}")
        
        # Frame principal
        main_frame = ctk.CTkFrame(modal, fg_color=STYLE["surface"], corner_radius=0)
        main_frame.pack(fill="both", expand=True, padx=0, pady=0)
        
        # Título
        title_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        title_frame.pack(fill="x", padx=25, pady=(20, 10))
        
        ctk.CTkLabel(
            title_frame,
            text="✏️ Editar Visita",
            font=FONT_SUBTITLE,
            text_color=STYLE["texto_oscuro"]
        ).pack(anchor="w")
        
        # Línea separadora
        separador = ctk.CTkFrame(main_frame, fg_color=STYLE["borde"], height=1)
        separador.pack(fill="x", padx=25, pady=(0, 10))
        # Mostrar advertencias de firmas (si vienen en los datos)
        try:
            if datos and isinstance(datos, dict) and datos.get('firma_warnings'):
                fw = str(datos.get('firma_warnings') or '')
                if fw:
                    try:
                        warn_frame = ctk.CTkFrame(main_frame, fg_color='transparent')
                        warn_frame.pack(fill='x', padx=25, pady=(0, 8))
                        ctk.CTkLabel(warn_frame, text=f"⚠️ {fw}", font=FONT_SMALL, text_color=STYLE.get('advertencia', '#ff1500'), wraplength=1100, anchor='w').pack(anchor='w')
                    except Exception:
                        pass
        except Exception:
            pass
        
        # Frame para contenido principal con scroll
        content_scroll = ctk.CTkScrollableFrame(
            main_frame, 
            fg_color="transparent",
            scrollbar_button_color=STYLE["primario"],
            scrollbar_button_hover_color=STYLE["primario"],
            height=350
        )
        content_scroll.pack(fill="both", expand=True, padx=25, pady=(5, 10))
        
        # Frame para contenido en grid (3 columnas para mejor organización)
        content_frame = ctk.CTkFrame(content_scroll, fg_color="transparent")
        content_frame.pack(fill="both", expand=True)
        
        # Configurar 4 columnas (última para inspectores)
        content_frame.grid_columnconfigure(0, weight=1)
        content_frame.grid_columnconfigure(1, weight=1)
        content_frame.grid_columnconfigure(2, weight=1)
        content_frame.grid_columnconfigure(3, weight=3)
        
        entries = {}
        # Variable closure para almacenar la dirección raw seleccionada
        selected_address_raw = {}
        # Cargar listas de normas e inspectores para helpers del modal
        try:
            normas_path = os.path.join(DATA_DIR, 'Normas.json')
            if os.path.exists(normas_path):
                with open(normas_path, 'r', encoding='utf-8') as nf:
                    normas_data = json.load(nf)
                    normas_list = [n.get('NOM') or n.get('NOMBRE') or str(n) for n in (normas_data or [])]
            else:
                normas_list = []
        except Exception:
            normas_list = []

        try:
            firmas_path = os.path.join(DATA_DIR, 'Firmas.json')
            if os.path.exists(firmas_path):
                with open(firmas_path, 'r', encoding='utf-8') as ff:
                    firmas_data = json.load(ff)
                    inspectores_list = [f.get('NOMBRE DE INSPECTOR') or f.get('NOMBRE') or '' for f in (firmas_data or [])]
                    # Mapa rápido nombre -> normas acreditadas
                    try:
                        firmas_map = {}
                        for f in (firmas_data or []):
                            name = f.get('NOMBRE DE INSPECTOR') or f.get('NOMBRE') or ''
                            normas_ac = f.get('Normas acreditadas') or f.get('Normas Acreditadas') or f.get('Normas') or []
                            firmas_map[name] = normas_ac or []
                    except Exception:
                        firmas_map = {}
            else:
                inspectores_list = []
        except Exception:
            inspectores_list = []
            firmas_map = {}
        
        # Definir campos organizados por columnas
        campos_por_columna = [
            [  # Columna 0: Información básica
                ("fecha_inicio", "Fecha Inicio"),
                ("fecha_termino", "Fecha Termino"),
                ("tipo_documento", "Tipo de documento"),
                ("folio_visita", "Folio Visita"),
                ("folio_acta", "Folio Acta"),
                ("folios_utilizados", "Folios Utilizados"),
            ],
            [  # Columna 1: normas (normas se mostrará aquí)
                
                ("norma", ""),
            ],
            [  # Columna 2: Cliente y estatus
                ("cliente", "Cliente"),
                ("direccion", "Domicilio registrado"),
                ("estatus", "Estatus"),
            ],
            [  # Columna 3: Inspectores (UI)
                # esta columna se llenará con la lista de inspectores (checkboxes)
            ]
        ]
        
        # Crear campos para cada columna
        col_frames = []
        for col_idx, campos in enumerate(campos_por_columna):
            col_frame = ctk.CTkFrame(content_frame, fg_color="transparent")
            col_frame.grid(row=0, column=col_idx, padx=10, pady=0, sticky="nsew")
            content_frame.grid_columnconfigure(col_idx, weight=1)
            col_frames.append(col_frame)
            
            for i, (key, label) in enumerate(campos):
                field_frame = ctk.CTkFrame(col_frame, fg_color="transparent")
                field_frame.pack(fill="x", pady=(0, 12))
                
                ctk.CTkLabel(
                    field_frame, 
                    text=label, 
                    anchor="w", 
                    font=FONT_SMALL,
                    text_color=STYLE["texto_oscuro"]
                ).pack(anchor="w", pady=(0, 5))
                
                if key == "tipo_documento":
                    # ComboBox para tipo de documento
                    opciones_tipo = ["Dictamen", "Negación de dictamen", "Constancia", "Negación de constancia"]
                    ent = ctk.CTkComboBox(field_frame, values=opciones_tipo, font=FONT_SMALL, state="readonly", height=35, corner_radius=8)
                    ent.pack(fill="x")
                    ent.set(datos.get("tipo_documento", "Dictamen"))
                    entries[key] = ent
                    continue

                if key == "norma":                    
                    ent = None
                    entries[key] = None
                    entries['_norma_summary'] = None
                    continue
                if key in ("fecha_inicio", "fecha_termino"):
                    # intentar usar DateEntry de tkcalendar si está disponible
                    DateEntry = None
                    try:
                        from tkcalendar import DateEntry as _DateEntry
                        DateEntry = _DateEntry
                    except Exception:
                        DateEntry = None

                    if DateEntry is not None:
                        try:
                            # crear un estilo ttk para que el DateEntry visualmente encaje con CTk
                            try:
                                style = ttk.Style()
                                style_name = f"CTkDate.{key}.TEntry"
                                style.configure(style_name, fieldbackground=STYLE.get('surface'), background=STYLE.get('surface'), foreground=STYLE.get('texto_oscuro'))
                            except Exception:
                                style_name = None
                            kwargs = {'date_pattern': 'dd/MM/yyyy', 'width': 16}
                            if style_name:
                                kwargs['style'] = style_name
                            ent = DateEntry(field_frame, **kwargs)
                            ent.pack(fill='x')
                            if datos and key in datos and datos.get(key):
                                try:
                                    ent.set_date(datos.get(key))
                                except Exception:
                                    try:
                                        ent.set_date(datetime.strptime(datos.get(key), '%d/%m/%Y'))
                                    except Exception:
                                        pass
                        except Exception:
                            ent = ctk.CTkEntry(field_frame, height=35, corner_radius=8, font=FONT_SMALL)
                            ent.pack(fill='x')
                            if datos and key in datos:
                                ent.insert(0, str(datos.get(key, '')))
                    else:
                        ent = ctk.CTkEntry(field_frame, height=35, corner_radius=8, font=FONT_SMALL)
                        ent.pack(fill='x')
                        if datos and key in datos:
                            ent.insert(0, str(datos.get(key, '')))
                    entries[key] = ent
                    continue
                if key == 'folios_utilizados':
                    # Mostrar folios usados en modo solo lectura con formato legible
                    ent = ctk.CTkEntry(field_frame, height=35, corner_radius=8, font=FONT_SMALL)
                    ent.pack(fill='x')
                    try:
                        val = ''
                        if datos and datos.get('folios_utilizados'):
                            val = str(datos.get('folios_utilizados'))
                        elif getattr(self, 'info_folios_actual', None):
                            val = str(self.info_folios_actual)
                        # si el valor numérico viene suelto, formatearlo
                        if val and not val.lower().startswith('folio') and re.match(r'^\d{1,6}(-|\s|$)', val):
                            # normalizar a 6 dígitos
                            try:
                                n = int(str(val).split()[0].split('-')[0])
                                val = f"Folio: {n:06d}"
                            except Exception:
                                pass
                        ent.insert(0, val)
                        ent.configure(state='disabled')
                    except Exception:
                        try:
                            ent.insert(0, '')
                            ent.configure(state='disabled')
                        except Exception:
                            pass
                    entries[key] = ent
                    continue
                if key == "cliente":
                    # Obtener lista de clientes (para dropdown)
                    clientes_lista = ['Seleccione un cliente...']
                    if hasattr(self, 'clientes_data') and self.clientes_data:
                        for cliente in self.clientes_data:
                            if not isinstance(cliente, dict):
                                continue
                            name = cliente.get('CLIENTE') or cliente.get('RAZÓN SOCIAL ') or cliente.get('RAZON SOCIAL') or cliente.get('RAZON_SOCIAL') or cliente.get('RFC') or cliente.get('NÚMERO_DE_CONTRATO')
                            if name:
                                clientes_lista.append(name)

                    # Crear combobox desplegable para clientes
                    ent = ctk.CTkComboBox(field_frame, values=clientes_lista, font=FONT_SMALL, dropdown_font=FONT_SMALL, state='readonly', height=35, corner_radius=8)
                    ent.pack(fill='x')
                    entries[key] = ent
                    # Callback cuando se seleccione un cliente en el modal
                    def _on_cliente_modal_select(val):
                        nombre = val
                        # buscar dict del cliente
                        encontrado = None
                        try:
                            for c in (self.clientes_data or []):
                                if not isinstance(c, dict):
                                    continue
                                name = c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or c.get('RAZON_SOCIAL') or c.get('RFC') or c.get('NÚMERO_DE_CONTRATO')
                                if name and name == nombre:
                                    encontrado = c
                                    break
                        except Exception:
                            encontrado = None

                        # Construir lista de domicilios (mismo heurístico usado en actualizar_cliente_seleccionado)
                        domicilios = []
                        raw = []
                        if encontrado:
                            try:
                                direcciones = encontrado.get('DIRECCIONES')
                                if isinstance(direcciones, list) and direcciones:
                                    for d in direcciones:
                                        if not isinstance(d, dict):
                                            continue
                                        parts = []
                                        for k in ('CALLE Y NO', 'CALLE', 'CALLE_Y_NO', 'CALLE_Y_NRO', 'NUMERO'):
                                            v = d.get(k) or d.get(k.upper()) if isinstance(d, dict) else None
                                            if v:
                                                parts.append(str(v))
                                        for k in ('COLONIA O POBLACION', 'COLONIA'):
                                            v = d.get(k)
                                            if v:
                                                parts.append(str(v))
                                        for k in ('MUNICIPIO O ALCADIA', 'MUNICIPIO'):
                                            v = d.get(k)
                                            if v:
                                                parts.append(str(v))
                                        if d.get('CIUDAD O ESTADO'):
                                            parts.append(str(d.get('CIUDAD O ESTADO')))
                                        if d.get('CP'):
                                            parts.append(str(d.get('CP')))
                                        addr = ", ".join(parts).strip()
                                        if addr:
                                            domicilios.append(addr)
                                            raw.append(d)

                                # fallback: intentar con campos a nivel superior
                                if not domicilios:
                                    parts = []
                                    for k in ('CALLE Y NO', 'CALLE', 'CALLE_Y_NO'):
                                        v = encontrado.get(k) or encontrado.get(k.upper())
                                        if v:
                                            parts.append(str(v))
                                    for k in ('COLONIA O POBLACION', 'COLONIA'):
                                        v = encontrado.get(k)
                                        if v:
                                            parts.append(str(v))
                                    for k in ('MUNICIPIO O ALCADIA', 'MUNICIPIO'):
                                        v = encontrado.get(k)
                                        if v:
                                            parts.append(str(v))
                                    if encontrado.get('CIUDAD O ESTADO'):
                                        parts.append(str(encontrado.get('CIUDAD O ESTADO')))
                                    if encontrado.get('CP') is not None:
                                        parts.append(str(encontrado.get('CP')))
                                    addr = ", ".join(parts).strip()
                                    if addr:
                                        domicilios.append(addr)
                                        raw.append({
                                            'CALLE Y NO': encontrado.get('CALLE Y NO') or encontrado.get('CALLE') or encontrado.get('CALLE_Y_NO') or '',
                                            'COLONIA O POBLACION': encontrado.get('COLONIA O POBLACION') or encontrado.get('COLONIA') or '',
                                            'MUNICIPIO O ALCADIA': encontrado.get('MUNICIPIO O ALCADIA') or encontrado.get('MUNICIPIO') or '',
                                            'CIUDAD O ESTADO': encontrado.get('CIUDAD O ESTADO') or encontrado.get('CIUDAD') or '',
                                            'CP': encontrado.get('CP')
                                        })
                            except Exception:
                                domicilios = []

                        if not domicilios:
                            domicilios = ["Domicilio no disponible"]
                            raw = [{'CALLE Y NO': '', 'COLONIA O POBLACION': '', 'MUNICIPIO O ALCADIA': '', 'CIUDAD O ESTADO': '', 'CP': ''}]

                        # configurar combobox de domicilios del modal
                        try:
                            vals = ['Seleccione un domicilio...'] + domicilios
                            if 'direccion' in entries and isinstance(entries['direccion'], ctk.CTkComboBox):
                                entries['direccion'].configure(values=vals, state='readonly')
                                entries['direccion'].set('Seleccione un domicilio...')
                                # almacenar raw list alineada y display list para referencias
                                entries['_domicilios_modal_raw'] = raw
                                entries['_domicilios_modal_display'] = domicilios

                                # handler cuando se selecciona un domicilio en el combobox
                                def _on_domicilio_select(val):
                                    try:
                                        display = entries.get('_domicilios_modal_display', []) or []
                                        rawlist = entries.get('_domicilios_modal_raw', []) or []
                                        if not display or not rawlist:
                                            selected_address_raw.clear()
                                            return
                                        if val == 'Seleccione un domicilio...':
                                            selected_address_raw.clear()
                                            return
                                        if val in display:
                                            idx = display.index(val)
                                            if idx < len(rawlist):
                                                selected_address_raw.clear()
                                                selected_address_raw.update(rawlist[idx])
                                    except Exception:
                                        pass

                                entries['direccion'].configure(command=_on_domicilio_select)

                                # si la visita ya tiene una dirección guardada, intentar seleccionarla
                                # considerar tanto 'direccion' como 'calle_numero' como posibles fuentes
                                saved_vals = []
                                if datos:
                                    if datos.get('direccion'):
                                        saved_vals.append(str(datos.get('direccion')))
                                    if datos.get('calle_numero'):
                                        saved_vals.append(str(datos.get('calle_numero')))

                                matched = False
                                for saved in saved_vals:
                                    if not saved:
                                        continue
                                    if saved in domicilios:
                                        try:
                                            entries['direccion'].set(saved)
                                            idx = domicilios.index(saved)
                                            selected_address_raw.clear()
                                            selected_address_raw.update(raw[idx])
                                            matched = True
                                            break
                                        except Exception:
                                            pass

                                # si no hubo match exacto, intentar empatar por componentes
                                if not matched and datos:
                                    parts = []
                                    # preferir campos disponibles en datos (soporta varias claves)
                                    for k in ('direccion', 'calle_numero', 'CALLE Y NO', 'CALLE'):
                                        v = datos.get(k)
                                        if v:
                                            parts.append(str(v))
                                            break
                                    if datos.get('colonia'):
                                        parts.append(str(datos.get('colonia')))
                                    if datos.get('municipio'):
                                        parts.append(str(datos.get('municipio')))
                                    if datos.get('ciudad_estado'):
                                        parts.append(str(datos.get('ciudad_estado')))
                                    if datos.get('cp'):
                                        parts.append(str(datos.get('cp')))
                                    built = ", ".join(parts).strip()
                                    if built and built in domicilios:
                                        try:
                                            entries['direccion'].set(built)
                                            idx = domicilios.index(built)
                                            selected_address_raw.clear()
                                            selected_address_raw.update(raw[idx])
                                        except Exception:
                                            pass
                                # Si sigue sin empatar exactamente, intentar empatar por fragmento de 'calle y no' o 'calle_numero'
                                if not matched and datos:
                                    fragment = None
                                    for k in ('calle_numero', 'CALLE Y NO', 'CALLE', 'direccion'):
                                        v = datos.get(k)
                                        if v:
                                            fragment = str(v).strip()
                                            break
                                    if fragment:
                                        for i, disp in enumerate(domicilios):
                                            try:
                                                if fragment and fragment.lower() in disp.lower():
                                                    entries['direccion'].set(disp)
                                                    selected_address_raw.clear()
                                                    if i < len(raw):
                                                        selected_address_raw.update(raw[i])
                                                    break
                                            except Exception:
                                                continue
                        except Exception:
                            pass

                    # enlazar callback
                    ent.configure(command=_on_cliente_modal_select)

                    # Establecer cliente si existe en datos
                    if datos and "cliente" in datos:
                        cliente_actual = datos.get("cliente", "")
                        if cliente_actual in clientes_lista:
                            ent.set(cliente_actual)
                            # forzar población de domicilios al abrir modal
                            try:
                                _on_cliente_modal_select(cliente_actual)
                            except Exception:
                                pass
                        else:
                            # intentar encontrar coincidencia en self.clientes_data por nombre
                            encontrado_cliente = None
                            try:
                                needle = cliente_actual.strip().lower()
                                for c in (self.clientes_data or []):
                                    try:
                                        name = c.get('CLIENTE') or c.get('RAZÓN SOCIAL ') or c.get('RAZON SOCIAL') or c.get('RAZON_SOCIAL') or c.get('RFC') or c.get('NÚMERO_DE_CONTRATO')
                                        if not name:
                                            continue
                                        name_s = str(name).strip().lower()
                                        if name_s == needle or needle in name_s or name_s in needle:
                                            encontrado_cliente = c
                                            break
                                    except Exception:
                                        continue
                            except Exception:
                                encontrado_cliente = None

                            if encontrado_cliente:
                                # poblar domicilios directamente desde el dict encontrado
                                try:
                                    parts_list = []
                                    domicilios = []
                                    raw = []
                                    direcciones = encontrado_cliente.get('DIRECCIONES')
                                    if isinstance(direcciones, list) and direcciones:
                                        for d in direcciones:
                                            if not isinstance(d, dict):
                                                continue
                                            parts = []
                                            for k in ('CALLE Y NO', 'CALLE', 'CALLE_Y_NO', 'CALLE_Y_NRO', 'NUMERO'):
                                                v = d.get(k) or d.get(k.upper()) if isinstance(d, dict) else None
                                                if v:
                                                    parts.append(str(v))
                                            for k in ('COLONIA O POBLACION', 'COLONIA'):
                                                v = d.get(k)
                                                if v:
                                                    parts.append(str(v))
                                            for k in ('MUNICIPIO O ALCADIA', 'MUNICIPIO'):
                                                v = d.get(k)
                                                if v:
                                                    parts.append(str(v))
                                            if d.get('CIUDAD O ESTADO'):
                                                parts.append(str(d.get('CIUDAD O ESTADO')))
                                            if d.get('CP') is not None:
                                                parts.append(str(d.get('CP')))
                                            addr = ", ".join(parts).strip()
                                            if addr:
                                                domicilios.append(addr)
                                                raw.append(d)

                                    if not domicilios:
                                        # fallback a nivel cliente
                                        parts = []
                                        for k in ('CALLE Y NO', 'CALLE', 'CALLE_Y_NO'):
                                            v = encontrado_cliente.get(k) or encontrado_cliente.get(k.upper())
                                            if v:
                                                parts.append(str(v))
                                        for k in ('COLONIA O POBLACION', 'COLONIA'):
                                            v = encontrado_cliente.get(k)
                                            if v:
                                                parts.append(str(v))
                                        for k in ('MUNICIPIO O ALCADIA', 'MUNICIPIO'):
                                            v = encontrado_cliente.get(k)
                                            if v:
                                                parts.append(str(v))
                                        if encontrado_cliente.get('CIUDAD O ESTADO'):
                                            parts.append(str(encontrado_cliente.get('CIUDAD O ESTADO')))
                                        if encontrado_cliente.get('CP') is not None:
                                            parts.append(str(encontrado_cliente.get('CP')))
                                        addr = ", ".join(parts).strip()
                                        if addr:
                                            domicilios.append(addr)
                                            raw.append({
                                                'CALLE Y NO': encontrado_cliente.get('CALLE Y NO') or encontrado_cliente.get('CALLE') or encontrado_cliente.get('CALLE_Y_NO') or '',
                                                'COLONIA O POBLACION': encontrado_cliente.get('COLONIA O POBLACION') or encontrado_cliente.get('COLONIA') or '',
                                                'MUNICIPIO O ALCADIA': encontrado_cliente.get('MUNICIPIO O ALCADIA') or encontrado_cliente.get('MUNICIPIO') or '',
                                                'CIUDAD O ESTADO': encontrado_cliente.get('CIUDAD O ESTADO') or encontrado_cliente.get('CIUDAD') or '',
                                                'CP': encontrado_cliente.get('CP')
                                            })

                                    vals = ['Seleccione un domicilio...'] + (domicilios or ['Domicilio no disponible'])
                                    if 'direccion' in entries and isinstance(entries['direccion'], ctk.CTkComboBox):
                                        entries['direccion'].configure(values=vals, state='readonly')
                                        entries['direccion'].set('Seleccione un domicilio...')
                                        entries['_domicilios_modal_raw'] = raw
                                        entries['_domicilios_modal_display'] = domicilios
                                except Exception:
                                    pass
                            else:
                                ent.set("Seleccione un cliente...")
                    else:
                        ent.set("Seleccione un cliente...")
                        
                elif key == "estatus":
                    # Combobox para estatus
                    ent = ctk.CTkComboBox(
                        field_frame,
                        values=["En proceso", "Completada", "Cancelada", "Pendiente"],
                        font=FONT_SMALL,
                        dropdown_font=FONT_SMALL,
                        state="readonly",
                        height=35,
                        corner_radius=8,
                        width=250
                    )
                    ent.pack(fill="x")
                    
                    if datos and "estatus" in datos:
                        ent.set(datos.get("estatus", "En proceso"))
                    else:
                        ent.set("En proceso")
                        
                else:
                    # Campo de texto normal, excepto casos especiales: 'direccion' y 'norma'
                    if key == 'direccion':
                        # combobox que se rellenará según cliente seleccionado
                        ent = ctk.CTkComboBox(
                            field_frame,
                            values=['Seleccione un domicilio...'],
                            font=FONT_SMALL,
                            dropdown_font=FONT_SMALL,
                            state='disabled',
                            height=35,
                            corner_radius=8
                        )
                        ent.pack(fill='x')
                        # si ya hay direccion en datos, mostrarla inmediatamente y crear mapping raw
                        if datos and datos.get('direccion'):
                            try:
                                v = str(datos.get('direccion'))
                                # preparar raw mapping a partir de campos en 'datos' cuando estén disponibles
                                raw_item = {
                                    'CALLE Y NO': datos.get('calle_numero') or (v.split(',')[0].strip() if v else ''),
                                    'COLONIA O POBLACION': datos.get('colonia') or (v.split(',')[1].strip() if len(v.split(','))>1 else ''),
                                    'MUNICIPIO O ALCADIA': datos.get('municipio') or (v.split(',')[2].strip() if len(v.split(','))>2 else ''),
                                    'CIUDAD O ESTADO': datos.get('ciudad_estado') or (v.split(',')[3].strip() if len(v.split(','))>3 else ''),
                                    'CP': datos.get('cp') or datos.get('CP') or (v.split(',')[-1].strip() if len(v.split(','))>0 else '')
                                }
                                entries['_domicilios_modal_raw'] = [raw_item]
                                entries['_domicilios_modal_display'] = [v]
                                ent.configure(values=[v], state='readonly')
                                ent.set(v)
                                # also set selected_address_raw so save picks it up
                                try:
                                    selected_address_raw.clear()
                                    selected_address_raw.update(raw_item)
                                except Exception:
                                    pass
                            except Exception:
                                pass
                    elif key == 'norma':
                        # Crear un contenedor aquí (justo debajo de Fecha Termino)
                        # y usarlo más adelante como padre del listado de normas.
                        try:
                            norma_container = ctk.CTkFrame(field_frame, fg_color='transparent')
                            # evitar valores negativos que rompen el escalado en algunos sistemas
                            norma_container.pack(fill='both', expand=True, pady=(0, 0))
                            entries['_norma_container'] = norma_container
                        except Exception:
                            entries['_norma_container'] = field_frame
                        ent = None
                    else:
                        ent = ctk.CTkEntry(
                            field_frame, 
                            height=35,
                            corner_radius=8, 
                            font=FONT_SMALL,
                            placeholder_text=f"Ingrese {label.lower()}" if key not in ["hora_inicio", "hora_termino"] else "HH:MM"
                        )
                        ent.pack(fill="x")
                        # Insertar datos si existen
                        if datos and key in datos:
                            ent.insert(0, str(datos.get(key, "")))
                
                entries[key] = ent
        
        # Helper: diferir la creación de los listados pesados (inspectores y normas)
        # para mostrar el modal rápidamente y poblar el contenido después.
        insp_placeholder = ctk.CTkFrame(col_frames[3], fg_color='transparent')
        insp_placeholder.pack(fill='both', expand=True)
        ctk.CTkLabel(insp_placeholder, text='Cargando...', font=FONT_SMALL, text_color=STYLE['texto_oscuro']).pack(anchor='center', pady=20)
        entries['_insp_placeholder'] = insp_placeholder

        # --- Comportamiento: ocultar/deshabilitar Folio Acta en el modal si el tipo es Constancia ---
        try:
            def _toggle_modal_folio_acta(val=None):
                try:
                    tipo_val = ''
                    try:
                        tipo_val = entries.get('tipo_documento').get().strip() if entries.get('tipo_documento') else ''
                    except Exception:
                        tipo_val = str(val or '').strip()
                    is_const = 'constancia' in (tipo_val or '').lower()
                    fa = entries.get('folio_acta')
                    if not fa:
                        return
                    parent = getattr(fa, 'master', None)
                    if is_const:
                        try:
                            fa.delete(0, 'end')
                        except Exception:
                            pass
                        try:
                            fa.configure(state='disabled')
                        except Exception:
                            pass
                        try:
                            if parent and getattr(parent, 'winfo_ismapped', lambda: False)():
                                parent.pack_forget()
                        except Exception:
                            pass
                    else:
                        try:
                            if parent and not getattr(parent, 'winfo_ismapped', lambda: False)():
                                parent.pack(fill='x', pady=(0, 10))
                        except Exception:
                            pass
                        try:
                            fa.configure(state='normal')
                        except Exception:
                            pass
                        try:
                            if fa.get().strip() == '':
                                fa.delete(0, 'end')
                                fa.insert(0, f"AC{self.current_folio}")
                        except Exception:
                            pass
                except Exception:
                    pass

            # asignar callback al combo del modal si existe
            if entries.get('tipo_documento') and hasattr(entries.get('tipo_documento'), 'configure'):
                try:
                    entries.get('tipo_documento').configure(command=lambda v: _toggle_modal_folio_acta(v))
                except Exception:
                    try:
                        # intentar trace si es StringVar-backed
                        v = getattr(entries.get('tipo_documento'), 'variable', None)
                        if v:
                            try:
                                v.trace_add('write', lambda *a: _toggle_modal_folio_acta())
                            except Exception:
                                pass
                    except Exception:
                        pass

            # aplicar estado inicial según datos (si vienen)
            try:
                if entries.get('tipo_documento'):
                    current = entries.get('tipo_documento').get()
                    _toggle_modal_folio_acta(current)
            except Exception:
                pass
        except Exception:
            pass

        def _populate_heavy_ui_chunked():
            try:
                # eliminar placeholder
                try:
                    insp_placeholder.destroy()
                except Exception:
                    pass

                # Crear contenedores visibles inmediatamente
                insp_frame = ctk.CTkFrame(col_frames[3], fg_color='transparent')
                insp_frame.pack(fill='both', expand=True)
                ctk.CTkLabel(insp_frame, text='Listado de Inspectores', font=FONT_SMALL, text_color=STYLE['texto_oscuro']).pack(anchor='w')
                scroll_insp = ctk.CTkScrollableFrame(insp_frame, height=100, fg_color='transparent')
                scroll_insp.pack(fill='both', expand=True, pady=(6,6), padx=(6,0))

                parent_container = entries.get('_norma_container') if entries.get('_norma_container') else col_frames[1]
                normas_frame = ctk.CTkFrame(parent_container, fg_color='transparent')
                normas_frame.pack(fill='both', expand=False, pady=(0, 0))
                ctk.CTkLabel(normas_frame, text='Listado de Normas', font=FONT_SMALL, text_color=STYLE['texto_oscuro']).pack(anchor='w', pady=(0,4))
                normas_list_frame = ctk.CTkFrame(normas_frame, fg_color='transparent')
                normas_list_frame.pack(fill='both', expand=True, pady=(0,2))

                # preparaciones comunes
                try:
                    existing_raw_insp = (datos.get('supervisores_tabla') or datos.get('nfirma1') or '') if datos else ''
                    selected_inspectores = [s.strip() for s in str(existing_raw_insp).split(',') if s.strip()]
                except Exception:
                    selected_inspectores = []

                last_insp_normas_label = ctk.CTkLabel(insp_frame, text='', font=("Inter", 11), text_color=STYLE['texto_oscuro'])
                last_insp_normas_label.pack(anchor='w', pady=(6,4))

                inspector_status_labels = {}

                def _on_insp_click(nombre, var):
                    try:
                        normas = firmas_map.get(nombre, []) if 'firmas_map' in locals() or 'firmas_map' in globals() else []
                        if normas:
                            lines = [f"{i}. {n}" for i, n in enumerate(normas, start=1)]
                            last_insp_normas_label.configure(text="\n".join(lines))
                        else:
                            last_insp_normas_label.configure(text='(Sin normas acreditadas)')
                    except Exception:
                        try:
                            last_insp_normas_label.configure(text='')
                        except Exception:
                            pass

                def update_inspector_statuses():
                    try:
                        norma_checks_local = entries.get('_norma_checks') or []
                        selected_norms = [nm for nm, v in norma_checks_local if getattr(v, 'get', lambda: '0')() in ('1', 'True', 'true')]
                        for nombre, lbl in inspector_status_labels.items():
                            try:
                                acc = set(firmas_map.get(nombre, []) or [])
                                ok = False
                                if selected_norms:
                                    ok = set(selected_norms).issubset(acc)
                                else:
                                    ok = False
                                if ok:
                                    lbl.configure(text='✓', text_color=STYLE['exito'])
                                else:
                                    lbl.configure(text='', text_color=STYLE['texto_oscuro'])
                            except Exception:
                                try:
                                    lbl.configure(text='', text_color=STYLE['texto_oscuro'])
                                except Exception:
                                    pass
                        # Actualizar resumen de normas (si existe widget de resumen)
                        try:
                            summary = entries.get('_norma_summary')
                            if summary is not None:
                                txt = ', '.join(selected_norms) if selected_norms else ''
                                try:
                                    summary.configure(state='normal')
                                except Exception:
                                    pass
                                try:
                                    # vaciar y escribir
                                    summary.delete(0, 'end')
                                    summary.insert(0, txt)
                                except Exception:
                                    try:
                                        summary_var = getattr(summary, 'set', None)
                                        if summary_var:
                                            summary.set(txt)
                                    except Exception:
                                        pass
                                try:
                                    summary.configure(state='disabled')
                                except Exception:
                                    pass
                        except Exception:
                            pass
                    except Exception:
                        pass

                # Guardar referencias en entries para que otros handlers las usen
                entries['_insp_scroll'] = scroll_insp
                entries['_normas_frame'] = normas_list_frame
                entries['_insp_status_labels'] = inspector_status_labels

                # Batching helpers
                insp_index = {'i': 0}
                norma_index = {'i': 0}
                BATCH = 10

                def _create_inspectores_batch():
                    try:
                        created = 0
                        while insp_index['i'] < len(inspectores_list or []) and created < BATCH:
                            nombre = inspectores_list[insp_index['i']]
                            try:
                                var = ctk.StringVar(value='0')
                                row = ctk.CTkFrame(scroll_insp, fg_color='transparent')
                                row.pack(fill='x', pady=(2,2), padx=(2,0))
                                row.grid_columnconfigure(0, weight=1)
                                chk = ctk.CTkCheckBox(row, text=nombre, variable=var, onvalue='1', offvalue='0', command=lambda n=nombre, v=var: _on_insp_click(n, v), font=("Inter", 11))
                                chk.grid(row=0, column=0, sticky='w')
                                status_lbl = ctk.CTkLabel(row, text='', font=("Inter", 11), text_color=STYLE['exito'])
                                status_lbl.grid(row=0, column=1, sticky='w', padx=(6,0))
                                inspector_status_labels[nombre] = status_lbl
                                try:
                                    var.trace_add('write', lambda *a, _n=nombre: update_inspector_statuses())
                                except Exception:
                                    pass
                                try:
                                    if nombre in selected_inspectores:
                                        var.set('1')
                                except Exception:
                                    pass
                                # append to list
                                lst = entries.get('_inspectores_checks') or []
                                lst.append((nombre, var))
                                entries['_inspectores_checks'] = lst
                            except Exception:
                                pass
                            insp_index['i'] += 1
                            created += 1
                        # schedule next batch if remaining
                        if insp_index['i'] < len(inspectores_list or []):
                            modal.after(10, _create_inspectores_batch)
                    except Exception:
                        pass

                def _create_normas_batch():
                    try:
                        created = 0
                        while norma_index['i'] < len(normas_list or []) and created < BATCH:
                            nm = normas_list[norma_index['i']]
                            try:
                                var = ctk.StringVar(value='0')
                                chk = ctk.CTkCheckBox(normas_list_frame, text=nm, variable=var, onvalue='1', offvalue='0')
                                chk.pack(anchor='w', fill='x')
                                try:
                                    var.trace_add('write', lambda *a: update_inspector_statuses())
                                except Exception:
                                    pass
                                if datos and nm in [n.strip() for n in (datos.get('norma') or '').split(',') if n.strip()]:
                                    try:
                                        var.set('1')
                                    except Exception:
                                        pass
                                lst = entries.get('_norma_checks') or []
                                lst.append((nm, var))
                                entries['_norma_checks'] = lst
                            except Exception:
                                pass
                            norma_index['i'] += 1
                            created += 1
                        if norma_index['i'] < len(normas_list or []):
                            modal.after(10, _create_normas_batch)
                    except Exception:
                        pass

                # Start batches quickly to keep UI responsive
                try:
                    modal.after(10, _create_inspectores_batch)
                    modal.after(10, _create_normas_batch)
                except Exception:
                    # fallback immediate creation if after fails
                    _create_inspectores_batch()
                    _create_normas_batch()

                # Población del combobox de clientes (deferred, lightweight)
                try:
                    cliente_widget = entries.get('cliente')
                    if cliente_widget and isinstance(cliente_widget, ctk.CTkComboBox):
                        clientes_lista = ['Seleccione un cliente...']
                        if hasattr(self, 'clientes_data') and self.clientes_data:
                            for cliente in self.clientes_data:
                                if not isinstance(cliente, dict):
                                    continue
                                name = cliente.get('CLIENTE') or cliente.get('RAZÓN SOCIAL ') or cliente.get('RAZON SOCIAL') or cliente.get('RAZON_SOCIAL') or cliente.get('RFC') or cliente.get('NÚMERO_DE_CONTRATO')
                                if name:
                                    clientes_lista.append(name)
                        cliente_widget.configure(values=clientes_lista, state='readonly')
                        if datos and "cliente" in datos:
                            try:
                                cliente_actual = datos.get('cliente', '')
                                if cliente_actual in clientes_lista:
                                    cliente_widget.set(cliente_actual)
                            except Exception:
                                pass
                except Exception:
                    pass

                try:
                    modal.update_idletasks()
                except Exception:
                    pass
            except Exception:
                pass

        # schedule deferred population so modal appears immediately; use tiny delay and chunking
        try:
            modal.after(1, _populate_heavy_ui_chunked)
        except Exception:
            try:
                _populate_heavy_ui_chunked()
            except Exception:
                pass

        # Frame para botones
        btn_frame = ctk.CTkFrame(main_frame, fg_color="transparent")
        btn_frame.pack(fill="x", pady=(15, 20), padx=25)
        
        def _guardar():
            # Recoger datos de todos los campos
            payload = {}
            for key, entry in entries.items():
                # Ignorar helpers internos (prefijo _ ) o valores que no sean widgets
                if str(key).startswith('_'):
                    continue
                # Si el objeto no tiene método get(), omitimos
                if not hasattr(entry, 'get'):
                    continue
                if key in ["cliente", "estatus"]:
                    # Para combobox, obtener el valor seleccionado
                    raw_value = entry.get()
                    if key == "cliente" and raw_value == "Seleccione un cliente...":
                        raw_value = ""
                    value = raw_value
                else:
                    raw_value = entry.get()
                    value = raw_value.strip() if isinstance(raw_value, str) else raw_value
                payload[key] = value
            
                # (normas) -- se recogen MÁS ABAJO fuera del bucle para evitar
                # recolecciones múltiples/duplicadas mientras iteramos `entries`.

                # Asegurar que la lista de supervisores se guarda a partir de los checkboxes de inspectores (si existen)
                try:
                    # Preferir checkboxes de inspectores si existen
                    insp_checks = entries.get('_inspectores_checks') or []
                    selected = [name for name, var in insp_checks if getattr(var, 'get', lambda: '0')() in ('1', 'True', 'true')]
                    if selected:
                        joined = ', '.join(selected)
                        payload['supervisores_tabla'] = joined
                        payload['nfirma1'] = joined
                    else:
                        # fallback: combinar inspectores previos con los del payload.nfirma1 (si el usuario escribió/insertó)
                        existing = []
                        try:
                            existing_raw = (datos.get('supervisores_tabla') or datos.get('nfirma1') or '') if datos else ''
                            existing = [s.strip() for s in str(existing_raw).split(',') if s.strip()]
                        except Exception:
                            existing = []
                        new_list = []
                        try:
                            nf = (payload.get('nfirma1') or '')
                            new_list = [s.strip() for s in str(nf).split(',') if s.strip()]
                        except Exception:
                            new_list = []
                        merged = []
                        for s in (existing + new_list):
                            if s and s not in merged:
                                merged.append(s)
                        if merged:
                            joined = ', '.join(merged)
                            payload['supervisores_tabla'] = joined
                            payload['nfirma1'] = joined
                except Exception:
                    pass

                # Recolectar normas seleccionadas (fuera del bucle) y deduplicar
            try:
                norma_checks = entries.get('_norma_checks') or []
                sel = []
                for nm, var in norma_checks:
                    try:
                        v = getattr(var, 'get', lambda: '0')()
                    except Exception:
                        v = '0'
                    if str(v).lower() in ('1', 'true', 'yes'):
                        if nm not in sel:
                            sel.append(nm)
                if sel:
                    payload['norma'] = ', '.join(sel)
                else:
                    # si no hay selección, dejar vacío (o mantener lo que ya exista)
                    payload['norma'] = payload.get('norma', '') or ''
            except Exception:
                pass

                # Validaciones
            if not payload.get("cliente"):
                messagebox.showwarning("Validación", "Por favor seleccione un cliente")
                return
            
            if not payload.get("estatus"):
                payload["estatus"] = "En proceso"
            
            # Conservar horas originales si el formulario no las incluye
            try:
                for h in ("hora_inicio", "hora_termino", "hora_inicio_24h", "hora_termino_24h"):
                    if (h not in payload or payload.get(h) in (None, "", " ")) and datos.get(h):
                        payload[h] = datos.get(h)
            except Exception:
                pass

            # Añadir componentes de dirección desde la selección modal si existen
            try:
                # Determinar valor desplegado del domicilio (si existe)
                dir_widget = entries.get('direccion') if isinstance(entries.get('direccion', None), ctk.CTkComboBox) else None
                dir_display = None
                if dir_widget:
                    try:
                        dir_display = dir_widget.get()
                    except Exception:
                        dir_display = None

                # Si existe la lista raw en entries, intentar mapear por índice
                raw_mapped = None
                try:
                    display_list = entries.get('_domicilios_modal_display') or []
                    raw_list = entries.get('_domicilios_modal_raw') or []
                    if dir_display and display_list and raw_list and dir_display in display_list:
                        idx = display_list.index(dir_display)
                        if idx < len(raw_list):
                            raw_mapped = raw_list[idx]
                except Exception:
                    raw_mapped = None

                # Priorizar raw_mapped, luego selected_address_raw (si fue seteado por el handler),
                # luego intentar parsear dir_display como respaldo
                source_raw = raw_mapped or (selected_address_raw if selected_address_raw else None)

                if source_raw:
                    # Guardar display completo en 'direccion' pero asegurar que 'calle_numero'
                    # almacene únicamente la 'CALLE Y NO' (o la mejor alternativa)
                    payload['direccion'] = dir_display or payload.get('direccion') or source_raw.get('CALLE Y NO') or source_raw.get('CALLE')
                    payload['calle_numero'] = source_raw.get('CALLE Y NO') or source_raw.get('CALLE') or payload.get('direccion') or payload.get('calle_numero')
                    payload['colonia'] = source_raw.get('COLONIA O POBLACION') or source_raw.get('COLONIA') or source_raw.get('colonia') or payload.get('colonia')
                    payload['municipio'] = source_raw.get('MUNICIPIO O ALCADIA') or source_raw.get('MUNICIPIO') or source_raw.get('municipio') or payload.get('municipio')
                    payload['ciudad_estado'] = source_raw.get('CIUDAD O ESTADO') or source_raw.get('CIUDAD') or source_raw.get('ciudad_estado') or payload.get('ciudad_estado')
                    payload['cp'] = str(source_raw.get('CP') or source_raw.get('cp') or payload.get('cp') or '')
                else:
                    # Si no hay raw, pero hay texto desplegado, intentar descomponerlo por comas
                    if dir_display and dir_display not in (None, '', 'Seleccione un domicilio...'):
                        payload['direccion'] = dir_display
                        parts = [p.strip() for p in dir_display.split(',') if p.strip()]
                        # heurística: último token suele ser CP o ciudad; asignar por posición
                        if parts:
                            payload['colonia'] = parts[1] if len(parts) > 1 else payload.get('colonia')
                            payload['municipio'] = parts[2] if len(parts) > 2 else payload.get('municipio')
                            payload['ciudad_estado'] = parts[3] if len(parts) > 3 else payload.get('ciudad_estado')
                            # intentar extraer CP numérico
                            for p in parts[::-1]:
                                s = ''.join(ch for ch in p if ch.isdigit())
                                if s:
                                    payload['cp'] = s
                                    break
            except Exception:
                pass
            # Antes de actualizar, sincronizar también atributos de instancia para UI principal
            try:
                if payload.get('direccion'):
                    self.direccion_seleccionada = payload.get('direccion')
                # sincronizar alias
                if payload.get('calle_numero') and not getattr(self, 'direccion_seleccionada', None):
                    self.direccion_seleccionada = payload.get('calle_numero')
                if payload.get('colonia'):
                    self.colonia_seleccionada = payload.get('colonia')
                if payload.get('municipio'):
                    self.municipio_seleccionado = payload.get('municipio')
                if payload.get('ciudad_estado'):
                    self.ciudad_seleccionada = payload.get('ciudad_estado')
                if payload.get('cp'):
                    self.cp_seleccionado = payload.get('cp')
                # mantener domicilio_seleccionado como display
                if payload.get('direccion'):
                    self.domicilio_seleccionado = payload.get('direccion')
            except Exception:
                pass

            # Log breve desde modal: mostrar sólo folio y cliente
            try:
                print(f"[INFO] Modal guardar: folio={payload.get('folio_visita','-')} cliente={payload.get('cliente','-')}")
            except Exception:
                try:
                    print("[INFO] Modal guardar")
                except Exception:
                    pass

            # usar id seguro (soporta _id, id o folio) para actualizar
            target_id = datos.get('_id') or datos.get('id') or datos.get('folio_visita') or datos.get('folio_acta')
            if not target_id:
                messagebox.showerror("Error", "No se pudo determinar el identificador de la visita para actualizar")
                return
            # Si el tipo es Constancia, eliminar/limpiar folio_acta del payload (no requerido)
            try:
                tipo_payload = (payload.get('tipo_documento') or payload.get('tipo') or '').strip().lower()
                if 'constancia' in tipo_payload:
                    payload['folio_acta'] = ''
            except Exception:
                pass

            self.hist_update_visita(target_id, payload)
            modal.destroy()
        
        # Botones mejorados
        ctk.CTkButton(
            btn_frame, 
            text="Cancelar", 
            command=modal.destroy,
            font=("Inter", 13),
            fg_color=STYLE["secundario"],
            hover_color="#1a1a1a",
            text_color=STYLE["texto_claro"],
            height=38,
            width=130,
            corner_radius=8
        ).pack(side="right", padx=(8, 0))
        
        ctk.CTkButton(
            btn_frame, 
            text="Guardar Cambios", 
            command=_guardar,
            font=("Inter", 13, "bold"),
            fg_color=STYLE["primario"],
            hover_color="#D4BF22",
            text_color=STYLE["secundario"],
            height=38,
            width=150,
            corner_radius=8
        ).pack(side="right")
        
        # Agregar un pequeño espaciador para empujar botones a la derecha
        ctk.CTkLabel(btn_frame, text="", fg_color="transparent").pack(side="left", expand=True)

    # -----------------------------------------------------------
    # NUEVOS MÉTODOS PARA DIAGNÓSTICO Y LIMPIEZA
    # -----------------------------------------------------------
    def verificar_integridad_datos(self):
        """Verifica la integridad de los datos cargados"""
        try:
            if not self.archivo_json_generado:
                messagebox.showwarning("Sin datos", "No hay archivo cargado para verificar")
                return
            
            with open(self.archivo_json_generado, 'r', encoding='utf-8') as f:
                datos = json.load(f)
            from collections import defaultdict
            solicitudes_by_sol = defaultdict(list)  # sol -> list of (idx, item, lista)
            solicitudes_indices = defaultdict(list)  # sol -> list of indices
            for idx, item in enumerate(datos or []):
                sol_raw = item.get('SOLICITUD') or item.get('Solicitud') or item.get('solicitud') or ''
                sol = str(sol_raw).strip() or '(Sin solicitud)'
                lista_raw = item.get('LISTA') or item.get('Lista') or item.get('lista') or ''
                lista = str(lista_raw).strip() or '(Sin lista)'
                solicitudes_by_sol[sol].append((idx, item, lista))
                solicitudes_indices[sol].append(idx)

            # Número de carpetas estimadas = solicitudes distintas
            total_carpetas = len(solicitudes_by_sol)
            carpetas_info = []
            total_dictamenes = 0

            possible_keys = ['TIPO DE DOCUMENTO', 'Tipo de documento', 'TIPO_DOCUMENTO', 'TIPO', 'TIPO_DOC', 'TIPO DE DOC', 'tipo', 'Tipo']
            tipo_map = {'D': 'Dictamen', 'C': 'Constancia', 'ND': 'Negación Dictamen', 'NC': 'Negación Constancia'}
            tipos_contador = defaultdict(int)
            tipos_no_identificados = 0
            for item in (datos or []):
                tipo_raw = None
                for k in possible_keys:
                    if k in item and item.get(k) not in (None, ''):
                        tipo_raw = item.get(k)
                        break
                if tipo_raw is None:
                    # intentar buscar por claves en mayúsculas/minúsculas dinámicamente
                    for k in list(item.keys()):
                        if 'TIPO' in k.upper() and item.get(k) not in (None, ''):
                            tipo_raw = item.get(k)
                            break

                t = str(tipo_raw).strip().upper() if tipo_raw is not None else ''
                # Normalizar casos como 'D', 'ND', 'C', 'NC' o palabras completas
                if t in tipo_map:
                    tipos_contador[t] += 1
                else:
                    # si viene la palabra completa, intentar mapear por iniciales
                    if t.startswith('NEG') and 'D' in t:
                        tipos_contador['ND'] += 1
                    elif t.startswith('NEG') and 'C' in t:
                        tipos_contador['NC'] += 1
                    elif t.startswith('D'):
                        tipos_contador['D'] += 1
                    elif t.startswith('C'):
                        tipos_contador['C'] += 1
                    elif t == '':
                        tipos_no_identificados += 1
                    else:
                        # valores inesperados los contamos como no identificados
                        tipos_no_identificados += 1

            # Construir reporte detallado
            total_tipos = sum(tipos_contador.values())
            has_tipo = total_tipos > 0

            # Preparar información de folios si está disponible (convertir a conjunto único por registro)
            folios_unicos_por_registro = None
            if hasattr(self, 'info_folios_actual') and isinstance(self.info_folios_actual, dict):
                info_f = self.info_folios_actual
                if info_f.get('total_folios', 0) > 0:
                    # construir un mapa registro->folio normalizado (formato de lista en info_folios puede contener números formateados)
                    folios_unicos_por_registro = []
                    for item in (datos or []):
                        folio_val = None
                        if 'FOLIO' in item and item.get('FOLIO') not in (None, ''):
                            folio_val = str(item.get('FOLIO')).strip()
                        else:
                            # intentar detectar cualquier clave que contenga 'FOLIO'
                            for kk in list(item.keys()):
                                if 'FOLIO' in kk.upper() and item.get(kk) not in (None, ''):
                                    folio_val = str(item.get(kk)).strip()
                                    break
                        if folio_val is None or folio_val == '' or folio_val.lower() in ('nan', 'none'):
                            folios_unicos_por_registro.append(None)
                        else:
                            try:
                                num = int(float(folio_val))
                                folios_unicos_por_registro.append(f"{num:06d}")
                            except Exception:
                                folios_unicos_por_registro.append(folio_val)

            # Si existe columna de tipo, computar totales basados en tipos reconocidos
            if has_tipo:
                total_detectados = total_tipos
                total_dictamenes = tipos_contador.get('D', 0) + tipos_contador.get('ND', 0)
            else:
                total_detectados = len(datos)

            if folios_unicos_por_registro:
                # contar folios únicos válidos
                total_dictamenes = len(set(x for x in folios_unicos_por_registro if x))
            carpetas_detalle = []

            def _get_type_key(item):
                tipo_raw = None
                for k in possible_keys:
                    if k in item and item.get(k) not in (None, ''):
                        tipo_raw = item.get(k)
                        break
                if tipo_raw is None:
                    for kk in list(item.keys()):
                        if 'TIPO' in kk.upper() and item.get(kk) not in (None, ''):
                            tipo_raw = item.get(kk)
                            break
                t = str(tipo_raw).strip().upper() if tipo_raw is not None else ''
                if t in tipo_map:
                    return t
                if t.startswith('NEG') and 'D' in t:
                    return 'ND'
                if t.startswith('NEG') and 'C' in t:
                    return 'NC'
                if t.startswith('D'):
                    return 'D'
                if t.startswith('C'):
                    return 'C'
                return None

            # Recorrer por solicitud
            for sol, entries in solicitudes_by_sol.items():
                # entries: list of (idx, item, lista)
                if has_tipo:
                    listas_set = set()
                    for idx, it, lista in entries:
                        if _get_type_key(it) is not None:
                            listas_set.add(lista)
                    cnt = len(listas_set)
                else:
                    # Preferir folios únicos cuando existan
                    if folios_unicos_por_registro:
                        folios_set = set()
                        for idx, it, lista in entries:
                            if idx < len(folios_unicos_por_registro):
                                fol = folios_unicos_por_registro[idx]
                                if fol:
                                    folios_set.add(fol)
                        if folios_set:
                            cnt = len(folios_set)
                        else:
                            # Fallback: contar listas únicas
                            cnt = len(set(lista for _, __, lista in entries))
                    else:
                        cnt = len(set(lista for _, __, lista in entries))
                carpetas_detalle.append((sol, cnt))

            carpetas_info = sorted(carpetas_detalle, key=lambda x: -x[1])
            # Recalcular total de dictámenes como suma de documentos por solicitud
            total_dictamenes = sum(c for _, c in carpetas_info)

            lines = []
            # Añadir años relevantes al reporte: año del folio (año en curso)
            # y año(s) de Solicitud de Servicio extraídos de las claves de solicitud (parte después de '/').
            try:
                current_year_full = datetime.now().year
                solicitud_years_set = set()
                for sol in solicitudes_by_sol.keys():
                    try:
                        s = str(sol or '').strip()
                        if '/' in s:
                            suf = s.split('/')[-1].strip()
                            # si viene como dos dígitos (ej. '25') convertir a 4 dígitos
                            if suf.isdigit() and len(suf) <= 2:
                                solicitud_years_set.add(str(2000 + int(suf)))
                            elif suf.isdigit() and len(suf) == 4:
                                solicitud_years_set.add(suf)
                    except Exception:
                        continue
                # Formatear líneas informativas
                lines.append(f"Año del folio (UDC): {current_year_full}")
                if solicitud_years_set:
                    sorted_sols = sorted(solicitud_years_set)
                    if len(sorted_sols) == 1:
                        lines.append(f"Año de Solicitud de Servicio: {sorted_sols[0]}")
                    else:
                        lines.append(f"Años de Solicitud de Servicio: {', '.join(sorted_sols)}")
                else:
                    lines.append("Año de Solicitud de Servicio: (no disponible)")
            except Exception:
                # no bloquear si falla el cálculo de años
                pass
            # Mostrar también folios únicos si existen
            if folios_unicos_por_registro:
                lines.append(f"🔢 Folios únicos detectados: {len(set(x for x in folios_unicos_por_registro if x))}")
            lines.append(f"🗂️ Carpetas estimadas: {total_carpetas}")
            lines.append(f"📋 Dictámenes estimados: {total_dictamenes}")
            lines.append("")
            lines.append("📂 Detalle por carpeta (nombre : documentos):")
            for name, count in carpetas_info:
                lines.append(f" - {name}: {count}")
            lines.append("")

            # Resumen de tipos de documento
            if sum(tipos_contador.values()) == 0 and tipos_no_identificados == 0:
                lines.append("ℹ️ No se detectó columna de tipo de documento.")
            else:
                total_tipos = sum(tipos_contador.values())
                # si hay registros no identificados, incluirlos en totales
                if tipos_no_identificados:
                    lines.append(f"❓ Registros sin tipo reconocido: {tipos_no_identificados}")
                if total_tipos > 0:
                    # Si todos los documentos pertenecen al mismo tipo conocido
                    if total_tipos == len(datos) and len(tipos_contador) == 1:
                        only_key = next(iter(tipos_contador))
                        lines.append(f"✅ Todos los documentos son: {tipo_map.get(only_key, only_key)}")
                    else:
                        lines.append("✅ Distribución por tipo de documento:")
                        for k, cnt in tipos_contador.items():
                            lines.append(f" - {tipo_map.get(k, k)}: {cnt}")

            # Intentar detectar firma(s) y validar acreditación
            try:
                from plantillaPDF import cargar_firmas, validar_acreditacion_inspector, cargar_normas
                firmas_map = cargar_firmas()
                # Buscar columna FIRMA en los datos cargados
                firma_keys = ['FIRMA', 'Firma', 'firma', 'CODIGO_FIRMA']
                found_codes = set()
                for item in (datos or []):
                    for fk in firma_keys:
                        if fk in item and item.get(fk) not in (None, ''):
                            found_codes.add(str(item.get(fk)).strip())
                # Si no se encontró en el archivo cargado, intentar buscar en tabla_de_relacion
                if not found_codes:
                    df_rel = None
                    try:
                        df_rel = cargar_tabla_relacion()
                    except Exception:
                        df_rel = None

                    if df_rel is not None and not df_rel.empty:
                        # detectar nombre de columna posible para firma
                        firma_col = None
                        for c in ('FIRMA', 'Firma', 'firma', 'CODIGO_FIRMA'):
                            if c in df_rel.columns:
                                firma_col = c
                                break
                        solicitud_col = None
                        for c in ('SOLICITUD', 'Solicitud', 'solicitud', 'NUMERO_SOLICITUD'):
                            if c in df_rel.columns:
                                solicitud_col = c
                                break
                        if firma_col and solicitud_col:
                            for sol in solicitudes_by_sol.keys():
                                matches = df_rel[df_rel[solicitud_col].astype(str).str.strip() == str(sol).strip()]
                                for _, r in matches.iterrows():
                                    v = r.get(firma_col)
                                    if v not in (None, ''):
                                        found_codes.add(str(v).strip())
                        # Additionally, as a robust fallback, collect ALL distinct non-empty values
                        # from any plausible FIRMA column in the tabla_de_relacion so the preview
                        # reflects every signature present in the source data.
                        try:
                            for c_try in ('FIRMA', 'Firma', 'firma', 'CODIGO_FIRMA', 'INSPECTOR', 'Inspector', 'inspector'):
                                if c_try in df_rel.columns:
                                    try:
                                        vals = df_rel[c_try].dropna().astype(str).str.strip()
                                        for vv in vals.unique():
                                            if vv and str(vv).strip().lower() != 'nan':
                                                found_codes.add(str(vv).strip())
                                    except Exception:
                                        continue
                        except Exception:
                            pass
                # Preparar reporte de firmas
                # Normalize and try to map found tokens to known firma CODES so validation succeeds
                try:
                    import unicodedata, difflib
                    def _norm_token(s):
                        try:
                            s2 = str(s or '').strip().upper()
                            s2 = unicodedata.normalize('NFKD', s2)
                            s2 = ''.join(ch for ch in s2 if not unicodedata.combining(ch))
                            # keep only alphanumerics for matching
                            s2 = ''.join(ch for ch in s2 if ch.isalnum())
                            return s2
                        except Exception:
                            return str(s or '').strip().upper()

                    norm_code_map = {}
                    norm_name_map = {}
                    try:
                        for c_code, info in (firmas_map or {}).items():
                            try:
                                nk = _norm_token(c_code)
                                if nk:
                                    norm_code_map[nk] = c_code
                                nombre = (info.get('nombre') or info.get('NOMBRE DE INSPECTOR') or '')
                                nn = _norm_token(nombre)
                                if nn and nn not in norm_name_map:
                                    norm_name_map[nn] = c_code
                                # also index name parts (surnames / tokens)
                                for part in str(nombre).split():
                                    p = _norm_token(part)
                                    if p and p not in norm_name_map:
                                        norm_name_map[p] = c_code
                            except Exception:
                                continue
                    except Exception:
                        norm_code_map = {}
                        norm_name_map = {}

                    suggestions_by_token = {}
                    effective_codes = set(found_codes)
                    try:
                        found_codes = set(effective_codes)
                    except Exception:
                        pass
                except Exception:
                    pass

                if found_codes:
                    # Construir lista de normas requeridas (puede haber varias)
                    norma_reqs = []
                    # intentar extraer todas las normas desde los datos
                    try:
                        for k in ('NORMA UVA', 'NORMA', 'NORMA_UVA'):
                            for it in (datos or []):
                                try:
                                    v = it.get(k)
                                except Exception:
                                    v = None
                                if v not in (None, ''):
                                    norma_reqs.append(str(v).strip())
                    except Exception:
                        pass

                    # Si no vino desde los datos o hay clasificaciones adicionales, intentar extraer CLASIF_UVA desde la tabla de relación
                    try:
                        clasif_vals = set()
                        if df_rel is not None and not df_rel.empty and 'solicitud_col' in locals():
                            for sol in (solicitudes_by_sol.keys()):
                                try:
                                    matches = df_rel[df_rel[solicitud_col].astype(str).str.strip() == str(sol).strip()]
                                    for _, r in matches.iterrows():
                                        for cc in ('CLASIF UVA', 'CLASIF_UVA', 'CLASIFUVA', 'CLASIF_UVA'):
                                            try:
                                                v = r.get(cc)
                                            except Exception:
                                                v = None
                                            if v not in (None, ''):
                                                clasif_vals.add(str(v).strip())
                                except Exception:
                                    continue
                        if clasif_vals:
                            try:
                                normas_map, normas_info = cargar_normas()
                            except Exception:
                                normas_map, normas_info = ({}, {})
                            import re
                            for cv in sorted(clasif_vals):
                                nums = re.findall(r"\d+", cv)
                                nom_text = cv
                                if nums:
                                    num = nums[0]
                                    try:
                                        if num in normas_map:
                                            nom_text = normas_map.get(num)
                                    except Exception:
                                        pass
                                norma_reqs.append(nom_text)
                    except Exception:
                        pass

                    # Normalizar y deduplicar
                    try:
                        norma_reqs = [str(n).strip() for n in norma_reqs if n and str(n).strip()]
                        # preservar orden pero eliminar duplicados
                        seen = set()
                        uniq_norms = []
                        for n in norma_reqs:
                            if n not in seen:
                                seen.add(n)
                                uniq_norms.append(n)
                        norma_reqs = uniq_norms
                    except Exception:
                        norma_reqs = []

                    any_accredited = False
                    any_no = False
                    lines.append('')
                    lines.append('✍️ Firma(s) detectada(s):')

                    # construir mapa norma_num -> códigos de firma asignados según los registros
                    # Usamos la representación NUMÉRICA principal de la norma (ej. '50') como clave
                    import re as _re
                    def _norm_to_num(s):
                        try:
                            ss = str(s or "").strip()
                            nums = _re.findall(r"\d+", ss)
                            return nums[0] if nums else ss
                        except Exception:
                            return str(s or "")

                    assigned_by_norm = {}
                    # prefill keys from norma_reqs normalized
                    if norma_reqs:
                        for nr in norma_reqs:
                            assigned_by_norm[_norm_to_num(nr)] = set()
                    try:
                        # cargar normas para mapear clasif numérica a texto si hace falta
                        try:
                            normas_map_tmp, normas_info_tmp = cargar_normas()
                        except Exception:
                            normas_map_tmp, normas_info_tmp = ({}, {})

                        # claves posibles donde aparece la norma en cada registro
                        norma_keys = ('NORMA UVA', 'NORMA', 'NORMA_UVA', 'CLASIF UVA', 'CLASIF_UVA', 'CLASIFUVA')
                        # claves posibles donde aparece el código de firma/inspector en el registro
                        firma_keys_all = ['FIRMA', 'Firma', 'firma', 'CODIGO_FIRMA', 'INSPECTOR', 'Inspector', 'inspector']

                        for it in (datos or []):
                            # determinar norma del registro (normalizar extrayendo número principal)
                            found_norm_text = None
                            found_norm_num = None
                            for nk in norma_keys:
                                try:
                                    v = it.get(nk)
                                except Exception:
                                    v = None
                                if v not in (None, ''):
                                    # si viene como número/clasif, mapear a texto usando normas_map_tmp
                                    s = str(v).strip()
                                    # intentar extraer número si es tipo '24' o 'NOM-24-...'
                                    import re
                                    nums = re.findall(r"\d+", s)
                                    if nums:
                                        num = nums[0]
                                        found_norm_num = num
                                        try:
                                            if num in normas_map_tmp:
                                                found_norm_text = normas_map_tmp.get(num)
                                            else:
                                                found_norm_text = s
                                        except Exception:
                                            found_norm_text = s
                                    else:
                                        found_norm_text = s
                                        found_norm_num = _norm_to_num(s)
                                    break

                            if not found_norm_text:
                                continue

                            # buscar códigos de firma en el registro
                            for fk in firma_keys_all:
                                try:
                                    val = it.get(fk)
                                except Exception:
                                    val = None
                                if val not in (None, ''):
                                    cand = str(val).strip()
                                    if cand:
                                        # asignar al mapa usando la clave numérica
                                        key = found_norm_num if found_norm_num is not None else _norm_to_num(found_norm_text)
                                        if key in assigned_by_norm:
                                            assigned_by_norm[key].add(cand)

                    except Exception:
                        # en caso de fallo, no interrumpir; assigned_by_norm puede quedar vacío
                        pass

                    # mapa para saber si cada norma tiene al menos una firma acreditada
                    norm_ok = {nr: False for nr in norma_reqs} if norma_reqs else {}

                    # Evaluar por norma: iterar por cada registro y mostrar por cada fila
                    # el inspector, el código de firma y la norma asociada (permitiendo repeticiones).
                    printed_codes = set()
                    # flags por norma para determinar si hubo evaluación y si alguna fue acreditada
                    norm_evaluated = {nr: False for nr in (norma_reqs or [])}
                    norm_ok = {nr: False for nr in (norma_reqs or [])} if norma_reqs else {}

                    # claves posibles donde aparece el código de firma/inspector en el registro
                    firma_keys_all = ['FIRMA', 'Firma', 'firma', 'CODIGO_FIRMA', 'INSPECTOR', 'Inspector', 'inspector']
                    import re as _re

                    # colecciones para deduplicar salidas pero mantener evaluación completa
                    seen_entries = set()
                    # cada elemento será tupla: (display_name, code, display_norm, ok)
                    output_entries = []

                    for it in (datos or []):
                        # determinar norma del registro (normalizar extrayendo número principal)
                        found_norm_text = None
                        found_norm_num = None
                        for nk in ('NORMA UVA', 'NORMA', 'NORMA_UVA', 'CLASIF UVA', 'CLASIF_UVA', 'CLASIFUVA'):
                            try:
                                v = it.get(nk)
                            except Exception:
                                v = None
                            if v not in (None, ''):
                                s = str(v).strip()
                                nums = _re.findall(r"\d+", s)
                                if nums:
                                    num = nums[0]
                                    found_norm_num = num
                                    try:
                                        if num in normas_map_tmp:
                                            found_norm_text = normas_map_tmp.get(num)
                                        else:
                                            found_norm_text = s
                                    except Exception:
                                        found_norm_text = s
                                else:
                                    found_norm_text = s
                                    found_norm_num = _norm_to_num(s)
                                break

                        if not found_norm_text:
                            continue

                        # preparar texto de requisito para validación (preferir mapa de normas si existe)
                        try:
                            req_text = normas_map_tmp.get(str(found_norm_num), str(found_norm_text)) if 'normas_map_tmp' in locals() else str(found_norm_text)
                        except Exception:
                            req_text = str(found_norm_text)

                        # extraer códigos de firma de la fila (permitir múltiples separados por , ; / | )
                        codes_in_row = []
                        for fk in firma_keys_all:
                            try:
                                val = it.get(fk)
                            except Exception:
                                val = None
                            if val not in (None, ''):
                                s = str(val)
                                parts = [p.strip() for p in _re.split(r"[,;/|]+", s) if p and p.strip()]
                                for p in parts:
                                    if p:
                                        codes_in_row.append(p)

                        if not codes_in_row:
                            # marcar norma como no evaluada si no hay códigos en esta fila
                            continue

                        # para cada código encontrado en esta fila, validar y preparar salida (no hacemos break)
                        for code in codes_in_row:
                            try:
                                nombre, img, ok = (None, None, False)
                                try:
                                    nombre, img, ok = validar_acreditacion_inspector(code, req_text, firmas_map)
                                except Exception:
                                    nombre, img, ok = (None, None, False)
                                display_name = nombre if nombre else code
                                # formar la representación de la norma a mostrar (usar found_norm_text)
                                display_norm = found_norm_text

                                # registrar resultados de evaluación para la norma
                                if ok:
                                    any_accredited = True
                                    for nr in (norma_reqs or []):
                                        if _norm_to_num(nr) == str(found_norm_num) or nr == found_norm_text:
                                            norm_ok[nr] = True
                                            norm_evaluated[nr] = True
                                            break
                                else:
                                    for nr in (norma_reqs or []):
                                        if _norm_to_num(nr) == str(found_norm_num) or nr == found_norm_text:
                                            norm_evaluated[nr] = True
                                            break

                                # siempre añadir el código a printed_codes para indicar que fue procesado
                                printed_codes.add(code)

                                # preparar la línea de salida y añadirla sólo si no está ya registrada
                                entry_key = (str(display_name).strip(), str(code).strip(), str(display_norm).strip())
                                if entry_key not in seen_entries:
                                    seen_entries.add(entry_key)
                                    output_entries.append((display_name, code, display_norm, bool(ok)))
                            except Exception:
                                continue

                    # agrupar por inspector+firma y listar normas en líneas indentadas para mejor legibilidad
                    grouped = {}
                    for name, code, norm, ok in output_entries:
                        key = (name, code)
                        grouped.setdefault(key, []).append((norm, ok))

                    for (name, code), norms in grouped.items():
                        # cabecera por inspector+firm3a
                        lines.append(f" - {name} ({code}):")
                        # listar normas atribuidas (una por línea, indentada)
                        for norm, ok in norms:
                            status = '✅ Acreditado' if ok else '❌ NO acreditado'
                            lines.append(f"    • {status}  · Norma: {norm}")

                    # después de procesar todas las filas, si alguna norma no fue evaluada o fue evaluada y ningún
                    # código resultó acreditado, marcar estado NO
                    for nr in (norma_reqs or []):
                        if not norm_evaluated.get(nr, False):
                            any_no = True
                        else:
                            if not norm_ok.get(nr, False):
                                any_no = True

                    # Determinar estado final del botón: si aparece cualquier 'NO acreditado' -> bloquear.
                    enabled = False
                    try:
                        if any_no:
                            enabled = False
                        else:
                            if norma_reqs:
                                enabled = all(bool(v) for v in norm_ok.values()) if norm_ok else False
                            else:
                                enabled = any_accredited
                    except Exception:
                        enabled = False

                    if hasattr(self, 'boton_generar_dictamen'):
                        try:
                            self.boton_generar_dictamen.configure(state='normal' if enabled else 'disabled')
                        except Exception:
                            pass
                    # Mostrar cualquier código/firma encontrada que no se haya listado aún
                    try:
                        remaining = set(found_codes) - printed_codes if found_codes else set()
                        if remaining:
                            lines.append('')
                            lines.append('✍️ Firmas encontradas (sin norma específica):')
                            for code in sorted(remaining):
                                try:
                                    # If possible, show accredited norms from the catalog instead of calling validator with empty req
                                    info = (firmas_map.get(code) or {}) if isinstance(firmas_map, dict) else {}
                                    nombre = info.get('nombre') or info.get('NOMBRE DE INSPECTOR') or None
                                    normas_ac = info.get('normas_acreditadas') or info.get('Normas acreditadas') or []
                                    display = nombre if nombre else code
                                    if normas_ac:
                                        lines.append(f" - {display} ({code}): Normas acreditadas: {', '.join(normas_ac)}")
                                    else:
                                        # If the code is not a known firma key, report it as NOT RECOGNIZED
                                        if isinstance(firmas_map, dict) and code not in firmas_map:
                                            # show the token exactly as entered and inform user it's unrecognized
                                            lines.append(f" - {code}: ❌ NO RECONOCIDA (no encontrada en Firmas.json)")
                                        else:
                                            # fallback to validator to get a simple yes/no
                                            try:
                                                _, _, ok = validar_acreditacion_inspector(code, '', firmas_map)
                                            except Exception:
                                                ok = False
                                            if ok:
                                                lines.append(f" - {display} ({code}): ✅ Acreditado")
                                            else:
                                                sugg = ''
                                                try:
                                                    if 'suggestions_by_token' in locals() and code in suggestions_by_token:
                                                        sugg = f" (sugerencia: {', '.join(suggestions_by_token.get(code) or [])})"
                                                except Exception:
                                                    sugg = ''
                                                lines.append(f" - {display} ({code}): ❌ NO acreditado{sugg}")
                                except Exception:
                                            pass
                    except Exception:
                        pass
                else:
                    lines.append('')
                    lines.append('✍️ Firma: (no detectada)')
                    if hasattr(self, 'boton_generar_dictamen'):
                        try:
                            self.boton_generar_dictamen.configure(state='disabled')
                        except Exception:
                            pass
            except Exception:
                # Si falla la validación de firmas, no bloquear la operación por defecto
                pass

            # --- Comprobar rutas de Pegado de Evidencia (si existen) ---
            try:
                pegado_paths = self._load_evidence_paths() or {}
                if pegado_paths:
                    # Construir mapa de búsqueda por solicitud
                    codigo_keys = ('CODIGO', 'Codigo', 'codigo', 'CODIGOS', 'Codigos')
                    solicitudes_imgs = {}
                    # helper: comprobar si hay archivos que contengan el código
                    import fnmatch
                    def _simple_normalize_key(s):
                        try:
                            return "".join(ch for ch in str(s or "") if ch.isalnum()).upper()
                        except Exception:
                            return str(s or "").upper()

                    def _search_destino_in_base(ruta_base, destino):
                        """Intento liviano de buscar un 'destino' (archivo o carpeta) dentro de la ruta base.
                        No falla si hay permisos restringidos; devuelve True si encuentra al menos una coincidencia."""
                        try:
                            # Comparar por nombre exacto (archivo o carpeta)
                            items = os.listdir(ruta_base)
                        except Exception:
                            items = []

                        base, ext = os.path.splitext(destino)
                        # Si destino apunta a un archivo (ext conocida), buscar archivo exacto
                        if ext:
                            for it in items:
                                if it.lower() == destino.lower():
                                    return True

                        # Buscar carpetas con nombre exacto
                        for it in items:
                            p = os.path.join(ruta_base, it)
                            if os.path.isdir(p) and it.strip().lower() == destino.strip().lower():
                                # carpeta encontrada
                                # comprobar si tiene imágenes dentro
                                try:
                                    for f in os.listdir(p):
                                        if os.path.splitext(f)[1].lower() in ('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.tif'):
                                            return True
                                except Exception:
                                    return True

                        # Buscar recursivamente archivos que empiecen por la base del destino o contengan el nombre
                        try:
                            for root, dirs, files in os.walk(ruta_base):
                                for f in files:
                                    if base and base.lower() in f.lower():
                                        if os.path.splitext(f)[1].lower() in ('.png', '.jpg', '.jpeg', '.webp', '.bmp', '.tif'):
                                            return True
                                for d in dirs:
                                    if d.lower() == destino.strip().lower():
                                        # carpeta exacta
                                        return True
                        except Exception:
                            pass
                        return False

                    def has_images_info(code):
                        """Devuelve (found:bool, reason:str).
                            reason: 'indice' si la evidencia se encontró usando el índice externo,
                                'ruta' si la evidencia se encontró explorando las rutas cargadas o
                                si el índice no aplica (modo simple/carpetas) y no se encontró.
                            En caso de found==False, reason indica por qué no se encontró:
                                - 'ruta' significa no encontrada en las rutas cargadas
                                - 'indice' significa que no hay entrada en el índice (solo relevante si modo=indice)
                        """
                        code_low = str(code).strip()
                        if not code_low:
                            return False, 'indice'

                        # 1) Búsqueda clásica en rutas: carpeta o archivo que contenga el código
                        for grp, base in pegado_paths.items():
                            bases = base if isinstance(base, (list, tuple)) else [base]
                            for b in bases:
                                try:
                                    for root, dirs, files in os.walk(b):
                                        if os.path.basename(root).lower() == code_low.lower():
                                            return True, 'ruta'
                                        for fname in files:
                                            if code_low.lower() in fname.lower():
                                                return True, 'ruta'
                                except Exception:
                                    continue

                        # 2) Intentar con índice (index_indice.json)
                        # Respetar la preferencia de modo de pegado guardada por la UI
                        try:
                            evidencia_cfg_path = os.path.join(DATA_DIR, 'evidence_paths.json')
                            modo_cfg = ''
                            if os.path.exists(evidencia_cfg_path):
                                try:
                                    with open(evidencia_cfg_path, 'r', encoding='utf-8') as _ef:
                                        _cfg = json.load(_ef) or {}
                                        modo_cfg = str(_cfg.get('modo_pegado', '')).strip().lower()
                                except Exception:
                                    modo_cfg = ''
                            use_index_pref = modo_cfg in ('indice', 'pegado indice', 'pegado_indice')
                        except Exception:
                            use_index_pref = False
                        try:
                            appdata = os.getenv('APPDATA') or ''
                            idx_path = os.path.join(appdata, 'ImagenesVC', 'index_indice.json')
                            if os.path.exists(idx_path):
                                try:
                                    with open(idx_path, 'r', encoding='utf-8') as f:
                                        idx = json.load(f) or {}
                                except Exception:
                                    idx = {}

                                cand_keys = {k: v for k, v in idx.items()}
                                norm_key = _simple_normalize_key(code_low)
                                possibles = []
                                if code_low in cand_keys:
                                    possibles.append(cand_keys[code_low])
                                if norm_key in cand_keys:
                                    possibles.append(cand_keys[norm_key])
                                try:
                                    raw = str(code_low)
                                    if raw.endswith('.0'):
                                        alt = raw[:-2]
                                        if alt in cand_keys:
                                            possibles.append(cand_keys[alt])
                                except Exception:
                                    pass

                                if possibles:
                                    # si el índice tiene destino(s), comprobar si alguno existe en las rutas persistidas
                                    found_any = False
                                    for destino in possibles:
                                        for grp, base in pegado_paths.items():
                                            bases = base if isinstance(base, (list, tuple)) else [base]
                                            for b in bases:
                                                try:
                                                    if _search_destino_in_base(b, destino):
                                                        return True, 'indice'
                                                except Exception:
                                                    continue
                                    # índice tenía entrada(s) pero no se encontró en rutas
                                    return False, 'ruta'
                                else:
                                    # índice no tiene la clave -> si el usuario NO usa modo 'indice'
                                    # esto debe reportarse como 'ruta' (no encontrada en las rutas cargadas),
                                    # no como 'falta en índice'. Solo reportar 'indice' cuando la
                                    # preferencia indica uso del índice.
                                    if use_index_pref:
                                        return False, 'indice'
                                    else:
                                        return False, 'ruta'
                            else:
                                # no existe índice -> considerar como no encontrado en índice
                                return False, 'indice'
                        except Exception:
                            return False, 'indice'

                    for sol, entries in solicitudes_by_sol.items():
                        docs_with = []
                        docs_without = []
                        for idx, item, lista in entries:
                            # intentar extraer todos los códigos posibles para este registro
                            # y preferir la columna ASIG cuando exista (LEDERY / BLUE STRIPES case)
                            codes = []
                            asig_val = None
                            for ak in ('ASIG', 'Asig', 'asig'):
                                try:
                                    if ak in item and item.get(ak) not in (None, ''):
                                        asig_val = str(item.get(ak)).strip()
                                        break
                                except Exception:
                                    continue

                            # Si hay ASIG, verificar si existe evidencia bajo esa carpeta/destino
                            found_by_asig = False
                            try:
                                if asig_val:
                                    for grp, base in (pegado_paths or {}).items():
                                        bases = base if isinstance(base, (list, tuple)) else [base]
                                        for b in bases:
                                            try:
                                                if _search_destino_in_base(b, asig_val):
                                                    found_by_asig = True
                                                    break
                                            except Exception:
                                                continue
                                        if found_by_asig:
                                            break
                            except Exception:
                                found_by_asig = False

                            if found_by_asig:
                                # Reportaremos que se pegarán evidencias desde ASIG
                                codes = [asig_val]
                            else:
                                # No se encontró por ASIG; volver a extraer códigos desde columnas CODIGO
                                for k in codigo_keys:
                                    if k in item and item.get(k) not in (None, ''):
                                        raw = str(item.get(k))
                                        for part in raw.split(','):
                                            p = part.strip()
                                            if p:
                                                codes.append(p)

                            # si no hay códigos, marcar como sin evidencia (sin códigos detectados)
                            if not codes:
                                docs_without.append((lista or f"registro_{idx}", []))
                                continue

                            # Para cada código, comprobar si tiene imágenes; recoger los códigos que sí/no
                            matched_codes = []
                            missing_codes = []
                            for c in codes:
                                try:
                                    found, reason = has_images_info(c)
                                    if found:
                                        matched_codes.append(c)
                                    else:
                                        if reason == 'indice':
                                            missing_codes.append(f"{c} (falta en índice)")
                                        else:
                                            missing_codes.append(f"{c} (falta en ruta)")
                                except Exception:
                                    missing_codes.append(f"{c} (falla comprobación)")

                            # Construir un identificador descriptivo tipo 'Dictamen_Lista_<clasif>_<estilo>_<sol>_<lista>'
                            try:
                                clasif = str(item.get('CLASIF UVA') or item.get('CLASIF_UVA') or item.get('NORMA UVA') or item.get('NORMA_UVA') or '')
                            except Exception:
                                clasif = ''
                            try:
                                estilo = str(item.get('ESTILO') or '')
                            except Exception:
                                estilo = ''
                            sol_clean = str(sol).replace('/', '_')
                            lista_part = str(lista) if lista is not None else f"registro_{idx}"
                            dictamen_name = f"Dictamen_Lista_{clasif}_{estilo}_{sol_clean}_{lista_part}"

                            # Si encontramos al menos una coincidencia, consideramos el documento como 'con imágenes'
                            if matched_codes:
                                docs_with.append((dictamen_name, matched_codes, missing_codes, codes))
                            else:
                                docs_without.append((dictamen_name, missing_codes, codes))

                        solicitudes_imgs[sol] = (docs_with, docs_without)

                    # Añadir resumen de pegado al reporte (se mostrará más abajo en un único diálogo)
                    lines.append("")
                    lines.append('📸 Estado de evidencias - imagenes encontradas')
                    # Mostrar las rutas donde se buscó evidencia (acortadas) para ayudar al usuario
                    try:
                        rutas = []
                        for grp, base in (pegado_paths or {}).items():
                            if isinstance(base, (list, tuple)):
                                rutas.extend(base)
                            else:
                                rutas.append(base)
                        if rutas:
                            short = []
                            for r in rutas[:3]:
                                rp = str(r)
                                if len(rp) > 80:
                                    rp = '...' + rp[-77:]
                                short.append(rp)
                            lines.append(f"Rutas buscadas: {', '.join(short)}{'...' if len(rutas)>3 else ''}")
                    except Exception:
                        pass

                    for sol, (with_list, without_list) in solicitudes_imgs.items():
                        # Resumen claro por solicitud en formato por Dictamen
                        lines.append("")
                        lines.append(f"Solicitud: {sol}")

                        # Procesar documentos con imágenes (se pegarán)
                        for entry in with_list:
                            try:
                                dictamen_name, matched_codes, missing_codes, all_codes = entry
                            except Exception:
                                # compatibilidad si estructura antigua
                                try:
                                    dictamen_name, matched_codes, missing_codes = entry
                                    all_codes = matched_codes + missing_codes
                                except Exception:
                                    continue

                            lines.append(f"{dictamen_name}")
                            # todos los códigos
                            try:
                                lines.append(f"  codigos {' '.join(str(x) for x in all_codes)}")
                            except Exception:
                                lines.append(f"  codigos {all_codes}")

                            # códigos que se pegarán
                            if matched_codes:
                                lines.append(f"  se pegaran en {' '.join(str(x) for x in matched_codes)}")
                            else:
                                lines.append(f"  se pegaran en (ninguno)")

                            # códigos que NO se pegarán con motivo
                            if missing_codes:
                                lines.append(f"  no se pegara {' '.join(str(x) for x in missing_codes)}")

                        # Procesar documentos SIN imágenes (todos serán no-pegados)
                        for entry in without_list:
                            try:
                                dictamen_name, missing_codes, all_codes = entry
                                matched_codes = []
                            except Exception:
                                try:
                                    dictamen_name, missing_codes = entry
                                    all_codes = missing_codes
                                except Exception:
                                    continue

                            lines.append(f"{dictamen_name}")
                            try:
                                lines.append(f"  codigos {' '.join(str(x) for x in all_codes)}")
                            except Exception:
                                lines.append(f"  codigos {all_codes}")

                            lines.append(f"  se pegaran en (ninguno)")
                            if missing_codes:
                                lines.append(f"  no se pegara {' '.join(str(x) for x in missing_codes)}")

                    # Sugerencias útiles para el usuario
                    lines.append("")
                    lines.append("Sugerencias:")
                    lines.append("  - Si faltan imágenes, verifique que las imágenes estén en alguna de las rutas listadas arriba.")
                    lines.append("  - Asegúrese de que los nombres de archivo o el nombre de la carpeta contengan el código/CODIGO indicado en la tabla de relación.")
                    lines.append("  - Puede configurar o limpiar rutas en 'Pegado de evidencia' -> 'Seleccionar carpeta' antes de generar los dictámenes.")
            except Exception:
                # No bloquear si falla la comprobación de rutas
                pass

            # Mostrar el informe (unificado: firmas + resumen de pegado si existe)
            try:
                # Crear modal scrollable para asegurar visibilidad aunque haya muchas solicitudes
                modal = ctk.CTkToplevel(self)
                modal.title("Reporte de Visita")
                modal.transient(self)
                modal.grab_set()

                sw = self.winfo_screenwidth()
                sh = self.winfo_screenheight()
                mw = min(int(sw * 0.8), 500)
                mh = min(int(sh * 0.8), 600)
                mx = (sw - mw) // 2
                my = (sh - mh) // 2
                try:
                    modal.geometry(f"{mw}x{mh}+{mx}+{my}")
                except Exception:
                    pass

                # Header visual acorde al STYLE
                header = ctk.CTkFrame(modal, fg_color=STYLE.get('fondo'))
                header.pack(fill='x')
                ctk.CTkLabel(header, text="📊 VISTA PREVIA DE LA VISITA", font=FONT_SUBTITLE, text_color=STYLE.get('texto_oscuro')).pack(side='left', padx=12, pady=8)
                

                container = ctk.CTkFrame(modal, fg_color=STYLE.get('surface'))
                container.pack(fill='both', expand=True, padx=8, pady=8)

                # Usar tk.Text con scrollbar para compatibilidad y rendimiento
                text_frame = tk.Frame(container, bg=STYLE.get('surface'))
                text_frame.pack(fill='both', expand=True)

                scrollbar = tk.Scrollbar(text_frame)
                scrollbar.pack(side='right', fill='y')

                # Configurar apariencia del Text para coincidir con el estilo del sistema
                text_widget = tk.Text(
                    text_frame,
                    wrap='word',
                    yscrollcommand=scrollbar.set,
                    bg=STYLE.get('surface'),
                    fg=STYLE.get('texto_oscuro'),
                    bd=0,
                    highlightthickness=0,
                    padx=6,
                    pady=6,
                    font=(FONT_LABEL[0], 12)
                )
                text_widget.pack(side='left', fill='both', expand=True)
                scrollbar.config(command=text_widget.yview)

                try:
                    text_widget.insert('1.0', "\n".join(lines))
                    text_widget.configure(state='disabled')
                except Exception:
                    try:
                        text_widget.insert('1.0', str(lines))
                        text_widget.configure(state='disabled')
                    except Exception:
                        pass

                # Botones de acción
                btns = ctk.CTkFrame(modal, fg_color='transparent')
                btns.pack(fill='x', padx=8, pady=(0,8))

                def _copy():
                    try:
                        self.clipboard_clear()
                        self.clipboard_append("\n".join(lines))
                        messagebox.showinfo("Copiado", "El texto del reporte fue copiado al portapapeles.")
                    except Exception:
                        pass

                ctk.CTkButton(btns, text="Copiar texto", command=_copy, width=140, fg_color=STYLE.get('primario'), text_color=STYLE.get('secundario')).pack(side='left', padx=6)
                ctk.CTkButton(btns, text="Aceptar", command=modal.destroy, width=140, fg_color=STYLE.get('surface'), text_color=STYLE.get('texto_oscuro'), border_color=STYLE.get('borde')).pack(side='right', padx=6)

                # Mantener modal en primer plano
                modal.lift()
                modal.focus_force()
            except Exception:
                # Fallback mínimo si falla el modal
                messagebox.showinfo("Reporte de Visita", "\n".join(lines))
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo verificar la visita:\n{e}")

    # ------------------ EVIDENCIAS: carga y persistencia ------------------
    def _load_evidence_paths(self):
        """Carga el archivo `data/evidence_paths.json` si existe."""
        try:
            data_file = os.path.join(APP_DIR, "data", "evidence_paths.json")
            if os.path.exists(data_file):
                with open(data_file, "r", encoding="utf-8") as f:
                    return json.load(f)
            return {}
        except Exception:
            return {}

    def _save_evidence_path(self, group, path, mode=None):
        """Guarda la ruta `path` bajo la clave `group` en `data/evidence_paths.json`.

        Si se especifica `mode`, lo guarda también bajo la clave especial
        `modo_pegado` (valor normalizado: 'simple', 'carpetas' o 'indice') para
        que el generador pueda respetar la preferencia de pegado.
        """
        data_file = os.path.join(APP_DIR, "data", "evidence_paths.json")
        os.makedirs(os.path.dirname(data_file), exist_ok=True)
        data = self._load_evidence_paths() or {}
        existing = data.get(group, [])
        if path not in existing:
            existing.append(path)
        data[group] = existing
        # Normalizar y guardar el modo si se proporcionó
        try:
            if mode:
                m = str(mode).strip().lower()
                if m in ("pegado simple", "pegado_simple", "simple"):
                    norm = "simple"
                elif m in ("pegado carpetas", "pegado_carpetas", "carpetas", "carpeta"):
                    norm = "carpetas"
                elif m in ("pegado indice", "pegado_indice", "indice", "índice"):
                    norm = "indice"
                else:
                    norm = m
                data['modo_pegado'] = norm
        except Exception:
            pass

        with open(data_file, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def configurar_carpeta_evidencias(self):
        """Abre un modal para elegir a qué grupo se guardará la ruta y seleccionar carpeta."""
        modal = ctk.CTkToplevel(self)
        modal.title("Configurar Carpeta de Evidencias")
        modal.geometry("520x260")
        modal.transient(self)
        modal.grab_set()

        ctk.CTkLabel(modal, text="Seleccione el grupo y luego elija la carpeta de evidencias:",
                     font=FONT_SMALL, text_color=STYLE["texto_oscuro"]).pack(anchor="w", padx=16, pady=(12, 8))

        var_grupo = ctk.StringVar(value="grupo_axo")
        opciones = [
            ("Grupo Axo (varios clientes)", "grupo_axo"),
            ("Bosch", "bosch"),
            ("Unilever", "unilever"),
        ]

        for texto, valor in opciones:
            ctk.CTkRadioButton(modal, text=texto, variable=var_grupo, value=valor).pack(anchor="w", padx=20, pady=6)

        # Mostrar rutas actualmente guardadas
        rutas_frame = ctk.CTkFrame(modal, fg_color="transparent")
        rutas_frame.pack(fill="both", expand=True, padx=12, pady=(6, 0))

        lbl_actual = ctk.CTkLabel(rutas_frame, text="Rutas guardadas:", font=FONT_SMALL, text_color=STYLE["texto_oscuro"]) 
        lbl_actual.pack(anchor="w")

        lista_rutas = ctk.CTkLabel(rutas_frame, text="(ninguna)", font=("Inter", 10), text_color=STYLE["texto_claro"], wraplength=480)
        lista_rutas.pack(anchor="w", pady=(4, 0))

        def _refrescar_rutas():
            data = self._load_evidence_paths() or {}
            lines = []
            for g, lst in data.items():
                lines.append(f"{g}: \n  " + "\n  ".join(lst))
            lista_rutas.configure(text="\n\n".join(lines) if lines else "(ninguna)")

        _refrescar_rutas()

        def elegir_carpeta():
            grp = var_grupo.get()
            carpeta = filedialog.askdirectory(title="Seleccionar carpeta de evidencias")
            if not carpeta:
                return
            try:
                self._save_evidence_path(grp, carpeta)
                messagebox.showinfo("Guardado", f"Ruta guardada para '{grp}':\n{carpeta}")
                _refrescar_rutas()
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo guardar la ruta:\n{e}")

        btn_frame = ctk.CTkFrame(modal, fg_color="transparent")
        btn_frame.pack(fill="x", pady=10)

        ctk.CTkButton(btn_frame, text="Seleccionar carpeta y guardar", command=elegir_carpeta,
                      fg_color=STYLE["primario"], text_color=STYLE["secundario"], height=36).pack(side="left", padx=12)
        ctk.CTkButton(btn_frame, text="Cerrar", command=modal.destroy, height=36).pack(side="right", padx=12)

    # ------------------ PEGADO EVIDENCIAS (botones UI) ------------------
    def _run_script_and_notify(self, fn):
        import traceback
        log_path = os.path.join(DATA_DIR, 'pegado.log')
        os.makedirs(os.path.dirname(log_path), exist_ok=True)
        try:
            with open(log_path, 'a', encoding='utf-8') as lg:
                lg.write(f"\n===== Pegado start: {datetime.now().isoformat()} =====\n")
            fn()
            try:
                with open(log_path, 'a', encoding='utf-8') as lg:
                    lg.write(f"Pegado finished: {datetime.now().isoformat()}\n")
            except Exception:
                pass
            try:
                messagebox.showinfo("Pegado", "Proceso de pegado finalizado. Revise el registro de fallos si corresponde.")
            except Exception:
                pass
        except Exception as e:
            tb = traceback.format_exc()
            try:
                with open(log_path, 'a', encoding='utf-8') as lg:
                    lg.write(f"ERROR: {e}\n{tb}\n")
            except Exception:
                pass
            try:
                messagebox.showerror("Error pegado", f"Error al ejecutar el proceso de pegado:\n{e}\nRevise {log_path} para más detalles.")
            except Exception:
                pass

    def _call_pegado_script(self, script_filename, func_name, ruta_docs=None, ruta_imgs=None):
        """Carga dinámicamente los módulos de Pegado sin persistir rutas del usuario.
        Si se proveen `ruta_docs` y `ruta_imgs`, inyecta una implementación de
        `obtener_rutas()` que devuelve esas rutas y convierte `guardar_config` en no-op.
        """
        try:
            base = os.path.join(APP_DIR, "Pegado de Evidenvia Fotografica")
            main_path = os.path.join(base, "main.py")

            added_to_path = False
            if base not in sys.path:
                sys.path.insert(0, base)
                added_to_path = True

            # Cargar module 'main' desde archivo y parchear
            spec_main = importlib.util.spec_from_file_location("main_pegado", main_path)
            mod_main = importlib.util.module_from_spec(spec_main)
            spec_main.loader.exec_module(mod_main)

            # Evitar persistir config
            try:
                mod_main.guardar_config = lambda x: None
            except Exception:
                pass

            if ruta_docs and ruta_imgs:
                mod_main.obtener_rutas = lambda: (ruta_docs, ruta_imgs)

            # Hacer disponible como 'main' para que los scripts que hacen `from main import ...` funcionen
            old_main = sys.modules.get('main')
            sys.modules['main'] = mod_main

            # Cargar y ejecutar el script solicitado
            script_path = os.path.join(base, script_filename)
            spec = importlib.util.spec_from_file_location("pegado_script", script_path)
            mod_script = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod_script)

            fn = getattr(mod_script, func_name, None)
            if not callable(fn):
                messagebox.showerror("Error", f"Función {func_name} no encontrada en {script_filename}")
                return

            # Ejecutar en hilo para no bloquear la UI
            threading.Thread(target=lambda: self._run_script_and_notify(fn), daemon=True).start()

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo ejecutar el pegado:\n{e}")
        finally:
            # Restaurar 'main' anterior si existía
            try:
                if old_main is not None:
                    sys.modules['main'] = old_main
                else:
                    sys.modules.pop('main', None)
                # Quitar la ruta añadida al sys.path
                try:
                    if added_to_path and base in sys.path:
                        sys.path.remove(base)
                except Exception:
                    pass
            except Exception:
                pass

    def handle_pegado_simple(self):
        # Guardar únicamente la ruta de imágenes para usarla posteriormente
        ruta_imgs = filedialog.askdirectory(title="Seleccionar carpeta de imágenes para evidencias (se guardará)")
        if not ruta_imgs:
            return
        try:
            # Guardar la ruta bajo un grupo genérico 'manual_pegado'
            self._save_evidence_path('manual_pegado', ruta_imgs, mode='simple')
            messagebox.showinfo("Pegado guardado", "Ruta de imágenes guardada. Cuando genere los dictámenes, se buscarán evidencias en esta carpeta.")
            try:
                self._update_pegado_status(mode="Pegado Simple", path=ruta_imgs)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar la ruta de evidencias:\n{e}")

    def handle_pegado_carpetas(self):
        ruta_imgs = filedialog.askdirectory(title="Seleccionar carpeta raíz de carpetas por código (se guardará)")
        if not ruta_imgs:
            return
        try:
            self._save_evidence_path('manual_pegado', ruta_imgs, mode='carpetas')
            messagebox.showinfo("Pegado guardado", "Ruta de carpetas guardada. Cuando genere los dictámenes, se buscarán evidencias en estas carpetas.")
            try:
                self._update_pegado_status(mode="Pegado Carpetas", path=ruta_imgs)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar la ruta de evidencias:\n{e}")

    def handle_pegado_indice(self):
        # Seleccionar carpeta donde están las imágenes (base para buscar carpetas/archivos)
        ruta_imgs = filedialog.askdirectory(title="Seleccionar carpeta de imágenes para índice (se guardará)")
        if not ruta_imgs:
            return

        # Preguntar por el archivo Excel que contiene la hoja 'CONCENTRADO'
        excel_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel con hoja CONCENTRADO",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls *.xlsb")]
        )

        try:
            # Guardar la ruta de evidencias para uso posterior
            self._save_evidence_path('manual_pegado', ruta_imgs, mode='indice')

            # Si el usuario proporcionó un Excel, construir el índice usando el script pegado_indice.py
            if excel_path:
                import importlib.util
                import sys
                mod_path = os.path.join(APP_DIR, "Pegado de Evidenvia Fotografica", "pegado_indice.py")
                base = os.path.dirname(mod_path)
                # Añadir temporalmente el directorio al sys.path para resolver imports locales (registro_fallos, main, etc.)
                added_to_path = False
                if base and base not in sys.path:
                    sys.path.insert(0, base)
                    added_to_path = True

                spec = importlib.util.spec_from_file_location("pegado_indice", mod_path)
                pegado_mod = importlib.util.module_from_spec(spec)

                try:
                    spec.loader.exec_module(pegado_mod)
                finally:
                    # Quitar la ruta añadida
                    try:
                        if added_to_path and base in sys.path:
                            sys.path.remove(base)
                    except Exception:
                        pass

                # Llamar a la función para construir el índice desde el Excel
                try:
                    indice = pegado_mod.construir_indice_desde_excel(excel_path)
                    messagebox.showinfo("Índice creado", f"Índice construido con {len(indice)} entradas. Ruta de imágenes guardada.")
                except Exception as e:
                    msg = str(e)
                    # Detectar error por falta de pyxlsb (xlsxb engine)
                    if "pyxlsb" in msg or "Missing optional dependency 'pyxlsb'" in msg or 'xlsb' in (excel_path or '').lower():
                        instalar = messagebox.askyesno(
                            "Falta dependencia opcional",
                            "El archivo seleccionado es de tipo .xlsb y falta la dependencia opcional 'pyxlsb'.\n¿Desea que el programa intente instalar 'pyxlsb' ahora?\n(Se usará pip en el mismo intérprete de Python que ejecuta la aplicación)."
                        )
                        if instalar:
                            try:
                                import subprocess, sys
                                proc = subprocess.run([sys.executable, "-m", "pip", "install", "pyxlsb"], capture_output=True, text=True)
                                if proc.returncode == 0:
                                    # Reintentar construir índice
                                    try:
                                        indice = pegado_mod.construir_indice_desde_excel(excel_path)
                                        messagebox.showinfo("Índice creado", f"Índice construido con {len(indice)} entradas. Ruta de imágenes guardada.")
                                    except Exception as e2:
                                        messagebox.showwarning("Índice parcial", f"Se intentó instalar 'pyxlsb' pero la construcción del índice falló:\n{e2}")
                                else:
                                    messagebox.showerror("Instalación fallida", f"No se pudo instalar 'pyxlsb'. Salida de pip:\n{proc.stdout}\n{proc.stderr}")
                            except Exception as ie:
                                messagebox.showerror("Error", f"Error al intentar instalar 'pyxlsb':\n{ie}")
                        else:
                            messagebox.showwarning("Índice parcial", f"Se guardó la ruta de imágenes, pero falló la construcción del índice desde el Excel:\n{e}")
                    else:
                        messagebox.showwarning("Índice parcial", f"Se guardó la ruta de imágenes, pero falló la construcción del índice desde el Excel:\n{e}")

                # Actualizar estado de pegado (después de intentar construir índice)
                try:
                    self._update_pegado_status(mode="Pegado Índice", path=ruta_imgs)
                except Exception:
                    pass
            else:
                messagebox.showinfo("Pegado guardado", "Ruta de imágenes guardada. Al generar dictámenes y subir el Excel de índice, se usarán estas imágenes.")

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar la ruta de evidencias:\n{e}")

    def handle_clear_evidence_paths(self):
        """Limpia las rutas de evidencias persistidas en `data/evidence_paths.json`.
        Pide confirmación al usuario antes de eliminar/limpiar.
        """
        try:
            data_file = os.path.join(APP_DIR, "data", "evidence_paths.json")
            if not os.path.exists(data_file):
                messagebox.showinfo("Limpiar rutas", "No hay rutas guardadas para limpiar.")
                return

            ok = messagebox.askyesno("Confirmar limpieza", "¿Desea eliminar todas las rutas de evidencias guardadas? Esta acción no se puede deshacer.")
            if not ok:
                return

            # Intentar eliminar el archivo; si falla, reescribir vacío
            try:
                os.remove(data_file)
            except Exception:
                try:
                    with open(data_file, "w", encoding="utf-8") as f:
                        json.dump({}, f, indent=2, ensure_ascii=False)
                except Exception as e:
                    messagebox.showerror("Error", f"No se pudo limpiar las rutas de evidencias:\n{e}")
                    return

            messagebox.showinfo("Rutas limpiadas", "Se han eliminado las rutas de evidencias guardadas.")
            try:
                self._update_pegado_status(mode=None, path=None)
            except Exception:
                pass
        except Exception as e:
            messagebox.showerror("Error", f"Error al limpiar rutas de evidencias:\n{e}")

    def _update_pegado_status(self, mode=None, path=None):
        """Actualiza la etiqueta de estado del pegado con la ruta y el modo seleccionados."""
        try:
            if not hasattr(self, 'pegado_status_label'):
                return
            display_mode = mode if mode else "Ninguno"
            if path:
                p = str(path)
                # acortar si es muy largo
                if len(p) > 80:
                    p = '...' + p[-77:]
                display_path = p
                color = STYLE["exito"]
                loaded = True
            else:
                display_path = "Ninguna"
                color = STYLE["advertencia"]
                loaded = False

            text = f"Ruta: {display_path}   ·   Modo: {display_mode}"
            self.pegado_status_label.configure(text=text, text_color=color)
            try:
                self.pegado_path_loaded_var.set(bool(loaded))
            except Exception:
                pass
        except Exception:
            pass
# Método auxiliar: auto-ajustar columnas del Treeview
    def _auto_resize_tree_columns(self, tree):
        """Ajusta el ancho de las columnas del Treeview según el contenido y encabezados."""
        try:
            font_obj = tkfont.Font()
        except Exception:
            try:
                font_obj = tkfont.Font(family="Inter", size=10)
            except Exception:
                font_obj = None

        for col in tree["columns"]:
            try:
                # empezar con el ancho del encabezado
                max_w = font_obj.measure(col) + 12 if font_obj else 100
            except Exception:
                max_w = 100
            for iid in tree.get_children():
                try:
                    val = tree.set(iid, col)
                    w = font_obj.measure(str(val)) + 12 if font_obj else 100
                    if w > max_w:
                        max_w = w
                except Exception:
                    continue
            # Establecer un mínimo razonable
            if max_w < 80:
                max_w = 80
            try:
                tree.column(col, width=max_w)
            except Exception:
                pass

    def _adjust_clientes_columns(self):
        """Ajusta anchos de columnas en `self.tree_clientes` para asegurar que
        la columna 'ACCIONES' permanezca visible dentro del ancho del contenedor.
        Se distribuye el espacio disponible entre las otras columnas con mínimos.
        """
        try:
            tree = getattr(self, 'tree_clientes', None)
            cont = getattr(self, 'tree_clientes_container', None)
            if not tree or not cont:
                return
            total_w = cont.winfo_width() or tree.winfo_width()
            if not total_w or total_w < 100:
                return

            padding = 8
            acc_min = 120
            cols = list(tree['columns'])
            if 'ACCIONES' not in cols:
                return
            other_cols = [c for c in cols if c != 'ACCIONES']

            # mínimos sugeridos por columna
            desired_mins = {
                'RFC': 90,
                'CLIENTE': 200,
                'NÚMERO DE CONTRATO': 140,
                'ACTIVIDAD': 100,
                'SERVICIO': 100
            }

            # espacio disponible para las columnas distintas de ACCIONES
            available = max(50, total_w - acc_min - padding)

            # suma de mínimos (solo para las columnas presentes)
            sum_mins = sum(desired_mins.get(c, 80) for c in other_cols)

            new_widths = {}
            if sum_mins <= available:
                # Asignar al menos el mínimo y expandir CLIENTE si sobra
                extra = available - sum_mins
                for c in other_cols:
                    base = desired_mins.get(c, 80)
                    add = 0
                    if c == 'CLIENTE' and extra > 0:
                        add = extra
                    new_widths[c] = max(60, int(base + add))
            else:
                # No hay espacio suficiente: escalar proporcionalmente pero respetar un mínimo duro
                min_hard = 60
                total_weight = sum(desired_mins.get(c, 80) for c in other_cols)
                for c in other_cols:
                    weight = desired_mins.get(c, 80)
                    w = int(max(min_hard, available * (weight / total_weight)))
                    new_widths[c] = w

            # Aplicar anchos calculados
            try:
                for c, w in new_widths.items():
                    try:
                        tree.column(c, width=w)
                    except Exception:
                        pass
                # ACCIONES fijo al mínimo
                try:
                    tree.column('ACCIONES', width=acc_min)
                except Exception:
                    pass
            except Exception:
                pass
        except Exception:
            pass

# ================== EJECUCIÓN ================== #
if __name__ == "__main__":
    # En Windows, habilitar DPI awareness antes de crear la ventana
    # se respeten cuando la app está empaquetada como .exe.
    if sys.platform.startswith("win"):
        try:
            import ctypes
            # Windows 7+ API
            ctypes.windll.user32.SetProcessDPIAware()
        except Exception:
            try:
                # Windows 8.1+ alternative
                ctypes.windll.shcore.SetProcessDpiAwareness(1)
            except Exception:
                pass

    app = SistemaDictamenesVC()
    try:
        # Forzar escala de Tk a 1.0 para evitar que el gestor de ventanas
        # aplique escalado automático que cambie el tamaño real en píxeles.
        app.tk.call('tk', 'scaling', 1.0)
    except Exception:
        pass
    app.mainloop()

